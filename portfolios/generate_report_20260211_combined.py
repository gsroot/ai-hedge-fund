#!/usr/bin/env python3
"""S&P 500 Combined Portfolio Report - 2026-02-11 (Investor Team + Analyst Team)"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
BULLISH_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BULLISH_FONT = Font(color="375623")
BEARISH_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
BEARISH_FONT = Font(color="C62828")
NEUTRAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
NEUTRAL_FONT = Font(color="616161")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
)

portfolio = [
    {"ticker": "DECK", "name": "Deckers Outdoor", "weight": 15.0, "signal": "강력매수", "confidence": 71, "return": 12.4, "cap": "$17B", "pe": 16.3, "roe": 40, "consensus": "6/7", "sector": "Consumer Disc."},
    {"ticker": "CF", "name": "CF Industries", "weight": 15.0, "signal": "매수", "confidence": 71, "return": 17.0, "cap": "$16B", "pe": 11.5, "roe": 22, "consensus": "4/7", "sector": "Materials"},
    {"ticker": "MU", "name": "Micron Technology", "weight": 12.0, "signal": "매수", "confidence": 77, "return": 15.8, "cap": "$420B", "pe": 36.5, "roe": 23, "consensus": "4/7", "sector": "Technology"},
    {"ticker": "PAYC", "name": "Paycom Software", "weight": 11.5, "signal": "강력매수", "confidence": 73, "return": 13.0, "cap": "$7.0B", "pe": 16.1, "roe": 29, "consensus": "5/7", "sector": "Technology"},
    {"ticker": "PYPL", "name": "PayPal Holdings", "weight": 11.5, "signal": "매수", "confidence": 73, "return": 11.8, "cap": "$39B", "pe": 7.6, "roe": 26, "consensus": "4/7", "sector": "Technology"},
    {"ticker": "ACGL", "name": "Arch Capital Group", "weight": 9.5, "signal": "매수", "confidence": 82, "return": 12.7, "cap": "$35B", "pe": 8.3, "roe": 20, "consensus": "4/7", "sector": "Financial"},
    {"ticker": "HIG", "name": "Hartford Financial", "weight": 9.0, "signal": "매수", "confidence": 77, "return": 12.3, "cap": "$40B", "pe": 10.5, "roe": 22, "consensus": "4/7", "sector": "Financial"},
    {"ticker": "TRV", "name": "Travelers Companies", "weight": 8.5, "signal": "강력매수", "confidence": 74, "return": 13.4, "cap": "$67B", "pe": 10.7, "roe": 21, "consensus": "5/7", "sector": "Financial"},
    {"ticker": "FITB", "name": "Fifth Third Bancorp", "weight": 8.0, "signal": "매수", "confidence": 64, "return": 13.1, "cap": "$49B", "pe": 15.4, "roe": 12, "consensus": "4/7", "sector": "Financial"},
]

rankings = [
    {"rank": 1, "ticker": "MOS", "name": "Mosaic Company", "total": 7.87, "fund": 7.8, "mom": 8.0, "ens": 7.1, "signal": "매수", "return": 17.1, "cap": "$9.6B", "pe": 7.7, "roe": 10, "included": False},
    {"rank": 2, "ticker": "CF", "name": "CF Industries", "total": 7.86, "fund": 8.0, "mom": 7.5, "ens": 7.8, "signal": "매수", "return": 17.0, "cap": "$16B", "pe": 11.5, "roe": 22, "included": True},
    {"rank": 3, "ticker": "MU", "name": "Micron Technology", "total": 7.51, "fund": 7.1, "mom": 8.5, "ens": 8.4, "signal": "매수", "return": 15.8, "cap": "$420B", "pe": 36.5, "roe": 23, "included": True},
    {"rank": 4, "ticker": "TER", "name": "Teradyne", "total": 7.42, "fund": 6.7, "mom": 9.0, "ens": 7.4, "signal": "매수", "return": 15.5, "cap": "$48B", "pe": 89.4, "roe": 20, "included": False},
    {"rank": 5, "ticker": "CTRA", "name": "Coterra Energy", "total": 7.38, "fund": 7.1, "mom": 8.0, "ens": 7.4, "signal": "매수", "return": 15.3, "cap": "$23B", "pe": 14.1, "roe": 12, "included": False},
    {"rank": 6, "ticker": "ADI", "name": "Analog Devices", "total": 7.15, "fund": 5.6, "mom": 9.0, "ens": 6.2, "signal": "매수", "return": 14.5, "cap": "$159B", "pe": 71.0, "roe": 7, "included": False},
    {"rank": 7, "ticker": "KEY", "name": "KeyCorp", "total": 7.12, "fund": 6.3, "mom": 9.0, "ens": 6.5, "signal": "매수", "return": 14.4, "cap": "$25B", "pe": 15.1, "roe": 10, "included": False},
    {"rank": 8, "ticker": "STX", "name": "Seagate Technology", "total": 6.95, "fund": 4.9, "mom": 9.5, "ens": 5.1, "signal": "매수", "return": 13.8, "cap": "$86B", "pe": 48.1, "roe": "N/A", "included": False},
    {"rank": 9, "ticker": "NEM", "name": "Newmont Corporation", "total": 6.94, "fund": 6.7, "mom": 7.5, "ens": 6.9, "signal": "매수", "return": 13.8, "cap": "$133B", "pe": 18.8, "roe": 23, "included": False},
    {"rank": 10, "ticker": "CFG", "name": "Citizens Financial", "total": 6.90, "fund": 6.0, "mom": 9.0, "ens": 5.8, "signal": "매수", "return": 13.6, "cap": "$29B", "pe": 17.5, "roe": 7, "included": False},
    {"rank": 11, "ticker": "TRV", "name": "Travelers Companies", "total": 6.82, "fund": 7.1, "mom": 5.0, "ens": 7.1, "signal": "매수", "return": 13.4, "cap": "$67B", "pe": 10.7, "roe": 21, "included": True},
    {"rank": 12, "ticker": "PNC", "name": "PNC Financial", "total": 6.81, "fund": 5.9, "mom": 9.0, "ens": 6.1, "signal": "매수", "return": 13.3, "cap": "$97B", "pe": 14.5, "roe": 12, "included": False},
    {"rank": 13, "ticker": "GLW", "name": "Corning", "total": 6.76, "fund": 4.9, "mom": 9.0, "ens": 5.4, "signal": "매수", "return": 13.2, "cap": "$110B", "pe": 72.0, "roe": 15, "included": False},
    {"rank": 14, "ticker": "FITB", "name": "Fifth Third Bancorp", "total": 6.74, "fund": 5.3, "mom": 8.5, "ens": 5.6, "signal": "매수", "return": 13.1, "cap": "$49B", "pe": 15.4, "roe": 12, "included": True},
    {"rank": 15, "ticker": "PAYC", "name": "Paycom Software", "total": 6.71, "fund": 7.6, "mom": 1.5, "ens": 7.4, "signal": "매수", "return": 13.0, "cap": "$7.0B", "pe": 16.1, "roe": 29, "included": True},
    {"rank": 16, "ticker": "LMT", "name": "Lockheed Martin", "total": 6.70, "fund": 4.8, "mom": 9.0, "ens": 5.1, "signal": "매수", "return": 13.0, "cap": "$146B", "pe": 29.7, "roe": 77, "included": False},
    {"rank": 17, "ticker": "TDY", "name": "Teledyne Technologies", "total": 6.68, "fund": 4.0, "mom": 10.0, "ens": 4.3, "signal": "매수", "return": 12.9, "cap": "$31B", "pe": 34.7, "roe": 9, "included": False},
    {"rank": 18, "ticker": "ACGL", "name": "Arch Capital Group", "total": 6.62, "fund": 6.9, "mom": 5.0, "ens": 7.2, "signal": "매수", "return": 12.7, "cap": "$35B", "pe": 8.3, "roe": 20, "included": True},
    {"rank": 19, "ticker": "LRCX", "name": "Lam Research", "total": 6.58, "fund": 6.2, "mom": 7.5, "ens": 7.4, "signal": "매수", "return": 12.5, "cap": "$285B", "pe": 47.2, "roe": 66, "included": False},
    {"rank": 20, "ticker": "GILD", "name": "Gilead Sciences", "total": 6.55, "fund": 5.9, "mom": 8.0, "ens": 6.7, "signal": "매수", "return": 12.4, "cap": "$183B", "pe": 23.5, "roe": 41, "included": False},
    {"rank": 21, "ticker": "DECK", "name": "Deckers Outdoor", "total": 6.53, "fund": 6.1, "mom": 9.0, "ens": 6.4, "signal": "매수", "return": 12.4, "cap": "$17B", "pe": 16.3, "roe": 40, "included": True},
    {"rank": 22, "ticker": "BMY", "name": "Bristol-Myers Squibb", "total": 6.52, "fund": 5.9, "mom": 8.0, "ens": 6.0, "signal": "매수", "return": 12.3, "cap": "$124B", "pe": 17.6, "roe": 40, "included": False},
    {"rank": 23, "ticker": "HIG", "name": "Hartford Financial", "total": 6.51, "fund": 7.0, "mom": 4.0, "ens": 7.2, "signal": "매수", "return": 12.3, "cap": "$40B", "pe": 10.5, "roe": 22, "included": True},
    {"rank": 24, "ticker": "TPR", "name": "Tapestry", "total": 6.44, "fund": 5.2, "mom": 8.0, "ens": 5.5, "signal": "매수", "return": 12.0, "cap": "$31B", "pe": 59.8, "roe": 55, "included": False},
    {"rank": 25, "ticker": "OKE", "name": "ONEOK", "total": 6.41, "fund": 6.2, "mom": 7.0, "ens": 6.2, "signal": "매수", "return": 11.9, "cap": "$53B", "pe": 15.3, "roe": 18, "included": False},
    {"rank": 26, "ticker": "PYPL", "name": "PayPal Holdings", "total": 6.38, "fund": 7.2, "mom": 1.5, "ens": 7.1, "signal": "매수", "return": 11.8, "cap": "$39B", "pe": 7.6, "roe": 26, "included": True},
    {"rank": 27, "ticker": "JNJ", "name": "Johnson & Johnson", "total": 6.30, "fund": 6.0, "mom": 7.0, "ens": 6.8, "signal": "매수", "return": 11.6, "cap": "$574B", "pe": 21.6, "roe": 36, "included": False},
    {"rank": 28, "ticker": "WDC", "name": "Western Digital", "total": 6.17, "fund": 4.7, "mom": 9.5, "ens": 5.5, "signal": "매수", "return": 11.1, "cap": "$90B", "pe": 27.0, "roe": 41, "included": False},
    {"rank": 29, "ticker": "USB", "name": "U.S. Bancorp", "total": 6.13, "fund": 5.3, "mom": 8.0, "ens": 5.7, "signal": "매수", "return": 11.0, "cap": "$94B", "pe": 13.0, "roe": 12, "included": False},
    {"rank": 30, "ticker": "FAST", "name": "Fastenal Company", "total": 6.12, "fund": 5.3, "mom": 8.0, "ens": 6.1, "signal": "매수", "return": 10.9, "cap": "$54B", "pe": 42.8, "roe": 33, "included": False},
]

# 30 stocks x 7 analysts: (Buffett, Lynch, Fisher, Growth, Fundamentals, Sentiment, Technical)
matrix = [
    ("MOS", "Mosaic Company", ("b",35),("b",40),("b",25),("b",35),("n",58),("n",50),("B",68), "-", 1),
    ("CF", "CF Industries", ("B",70),("n",55),("b",30),("n",55),("B",78),("B",62),("B",72), "매수", 4),
    ("MU", "Micron Technology", ("n",55),("b",35),("B",75),("B",85),("B",72),("b",58),("B",75), "매수", 4),
    ("TER", "Teradyne", ("b",40),("b",30),("n",55),("B",72),("n",62),("n",48),("n",55), "-", 1),
    ("CTRA", "Coterra Energy", ("n",50),("n",60),("b",20),("b",40),("B",70),("n",52),("B",65), "-", 2),
    ("ADI", "Analog Devices", ("b",45),("b",35),("n",60),("n",58),("b",48),("b",55),("B",70), "-", 1),
    ("KEY", "KeyCorp", ("n",45),("b",40),("b",25),("b",30),("n",55),("b",60),("B",68), "-", 1),
    ("STX", "Seagate Technology", ("b",35),("b",30),("n",50),("B",78),("n",60),("b",65),("B",78), "-", 2),
    ("NEM", "Newmont Corp.", ("n",50),("B",70),("b",15),("B",80),("B",75),("b",57),("n",50), "-", 3),
    ("CFG", "Citizens Financial", ("n",48),("n",55),("b",30),("n",50),("n",54),("n",45),("B",68), "-", 1),
    ("TRV", "Travelers", ("B",80),("B",75),("n",45),("B",68),("B",80),("B",68),("n",45), "강력매수", 5),
    ("PNC", "PNC Financial", ("B",65),("n",60),("b",35),("n",52),("B",68),("n",48),("B",70), "-", 3),
    ("GLW", "Corning", ("b",40),("b",35),("n",55),("n",48),("b",45),("b",62),("B",70), "-", 1),
    ("FITB", "Fifth Third", ("B",62),("n",55),("b",35),("n",54),("B",70),("B",55),("B",67), "매수", 4),
    ("PAYC", "Paycom Software", ("B",75),("B",80),("B",70),("B",65),("B",76),("b",58),("b",40), "강력매수", 5),
    ("LMT", "Lockheed Martin", ("n",55),("n",50),("n",50),("n",58),("n",65),("n",50),("B",68), "-", 1),
    ("TDY", "Teledyne", ("n",52),("n",55),("B",65),("n",55),("n",62),("n",52),("B",75), "-", 2),
    ("ACGL", "Arch Capital", ("B",85),("B",85),("n",55),("B",75),("B",82),("b",60),("n",48), "매수", 4),
    ("LRCX", "Lam Research", ("b",42),("b",35),("B",80),("B",82),("B",74),("n",48),("n",60), "-", 3),
    ("GILD", "Gilead Sciences", ("n",58),("n",60),("n",50),("n",45),("n",58),("b",55),("n",58), "-", 0),
    ("DECK", "Deckers Outdoor", ("B",78),("B",75),("B",70),("B",62),("B",72),("n",45),("B",68), "강력매수", 6),
    ("BMY", "Bristol-Myers", ("n",50),("n",55),("n",45),("n",42),("n",60),("B",65),("n",55), "-", 1),
    ("HIG", "Hartford Fin.", ("B",82),("B",80),("n",50),("B",70),("B",77),("b",70),("b",42), "매수", 4),
    ("TPR", "Tapestry", ("b",38),("b",35),("n",40),("n",48),("b",42),("b",58),("B",65), "-", 1),
    ("OKE", "ONEOK", ("n",60),("n",65),("b",30),("B",72),("n",62),("B",60),("n",52), "-", 2),
    ("PYPL", "PayPal Holdings", ("B",72),("B",75),("B",65),("n",52),("B",78),("b",63),("b",35), "매수", 4),
    ("JNJ", "Johnson & Johnson", ("B",85),("n",60),("B",75),("n",55),("B",73),("n",52),("n",52), "-", 3),
    ("WDC", "Western Digital", ("n",48),("b",30),("n",50),("B",88),("B",70),("b",68),("B",72), "-", 3),
    ("USB", "U.S. Bancorp", ("B",68),("n",55),("b",30),("n",50),("n",58),("b",62),("n",55), "-", 1),
    ("FAST", "Fastenal", ("n",55),("b",40),("n",55),("n",58),("n",55),("n",48),("n",58), "-", 0),
]

details = [
    ("DECK","Deckers Outdoor","W.Buffett","B",78,"ROE 40% 탁월, 브랜드 파워(UGG/HOKA) 강력, 현금 21억으로 부채 초과"),
    ("DECK","Deckers Outdoor","P.Lynch","B",75,"P/E 16.3 ROE 40% GARP 매력, 매출+16% 순이익+27% 성장"),
    ("DECK","Deckers Outdoor","P.Fisher","B",70,"브랜드 파워 강력, R&D 없어도 제품 혁신 지속하는 특이 케이스"),
    ("DECK","Deckers Outdoor","Growth","B",62,"매출+16.3% 순이익+27.2% 가속, 프리미엄 브랜드와 마진 확대"),
    ("DECK","Deckers Outdoor","Fundamentals","B",72,"ROE 40% P/E 16.4 매력적, 유동비율 2.86 재무 건전"),
    ("DECK","Deckers Outdoor","Sentiment","n",45,"이사회 보상만 존재, 의미있는 매수 없음, 수요 신호 혼조"),
    ("DECK","Deckers Outdoor","Technical","B",68,"모멘텀 9.0, $105→$135 소비재 리더, 강한 상승 추세"),
    ("TRV","Travelers","W.Buffett","B",80,"보험업 ROE 21% P/E 10.7 저평가, 일관된 언더라이팅 수익성"),
    ("TRV","Travelers","P.Lynch","B",75,"P/E 10.7 저평가 보험 stalwart, FCF 수익률 15% 우수"),
    ("TRV","Travelers","P.Fisher","n",45,"언더라이팅 우수하나 R&D 없는 리스크 관리 비즈니스"),
    ("TRV","Travelers","Growth","B",68,"매출+12.4% 순이익+67.1% 가속, 보험 가격 결정력"),
    ("TRV","Travelers","Fundamentals","B",80,"ROE 21% P/E 10.9 저평가, FCF 수익률 15.5%"),
    ("TRV","Travelers","Sentiment","B",68,"경영진 낮은 행사가 옵션 행사로 자신감 표시"),
    ("TRV","Travelers","Technical","n",45,"모멘텀 5.0 약세, $260-$280 박스권"),
    ("PAYC","Paycom Software","W.Buffett","B",75,"ROE 29% gross margin 87% SaaS 반복매출, D/E 0.05"),
    ("PAYC","Paycom Software","P.Lynch","B",80,"P/E 16.1 ROE 29% GARP, 순이익+47% 시총$7B 10배주 잠재력"),
    ("PAYC","Paycom Software","P.Fisher","B",70,"R&D 13% HR SW 혁신, 86% 매출총이익률"),
    ("PAYC","Paycom Software","Growth","B",65,"매출+11.2% 순이익+47.3% 가속, SaaS 영업 레버리지"),
    ("PAYC","Paycom Software","Fundamentals","B",76,"ROE 29% 매출총이익률 87% ROIC 40% 최소 부채"),
    ("PAYC","Paycom Software","Sentiment","b",58,"경영진 지속적 매도, 급여 소프트웨어 경쟁 심화"),
    ("PAYC","Paycom Software","Technical","b",40,"모멘텀 1.5 극약세, $263→$207 하락 추세"),
    ("CF","CF Industries","W.Buffett","B",70,"ROE 22% P/E 11.5 합리적, 질소 비료 시장 지배력과 현금 창출력"),
    ("CF","CF Industries","Fundamentals","B",78,"ROE 22% P/E 11.5 유동비율 2.27 FCF 수익률 8.7%"),
    ("CF","CF Industries","Sentiment","B",62,"경영진 지분 보상 리텐션 집중, 질소 가격 개선 신호"),
    ("CF","CF Industries","Technical","B",72,"모멘텀 7.5, $90→$102 강한 상승, 저항선 돌파"),
    ("MU","Micron Technology","P.Fisher","B",75,"AI 메모리 수요 확대, R&D 집중도 높고 기술 우위 보유"),
    ("MU","Micron Technology","Growth","B",85,"매출+48.7% AI 메모리 수요와 강력한 영업 레버리지"),
    ("MU","Micron Technology","Fundamentals","B",72,"ROE 23% 유동비율 2.46 이자보상배율 21.2"),
    ("MU","Micron Technology","Technical","B",75,"모멘텀 8.5, $102→$127 강한 상승 추세"),
    ("ACGL","Arch Capital","W.Buffett","B",85,"보험업 ROE 20% P/E 8.3 저평가, 일관된 언더라이팅 수익성"),
    ("ACGL","Arch Capital","P.Lynch","B",85,"P/E 8.3 매우 저평가 boring business, 순이익+191% 폭증"),
    ("ACGL","Arch Capital","Growth","B",75,"매출+27.3% 급증, 보험 가격 사이클 순풍"),
    ("ACGL","Arch Capital","Fundamentals","B",82,"ROE 20% P/E 8.4 부채비율 0.11 FCF 수익률 우수"),
    ("HIG","Hartford Fin.","W.Buffett","B",82,"보험업 ROE 22% P/E 10.5 저평가, 건전한 자본 구조"),
    ("HIG","Hartford Fin.","P.Lynch","B",80,"P/E 10.5 저평가 보험 stalwart 순이익+24%"),
    ("HIG","Hartford Fin.","Growth","B",70,"매출+20.8% 순이익+24.2% 가속, 보험 가격 결정력"),
    ("HIG","Hartford Fin.","Fundamentals","B",77,"ROE 22% P/E 10.6 이자보상배율 20.3 부채비율 0.23"),
    ("PYPL","PayPal Holdings","W.Buffett","B",72,"ROE 26% P/E 7.6 저평가, 네트워크 효과"),
    ("PYPL","PayPal Holdings","P.Lynch","B",75,"P/E 7.6 극단적 저평가 핀테크 터너라운드 PEG<0.6"),
    ("PYPL","PayPal Holdings","P.Fisher","B",65,"R&D 9% 핀테크 기술 투자, ROE 26% 현금흐름 강력"),
    ("PYPL","PayPal Holdings","Fundamentals","B",78,"ROE 26% P/E 7.7 FCF 수익률 8.4% 이자보상배율 15.3"),
    ("FITB","Fifth Third","W.Buffett","B",62,"ROE 12% P/E 15.4 합리적, 배당 성장 지속"),
    ("FITB","Fifth Third","Fundamentals","B",70,"ROE 12% P/E 15.4 영업이익률 41% 배당 2.9%"),
    ("FITB","Fifth Third","Sentiment","B",55,"2026년 2월 대규모 지분 보상 프로그램 개시"),
    ("FITB","Fifth Third","Technical","B",67,"모멘텀 8.5, $37→$44 지역은행 회복 거래량 지지"),
]


def style_cell(ws, row, col, value, fill=None, font=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    return cell


def create_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "요약"
    # Summary
    ws.merge_cells('A1:B1')
    ws['A1'] = "AI Hedge Fund 통합 포트폴리오 리포트"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = Alignment(horizontal='center')
    for i, (k, v) in enumerate([
        ("분석 일자", "2026-02-11"), ("분석 대상", "S&P 500 상위 30개"),
        ("분석 전략", "하이브리드 (펀더멘털 70% + 모멘텀 30%)"),
        ("분석가 관점", "W.Buffett, P.Lynch, P.Fisher, Growth, Fundamentals, Sentiment, Technical (7명)"),
    ], 3):
        ws[f'A{i}'], ws[f'B{i}'] = k, v
    ws['A8'] = "포트폴리오 통계"
    ws['A8'].font = Font(bold=True)
    for i, (k, v) in enumerate([
        ("편입 종목 수", "9 / 30 (30.0%)"), ("평균 신뢰도", "74%"),
        ("평균 예상 수익률", "+13.5%"), ("강력매수 비중", "35.0%"), ("매수 비중", "65.0%"),
    ], 9):
        ws[f'A{i}'], ws[f'B{i}'] = k, v
    ws['A15'] = "섹터 분포"
    ws['A15'].font = Font(bold=True)
    for i, (k, v) in enumerate([
        ("Technology", "35.0% (MU, PAYC, PYPL)"), ("Financial", "35.0% (ACGL, HIG, TRV, FITB)"),
        ("Consumer Discretionary", "15.0% (DECK)"), ("Materials", "15.0% (CF)"),
    ], 16):
        ws[f'A{i}'], ws[f'B{i}'] = k, v
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 55

    # Portfolio sheet
    ws2 = wb.create_sheet("포트폴리오")
    headers = ["#", "종목코드", "회사명", "비중", "신호", "신뢰도", "예상수익률", "시가총액", "P/E", "ROE", "합의", "섹터"]
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    for r, item in enumerate(portfolio, 2):
        ws2.cell(row=r, column=1, value=r-1).border = THIN_BORDER
        ws2.cell(row=r, column=2, value=item['ticker']).border = THIN_BORDER
        ws2.cell(row=r, column=3, value=item['name']).border = THIN_BORDER
        ws2.cell(row=r, column=4, value=f"{item['weight']}%").border = THIN_BORDER
        sig = ws2.cell(row=r, column=5, value=f"{'🟢' if item['signal']=='강력매수' else '🔵'} {item['signal']}")
        sig.border = THIN_BORDER
        if item['signal'] == '강력매수':
            sig.fill = BULLISH_FILL
        ws2.cell(row=r, column=6, value=f"{item['confidence']}%").border = THIN_BORDER
        ws2.cell(row=r, column=7, value=f"+{item['return']}%").border = THIN_BORDER
        ws2.cell(row=r, column=8, value=item['cap']).border = THIN_BORDER
        ws2.cell(row=r, column=9, value=item['pe']).border = THIN_BORDER
        ws2.cell(row=r, column=10, value=f"{item['roe']}%").border = THIN_BORDER
        ws2.cell(row=r, column=11, value=item['consensus']).border = THIN_BORDER
        ws2.cell(row=r, column=12, value=item['sector']).border = THIN_BORDER
    tr = len(portfolio) + 2
    for c in range(1, 13):
        ws2.cell(row=tr, column=c).fill = TOTAL_FILL
        ws2.cell(row=tr, column=c).border = THIN_BORDER
    ws2.cell(row=tr, column=2, value="합계").font = Font(bold=True)
    ws2.cell(row=tr, column=4, value="100.0%").font = Font(bold=True)
    ws2.cell(row=tr, column=6, value="avg 74%")
    ws2.cell(row=tr, column=7, value="avg +13.5%")
    for i, w in enumerate([5,10,22,8,14,10,12,10,8,8,8,18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # Rankings sheet
    ws3 = wb.create_sheet("순위")
    rh = ["순위","종목코드","회사명","종합점수","펀더멘털","모멘텀","앙상블","신호","예상수익률","시가총액","P/E","ROE","편입여부"]
    for c, h in enumerate(rh, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    for r, item in enumerate(rankings, 2):
        ws3.cell(row=r, column=1, value=item['rank']).border = THIN_BORDER
        ws3.cell(row=r, column=2, value=item['ticker']).border = THIN_BORDER
        ws3.cell(row=r, column=3, value=item['name']).border = THIN_BORDER
        ws3.cell(row=r, column=4, value=item['total']).border = THIN_BORDER
        ws3.cell(row=r, column=5, value=item['fund']).border = THIN_BORDER
        ws3.cell(row=r, column=6, value=item['mom']).border = THIN_BORDER
        ws3.cell(row=r, column=7, value=item['ens']).border = THIN_BORDER
        ws3.cell(row=r, column=8, value=item['signal']).border = THIN_BORDER
        ws3.cell(row=r, column=9, value=f"+{item['return']}%").border = THIN_BORDER
        ws3.cell(row=r, column=10, value=item['cap']).border = THIN_BORDER
        ws3.cell(row=r, column=11, value=item['pe']).border = THIN_BORDER
        roe_val = f"{item['roe']}%" if item['roe'] != "N/A" else "N/A"
        ws3.cell(row=r, column=12, value=roe_val).border = THIN_BORDER
        inc = ws3.cell(row=r, column=13, value="예" if item['included'] else "아니오")
        inc.border = THIN_BORDER
        if item['included']:
            for ci in range(1, 14):
                ws3.cell(row=r, column=ci).fill = BULLISH_FILL
    for i, w in enumerate([6,10,22,10,10,8,8,10,12,10,8,8,10], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    # Matrix sheet (7 analysts)
    ws4 = wb.create_sheet("분석가 매트릭스")
    mh = ["종목코드","회사명","W.Buffett","P.Lynch","P.Fisher","Growth","Fundamentals","Sentiment","Technical","종합","Bullish"]
    for c, h in enumerate(mh, 1):
        cell = ws4.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    sig_map = {"B": ("매수", BULLISH_FILL, BULLISH_FONT), "b": ("매도", BEARISH_FILL, BEARISH_FONT), "n": ("중립", NEUTRAL_FILL, NEUTRAL_FONT)}
    for r, (ticker, name, buf, lyn, fis, gro, fun, sen, tec, combined, bc) in enumerate(matrix, 2):
        ws4.cell(row=r, column=1, value=ticker).border = THIN_BORDER
        ws4.cell(row=r, column=2, value=name).border = THIN_BORDER
        for ci, (sig, conf) in enumerate([buf, lyn, fis, gro, fun, sen, tec], 3):
            label, fill, font = sig_map[sig]
            cell = ws4.cell(row=r, column=ci, value=f"{label}({conf})")
            cell.border = THIN_BORDER
            cell.fill = fill
            cell.font = font
        ws4.cell(row=r, column=10, value=combined).border = THIN_BORDER
        ws4.cell(row=r, column=11, value=f"{bc}/7").border = THIN_BORDER
    for i, w in enumerate([10,20,13,13,13,13,13,13,13,10,8], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.freeze_panes = "A2"

    # Details sheet
    ws5 = wb.create_sheet("분석가 상세")
    dh = ["종목코드","회사명","분석가","신호","신뢰도","분석근거"]
    for c, h in enumerate(dh, 1):
        cell = ws5.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    for r, (ticker, name, analyst, sig, conf, reason) in enumerate(details, 2):
        ws5.cell(row=r, column=1, value=ticker).border = THIN_BORDER
        ws5.cell(row=r, column=2, value=name).border = THIN_BORDER
        ws5.cell(row=r, column=3, value=analyst).border = THIN_BORDER
        label, fill, font = sig_map[sig]
        cell = ws5.cell(row=r, column=4, value=label)
        cell.border = THIN_BORDER
        cell.fill = fill
        cell.font = font
        ws5.cell(row=r, column=5, value=conf).border = THIN_BORDER
        ws5.cell(row=r, column=6, value=reason).border = THIN_BORDER
    for i, w in enumerate([10,18,14,8,8,65], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w
    ws5.freeze_panes = "A2"

    # Risk sheet
    ws6 = wb.create_sheet("리스크 분석")
    ws6.merge_cells('A1:B1')
    ws6['A1'] = "리스크 분석"
    ws6['A1'].font = TITLE_FONT
    ws6['A1'].fill = TITLE_FILL
    ws6['A3'] = "집중도 지표"
    ws6['A3'].font = Font(bold=True)
    for i, (k, v) in enumerate([
        ("상위 1종목 비중", "15.0% (DECK/CF)"),
        ("상위 3종목 비중", "42.0%"),
        ("상위 5종목 비중", "65.0%"),
    ], 4):
        ws6[f'A{i}'], ws6[f'B{i}'] = k, v
    ws6['A8'] = "섹터 집중도"
    ws6['A8'].font = Font(bold=True)
    for i, (k, v) in enumerate([
        ("최대 섹터", "Technology / Financial (각 35.0%)"),
        ("상위 2섹터 비중", "70.0%"),
    ], 9):
        ws6[f'A{i}'], ws6[f'B{i}'] = k, v
    ws6['A12'] = "분석가 합의 품질"
    ws6['A12'].font = Font(bold=True)
    for i, (k, v) in enumerate([
        ("최다 합의 (6/7)", "1개 (DECK)"),
        ("강한 합의 (5/7)", "2개 (TRV, PAYC)"),
        ("과반 합의 (4/7)", "6개 (CF, MU, FITB, ACGL, HIG, PYPL)"),
        ("의견 분산 주의", "PAYC: Investor 3/3 B vs Analyst Sent/Tech bearish"),
    ], 13):
        ws6[f'A{i}'], ws6[f'B{i}'] = k, v
    ws6.column_dimensions['A'].width = 25
    ws6.column_dimensions['B'].width = 55
    return wb


if __name__ == "__main__":
    wb = create_workbook()
    path = "/Users/sgkim/Projects/ai-hedge-fund/portfolios/sp500_20260211_combined.xlsx"
    wb.save(path)
    print(f"Excel report saved: {path}")
