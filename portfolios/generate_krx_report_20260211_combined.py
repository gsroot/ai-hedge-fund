#!/usr/bin/env python3
"""KRX Combined Portfolio Report - Investor Team + Analyst Team (7 evaluators) - 2026-02-11"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style constants ──
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
BULLISH_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BULLISH_FONT = Font(color="375623")
BEARISH_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
BEARISH_FONT = Font(color="C62828")
NEUTRAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
NEUTRAL_FONT = Font(color="616161")
SUMMARY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
)

EVALUATORS = ["buffett", "lynch", "fisher", "growth", "fundamentals", "sentiment", "technical"]
EVAL_LABELS = {
    "buffett": "W.Buffett", "lynch": "P.Lynch", "fisher": "P.Fisher",
    "growth": "Growth", "fundamentals": "Fundamentals", "sentiment": "Sentiment", "technical": "Technical",
}

# ── Combined Portfolio (9 stocks, >=4/7 bullish) ──
PORTFOLIO = [
    {"rank": 1, "ticker": "018290", "name": "브이티", "weight": 12.1, "signal": "강력매수", "confidence": 80, "ret": 26.1, "mcap": "₩7,233억", "pe": 7.5, "roe": 43.0, "peg": 0.53, "consensus": "6/7", "sector": "화장품/뷰티"},
    {"rank": 2, "ticker": "033100", "name": "제룡전기", "weight": 11.4, "signal": "강력매수", "confidence": 76, "ret": 61.3, "mcap": "₩8,609억", "pe": 10.8, "roe": 41.0, "peg": 0.50, "consensus": "6/7", "sector": "전력기기"},
    {"rank": 3, "ticker": "137400", "name": "피엔티", "weight": 11.5, "signal": "강력매수", "confidence": 76, "ret": 48.9, "mcap": "₩12,106억", "pe": 8.2, "roe": 23.0, "peg": 0.49, "consensus": "6/7", "sector": "이차전지장비"},
    {"rank": 4, "ticker": "192080", "name": "더블유게임즈", "weight": 11.2, "signal": "강력매수", "confidence": 74, "ret": 15.5, "mcap": "₩11,815억", "pe": 5.9, "roe": 16.0, "peg": 0.48, "consensus": "6/7", "sector": "게임"},
    {"rank": 5, "ticker": "031980", "name": "피에스케이홀딩스", "weight": 11.0, "signal": "매수", "confidence": 73, "ret": 43.3, "mcap": "₩16,086억", "pe": 16.8, "roe": 22.0, "peg": 0.85, "consensus": "5/7", "sector": "반도체장비"},
    {"rank": 6, "ticker": "211050", "name": "인카금융서비스", "weight": 10.9, "signal": "매수", "confidence": 72, "ret": 22.9, "mcap": "₩6,941억", "pe": 11.2, "roe": 42.0, "peg": 0.80, "consensus": "5/7", "sector": "금융"},
    {"rank": 7, "ticker": "067160", "name": "SOOP", "weight": 10.7, "signal": "매수", "confidence": 71, "ret": 15.5, "mcap": "₩8,230억", "pe": 7.5, "roe": 27.0, "peg": 0.60, "consensus": "5/7", "sector": "미디어/IT"},
    {"rank": 8, "ticker": "348210", "name": "넥스틴", "weight": 10.6, "signal": "매수", "confidence": 70, "ret": 34.4, "mcap": "₩8,172억", "pe": 20.8, "roe": 24.0, "peg": 1.42, "consensus": "5/7", "sector": "반도체장비"},
    {"rank": 9, "ticker": "011200", "name": "HMM", "weight": 10.6, "signal": "강력매수", "confidence": 70, "ret": 14.7, "mcap": "₩20조", "pe": 4.2, "roe": 14.0, "peg": 0.31, "consensus": "6/7", "sector": "해운"},
]

# ── Rankings (30 stocks from latest predict) ──
RANKINGS = [
    ("033100", "제룡전기", 9.17, 8.5, 10.0, None, "강력매수", 61.3, "₩8,609억", 10.8, None, 41, None, 0.50, True),
    ("031980", "피에스케이홀딩스", 8.65, 7.5, 10.0, None, "강력매수", 43.3, "₩16,086억", 16.8, None, 22, None, 0.85, True),
    ("137400", "피엔티", 8.65, 7.5, 10.0, None, "강력매수", 48.9, "₩12,106억", 8.2, None, 23, None, 0.49, True),
    ("036930", "주성엔지니어링", 8.30, 7.0, 10.0, None, "강력매수", 34.9, "₩2.3조", 21.8, None, 19, None, None, False),
    ("348210", "넥스틴", 8.22, 7.5, 9.2, None, "강력매수", 34.4, "₩8,172억", 20.8, None, 24, None, 1.42, True),
    ("214150", "클래시스", 8.00, 8.0, 8.0, None, "강력매수", 21.0, "₩4.4조", 44.6, None, 22, None, None, False),
    ("018290", "브이티", 7.85, 6.5, 9.5, None, "매수", 26.1, "₩7,233억", 7.5, None, 43, None, 0.53, True),
    ("095340", "ISC", 7.75, 6.5, 9.3, None, "매수", 35.9, "₩3.6조", 64.0, None, 10, None, None, False),
    ("000660", "SK하이닉스", 7.72, 7.0, 8.6, None, "매수", 17.4, "₩636조", 30.4, None, 27, None, None, False),
    ("211050", "인카금융서비스", 7.65, 7.5, 7.8, None, "매수", 22.9, "₩6,941억", 11.2, None, 42, None, 0.80, True),
    ("178320", "서진시스템", 7.55, 7.5, 7.6, None, "매수", 22.5, "₩2.2조", 22.4, None, 10, None, None, False),
    ("011200", "HMM", 7.30, 7.0, 7.6, None, "매수", 14.7, "₩20조", 4.2, None, 14, None, 0.31, True),
    ("259960", "크래프톤", 7.28, 7.0, 7.6, None, "매수", 8.2, "₩12조", 8.9, None, 19, None, None, False),
    ("067160", "SOOP", 7.28, 8.0, 6.3, None, "매수", 15.5, "₩8,230억", 7.5, None, 27, None, 0.60, True),
    ("192080", "더블유게임즈", 7.25, 7.5, 6.8, None, "매수", 15.5, "₩11,815억", 5.9, None, 16, None, 0.48, True),
    ("278470", "카카오페이", 7.00, 5.0, 9.5, None, "매수", 13.2, "₩11조", 99.1, None, 33, None, None, False),
    ("101490", "에스앤에스텍", 6.95, 6.5, 7.5, None, "매수", 19.2, "₩19,116억", 61.7, None, 12, None, None, False),
    ("036540", "SFA반도체", 6.85, 5.5, 8.5, None, "매수", 25.7, "₩5,200억", 15.0, None, 12, None, None, False),
    ("005385", "현대차 우", 6.72, 6.5, 7.0, None, "매수", 12.4, "₩20조", 5.0, None, 12, None, None, False),
    ("323280", "태성", 6.72, 5.0, 8.8, None, "매수", 29.4, "₩2.4조", 345.7, None, 15, None, None, False),
    ("003690", "코리안리", 6.55, 6.0, 7.2, None, "매수", 10.3, "₩1.5조", 8.0, None, 10, None, None, False),
    ("005930", "삼성전자", 6.50, 5.5, 7.7, None, "매수", 8.1, "₩994조", 33.9, None, 9, None, None, False),
    ("047050", "포스코인터내셔널", 6.50, 5.5, 7.7, None, "매수", 11.8, "₩5.0조", 10.0, None, 15, None, None, False),
    ("035420", "NAVER", 6.50, 5.5, 7.7, None, "매수", 10.5, "₩35조", 25.0, None, 8, None, None, False),
    ("010140", "삼성중공업", 6.40, 5.0, 8.0, None, "매수", 14.2, "₩8.0조", 35.0, None, 5, None, None, False),
    ("003850", "보령", 6.35, 6.0, 6.8, None, "매수", 12.8, "₩8,000억", 12.0, None, 10, None, None, False),
    ("009150", "삼성전기", 6.22, 5.5, 7.1, None, "매수", 10.1, "₩12조", 18.0, None, 8, None, None, False),
    ("042670", "두산퓨얼셀", 6.15, 4.5, 8.0, None, "매수", 18.5, "₩5,000억", None, None, None, None, None, False),
    ("028260", "삼성물산", 6.10, 5.5, 6.8, None, "매수", 8.7, "₩20조", 12.0, None, 6, None, None, False),
    ("017370", "우신시스템", 6.00, 4.0, 8.5, None, "매수", 22.1, "₩3,000억", 10.0, None, 8, None, None, False),
]

# ── Combined Matrix (19 overlapping stocks × 7 evaluators) ──
# Investor team signals from Buffett/Lynch/Fisher analysis
# Analyst team signals from Growth/Fundamentals/Sentiment/Technical analysis
COMBINED_MATRIX = {
    "033100": {"buffett": ("매수", 78), "lynch": ("매수", 82), "fisher": ("매수", 78), "growth": ("매수", 72), "fundamentals": ("매수", 78), "sentiment": ("중립", 52), "technical": ("매수", 92), "combined": "강력매수", "conf": 76, "bullish": 6},
    "031980": {"buffett": ("중립", 65), "lynch": ("매수", 75), "fisher": ("매수", 82), "growth": ("매수", 75), "fundamentals": ("매수", 72), "sentiment": ("중립", 52), "technical": ("매수", 88), "combined": "매수", "conf": 73, "bullish": 5},
    "137400": {"buffett": ("매수", 72), "lynch": ("매수", 88), "fisher": ("매수", 85), "growth": ("매수", 70), "fundamentals": ("매수", 75), "sentiment": ("중립", 52), "technical": ("매수", 90), "combined": "강력매수", "conf": 76, "bullish": 6},
    "036930": {"buffett": ("중립", 62), "lynch": ("중립", 62), "fisher": ("중립", 68), "growth": ("매수", 68), "fundamentals": ("중립", 58), "sentiment": ("중립", 52), "technical": ("매수", 87), "combined": "중립", "conf": 65, "bullish": 2},
    "348210": {"buffett": ("매수", 68), "lynch": ("중립", 65), "fisher": ("매수", 76), "growth": ("매수", 70), "fundamentals": ("매수", 70), "sentiment": ("중립", 52), "technical": ("매수", 86), "combined": "매수", "conf": 70, "bullish": 5},
    "214150": {"buffett": ("중립", 58), "lynch": ("매도", 40), "fisher": ("중립", 55), "growth": ("중립", 55), "fundamentals": ("매도", 48), "sentiment": ("매수", 62), "technical": ("매수", 78), "combined": "중립", "conf": 57, "bullish": 2},
    "018290": {"buffett": ("매수", 82), "lynch": ("매수", 85), "fisher": ("매수", 88), "growth": ("매수", 78), "fundamentals": ("매수", 80), "sentiment": ("매도", 65), "technical": ("매수", 84), "combined": "강력매수", "conf": 80, "bullish": 6},
    "095340": {"buffett": ("매도", 42), "lynch": ("매도", 35), "fisher": ("매도", 45), "growth": ("중립", 55), "fundamentals": ("매도", 45), "sentiment": ("중립", 52), "technical": ("매수", 83), "combined": "매도", "conf": 51, "bullish": 1},
    "000660": {"buffett": ("중립", 70), "lynch": ("중립", 45), "fisher": ("매수", 80), "growth": ("매수", 80), "fundamentals": ("중립", 60), "sentiment": ("중립", 52), "technical": ("매수", 80), "combined": "중립", "conf": 67, "bullish": 3},
    "211050": {"buffett": ("매수", 75), "lynch": ("중립", 70), "fisher": ("매수", 81), "growth": ("매수", 75), "fundamentals": ("매수", 76), "sentiment": ("중립", 52), "technical": ("매수", 76), "combined": "매수", "conf": 72, "bullish": 5},
    "178320": {"buffett": ("중립", 48), "lynch": ("중립", 60), "fisher": ("중립", 66), "growth": ("중립", 55), "fundamentals": ("중립", 55), "sentiment": ("매도", 58), "technical": ("매수", 75), "combined": "중립", "conf": 60, "bullish": 1},
    "011200": {"buffett": ("매수", 68), "lynch": ("매수", 78), "fisher": ("매수", 70), "growth": ("매수", 72), "fundamentals": ("매수", 73), "sentiment": ("매도", 60), "technical": ("매수", 72), "combined": "강력매수", "conf": 70, "bullish": 6},
    "259960": {"buffett": ("중립", 65), "lynch": ("중립", 68), "fisher": ("매수", 75), "growth": ("매수", 76), "fundamentals": ("매수", 71), "sentiment": ("중립", 52), "technical": ("중립", 68), "combined": "중립", "conf": 68, "bullish": 3},
    "067160": {"buffett": ("매수", 72), "lynch": ("매수", 80), "fisher": ("매수", 77), "growth": ("매수", 74), "fundamentals": ("매수", 74), "sentiment": ("중립", 52), "technical": ("중립", 70), "combined": "매수", "conf": 71, "bullish": 5},
    "192080": {"buffett": ("매수", 70), "lynch": ("매수", 83), "fisher": ("매수", 79), "growth": ("매수", 73), "fundamentals": ("매수", 76), "sentiment": ("매수", 65), "technical": ("중립", 69), "combined": "강력매수", "conf": 74, "bullish": 6},
    "278470": {"buffett": ("매도", 38), "lynch": ("매도", 25), "fisher": ("매도", 40), "growth": ("매도", 45), "fundamentals": ("매도", 42), "sentiment": ("중립", 52), "technical": ("매수", 77), "combined": "매도", "conf": 46, "bullish": 1},
    "101490": {"buffett": ("매도", 40), "lynch": ("매도", 30), "fisher": ("매도", 42), "growth": ("중립", 55), "fundamentals": ("매도", 46), "sentiment": ("중립", 52), "technical": ("매수", 74), "combined": "매도", "conf": 48, "bullish": 1},
    "323280": {"buffett": ("매도", 35), "lynch": ("매도", 20), "fisher": ("매도", 35), "growth": ("매도", 42), "fundamentals": ("매도", 38), "sentiment": ("중립", 52), "technical": ("매수", 81), "combined": "매도", "conf": 43, "bullish": 1},
    "005930": {"buffett": ("중립", 65), "lynch": ("매도", 38), "fisher": ("중립", 62), "growth": ("중립", 55), "fundamentals": ("중립", 55), "sentiment": ("중립", 52), "technical": ("중립", 72), "combined": "중립", "conf": 57, "bullish": 0},
}

# ── Combined Detail (9 portfolio stocks × 7 evaluators = 63 rows) ──
COMBINED_DETAIL = [
    # 018290 브이티 — 6/7 강력매수
    ("018290", "브이티", "W.Buffett", "매수", 82, "ROE 43%로 탁월, P/E 7.5로 매우 저평가, 화장품 브랜드 해자 가능"),
    ("018290", "브이티", "P.Lynch", "매수", 85, "PEG 0.53 극도로 저평가, ROE 43%, 시총 7,233억 10배주 가능"),
    ("018290", "브이티", "P.Fisher", "매수", 88, "VT 브랜드 글로벌 확장 성공, 경영진 실행력 검증, R&D 지속 투자"),
    ("018290", "브이티", "Growth", "매수", 78, "글로벌 K-뷰티 시장 확대로 매출 고성장 지속, 브랜드 인지도 급상승"),
    ("018290", "브이티", "Fundamentals", "매수", 80, "ROE 43%로 업계 최고 수준, P/E 7.5 극도 저평가, 현금흐름 우수"),
    ("018290", "브이티", "Sentiment", "매도", 65, "밸류에이션 급등 후 차익실현 압력, 외국인 순매도 흐름 관찰"),
    ("018290", "브이티", "Technical", "매수", 84, "모멘텀 9.5로 강한 상승 모멘텀, 예상수익률 26.1%로 추세 지속"),
    # 033100 제룡전기 — 6/7 강력매수
    ("033100", "제룡전기", "W.Buffett", "매수", 78, "ROE 41%로 탁월한 자본효율성, P/E 10.8로 저평가, 전력기기 수요 예측 가능"),
    ("033100", "제룡전기", "P.Lynch", "매수", 82, "PEG 0.5로 매우 저평가된 fast grower, 시총 8,609억 10배주 잠재력"),
    ("033100", "제룡전기", "P.Fisher", "매수", 78, "ROE 41%로 경영진 우수성 입증, 탄소중립 시대 전력 인프라 수요 증가"),
    ("033100", "제룡전기", "Growth", "매수", 72, "전력기기 수요 급증으로 매출/이익 고성장 지속, 탄소중립 시대 수혜"),
    ("033100", "제룡전기", "Fundamentals", "매수", 78, "ROE 41%로 탁월한 자본효율성, P/E 10.8 저평가, 재무건전성 양호"),
    ("033100", "제룡전기", "Sentiment", "중립", 52, "전력기기 섹터 전반적 긍정적이나 개별 종목 뉴스 흐름 제한적"),
    ("033100", "제룡전기", "Technical", "매수", 92, "모멘텀 만점(10.0) + 예상수익률 61.3%로 강력한 상승 추세"),
    # 137400 피엔티 — 6/7 강력매수
    ("137400", "피엔티", "W.Buffett", "매수", 72, "ROE 23% 우수, P/E 8.2로 저평가, 이차전지 장비 수요 증가"),
    ("137400", "피엔티", "P.Lynch", "매수", 88, "PEG 0.49 극도로 저평가된 fast grower, 이차전지 성장 스토리 명확"),
    ("137400", "피엔티", "P.Fisher", "매수", 85, "이차전지 장비 수요 폭발적 증가, 기술 리더십으로 장기 성장동력 확보"),
    ("137400", "피엔티", "Growth", "매수", 70, "이차전지 장비 수요 폭발적 증가, 매출/이익 고성장 모멘텀 지속"),
    ("137400", "피엔티", "Fundamentals", "매수", 75, "ROE 23%, P/E 8.2로 저평가, PEG 0.49로 성장 대비 가격 매력적"),
    ("137400", "피엔티", "Sentiment", "중립", 52, "이차전지 섹터 불확실성 존재하나 장비주 투자 심리 안정적"),
    ("137400", "피엔티", "Technical", "매수", 90, "모멘텀 만점(10.0) + 예상수익률 48.9%로 매우 강한 상승 추세"),
    # 192080 더블유게임즈 — 6/7 강력매수
    ("192080", "더블유게임즈", "W.Buffett", "매수", 70, "P/E 5.9 매우 저평가, 소셜 카지노 반복 수익 모델"),
    ("192080", "더블유게임즈", "P.Lynch", "매수", 83, "PEG 0.48 극도로 저평가, 현금흐름 우수"),
    ("192080", "더블유게임즈", "P.Fisher", "매수", 79, "소셜카지노 글로벌 시장 1위, 경쟁우위 지속"),
    ("192080", "더블유게임즈", "Growth", "매수", 73, "소셜카지노 글로벌 시장 확대, 안정적 매출 성장 유지 중"),
    ("192080", "더블유게임즈", "Fundamentals", "매수", 76, "ROE 16%, P/E 5.9로 저평가, 반복 수익 모델로 안정적 현금흐름"),
    ("192080", "더블유게임즈", "Sentiment", "매수", 65, "글로벌 게임 시장 긍정적 전망, 투자심리 양호"),
    ("192080", "더블유게임즈", "Technical", "중립", 69, "모멘텀 6.8로 중립, 이동평균선 근처 횡보 중"),
    # 031980 피에스케이홀딩스 — 5/7 매수
    ("031980", "피에스케이홀딩스", "W.Buffett", "중립", 65, "ROE 22% 양호하나 반도체 장비 경기 민감도 높아 예측성 낮음"),
    ("031980", "피에스케이홀딩스", "P.Lynch", "매수", 75, "PEG 0.85 저평가된 stalwart, 반도체장비 수혜 명확"),
    ("031980", "피에스케이홀딩스", "P.Fisher", "매수", 82, "글로벌 반도체 투자 확대로 지속 성장, 기술 진입장벽 높음"),
    ("031980", "피에스케이홀딩스", "Growth", "매수", 75, "반도체장비 매출 성장 지속, HBM 수혜주로 고성장 전망"),
    ("031980", "피에스케이홀딩스", "Fundamentals", "매수", 72, "ROE 22% 양호, P/E 16.8 적정, 안정적 수익구조"),
    ("031980", "피에스케이홀딩스", "Sentiment", "중립", 52, "반도체 섹터 전반적 관심 높으나 개별 뉴스 흐름 중립적"),
    ("031980", "피에스케이홀딩스", "Technical", "매수", 88, "모멘텀 만점(10.0) + 예상수익률 43.3%로 강한 상승 모멘텀"),
    # 211050 인카금융서비스 — 5/7 매수
    ("211050", "인카금융서비스", "W.Buffett", "매수", 75, "ROE 42% 탁월, P/E 11.2 저평가, 안정적 수익 모델"),
    ("211050", "인카금융서비스", "P.Lynch", "중립", 70, "PEG 0.80 저평가이나 보험대리점 경쟁 심화 위험"),
    ("211050", "인카금융서비스", "P.Fisher", "매수", 81, "ROE 42% 탁월, 경영진 주주친화적 배당 정책"),
    ("211050", "인카금융서비스", "Growth", "매수", 75, "보험대리점 시장 구조적 성장, 실적 성장률 양호"),
    ("211050", "인카금융서비스", "Fundamentals", "매수", 76, "ROE 42%로 탁월, 안정적 배당 수익률, 수익 모델 견고"),
    ("211050", "인카금융서비스", "Sentiment", "중립", 52, "보험 섹터 전반적 중립, 내부자 거래 특이사항 없음"),
    ("211050", "인카금융서비스", "Technical", "매수", 76, "모멘텀 7.8 + 예상수익률 22.9%로 중상 수준 상승세"),
    # 067160 SOOP — 5/7 매수
    ("067160", "SOOP", "W.Buffett", "매수", 72, "ROE 27% 우수, P/E 7.5 저평가, 네트워크 효과로 해자 가능"),
    ("067160", "SOOP", "P.Lynch", "매수", 80, "PEG 0.60 저평가, 시총 8,230억 10배주 잠재력"),
    ("067160", "SOOP", "P.Fisher", "매수", 77, "인터넷 방송 플랫폼 독점적 지위, 콘텐츠 IP 확보"),
    ("067160", "SOOP", "Growth", "매수", 74, "라이브 스트리밍 시장 확대, 매출 성장 모멘텀 지속"),
    ("067160", "SOOP", "Fundamentals", "매수", 74, "ROE 27%, P/E 7.5로 저평가, 안정적 수익구조"),
    ("067160", "SOOP", "Sentiment", "중립", 52, "미디어 섹터 전반적 중립, 개별 종목 뉴스 제한적"),
    ("067160", "SOOP", "Technical", "중립", 70, "모멘텀 6.3으로 중립, 펀더멘털 강하나 기술적 모멘텀 약세"),
    # 348210 넥스틴 — 5/7 매수
    ("348210", "넥스틴", "W.Buffett", "매수", 68, "ROE 24% 우수, P/E 20.8 합리적, 반도체 검사장비 기술력"),
    ("348210", "넥스틴", "P.Lynch", "중립", 65, "PEG 1.42 적정가, ROE 24% 우수하나 밸류에이션 제약"),
    ("348210", "넥스틴", "P.Fisher", "매수", 76, "반도체 검사장비 특화 기술로 진입장벽 높음, 마진 개선 가능"),
    ("348210", "넥스틴", "Growth", "매수", 70, "반도체 검사장비 성장성 양호, 국내외 수주 확대 전망"),
    ("348210", "넥스틴", "Fundamentals", "매수", 70, "ROE 24% 양호, 재무구조 건전, 수익성 개선 추세"),
    ("348210", "넥스틴", "Sentiment", "중립", 52, "반도체 검사장비 개별 뉴스 부족, 시장 관심 제한적"),
    ("348210", "넥스틴", "Technical", "매수", 86, "모멘텀 9.2 + 예상수익률 34.4%로 강한 상승세 지속"),
    # 011200 HMM — 6/7 강력매수
    ("011200", "HMM", "W.Buffett", "매수", 68, "P/E 4.2 극단적 저평가, 해운 사이클 하단에서 안전마진 확보"),
    ("011200", "HMM", "P.Lynch", "매수", 78, "PEG 0.31 극도로 저평가된 cyclical, 해운 사이클 타이밍"),
    ("011200", "HMM", "P.Fisher", "매수", 70, "P/E 4.2 초저평가, 해운 업황 개선 시 수혜"),
    ("011200", "HMM", "Growth", "매수", 72, "해운 업황 회복기 진입, 매출/영업이익 성장 기대"),
    ("011200", "HMM", "Fundamentals", "매수", 73, "P/E 4.2로 극단적 저평가, 자산가치 대비 매력적"),
    ("011200", "HMM", "Sentiment", "매도", 60, "해운 업황 불확실성, 글로벌 무역 리스크 우려로 투자심리 약세"),
    ("011200", "HMM", "Technical", "매수", 72, "모멘텀 7.6 + 예상수익률 14.7%로 중간 수준 상승세"),
]


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22


def signal_style(signal):
    if signal in ("매수", "강력매수"):
        return BULLISH_FILL, BULLISH_FONT
    elif signal == "매도":
        return BEARISH_FILL, BEARISH_FONT
    return NEUTRAL_FILL, NEUTRAL_FONT


def apply_borders(ws, row, max_col):
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER


def create_sheet1_summary(wb):
    ws = wb.active
    ws.title = "요약"
    set_col_widths(ws, [25, 35])
    ws.merge_cells("A1:B1")
    cell = ws.cell(row=1, column=1, value="AI Hedge Fund 통합 포트폴리오 리포트")
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    ws.cell(row=1, column=2).fill = TITLE_FILL
    ws.row_dimensions[1].height = 30

    data = [
        (3, "분석 일자", "2026-02-11"),
        (4, "분석 대상", "KRX 상위 30개 (교차 분석 19개)"),
        (5, "분석 전략", "하이브리드 (펀더멘털 70% + 모멘텀 30%)"),
        (6, "투자자 팀", "W.Buffett, P.Lynch, P.Fisher"),
        (7, "분석가 팀", "Growth, Fundamentals, Sentiment, Technical"),
        (8, "편입 기준", "7인 중 과반(≥4/7) bullish"),
        (10, "포트폴리오 통계", ""),
        (11, "편입 종목 수", "9 / 19 (47.4%)"),
        (12, "평균 신뢰도", "74%"),
        (13, "평균 예상 수익률", "+31.6%"),
        (14, "강력매수 비중", "56.8%"),
        (15, "매수 비중", "43.2%"),
        (16, "가중평균 P/E", "10.3"),
        (17, "가중평균 ROE", "28.2%"),
        (19, "시가총액 분포", ""),
        (20, "대형주 (₩10조+)", "10.6%"),
        (21, "중형주 (₩1-10조)", "33.7%"),
        (22, "소형주 (<₩1조)", "55.7%"),
        (24, "섹터 분포", ""),
        (25, "반도체장비", "21.6%"),
        (26, "화장품/뷰티", "12.1%"),
        (27, "이차전지장비", "11.5%"),
        (28, "전력기기", "11.4%"),
        (29, "게임", "11.2%"),
        (30, "금융", "10.9%"),
        (31, "미디어/IT", "10.7%"),
        (32, "해운", "10.6%"),
    ]
    for r, label, value in data:
        if value == "":
            ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=12, color="1F4E79")
        else:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=value)


def create_sheet2_portfolio(wb):
    ws = wb.create_sheet("포트폴리오")
    headers = ["#", "종목코드", "회사명", "비중", "신호", "신뢰도", "예상수익률", "시가총액", "P/E", "ROE", "PEG", "합의", "섹터"]
    set_col_widths(ws, [5, 10, 16, 8, 10, 8, 10, 12, 8, 8, 8, 8, 14])
    write_header_row(ws, 1, headers)

    for i, s in enumerate(PORTFOLIO, 2):
        ws.cell(row=i, column=1, value=s["rank"])
        ws.cell(row=i, column=2, value=s["ticker"])
        ws.cell(row=i, column=3, value=s["name"])
        ws.cell(row=i, column=4, value=s["weight"] / 100).number_format = "0.0%"
        sig_cell = ws.cell(row=i, column=5, value=s["signal"])
        fill, font = signal_style(s["signal"])
        sig_cell.fill = fill
        sig_cell.font = font
        ws.cell(row=i, column=6, value=s["confidence"])
        ws.cell(row=i, column=7, value=s["ret"] / 100).number_format = "+0.0%;-0.0%"
        ws.cell(row=i, column=8, value=s["mcap"])
        ws.cell(row=i, column=9, value=s["pe"]).number_format = "0.0"
        ws.cell(row=i, column=10, value=s["roe"] / 100).number_format = "0.0%"
        ws.cell(row=i, column=11, value=s["peg"]).number_format = "0.00"
        ws.cell(row=i, column=12, value=s["consensus"])
        ws.cell(row=i, column=13, value=s["sector"])
        apply_borders(ws, i, 13)
        ws.row_dimensions[i].height = 18

    r = len(PORTFOLIO) + 2
    ws.cell(row=r, column=3, value="합계/평균").font = Font(bold=True)
    ws.cell(row=r, column=4, value=1.0).number_format = "0.0%"
    ws.cell(row=r, column=6, value=74)
    ws.cell(row=r, column=7, value=0.316).number_format = "+0.0%;-0.0%"
    for c in range(1, 14):
        ws.cell(row=r, column=c).fill = SUMMARY_FILL
        ws.cell(row=r, column=c).font = Font(bold=True)
        ws.cell(row=r, column=c).border = THIN_BORDER
    ws.auto_filter.ref = f"A1:M{len(PORTFOLIO) + 1}"
    ws.freeze_panes = "A2"


def create_sheet3_ranking(wb):
    ws = wb.create_sheet("순위")
    headers = ["순위", "종목코드", "회사명", "종합점수", "펀더멘털", "모멘텀", "앙상블", "신호", "예상수익률", "시가총액", "P/E", "P/B", "ROE", "매출성장률", "PEG", "편입여부"]
    set_col_widths(ws, [6, 10, 16, 8, 8, 8, 8, 10, 10, 12, 8, 8, 8, 10, 8, 8])
    write_header_row(ws, 1, headers)
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for i, r in enumerate(RANKINGS, 2):
        ticker, name, score, fund, mom, ens, sig, ret, mcap, pe, pb, roe, rev_g, peg, included = r
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=ticker)
        ws.cell(row=i, column=3, value=name)
        ws.cell(row=i, column=4, value=score).number_format = "0.00"
        ws.cell(row=i, column=5, value=fund).number_format = "0.0"
        ws.cell(row=i, column=6, value=mom).number_format = "0.0"
        ws.cell(row=i, column=7, value=ens)
        ws.cell(row=i, column=8, value=sig)
        ws.cell(row=i, column=9, value=ret / 100 if ret else None)
        if ws.cell(row=i, column=9).value is not None:
            ws.cell(row=i, column=9).number_format = "+0.0%;-0.0%"
        ws.cell(row=i, column=10, value=mcap)
        if pe is not None:
            ws.cell(row=i, column=11, value=pe).number_format = "0.0"
        ws.cell(row=i, column=12, value=pb)
        if roe is not None:
            ws.cell(row=i, column=13, value=roe / 100).number_format = "0.0%"
        ws.cell(row=i, column=14, value=rev_g)
        ws.cell(row=i, column=15, value=peg)
        ws.cell(row=i, column=16, value="예" if included else "아니오")
        for c in range(1, 17):
            ws.cell(row=i, column=c).border = THIN_BORDER
            if included:
                ws.cell(row=i, column=c).fill = green_fill
        ws.row_dimensions[i].height = 18
    ws.auto_filter.ref = f"A1:P{len(RANKINGS) + 1}"
    ws.freeze_panes = "A2"


def create_sheet4_matrix(wb):
    ws = wb.create_sheet("통합 매트릭스")
    headers = ["종목코드", "회사명"] + [EVAL_LABELS[e] for e in EVALUATORS] + ["종합신호", "신뢰도", "Bullish"]
    set_col_widths(ws, [10, 16, 12, 12, 12, 12, 14, 12, 12, 10, 8, 8])
    write_header_row(ws, 1, headers)
    # Order: by combined matrix keys matching RANKINGS order
    tickers_ordered = [r[0] for r in RANKINGS if r[0] in COMBINED_MATRIX]

    for i, ticker in enumerate(tickers_ordered, 2):
        name = next(r[1] for r in RANKINGS if r[0] == ticker)
        m = COMBINED_MATRIX[ticker]
        ws.cell(row=i, column=1, value=ticker)
        ws.cell(row=i, column=2, value=name)
        for ci, key in enumerate(EVALUATORS, 3):
            sig, conf = m[key]
            cell = ws.cell(row=i, column=ci, value=f"{sig}({conf})")
            fill, font = signal_style(sig)
            cell.fill = fill
            cell.font = font
        col_sig = 3 + len(EVALUATORS)  # column 10
        ws.cell(row=i, column=col_sig, value=m["combined"])
        f_sig, fn_sig = signal_style(m["combined"])
        ws.cell(row=i, column=col_sig).fill = f_sig
        ws.cell(row=i, column=col_sig).font = fn_sig
        ws.cell(row=i, column=col_sig + 1, value=m["conf"])
        ws.cell(row=i, column=col_sig + 2, value=f"{m['bullish']}/7")
        apply_borders(ws, i, col_sig + 2)
        ws.row_dimensions[i].height = 18
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(tickers_ordered) + 1}"
    ws.freeze_panes = "A2"


def create_sheet5_detail(wb):
    ws = wb.create_sheet("통합 상세")
    headers = ["종목코드", "회사명", "평가자", "팀", "신호", "신뢰도", "분석근거"]
    set_col_widths(ws, [10, 16, 14, 10, 8, 8, 55])
    write_header_row(ws, 1, headers)

    team_map = {"W.Buffett": "투자자", "P.Lynch": "투자자", "P.Fisher": "투자자",
                "Growth": "분석가", "Fundamentals": "분석가", "Sentiment": "분석가", "Technical": "분석가"}

    for i, (ticker, name, evaluator, sig, conf, reason) in enumerate(COMBINED_DETAIL, 2):
        ws.cell(row=i, column=1, value=ticker)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=evaluator)
        ws.cell(row=i, column=4, value=team_map.get(evaluator, ""))
        sig_cell = ws.cell(row=i, column=5, value=sig)
        fill, font = signal_style(sig)
        sig_cell.fill = fill
        sig_cell.font = font
        ws.cell(row=i, column=6, value=conf)
        ws.cell(row=i, column=7, value=reason)
        apply_borders(ws, i, 7)
        ws.row_dimensions[i].height = 18
    ws.freeze_panes = "A2"


def create_sheet6_risk(wb):
    ws = wb.create_sheet("리스크 분석")
    set_col_widths(ws, [28, 25])
    ws.merge_cells("A1:B1")
    cell = ws.cell(row=1, column=1, value="리스크 분석")
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    ws.cell(row=1, column=2).fill = TITLE_FILL

    data = [
        (3, "집중도 지표", "", True),
        (4, "상위 1종목 비중", "12.1% (브이티)"),
        (5, "상위 3종목 비중", "35.0%"),
        (6, "상위 5종목 비중", "57.2%"),
        (7, "상위 9종목 비중", "100.0%"),
        (8, "HHI (허핀달 지수)", "0.111 (9종목 거의 균등배분)"),
        (10, "섹터 집중도", "", True),
        (11, "최대 섹터", "반도체장비 (21.6%)"),
        (12, "상위 3섹터 비중", "45.2%"),
        (14, "7인 합의 품질", "", True),
        (15, "만장일치(7/7) 비율", "0.0%"),
        (16, "강력합의(6/7) 비율", "55.6% (5종목)"),
        (17, "다수합의(5/7) 비율", "44.4% (4종목)"),
        (18, "의견 분산 종목", "018290, 011200 (Sentiment 매도 vs 나머지 매수)"),
        (20, "비편입 사유 분포", "", True),
        (21, "교차분석 미달(≤3/7)", "10개 (19개 교차분석 중)"),
        (22, "투자자 미분석", "11개 (분석가 팀만 분석)"),
        (24, "시가총액 리스크", "", True),
        (25, "소형주 비중", "55.7% (유동성 주의)"),
        (27, "팀 간 비교", "", True),
        (28, "투자자 팀 전용 편입", "067160(SOOP) — 분석가 팀에서는 2/4로 미달"),
        (29, "분석가 팀 전용 편입", "없음 — 분석가 팀 편입 8종목 모두 통합에도 편입"),
        (30, "양팀 공통 편입", "8종목 (018290, 033100, 137400, 031980 등)"),
        (32, "팀별 특성", "", True),
        (33, "투자자 팀", "질적 분석 중심, 가치평가(PEG/ROE) 중시"),
        (34, "분석가 팀", "정량 분석 중심, Sentiment 매우 보수적"),
        (35, "통합 시너지", "7인 교차검증으로 신뢰도 향상, 편향 완화"),
    ]
    for r, label, value, *rest in data:
        is_header = rest[0] if rest else False
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=12, color="1F4E79") if is_header else Font()
        if value:
            ws.cell(row=r, column=2, value=value)


def main():
    wb = openpyxl.Workbook()
    create_sheet1_summary(wb)
    create_sheet2_portfolio(wb)
    create_sheet3_ranking(wb)
    create_sheet4_matrix(wb)
    create_sheet5_detail(wb)
    create_sheet6_risk(wb)

    path = "portfolios/krx_20260211_combined.xlsx"
    wb.save(path)
    print(f"Excel report saved: {path}")


if __name__ == "__main__":
    main()
