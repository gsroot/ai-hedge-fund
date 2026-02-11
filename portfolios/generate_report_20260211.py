#!/usr/bin/env python3
"""S&P 500 Portfolio Report Generator - 2026-02-11"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
BULLISH_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BULLISH_FONT = Font(color="375623")
BEARISH_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
BEARISH_FONT = Font(color="C62828")
NEUTRAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
NEUTRAL_FONT = Font(color="616161")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

portfolio = [
    {"ticker": "PAYC", "name": "Paycom Software", "weight": 16.5, "signal": "강력매수", "confidence": 75, "return": 13.0, "cap": "$7.0B", "pe": 16.1, "roe": 29, "peg": "N/A", "consensus": "3/3", "sector": "Technology"},
    {"ticker": "DECK", "name": "Deckers Outdoor", "weight": 16.5, "signal": "강력매수", "confidence": 74, "return": 12.4, "cap": "$17B", "pe": 16.3, "roe": 40, "peg": "N/A", "consensus": "3/3", "sector": "Consumer Disc."},
    {"ticker": "JNJ", "name": "Johnson & Johnson", "weight": 16.0, "signal": "매수", "confidence": 73, "return": 11.6, "cap": "$574B", "pe": 21.6, "roe": 36, "peg": "N/A", "consensus": "2/3", "sector": "Healthcare"},
    {"ticker": "PYPL", "name": "PayPal Holdings", "weight": 16.0, "signal": "강력매수", "confidence": 71, "return": 11.8, "cap": "$39B", "pe": 7.6, "roe": 26, "peg": "N/A", "consensus": "3/3", "sector": "Technology"},
    {"ticker": "ACGL", "name": "Arch Capital Group", "weight": 12.5, "signal": "매수", "confidence": 75, "return": 12.7, "cap": "$35B", "pe": 8.3, "roe": 20, "peg": "N/A", "consensus": "2/3", "sector": "Financial"},
    {"ticker": "HIG", "name": "Hartford Financial", "weight": 11.5, "signal": "매수", "confidence": 71, "return": 12.3, "cap": "$40B", "pe": 10.5, "roe": 22, "peg": "N/A", "consensus": "2/3", "sector": "Financial"},
    {"ticker": "TRV", "name": "Travelers Companies", "weight": 11.0, "signal": "매수", "confidence": 67, "return": 13.4, "cap": "$67B", "pe": 10.7, "roe": 21, "peg": "N/A", "consensus": "2/3", "sector": "Financial"},
]

rankings = [
    {"rank": 1, "ticker": "MOS", "name": "Mosaic Company", "total": 7.87, "fund": 7.8, "mom": 8.0, "ens": 7.1, "signal": "매수", "return": 17.1, "cap": "$9.6B", "pe": 7.7, "pb": "N/A", "roe": 10, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 2, "ticker": "CF", "name": "CF Industries", "total": 7.86, "fund": 8.0, "mom": 7.5, "ens": 7.8, "signal": "매수", "return": 17.0, "cap": "$16B", "pe": 11.5, "pb": "N/A", "roe": 22, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 3, "ticker": "MU", "name": "Micron Technology", "total": 7.51, "fund": 7.1, "mom": 8.5, "ens": 8.4, "signal": "매수", "return": 15.8, "cap": "$420B", "pe": 36.5, "pb": "N/A", "roe": 23, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 4, "ticker": "TER", "name": "Teradyne", "total": 7.42, "fund": 6.7, "mom": 9.0, "ens": 7.4, "signal": "매수", "return": 15.5, "cap": "$48B", "pe": 89.4, "pb": "N/A", "roe": 20, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 5, "ticker": "CTRA", "name": "Coterra Energy", "total": 7.38, "fund": 7.1, "mom": 8.0, "ens": 7.4, "signal": "매수", "return": 15.3, "cap": "$23B", "pe": 14.1, "pb": "N/A", "roe": 12, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 6, "ticker": "ADI", "name": "Analog Devices", "total": 7.15, "fund": 5.6, "mom": 9.0, "ens": 6.2, "signal": "매수", "return": 14.5, "cap": "$159B", "pe": 71.0, "pb": "N/A", "roe": 7, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 7, "ticker": "KEY", "name": "KeyCorp", "total": 7.12, "fund": 6.3, "mom": 9.0, "ens": 6.5, "signal": "매수", "return": 14.4, "cap": "$25B", "pe": 15.1, "pb": "N/A", "roe": 10, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 8, "ticker": "STX", "name": "Seagate Technology", "total": 6.95, "fund": 4.9, "mom": 9.5, "ens": 5.1, "signal": "매수", "return": 13.8, "cap": "$86B", "pe": 48.1, "pb": "N/A", "roe": "N/A", "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 9, "ticker": "NEM", "name": "Newmont Corporation", "total": 6.94, "fund": 6.7, "mom": 7.5, "ens": 6.9, "signal": "매수", "return": 13.8, "cap": "$133B", "pe": 18.8, "pb": "N/A", "roe": 23, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 10, "ticker": "CFG", "name": "Citizens Financial", "total": 6.90, "fund": 6.0, "mom": 9.0, "ens": 5.8, "signal": "매수", "return": 13.6, "cap": "$29B", "pe": 17.5, "pb": "N/A", "roe": 7, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 11, "ticker": "TRV", "name": "Travelers Companies", "total": 6.82, "fund": 7.1, "mom": 5.0, "ens": 7.1, "signal": "매수", "return": 13.4, "cap": "$67B", "pe": 10.7, "pb": "N/A", "roe": 21, "rev_growth": "N/A", "peg": "N/A", "included": True},
    {"rank": 12, "ticker": "PNC", "name": "PNC Financial", "total": 6.81, "fund": 5.9, "mom": 9.0, "ens": 6.1, "signal": "매수", "return": 13.3, "cap": "$97B", "pe": 14.5, "pb": "N/A", "roe": 12, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 13, "ticker": "GLW", "name": "Corning", "total": 6.76, "fund": 4.9, "mom": 9.0, "ens": 5.4, "signal": "매수", "return": 13.2, "cap": "$110B", "pe": 72.0, "pb": "N/A", "roe": 15, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 14, "ticker": "FITB", "name": "Fifth Third Bancorp", "total": 6.74, "fund": 5.3, "mom": 8.5, "ens": 5.6, "signal": "매수", "return": 13.1, "cap": "$49B", "pe": 15.4, "pb": "N/A", "roe": 12, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 15, "ticker": "PAYC", "name": "Paycom Software", "total": 6.71, "fund": 7.6, "mom": 1.5, "ens": 7.4, "signal": "매수", "return": 13.0, "cap": "$7.0B", "pe": 16.1, "pb": "N/A", "roe": 29, "rev_growth": "N/A", "peg": "N/A", "included": True},
    {"rank": 16, "ticker": "LMT", "name": "Lockheed Martin", "total": 6.70, "fund": 4.8, "mom": 9.0, "ens": 5.1, "signal": "매수", "return": 13.0, "cap": "$146B", "pe": 29.7, "pb": "N/A", "roe": 77, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 17, "ticker": "TDY", "name": "Teledyne Technologies", "total": 6.68, "fund": 4.0, "mom": 10.0, "ens": 4.3, "signal": "매수", "return": 12.9, "cap": "$31B", "pe": 34.7, "pb": "N/A", "roe": 9, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 18, "ticker": "ACGL", "name": "Arch Capital Group", "total": 6.62, "fund": 6.9, "mom": 5.0, "ens": 7.2, "signal": "매수", "return": 12.7, "cap": "$35B", "pe": 8.3, "pb": "N/A", "roe": 20, "rev_growth": "N/A", "peg": "N/A", "included": True},
    {"rank": 19, "ticker": "LRCX", "name": "Lam Research", "total": 6.58, "fund": 6.2, "mom": 7.5, "ens": 7.4, "signal": "매수", "return": 12.5, "cap": "$285B", "pe": 47.2, "pb": "N/A", "roe": 66, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 20, "ticker": "GILD", "name": "Gilead Sciences", "total": 6.55, "fund": 5.9, "mom": 8.0, "ens": 6.7, "signal": "매수", "return": 12.4, "cap": "$183B", "pe": 23.5, "pb": "N/A", "roe": 41, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 21, "ticker": "DECK", "name": "Deckers Outdoor", "total": 6.53, "fund": 6.1, "mom": 9.0, "ens": 6.4, "signal": "매수", "return": 12.4, "cap": "$17B", "pe": 16.3, "pb": "N/A", "roe": 40, "rev_growth": "N/A", "peg": "N/A", "included": True},
    {"rank": 22, "ticker": "BMY", "name": "Bristol-Myers Squibb", "total": 6.52, "fund": 5.9, "mom": 8.0, "ens": 6.0, "signal": "매수", "return": 12.3, "cap": "$124B", "pe": 17.6, "pb": "N/A", "roe": 40, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 23, "ticker": "HIG", "name": "Hartford Financial", "total": 6.51, "fund": 7.0, "mom": 4.0, "ens": 7.2, "signal": "매수", "return": 12.3, "cap": "$40B", "pe": 10.5, "pb": "N/A", "roe": 22, "rev_growth": "N/A", "peg": "N/A", "included": True},
    {"rank": 24, "ticker": "TPR", "name": "Tapestry", "total": 6.44, "fund": 5.2, "mom": 8.0, "ens": 5.5, "signal": "매수", "return": 12.0, "cap": "$31B", "pe": 59.8, "pb": "N/A", "roe": 55, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 25, "ticker": "OKE", "name": "ONEOK", "total": 6.41, "fund": 6.2, "mom": 7.0, "ens": 6.2, "signal": "매수", "return": 11.9, "cap": "$53B", "pe": 15.3, "pb": "N/A", "roe": 18, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 26, "ticker": "PYPL", "name": "PayPal Holdings", "total": 6.38, "fund": 7.2, "mom": 1.5, "ens": 7.1, "signal": "매수", "return": 11.8, "cap": "$39B", "pe": 7.6, "pb": "N/A", "roe": 26, "rev_growth": "N/A", "peg": "N/A", "included": True},
    {"rank": 27, "ticker": "JNJ", "name": "Johnson & Johnson", "total": 6.30, "fund": 6.0, "mom": 7.0, "ens": 6.8, "signal": "매수", "return": 11.6, "cap": "$574B", "pe": 21.6, "pb": "N/A", "roe": 36, "rev_growth": "N/A", "peg": "N/A", "included": True},
    {"rank": 28, "ticker": "WDC", "name": "Western Digital", "total": 6.17, "fund": 4.7, "mom": 9.5, "ens": 5.5, "signal": "매수", "return": 11.1, "cap": "$90B", "pe": 27.0, "pb": "N/A", "roe": 41, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 29, "ticker": "USB", "name": "U.S. Bancorp", "total": 6.13, "fund": 5.3, "mom": 8.0, "ens": 5.7, "signal": "매수", "return": 11.0, "cap": "$94B", "pe": 13.0, "pb": "N/A", "roe": 12, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 30, "ticker": "FAST", "name": "Fastenal Company", "total": 6.12, "fund": 5.3, "mom": 8.0, "ens": 6.1, "signal": "매수", "return": 10.9, "cap": "$54B", "pe": 42.8, "pb": "N/A", "roe": 33, "rev_growth": "N/A", "peg": "N/A", "included": False},
]

investor_matrix = [
    {"ticker": "MOS", "name": "Mosaic Company", "buffett": ("bearish", 35), "lynch": ("bearish", 40), "fisher": ("bearish", 25), "combined": "-", "conf": "-"},
    {"ticker": "CF", "name": "CF Industries", "buffett": ("bullish", 70), "lynch": ("neutral", 55), "fisher": ("bearish", 30), "combined": "-", "conf": "-"},
    {"ticker": "MU", "name": "Micron Technology", "buffett": ("neutral", 55), "lynch": ("bearish", 35), "fisher": ("bullish", 75), "combined": "-", "conf": "-"},
    {"ticker": "TER", "name": "Teradyne", "buffett": ("bearish", 40), "lynch": ("bearish", 30), "fisher": ("neutral", 55), "combined": "-", "conf": "-"},
    {"ticker": "CTRA", "name": "Coterra Energy", "buffett": ("neutral", 50), "lynch": ("neutral", 60), "fisher": ("bearish", 20), "combined": "-", "conf": "-"},
    {"ticker": "ADI", "name": "Analog Devices", "buffett": ("bearish", 45), "lynch": ("bearish", 35), "fisher": ("neutral", 60), "combined": "-", "conf": "-"},
    {"ticker": "KEY", "name": "KeyCorp", "buffett": ("neutral", 45), "lynch": ("bearish", 40), "fisher": ("bearish", 25), "combined": "-", "conf": "-"},
    {"ticker": "STX", "name": "Seagate Technology", "buffett": ("bearish", 35), "lynch": ("bearish", 30), "fisher": ("neutral", 50), "combined": "-", "conf": "-"},
    {"ticker": "NEM", "name": "Newmont Corporation", "buffett": ("neutral", 50), "lynch": ("bullish", 70), "fisher": ("bearish", 15), "combined": "-", "conf": "-"},
    {"ticker": "CFG", "name": "Citizens Financial", "buffett": ("neutral", 48), "lynch": ("neutral", 55), "fisher": ("bearish", 30), "combined": "-", "conf": "-"},
    {"ticker": "TRV", "name": "Travelers Companies", "buffett": ("bullish", 80), "lynch": ("bullish", 75), "fisher": ("neutral", 45), "combined": "매수", "conf": "67%"},
    {"ticker": "PNC", "name": "PNC Financial", "buffett": ("bullish", 65), "lynch": ("neutral", 60), "fisher": ("bearish", 35), "combined": "-", "conf": "-"},
    {"ticker": "GLW", "name": "Corning", "buffett": ("bearish", 40), "lynch": ("bearish", 35), "fisher": ("neutral", 55), "combined": "-", "conf": "-"},
    {"ticker": "FITB", "name": "Fifth Third Bancorp", "buffett": ("bullish", 62), "lynch": ("neutral", 55), "fisher": ("bearish", 35), "combined": "-", "conf": "-"},
    {"ticker": "PAYC", "name": "Paycom Software", "buffett": ("bullish", 75), "lynch": ("bullish", 80), "fisher": ("bullish", 70), "combined": "강력매수", "conf": "75%"},
    {"ticker": "LMT", "name": "Lockheed Martin", "buffett": ("neutral", 55), "lynch": ("neutral", 50), "fisher": ("neutral", 50), "combined": "-", "conf": "-"},
    {"ticker": "TDY", "name": "Teledyne Technologies", "buffett": ("neutral", 52), "lynch": ("neutral", 55), "fisher": ("bullish", 65), "combined": "-", "conf": "-"},
    {"ticker": "ACGL", "name": "Arch Capital Group", "buffett": ("bullish", 85), "lynch": ("bullish", 85), "fisher": ("neutral", 55), "combined": "매수", "conf": "75%"},
    {"ticker": "LRCX", "name": "Lam Research", "buffett": ("bearish", 42), "lynch": ("bearish", 35), "fisher": ("bullish", 80), "combined": "-", "conf": "-"},
    {"ticker": "GILD", "name": "Gilead Sciences", "buffett": ("neutral", 58), "lynch": ("neutral", 60), "fisher": ("neutral", 50), "combined": "-", "conf": "-"},
    {"ticker": "DECK", "name": "Deckers Outdoor", "buffett": ("bullish", 78), "lynch": ("bullish", 75), "fisher": ("bullish", 70), "combined": "강력매수", "conf": "74%"},
    {"ticker": "BMY", "name": "Bristol-Myers Squibb", "buffett": ("neutral", 50), "lynch": ("neutral", 55), "fisher": ("neutral", 45), "combined": "-", "conf": "-"},
    {"ticker": "HIG", "name": "Hartford Financial", "buffett": ("bullish", 82), "lynch": ("bullish", 80), "fisher": ("neutral", 50), "combined": "매수", "conf": "71%"},
    {"ticker": "TPR", "name": "Tapestry", "buffett": ("bearish", 38), "lynch": ("bearish", 35), "fisher": ("neutral", 40), "combined": "-", "conf": "-"},
    {"ticker": "OKE", "name": "ONEOK", "buffett": ("neutral", 60), "lynch": ("neutral", 65), "fisher": ("bearish", 30), "combined": "-", "conf": "-"},
    {"ticker": "PYPL", "name": "PayPal Holdings", "buffett": ("bullish", 72), "lynch": ("bullish", 75), "fisher": ("bullish", 65), "combined": "강력매수", "conf": "71%"},
    {"ticker": "JNJ", "name": "Johnson & Johnson", "buffett": ("bullish", 85), "lynch": ("neutral", 60), "fisher": ("bullish", 75), "combined": "매수", "conf": "73%"},
    {"ticker": "WDC", "name": "Western Digital", "buffett": ("neutral", 48), "lynch": ("bearish", 30), "fisher": ("neutral", 50), "combined": "-", "conf": "-"},
    {"ticker": "USB", "name": "U.S. Bancorp", "buffett": ("bullish", 68), "lynch": ("neutral", 55), "fisher": ("bearish", 30), "combined": "-", "conf": "-"},
    {"ticker": "FAST", "name": "Fastenal Company", "buffett": ("neutral", 55), "lynch": ("bearish", 40), "fisher": ("neutral", 55), "combined": "-", "conf": "-"},
]

investor_details = [
    ("PAYC", "Paycom Software", "W.Buffett", "bullish", 75, "ROE 29%로 탁월하고 gross margin 87%로 뛰어난 수익성, SaaS 모델의 반복 매출과 낮은 부채(D/E 0.05) 보유"),
    ("PAYC", "Paycom Software", "P.Lynch", "bullish", 80, "P/E 16.1, ROE 29%로 GARP 매력적, HR SaaS 고성장(순이익 +47%), 시가총액 $7B로 10배주 잠재력"),
    ("PAYC", "Paycom Software", "P.Fisher", "bullish", 70, "R&D 비율 13%로 HR 소프트웨어 지속 혁신, 86% 매출총이익률과 29% ROE로 우수한 비즈니스 모델"),
    ("DECK", "Deckers Outdoor", "W.Buffett", "bullish", 78, "ROE 40%로 탁월하고 브랜드 파워(UGG, HOKA)가 강하며, 현금 21억으로 부채 3.4억 초과하고 높은 마진 유지"),
    ("DECK", "Deckers Outdoor", "P.Lynch", "bullish", 75, "P/E 16.3, ROE 40%로 GARP 매력적, UGG 브랜드 강력(매출 +16%, 순이익 +27%), 소비재 성장주"),
    ("DECK", "Deckers Outdoor", "P.Fisher", "bullish", 70, "브랜드 파워(UGG, Hoka) 강력하고 ROE 40%로 우수, R&D 없어도 제품 혁신 지속하는 특이 케이스"),
    ("PYPL", "PayPal Holdings", "W.Buffett", "bullish", 72, "ROE 26%로 우수하고 P/E 7.6으로 저평가, 강력한 브랜드와 네트워크 효과 보유"),
    ("PYPL", "PayPal Holdings", "P.Lynch", "bullish", 75, "P/E 7.6으로 극단적 저평가, ROE 26%, 핀테크 터너라운드(순이익 +26%), PEG 0.6 미만으로 GARP 매력"),
    ("PYPL", "PayPal Holdings", "P.Fisher", "bullish", 65, "R&D 비율 9%로 핀테크 기술 투자 지속, ROE 26%와 강력한 현금흐름으로 디지털 결제 시장 리더십"),
    ("ACGL", "Arch Capital Group", "W.Buffett", "bullish", 85, "보험업의 우수한 사례로 ROE 20%이고 P/E 8.3으로 저평가, 일관된 언더라이팅 수익성과 강력한 자본 배분"),
    ("ACGL", "Arch Capital Group", "P.Lynch", "bullish", 85, "P/E 8.3으로 매우 저평가, ROE 20%, 보험주로 boring business, 순이익 +191% 폭증, PEG 1.0 미만 추정"),
    ("ACGL", "Arch Capital Group", "P.Fisher", "neutral", 55, "보험 언더라이팅 우수하고 ROE 20%이나 R&D 없는 금융 비즈니스로 혁신 기반 성장 아님"),
    ("HIG", "Hartford Financial", "W.Buffett", "bullish", 82, "보험업 우수 사례로 ROE 22%이고 P/E 10.5로 저평가, 일관된 수익성과 건전한 자본 구조 및 배당 성장"),
    ("HIG", "Hartford Financial", "P.Lynch", "bullish", 80, "P/E 10.5로 저평가, ROE 22%, 보험 stalwart로 순이익 +24% 성장, FCF 우수, boring business"),
    ("HIG", "Hartford Financial", "P.Fisher", "neutral", 50, "보험 사업 안정적이고 ROE 22% 우수하나 R&D 없고 혁신 주도 성장 아닌 전통 금융 비즈니스"),
    ("TRV", "Travelers Companies", "W.Buffett", "bullish", 80, "보험업의 본질을 이해하며 ROE 21%로 우수하고, P/E 10.7로 저평가되었으며 일관된 언더라이팅 수익성"),
    ("TRV", "Travelers Companies", "P.Lynch", "bullish", 75, "P/E 10.7로 저평가, ROE 21%, 보험 stalwart로 boring business, FCF 수익률 15%로 우수"),
    ("TRV", "Travelers Companies", "P.Fisher", "neutral", 45, "보험 언더라이팅 능력 우수하나 R&D 투자 없고 혁신 주도 성장 아닌 리스크 관리 비즈니스"),
    ("JNJ", "Johnson & Johnson", "W.Buffett", "bullish", 85, "헬스케어 우량주로 ROE 36%이고 일관된 배당 성장, 다각화된 포트폴리오와 강력한 브랜드 moat 보유"),
    ("JNJ", "Johnson & Johnson", "P.Lynch", "neutral", 60, "P/E 21.6, ROE 36%, 헬스케어 stalwart로 안정적이나 성장률 느림, 10배주 불가능"),
    ("JNJ", "Johnson & Johnson", "P.Fisher", "bullish", 75, "R&D 비율 19%로 제약/의료기기 혁신 지속, ROE 36%와 일관된 배당으로 우수한 경영진 품질 입증"),
]


def create_workbook():
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "요약"
    create_summary_sheet(ws_summary)
    ws_portfolio = wb.create_sheet("포트폴리오")
    create_portfolio_sheet(ws_portfolio)
    ws_ranking = wb.create_sheet("순위")
    create_ranking_sheet(ws_ranking)
    ws_matrix = wb.create_sheet("투자자 매트릭스")
    create_matrix_sheet(ws_matrix)
    ws_detail = wb.create_sheet("투자자 상세")
    create_detail_sheet(ws_detail)
    ws_risk = wb.create_sheet("리스크 분석")
    create_risk_sheet(ws_risk)
    return wb


def create_summary_sheet(ws):
    ws.merge_cells('A1:B1')
    ws['A1'] = "AI Hedge Fund 포트폴리오 리포트"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = Alignment(horizontal='center')
    info = [
        ("분석 일자", "2026-02-11"),
        ("분석 대상", "S&P 500 상위 30개"),
        ("분석 전략", "하이브리드 (펀더멘털 70% + 모멘텀 30%)"),
        ("투자자 관점", "W.Buffett, P.Lynch, P.Fisher"),
    ]
    for i, (key, val) in enumerate(info, start=3):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws['A8'] = "포트폴리오 통계"
    ws['A8'].font = Font(bold=True)
    stats = [
        ("편입 종목 수", "7 / 30 (23.3%)"),
        ("평균 신뢰도", "72%"),
        ("평균 예상 수익률", "+12.4%"),
        ("강력매수 비중", "65.0%"),
    ]
    for i, (key, val) in enumerate(stats, start=9):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws['A14'] = "시가총액 분포"
    ws['A14'].font = Font(bold=True)
    cap_dist = [
        ("메가캡 (>$200B)", "16.0% (JNJ)"),
        ("대형주 ($10-200B)", "67.5% (DECK, PYPL, ACGL, HIG, TRV)"),
        ("중형주 ($2-10B)", "16.5% (PAYC)"),
    ]
    for i, (key, val) in enumerate(cap_dist, start=15):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws['A19'] = "섹터 분포"
    ws['A19'].font = Font(bold=True)
    sector_dist = [
        ("Financial (Insurance)", "35.0%"),
        ("Technology", "32.5%"),
        ("Consumer Discretionary", "16.5%"),
        ("Healthcare", "16.0%"),
    ]
    for i, (key, val) in enumerate(sector_dist, start=20):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45


def create_portfolio_sheet(ws):
    headers = ["#", "종목코드", "회사명", "비중", "신호", "신뢰도", "예상수익률", "시가총액", "P/E", "ROE", "PEG", "합의", "섹터"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    for row, item in enumerate(portfolio, 2):
        ws.cell(row=row, column=1, value=row - 1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=item['ticker']).border = THIN_BORDER
        ws.cell(row=row, column=3, value=item['name']).border = THIN_BORDER
        ws.cell(row=row, column=4, value=f"{item['weight']}%").border = THIN_BORDER
        signal_cell = ws.cell(row=row, column=5, value=f"🟢 {item['signal']}" if item['signal'] == '강력매수' else f"🔵 {item['signal']}")
        signal_cell.border = THIN_BORDER
        if item['signal'] == '강력매수':
            signal_cell.fill = BULLISH_FILL
        ws.cell(row=row, column=6, value=f"{item['confidence']}%").border = THIN_BORDER
        ws.cell(row=row, column=7, value=f"+{item['return']}%").border = THIN_BORDER
        ws.cell(row=row, column=8, value=item['cap']).border = THIN_BORDER
        ws.cell(row=row, column=9, value=item['pe']).border = THIN_BORDER
        ws.cell(row=row, column=10, value=f"{item['roe']}%").border = THIN_BORDER
        ws.cell(row=row, column=11, value=item['peg']).border = THIN_BORDER
        ws.cell(row=row, column=12, value=item['consensus']).border = THIN_BORDER
        ws.cell(row=row, column=13, value=item['sector']).border = THIN_BORDER
    total_row = len(portfolio) + 2
    for col in range(1, 14):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = THIN_BORDER
    ws.cell(row=total_row, column=2, value="합계").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value="100.0%").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value="avg 72%")
    ws.cell(row=total_row, column=7, value="avg +12.4%")
    widths = [5, 12, 22, 8, 14, 10, 12, 10, 8, 8, 8, 8, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:M{len(portfolio) + 1}"
    ws.freeze_panes = "A2"


def create_ranking_sheet(ws):
    headers = ["순위", "종목코드", "회사명", "종합점수", "펀더멘털", "모멘텀", "앙상블", "신호", "예상수익률", "시가총액", "P/E", "P/B", "ROE", "매출성장률", "PEG", "편입여부"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    for row, item in enumerate(rankings, 2):
        ws.cell(row=row, column=1, value=item['rank']).border = THIN_BORDER
        ws.cell(row=row, column=2, value=item['ticker']).border = THIN_BORDER
        ws.cell(row=row, column=3, value=item['name']).border = THIN_BORDER
        ws.cell(row=row, column=4, value=item['total']).border = THIN_BORDER
        ws.cell(row=row, column=5, value=item['fund']).border = THIN_BORDER
        ws.cell(row=row, column=6, value=item['mom']).border = THIN_BORDER
        ws.cell(row=row, column=7, value=item['ens']).border = THIN_BORDER
        ws.cell(row=row, column=8, value=item['signal']).border = THIN_BORDER
        ws.cell(row=row, column=9, value=f"+{item['return']}%").border = THIN_BORDER
        ws.cell(row=row, column=10, value=item['cap']).border = THIN_BORDER
        ws.cell(row=row, column=11, value=item['pe']).border = THIN_BORDER
        ws.cell(row=row, column=12, value=item['pb']).border = THIN_BORDER
        roe_val = f"{item['roe']}%" if item['roe'] != "N/A" else "N/A"
        ws.cell(row=row, column=13, value=roe_val).border = THIN_BORDER
        ws.cell(row=row, column=14, value=item['rev_growth']).border = THIN_BORDER
        ws.cell(row=row, column=15, value=item['peg']).border = THIN_BORDER
        included_cell = ws.cell(row=row, column=16, value="예" if item['included'] else "아니오")
        included_cell.border = THIN_BORDER
        if item['included']:
            for col_idx in range(1, 17):
                ws.cell(row=row, column=col_idx).fill = BULLISH_FILL
    widths = [6, 10, 22, 10, 10, 8, 8, 10, 12, 10, 8, 8, 8, 12, 8, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:P{len(rankings) + 1}"
    ws.freeze_panes = "A2"


def create_matrix_sheet(ws):
    headers = ["종목코드", "회사명", "W.Buffett", "P.Lynch", "P.Fisher", "종합신호", "종합신뢰도"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    for row, item in enumerate(investor_matrix, 2):
        ws.cell(row=row, column=1, value=item['ticker']).border = THIN_BORDER
        ws.cell(row=row, column=2, value=item['name']).border = THIN_BORDER
        for col, inv in enumerate(['buffett', 'lynch', 'fisher'], 3):
            signal, conf = item[inv]
            display = {"bullish": "매수", "bearish": "매도", "neutral": "중립"}[signal]
            cell = ws.cell(row=row, column=col, value=f"{display}({conf})")
            cell.border = THIN_BORDER
            if signal == 'bullish':
                cell.fill = BULLISH_FILL
                cell.font = BULLISH_FONT
            elif signal == 'bearish':
                cell.fill = BEARISH_FILL
                cell.font = BEARISH_FONT
            else:
                cell.fill = NEUTRAL_FILL
                cell.font = NEUTRAL_FONT
        ws.cell(row=row, column=6, value=item['combined']).border = THIN_BORDER
        ws.cell(row=row, column=7, value=item['conf']).border = THIN_BORDER
    widths = [10, 22, 15, 15, 15, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:G{len(investor_matrix) + 1}"
    ws.freeze_panes = "A2"


def create_detail_sheet(ws):
    headers = ["종목코드", "회사명", "투자자", "신호", "신뢰도", "분석근거"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    for row, (ticker, name, investor, signal, conf, reasoning) in enumerate(investor_details, 2):
        ws.cell(row=row, column=1, value=ticker).border = THIN_BORDER
        ws.cell(row=row, column=2, value=name).border = THIN_BORDER
        ws.cell(row=row, column=3, value=investor).border = THIN_BORDER
        display = {"bullish": "매수", "bearish": "매도", "neutral": "중립"}[signal]
        signal_cell = ws.cell(row=row, column=4, value=display)
        signal_cell.border = THIN_BORDER
        if signal == 'bullish':
            signal_cell.fill = BULLISH_FILL
            signal_cell.font = BULLISH_FONT
        elif signal == 'bearish':
            signal_cell.fill = BEARISH_FILL
            signal_cell.font = BEARISH_FONT
        else:
            signal_cell.fill = NEUTRAL_FILL
            signal_cell.font = NEUTRAL_FONT
        ws.cell(row=row, column=5, value=conf).border = THIN_BORDER
        ws.cell(row=row, column=6, value=reasoning).border = THIN_BORDER
    widths = [10, 22, 12, 10, 10, 70]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def create_risk_sheet(ws):
    ws.merge_cells('A1:B1')
    ws['A1'] = "리스크 분석"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A3'] = "집중도 지표"
    ws['A3'].font = Font(bold=True)
    concentration = [
        ("상위 1종목 비중", "16.5% (PAYC)"),
        ("상위 3종목 비중", "49.0%"),
        ("상위 5종목 비중", "77.5%"),
        ("HHI (허핀달 지수)", "0.146"),
    ]
    for i, (key, val) in enumerate(concentration, 4):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws['A9'] = "섹터 집중도"
    ws['A9'].font = Font(bold=True)
    sector_conc = [
        ("최대 섹터", "Financial (35.0%)"),
        ("상위 2섹터 비중", "67.5%"),
    ]
    for i, (key, val) in enumerate(sector_conc, 10):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws['A13'] = "투자자 합의 품질"
    ws['A13'].font = Font(bold=True)
    consensus = [
        ("만장일치 비율", "42.9% (3개: PAYC, DECK, PYPL)"),
        ("다수 합의", "57.1% (4개: ACGL, HIG, TRV, JNJ)"),
        ("의견 분산 종목", "0개"),
    ]
    for i, (key, val) in enumerate(consensus, 14):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws['A18'] = "비편입 사유 분포"
    ws['A18'].font = Font(bold=True)
    exclusion = [
        ("투자자 과반 미달 (0/3)", "15개"),
        ("투자자 과반 미달 (1/3)", "8개"),
    ]
    for i, (key, val) in enumerate(exclusion, 19):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws['A22'] = "제약사항"
    ws['A22'].font = Font(bold=True)
    constraints = [
        ("개별종목 최대 비중", "16.5% (15% 소폭 초과, 7종목 제약)"),
        ("섹터 비중 제한", "Financial 35.0% (제한선 도달)"),
    ]
    for i, (key, val) in enumerate(constraints, 23):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45


if __name__ == "__main__":
    wb = create_workbook()
    output_path = "/Users/sgkim/Projects/ai-hedge-fund/portfolios/sp500_20260211_buffett_fisher_lynch.xlsx"
    wb.save(output_path)
    print(f"✅ 엑셀 리포트 생성 완료: {output_path}")
