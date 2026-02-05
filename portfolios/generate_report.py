#!/usr/bin/env python3
"""S&P 500 Portfolio Report Generator - 2026-02-05"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference

# 스타일 정의
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
BULLISH_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BULLISH_FONT = Font(color="375623")
BEARISH_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
BEARISH_FONT = Font(color="C62828")
NEUTRAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
NEUTRAL_FONT = Font(color="616161")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# 포트폴리오 데이터
portfolio = [
    {"ticker": "MPWR", "name": "Monolithic Power", "weight": 15.0, "signal": "강력매수", "confidence": 87, "return": 10.8, "cap": "$35B", "pe": 75.9, "roe": 21, "peg": "N/A", "consensus": "3/3", "sector": "Technology"},
    {"ticker": "PAYC", "name": "Paycom Software", "weight": 15.0, "signal": "강력매수", "confidence": 82, "return": 9.6, "cap": "$9.9B", "pe": 21.7, "roe": 23, "peg": "N/A", "consensus": "3/3", "sector": "Technology"},
    {"ticker": "GOOGL", "name": "Alphabet Inc", "weight": 15.0, "signal": "강력매수", "confidence": 90, "return": 9.6, "cap": "$265B", "pe": 21.9, "roe": 33, "peg": 1.5, "consensus": "2/3", "sector": "Communication"},
    {"ticker": "UHS", "name": "Universal Health", "weight": 14.6, "signal": "매수", "confidence": 82, "return": 9.5, "cap": "$19B", "pe": 22.5, "roe": 15, "peg": 0.22, "consensus": "1/3*", "sector": "Healthcare"},
    {"ticker": "JNJ", "name": "Johnson & Johnson", "weight": 14.0, "signal": "매수", "confidence": 84, "return": 10.0, "cap": "$307B", "pe": 13.8, "roe": 22, "peg": "N/A", "consensus": "2/3", "sector": "Healthcare"},
    {"ticker": "LRCX", "name": "Lam Research", "weight": 13.3, "signal": "매수", "confidence": 80, "return": 10.0, "cap": "$218B", "pe": 20.3, "roe": 65, "peg": 0.3, "consensus": "2/3", "sector": "Technology"},
    {"ticker": "TER", "name": "Teradyne", "weight": 13.1, "signal": "매수", "confidence": 79, "return": 14.5, "cap": "$16B", "pe": 81.0, "roe": 12, "peg": 1.07, "consensus": "2/3", "sector": "Semiconductors"},
]

# 순위 데이터 (상위 30개)
rankings = [
    {"rank": 1, "ticker": "CF", "name": "CF Industries", "total": 8.79, "fund": 6.0, "mom": 9.0, "ens": 5.8, "signal": "강력매수", "return": 20.3, "cap": "$18B", "pe": 13.6, "pb": 2.8, "roe": 19, "rev_growth": 21, "peg": 0.4, "included": False},
    {"rank": 2, "ticker": "TER", "name": "Teradyne", "total": 7.13, "fund": 5.5, "mom": 8.0, "ens": 5.2, "signal": "강력매수", "return": 14.5, "cap": "$16B", "pe": 81.0, "pb": 8.5, "roe": 12, "rev_growth": 44, "peg": 1.07, "included": True},
    {"rank": 3, "ticker": "MU", "name": "Micron Technology", "total": 6.90, "fund": 5.8, "mom": 7.5, "ens": 5.4, "signal": "매수", "return": 13.6, "cap": "$427B", "pe": 16.7, "pb": 2.8, "roe": 13, "rev_growth": 57, "peg": 0.22, "included": False},
    {"rank": 4, "ticker": "CTRA", "name": "Coterra Energy", "total": 6.61, "fund": 5.2, "mom": 7.0, "ens": 5.0, "signal": "매수", "return": 12.6, "cap": "$21B", "pe": 14.9, "pb": 2.0, "roe": 16, "rev_growth": 35, "peg": 0.48, "included": False},
    {"rank": 5, "ticker": "OKE", "name": "ONEOK", "total": 6.53, "fund": 5.0, "mom": 7.5, "ens": 4.8, "signal": "매수", "return": 12.4, "cap": "$73B", "pe": 23.6, "pb": 4.2, "roe": 29, "rev_growth": 72, "peg": 0.41, "included": False},
    {"rank": 6, "ticker": "MOS", "name": "Mosaic Company", "total": 6.50, "fund": 6.5, "mom": 6.0, "ens": 5.5, "signal": "매수", "return": 12.3, "cap": "$7.6B", "pe": 8.9, "pb": 0.7, "roe": 5, "rev_growth": 237, "peg": 0.03, "included": False},
    {"rank": 7, "ticker": "NEM", "name": "Newmont Corp", "total": 6.27, "fund": 4.8, "mom": 7.0, "ens": 4.5, "signal": "매수", "return": 11.4, "cap": "$59B", "pe": 15.7, "pb": 1.8, "roe": 8, "rev_growth": 20, "peg": 0.18, "included": False},
    {"rank": 8, "ticker": "LMT", "name": "Lockheed Martin", "total": 6.27, "fund": 5.0, "mom": 6.5, "ens": 4.8, "signal": "매수", "return": 11.4, "cap": "$135B", "pe": 22.6, "pb": 17.0, "roe": 86, "rev_growth": 9, "peg": 0.19, "included": False},
    {"rank": 9, "ticker": "FANG", "name": "Diamondback Energy", "total": 6.20, "fund": 5.5, "mom": 6.0, "ens": 5.0, "signal": "매수", "return": 11.2, "cap": "$50B", "pe": 12.6, "pb": 1.8, "roe": 20, "rev_growth": 42, "peg": 0.21, "included": False},
    {"rank": 10, "ticker": "ADI", "name": "Analog Devices", "total": 6.17, "fund": 4.5, "mom": 7.5, "ens": 4.2, "signal": "매수", "return": 11.1, "cap": "$95B", "pe": 59.7, "pb": 3.2, "roe": 5, "rev_growth": 26, "peg": 1.06, "included": False},
    {"rank": 11, "ticker": "MPWR", "name": "Monolithic Power", "total": 6.08, "fund": 5.2, "mom": 6.5, "ens": 5.0, "signal": "매수", "return": 10.8, "cap": "$35B", "pe": 75.9, "pb": 16.0, "roe": 21, "rev_growth": 19, "peg": "N/A", "included": True},
    {"rank": 12, "ticker": "HIG", "name": "Hartford Financial", "total": 5.97, "fund": 5.5, "mom": 5.5, "ens": 5.2, "signal": "매수", "return": 10.4, "cap": "$46B", "pe": 14.3, "pb": 2.5, "roe": 17, "rev_growth": 7, "peg": "N/A", "included": False},
    {"rank": 13, "ticker": "GL", "name": "Globe Life", "total": 5.95, "fund": 5.8, "mom": 5.0, "ens": 5.5, "signal": "매수", "return": 10.3, "cap": "$15B", "pe": 22.5, "pb": 2.8, "roe": 13, "rev_growth": 4, "peg": "N/A", "included": False},
    {"rank": 14, "ticker": "GLW", "name": "Corning Inc", "total": 5.94, "fund": 4.2, "mom": 7.0, "ens": 4.0, "signal": "매수", "return": 10.3, "cap": "$56B", "pe": 83.1, "pb": 2.8, "roe": 6, "rev_growth": 7, "peg": 1.1, "included": False},
    {"rank": 15, "ticker": "PAYX", "name": "Paychex", "total": 5.89, "fund": 5.5, "mom": 5.5, "ens": 5.2, "signal": "매수", "return": 10.1, "cap": "$53B", "pe": 29.7, "pb": 12.0, "roe": 46, "rev_growth": 5, "peg": "N/A", "included": False},
    {"rank": 16, "ticker": "ENPH", "name": "Enphase Energy", "total": 5.89, "fund": 4.0, "mom": 7.5, "ens": 3.8, "signal": "매수", "return": 10.1, "cap": "$9.1B", "pe": 76.5, "pb": 6.0, "roe": 20, "rev_growth": 8, "peg": "N/A", "included": False},
    {"rank": 17, "ticker": "JNJ", "name": "Johnson & Johnson", "total": 5.86, "fund": 5.5, "mom": 5.5, "ens": 5.2, "signal": "매수", "return": 10.0, "cap": "$307B", "pe": 13.8, "pb": 5.0, "roe": 22, "rev_growth": 9, "peg": "N/A", "included": True},
    {"rank": 18, "ticker": "LRCX", "name": "Lam Research", "total": 5.85, "fund": 5.0, "mom": 6.0, "ens": 4.8, "signal": "매수", "return": 10.0, "cap": "$218B", "pe": 20.3, "pb": 13.0, "roe": 65, "rev_growth": 22, "peg": 0.3, "included": True},
    {"rank": 19, "ticker": "DVN", "name": "Devon Energy", "total": 5.79, "fund": 5.5, "mom": 5.5, "ens": 5.0, "signal": "매수", "return": 9.8, "cap": "$28B", "pe": 10.3, "pb": 2.0, "roe": 19, "rev_growth": -15, "peg": "N/A", "included": False},
    {"rank": 20, "ticker": "FMC", "name": "FMC Corporation", "total": 5.78, "fund": 5.0, "mom": 6.0, "ens": 4.5, "signal": "매수", "return": 9.7, "cap": "$7.9B", "pe": 11.1, "pb": 1.5, "roe": 16, "rev_growth": -49, "peg": "N/A", "included": False},
    {"rank": 21, "ticker": "PAYC", "name": "Paycom Software", "total": 5.74, "fund": 5.2, "mom": 5.5, "ens": 5.0, "signal": "매수", "return": 9.6, "cap": "$9.9B", "pe": 21.7, "pb": 5.0, "roe": 23, "rev_growth": 9, "peg": "N/A", "included": True},
    {"rank": 22, "ticker": "GOOGL", "name": "Alphabet Inc", "total": 5.74, "fund": 5.5, "mom": 5.0, "ens": 5.2, "signal": "매수", "return": 9.6, "cap": "$265B", "pe": 21.9, "pb": 7.0, "roe": 33, "rev_growth": 18, "peg": 1.5, "included": True},
    {"rank": 23, "ticker": "CAT", "name": "Caterpillar", "total": 5.73, "fund": 5.0, "mom": 5.5, "ens": 4.8, "signal": "매수", "return": 9.6, "cap": "$186B", "pe": 18.2, "pb": 11.0, "roe": 60, "rev_growth": 6, "peg": "N/A", "included": False},
    {"rank": 24, "ticker": "UHS", "name": "Universal Health", "total": 5.73, "fund": 5.5, "mom": 5.0, "ens": 5.2, "signal": "매수", "return": 9.5, "cap": "$19B", "pe": 22.5, "pb": 2.8, "roe": 15, "rev_growth": 44, "peg": 0.22, "included": True},
    {"rank": 25, "ticker": "TRV", "name": "Travelers", "total": 5.72, "fund": 5.5, "mom": 5.0, "ens": 5.2, "signal": "매수", "return": 9.5, "cap": "$66B", "pe": 14.1, "pb": 2.1, "roe": 17, "rev_growth": 7, "peg": "N/A", "included": False},
    {"rank": 26, "ticker": "APA", "name": "APA Corporation", "total": 5.69, "fund": 4.5, "mom": 6.5, "ens": 4.2, "signal": "매수", "return": 9.4, "cap": "$8.9B", "pe": 33.7, "pb": 1.5, "roe": 7, "rev_growth": -16, "peg": "N/A", "included": False},
    {"rank": 27, "ticker": "CVX", "name": "Chevron", "total": 5.68, "fund": 5.0, "mom": 5.5, "ens": 4.8, "signal": "매수", "return": 9.4, "cap": "$327B", "pe": 18.7, "pb": 2.0, "roe": 12, "rev_growth": -6, "peg": "N/A", "included": False},
    {"rank": 28, "ticker": "KMI", "name": "Kinder Morgan", "total": 5.67, "fund": 5.5, "mom": 5.0, "ens": 5.2, "signal": "매수", "return": 9.3, "cap": "$64B", "pe": 23.9, "pb": 2.2, "roe": 9, "rev_growth": 49, "peg": "N/A", "included": False},
    {"rank": 29, "ticker": "VIAV", "name": "Viavi Solutions", "total": 5.64, "fund": 4.5, "mom": 6.0, "ens": 4.2, "signal": "매수", "return": 9.2, "cap": "$2.6B", "pe": -83.7, "pb": 1.5, "roe": -8, "rev_growth": 17, "peg": "N/A", "included": False},
    {"rank": 30, "ticker": "EOG", "name": "EOG Resources", "total": 5.64, "fund": 5.2, "mom": 5.5, "ens": 5.0, "signal": "매수", "return": 9.2, "cap": "$66B", "pe": 13.7, "pb": 3.0, "roe": 20, "rev_growth": -12, "peg": "N/A", "included": False},
]

# 투자자 매트릭스 데이터
investor_matrix = [
    {"ticker": "MPWR", "name": "Monolithic Power", "buffett": ("bullish", 92), "lynch": ("bullish", 82), "fisher": ("bullish", 88), "combined": "강력매수", "conf": 87},
    {"ticker": "PAYC", "name": "Paycom Software", "buffett": ("bullish", 82), "lynch": ("bullish", 78), "fisher": ("bullish", 85), "combined": "강력매수", "conf": 82},
    {"ticker": "GOOGL", "name": "Alphabet Inc", "buffett": ("bullish", 88), "lynch": ("neutral", 52), "fisher": ("bullish", 92), "combined": "강력매수", "conf": 90},
    {"ticker": "JNJ", "name": "Johnson & Johnson", "buffett": ("bullish", 90), "lynch": ("neutral", 60), "fisher": ("bullish", 78), "combined": "매수", "conf": 84},
    {"ticker": "LRCX", "name": "Lam Research", "buffett": ("neutral", 52), "lynch": ("bullish", 78), "fisher": ("bullish", 82), "combined": "매수", "conf": 80},
    {"ticker": "TER", "name": "Teradyne", "buffett": ("neutral", 48), "lynch": ("bullish", 75), "fisher": ("bullish", 82), "combined": "매수", "conf": 79},
    {"ticker": "UHS", "name": "Universal Health", "buffett": ("neutral", 58), "lynch": ("bullish", 82), "fisher": ("neutral", 50), "combined": "매수", "conf": 82},
]

# 투자자 상세 분석
investor_details = [
    ("MPWR", "Monolithic Power", "W.Buffett", "bullish", 92, "64% ROE와 무부채(D/E 0.005)로 탁월한 해자. 전력 반도체 프랜차이즈"),
    ("MPWR", "Monolithic Power", "P.Lynch", "bullish", 82, "Fast Grower (EPS 23.4%, 매출 18.9%), AI/전력 반도체 수혜"),
    ("MPWR", "Monolithic Power", "P.Fisher", "bullish", 88, "R&D 14.7% 투자, 5년 CAGR 16.3%, Fisher 기술 혁신 기업"),
    ("PAYC", "Paycom Software", "W.Buffett", "bullish", 82, "ROE 28.6%, D/E 0.05 거의 무부채. 급여 소프트웨어 전환 비용 moat"),
    ("PAYC", "Paycom Software", "P.Lynch", "bullish", 78, "PER 15.5 저평가, EPS 51% 성장, 시총 $7.3B 10배 잠재력"),
    ("PAYC", "Paycom Software", "P.Fisher", "bullish", 85, "R&D 12.9% 투자, FCF 전환율 57.6%, HR 기술 혁신"),
    ("GOOGL", "Alphabet Inc", "W.Buffett", "bullish", 88, "검색 독점 + 네트워크 효과 최강 moat, ROE 35.5%, FCF $73B"),
    ("GOOGL", "Alphabet Inc", "P.Lynch", "neutral", 52, "메가캡 $4T로 10배 성장 불가능, PER 33.6 고평가"),
    ("GOOGL", "Alphabet Inc", "P.Fisher", "bullish", 92, "R&D 14.1% 최고 수준, AI 기술 리더십, Fisher 완벽 부합"),
    ("JNJ", "Johnson & Johnson", "W.Buffett", "bullish", 90, "헬스케어 브랜드 moat, ROE 36%, 연간 FCF $18B+ 복리 기계"),
    ("JNJ", "Johnson & Johnson", "P.Lynch", "neutral", 60, "Stalwart 초대형주, 안정적이나 10배주 불가능"),
    ("JNJ", "Johnson & Johnson", "P.Fisher", "bullish", 78, "R&D 19.4% 투자, 혁신 파이프라인 강력, 장기 성장 지속"),
    ("LRCX", "Lam Research", "W.Buffett", "neutral", 52, "ROE 66% 탁월하나 고주기적, P/E 47배 확장"),
    ("LRCX", "Lam Research", "P.Lynch", "bullish", 78, "Fast Grower, EPS 33.8%, Forward PEG < 0.3 저평가"),
    ("LRCX", "Lam Research", "P.Fisher", "bullish", 82, "R&D 11.4%, 첨단공정 반도체 장비 선도, 장기 성장성"),
    ("TER", "Teradyne", "W.Buffett", "neutral", 48, "반도체 테스트 리더이나 P/E 81배 과대평가, 안전마진 없음"),
    ("TER", "Teradyne", "P.Lynch", "bullish", 75, "AI 반도체 테스트 수요 Fast Grower, PEG 1.07 적정가"),
    ("TER", "Teradyne", "P.Fisher", "bullish", 82, "R&D 16.4% 투자, 영업이익률 28% 유지, 장기 성장 가시성"),
    ("UHS", "Universal Health", "W.Buffett", "neutral", 58, "ROE 20% 양호, 낮은 마진 8% 부담, 규제 리스크"),
    ("UHS", "Universal Health", "P.Lynch", "bullish", 82, "PER 9.85 초저평가, EPS 44.2% 고성장, 추정 PEG 0.22"),
    ("UHS", "Universal Health", "P.Fisher", "neutral", 50, "R&D 투자 없음, 혁신 기반 성장주로 부적합"),
]


def create_workbook():
    wb = Workbook()

    # 시트 1: 요약
    ws_summary = wb.active
    ws_summary.title = "요약"
    create_summary_sheet(ws_summary)

    # 시트 2: 포트폴리오
    ws_portfolio = wb.create_sheet("포트폴리오")
    create_portfolio_sheet(ws_portfolio)

    # 시트 3: 순위
    ws_ranking = wb.create_sheet("순위")
    create_ranking_sheet(ws_ranking)

    # 시트 4: 투자자 매트릭스
    ws_matrix = wb.create_sheet("투자자 매트릭스")
    create_matrix_sheet(ws_matrix)

    # 시트 5: 투자자 상세
    ws_detail = wb.create_sheet("투자자 상세")
    create_detail_sheet(ws_detail)

    # 시트 6: 리스크 분석
    ws_risk = wb.create_sheet("리스크 분석")
    create_risk_sheet(ws_risk)

    return wb


def create_summary_sheet(ws):
    # 제목
    ws.merge_cells('A1:B1')
    ws['A1'] = "AI Hedge Fund 포트폴리오 리포트"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = Alignment(horizontal='center')

    # 기본 정보
    info = [
        ("분석 일자", "2026-02-05"),
        ("분석 대상", "S&P 500 상위 30개"),
        ("분석 전략", "하이브리드"),
        ("투자자 관점", "W.Buffett, P.Lynch, P.Fisher"),
    ]
    for i, (key, val) in enumerate(info, start=3):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val

    # 포트폴리오 통계
    ws['A8'] = "포트폴리오 통계"
    ws['A8'].font = Font(bold=True)
    stats = [
        ("편입 종목 수", "7 / 30 (23.3%)"),
        ("평균 신뢰도", "83%"),
        ("평균 예상 수익률", "+10.6%"),
        ("강력매수 비중", "45.0%"),
    ]
    for i, (key, val) in enumerate(stats, start=9):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val

    # 시가총액 분포
    ws['A14'] = "시가총액 분포"
    ws['A14'].font = Font(bold=True)
    cap_dist = [
        ("메가캡 (>$200B)", "42.3%"),
        ("대형주 ($10-200B)", "42.7%"),
        ("중형주 ($2-10B)", "15.0%"),
    ]
    for i, (key, val) in enumerate(cap_dist, start=15):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val

    # 섹터 분포
    ws['A19'] = "섹터 분포"
    ws['A19'].font = Font(bold=True)
    sector_dist = [
        ("Technology", "41.4%"),
        ("Healthcare", "28.6%"),
        ("Communication", "15.0%"),
        ("Semiconductors", "13.1%"),
    ]
    for i, (key, val) in enumerate(sector_dist, start=20):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val

    # 열 너비 조정
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30


def create_portfolio_sheet(ws):
    headers = ["#", "종목코드", "회사명", "비중", "신호", "신뢰도", "예상수익률", "시가총액", "P/E", "ROE", "PEG", "합의", "섹터"]

    # 헤더 작성
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # 데이터 작성
    for row, item in enumerate(portfolio, 2):
        ws.cell(row=row, column=1, value=row-1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=item['ticker']).border = THIN_BORDER
        ws.cell(row=row, column=3, value=item['name']).border = THIN_BORDER
        ws.cell(row=row, column=4, value=f"{item['weight']}%").border = THIN_BORDER

        signal_cell = ws.cell(row=row, column=5, value=f"🟢 {item['signal']}" if item['signal'] == '강력매수' else f"🔵 {item['signal']}")
        signal_cell.border = THIN_BORDER
        if item['signal'] == '강력매수':
            signal_cell.fill = BULLISH_FILL

        ws.cell(row=row, column=6, value=f"{item['confidence']}%").border = THIN_BORDER
        ws.cell(row=row, column=7, value=f"+{item['return']}%").border = THIN_BORDER
        ws.cell(row=row, column=8, value=item['cap']).border = THIN_BORDER
        ws.cell(row=row, column=9, value=item['pe']).border = THIN_BORDER
        ws.cell(row=row, column=10, value=f"{item['roe']}%").border = THIN_BORDER
        ws.cell(row=row, column=11, value=item['peg']).border = THIN_BORDER
        ws.cell(row=row, column=12, value=item['consensus']).border = THIN_BORDER
        ws.cell(row=row, column=13, value=item['sector']).border = THIN_BORDER

    # 합계 행
    total_row = len(portfolio) + 2
    ws.cell(row=total_row, column=1, value="").fill = TOTAL_FILL
    ws.cell(row=total_row, column=2, value="합계").font = Font(bold=True)
    ws.cell(row=total_row, column=2).fill = TOTAL_FILL
    ws.cell(row=total_row, column=3, value="").fill = TOTAL_FILL
    ws.cell(row=total_row, column=4, value="100.0%").font = Font(bold=True)
    ws.cell(row=total_row, column=4).fill = TOTAL_FILL
    ws.cell(row=total_row, column=5, value="").fill = TOTAL_FILL
    ws.cell(row=total_row, column=6, value="avg 83%").fill = TOTAL_FILL
    ws.cell(row=total_row, column=7, value="avg +10.6%").fill = TOTAL_FILL
    for col in range(8, 14):
        ws.cell(row=total_row, column=col, value="").fill = TOTAL_FILL

    # 열 너비 조정
    widths = [5, 12, 20, 8, 12, 10, 12, 10, 8, 8, 8, 8, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 필터 및 틀 고정
    ws.auto_filter.ref = f"A1:M{len(portfolio)+1}"
    ws.freeze_panes = "A2"


def create_ranking_sheet(ws):
    headers = ["순위", "종목코드", "회사명", "종합점수", "펀더멘털", "모멘텀", "앙상블", "신호", "예상수익률", "시가총액", "P/E", "P/B", "ROE", "매출성장률", "PEG", "편입여부"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for row, item in enumerate(rankings, 2):
        ws.cell(row=row, column=1, value=item['rank']).border = THIN_BORDER
        ws.cell(row=row, column=2, value=item['ticker']).border = THIN_BORDER
        ws.cell(row=row, column=3, value=item['name']).border = THIN_BORDER
        ws.cell(row=row, column=4, value=item['total']).border = THIN_BORDER
        ws.cell(row=row, column=5, value=item['fund']).border = THIN_BORDER
        ws.cell(row=row, column=6, value=item['mom']).border = THIN_BORDER
        ws.cell(row=row, column=7, value=item['ens']).border = THIN_BORDER
        ws.cell(row=row, column=8, value=item['signal']).border = THIN_BORDER
        ws.cell(row=row, column=9, value=f"+{item['return']}%").border = THIN_BORDER
        ws.cell(row=row, column=10, value=item['cap']).border = THIN_BORDER
        ws.cell(row=row, column=11, value=item['pe']).border = THIN_BORDER
        ws.cell(row=row, column=12, value=item['pb']).border = THIN_BORDER
        ws.cell(row=row, column=13, value=f"{item['roe']}%").border = THIN_BORDER
        ws.cell(row=row, column=14, value=f"{item['rev_growth']}%").border = THIN_BORDER
        ws.cell(row=row, column=15, value=item['peg']).border = THIN_BORDER

        included_cell = ws.cell(row=row, column=16, value="예" if item['included'] else "아니오")
        included_cell.border = THIN_BORDER
        if item['included']:
            for col in range(1, 17):
                ws.cell(row=row, column=col).fill = BULLISH_FILL

    # 열 너비 조정
    widths = [6, 10, 18, 10, 10, 8, 8, 10, 12, 10, 8, 8, 8, 12, 8, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A1:P{len(rankings)+1}"
    ws.freeze_panes = "A2"


def create_matrix_sheet(ws):
    headers = ["종목코드", "회사명", "W.Buffett", "P.Lynch", "P.Fisher", "종합신호", "종합신뢰도"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for row, item in enumerate(investor_matrix, 2):
        ws.cell(row=row, column=1, value=item['ticker']).border = THIN_BORDER
        ws.cell(row=row, column=2, value=item['name']).border = THIN_BORDER

        # 투자자별 신호
        for col, inv in enumerate(['buffett', 'lynch', 'fisher'], 3):
            signal, conf = item[inv]
            cell = ws.cell(row=row, column=col, value=f"{signal}({conf})")
            cell.border = THIN_BORDER
            if signal == 'bullish':
                cell.fill = BULLISH_FILL
                cell.font = BULLISH_FONT
            elif signal == 'bearish':
                cell.fill = BEARISH_FILL
                cell.font = BEARISH_FONT
            else:
                cell.fill = NEUTRAL_FILL
                cell.font = NEUTRAL_FONT

        ws.cell(row=row, column=6, value=item['combined']).border = THIN_BORDER
        ws.cell(row=row, column=7, value=f"{item['conf']}%").border = THIN_BORDER

    widths = [10, 20, 15, 15, 15, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A1:G{len(investor_matrix)+1}"
    ws.freeze_panes = "A2"


def create_detail_sheet(ws):
    headers = ["종목코드", "회사명", "투자자", "신호", "신뢰도", "분석근거"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for row, (ticker, name, investor, signal, conf, reasoning) in enumerate(investor_details, 2):
        ws.cell(row=row, column=1, value=ticker).border = THIN_BORDER
        ws.cell(row=row, column=2, value=name).border = THIN_BORDER
        ws.cell(row=row, column=3, value=investor).border = THIN_BORDER

        signal_cell = ws.cell(row=row, column=4, value=signal)
        signal_cell.border = THIN_BORDER
        if signal == 'bullish':
            signal_cell.fill = BULLISH_FILL
            signal_cell.font = BULLISH_FONT
        elif signal == 'bearish':
            signal_cell.fill = BEARISH_FILL
            signal_cell.font = BEARISH_FONT
        else:
            signal_cell.fill = NEUTRAL_FILL
            signal_cell.font = NEUTRAL_FONT

        ws.cell(row=row, column=5, value=conf).border = THIN_BORDER
        ws.cell(row=row, column=6, value=reasoning).border = THIN_BORDER

    widths = [10, 20, 12, 10, 10, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def create_risk_sheet(ws):
    ws.merge_cells('A1:B1')
    ws['A1'] = "리스크 분석"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL

    ws['A3'] = "집중도 지표"
    ws['A3'].font = Font(bold=True)
    concentration = [
        ("상위 1종목 비중", "15.0%"),
        ("상위 3종목 비중", "45.0%"),
        ("상위 5종목 비중", "72.3%"),
        ("HHI (허핀달 지수)", "0.147"),
    ]
    for i, (key, val) in enumerate(concentration, 4):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val

    ws['A9'] = "섹터 집중도"
    ws['A9'].font = Font(bold=True)
    sector_conc = [
        ("최대 섹터", "Technology (41.4%)"),
        ("상위 3섹터 비중", "85.1%"),
    ]
    for i, (key, val) in enumerate(sector_conc, 10):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val

    ws['A13'] = "투자자 합의 품질"
    ws['A13'].font = Font(bold=True)
    consensus = [
        ("만장일치 비율", "28.6% (2개)"),
        ("의견 분산 종목 수", "0개"),
    ]
    for i, (key, val) in enumerate(consensus, 14):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val

    ws['A17'] = "비편입 사유 분포"
    ws['A17'].font = Font(bold=True)
    exclusion = [
        ("투자자 과반 미달", "23개"),
        ("최소 비중 미달", "0개"),
    ]
    for i, (key, val) in enumerate(exclusion, 18):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25


if __name__ == "__main__":
    wb = create_workbook()
    output_path = "/home/ubuntu/projects/ai-hedge-fund/portfolios/sp500_20260205_buffett_fisher_lynch.xlsx"
    wb.save(output_path)
    print(f"✅ 엑셀 리포트 생성 완료: {output_path}")
