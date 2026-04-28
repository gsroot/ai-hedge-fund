#!/usr/bin/env python3
"""Generate KRX portfolio report from the latest predict JSON."""

from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
PREDICT_JSON = Path("/tmp/krx_analysis_20260424.json")
OUTPUT_XLSX = ROOT / "portfolios" / "krx_20260424_buffett_fisher_lynch.xlsx"
OUTPUT_TXT = ROOT / "portfolios" / "krx_20260424_buffett_fisher_lynch.txt"
TOP_N = 30
INDEX_NAME = "krx"
INVESTORS = ["buffett", "lynch", "fisher"]

sys.path.insert(0, str(ROOT / ".claude" / "skills" / "predict" / "scripts"))
from factor_scoring import get_market_cap_category  # noqa: E402
from korean_data_fetcher import get_market_cap_kr  # noqa: E402


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="2E75B6")
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
BUY_FILL = PatternFill("solid", fgColor="E2EFDA")
BUY_FONT = Font(name="맑은 고딕", color="375623")
SELL_FILL = PatternFill("solid", fgColor="FCE4EC")
SELL_FONT = Font(name="맑은 고딕", color="C62828")
NEUTRAL_FILL = PatternFill("solid", fgColor="F5F5F5")
NEUTRAL_FONT = Font(name="맑은 고딕", color="616161")
TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

INVESTOR_LABELS = {
    "buffett": "Warren Buffett",
    "lynch": "Peter Lynch",
    "fisher": "Phil Fisher",
}
INVESTOR_SHORT = {
    "buffett": "W.Buffett",
    "lynch": "P.Lynch",
    "fisher": "P.Fisher",
}
SIGNAL_KR = {
    "bullish": "매수",
    "neutral": "중립",
    "bearish": "매도",
}
COMBINED_KR = {
    "strong_buy": "강력매수",
    "buy": "매수",
    "hold": "보유",
    "sell": "매도",
}
CAP_LABELS = {
    "mega": "메가캡",
    "large": "대형주",
    "mid": "중형주",
    "small": "소형주",
    None: "시총 N/A",
    "cash": "현금",
}


def investor_signal(score: float) -> str:
    if score >= 7:
        return "bullish"
    if score <= 3:
        return "bearish"
    return "neutral"


def investor_confidence(score: float, signal: str) -> int:
    if signal == "bullish":
        return min(95, round(60 + (score - 7) / 3 * 35))
    if signal == "bearish":
        return min(90, round(60 + (3 - score) / 3 * 30))
    return round(55 + abs(score - 5) * 4)


def fmt_pct(value: float | None, signed: bool = False) -> str:
    if value is None:
        return "N/A"
    return f"{value:+.1f}%" if signed else f"{value:.1f}%"


def fmt_num(value: float | None, digits: int = 1) -> str:
    if value is None:
        return "N/A"
    return f"{value:.{digits}f}"


def clean_warning(text: str) -> str:
    return re.sub(r"^⚠️\s*", "", text)


def reasoning_for(investor: str, row: dict, inv_signal: str) -> str:
    metrics = row.get("metrics", {})
    scores = row.get("scores", {})
    factors = row.get("factors") or []
    roe = metrics.get("roe")
    growth = metrics.get("revenue_growth")
    momentum = scores.get("enhanced_momentum")
    factor_hint = "; ".join(factors[:2]) if factors else "핵심 재무 데이터 제한"

    if investor == "buffett":
        if inv_signal == "bullish":
            return f"ROE {fmt_pct(roe)}와 품질 지표가 강해 경제적 해자 관점에서 우호적입니다. 다만 P/E가 N/A라 안전마진 확인은 제한됩니다."
        if inv_signal == "bearish":
            return f"장기 보유 관점의 품질·예측가능성 점수가 낮습니다. {factor_hint}에도 불구하고 안전마진 근거가 부족합니다."
        return f"일부 품질 요인은 확인되지만 버핏식 확신에는 부족합니다. {factor_hint}."

    if investor == "lynch":
        if inv_signal == "bullish":
            return f"매출 성장 {fmt_pct(growth)}과 성장 스토리가 뚜렷해 GARP 후보입니다. PEG 데이터가 N/A라 가격 대비 성장 검증은 보수적으로 봅니다."
        if inv_signal == "bearish":
            return f"PEG·밸류에이션 확인이 부족하고 린치식 10배주 조건을 충족하지 못했습니다. 성장률 {fmt_pct(growth)}만으로는 과반 매수 근거가 약합니다."
        return f"성장 단서는 있으나 PEG와 밸류에이션 검증이 제한적입니다. {factor_hint}."

    if investor == "fisher":
        if inv_signal == "bullish":
            return f"성장 품질과 장기 경쟁력 신호가 양호합니다. 매출 성장 {fmt_pct(growth)}, 모멘텀 {fmt_num(momentum)}점이 피셔 관점의 긍정 근거입니다."
        if inv_signal == "bearish":
            return f"장기 성장 품질과 경영·R&D 우위의 증거가 충분하지 않습니다. {factor_hint}."
        return f"성장성은 일부 확인되지만 피셔식 장기 확신에는 추가 검증이 필요합니다. {factor_hint}."

    return factor_hint


def build_rows(data: dict) -> list[dict]:
    rows = []
    for item in data["rankings"][:TOP_N]:
        cap_value = item.get("market_cap", {}).get("value")
        if not cap_value:
            cap_value = get_market_cap_kr(item["ticker"], data["analysis_date"])
        cap_category, cap_display = get_market_cap_category(cap_value, currency="KRW")

        analyses = {}
        bullish_count = 0
        bearish_count = 0
        for investor in INVESTORS:
            score = item["investor_scores"].get(investor, 0)
            sig = investor_signal(score)
            conf = investor_confidence(score, sig)
            if sig == "bullish":
                bullish_count += 1
            elif sig == "bearish":
                bearish_count += 1
            analyses[investor] = {
                "score": score,
                "signal": sig,
                "confidence": conf,
                "reasoning": reasoning_for(investor, item, sig),
            }

        avg_confidence = round(sum(a["confidence"] for a in analyses.values()) / len(INVESTORS))
        if bullish_count == len(INVESTORS):
            combined_signal = "strong_buy"
        elif bullish_count >= 2:
            combined_signal = "buy"
        elif bearish_count >= 2:
            combined_signal = "sell"
        else:
            combined_signal = "hold"

        row = {
            **item,
            "market_cap": {
                "value": cap_value,
                "display": cap_display,
                "category": cap_category,
            },
            "investor_analysis": analyses,
            "bullish_count": bullish_count,
            "bearish_count": bearish_count,
            "consensus_ratio": bullish_count / len(INVESTORS),
            "combined_signal": combined_signal,
            "combined_confidence": avg_confidence,
            "included": bullish_count >= 2,
            "weight": 0.0,
        }
        rows.append(row)
    return rows


def allocate_weights(included: list[dict]) -> float:
    if not included:
        return 100.0

    for row in included:
        row["_weight_base"] = row["combined_confidence"] * (1 + (row.get("predicted_return_1y") or 0) / 100)

    remaining = 100.0
    active = included[:]
    capped = []
    while active:
        total_base = sum(row["_weight_base"] for row in active)
        if total_base <= 0:
            break
        over = [row for row in active if row["_weight_base"] / total_base * remaining > 15.0]
        if not over:
            for row in active:
                row["weight"] = row["_weight_base"] / total_base * remaining
            break
        for row in over:
            row["weight"] = 15.0
        capped.extend(over)
        remaining -= 15.0 * len(over)
        active = [row for row in active if row not in over]

    for row in included:
        if row["weight"] < 2.0:
            row["included"] = False
            row["weight"] = 0.0

    # The skill defines sector from market_cap.category when no explicit sector
    # is present. Enforce the 35% concentration cap on that category.
    by_category = defaultdict(list)
    for row in included:
        if row["included"]:
            by_category[row["market_cap"]["category"]].append(row)
    for category_rows in by_category.values():
        category_weight = sum(row["weight"] for row in category_rows)
        if category_weight > 35.0:
            scale = 35.0 / category_weight
            for row in category_rows:
                row["weight"] *= scale

    cash_weight = max(0.0, 100.0 - sum(row["weight"] for row in included if row["included"]))
    for row in included:
        row.pop("_weight_base", None)
    return cash_weight


def portfolio_rows(rows: list[dict]) -> tuple[list[dict], float]:
    included = [row for row in rows if row["included"]]
    cash_weight = allocate_weights(included)
    included = [row for row in included if row["included"]]
    included.sort(key=lambda row: row["weight"], reverse=True)
    return included, cash_weight


def weighted_avg(rows: list[dict], key: str) -> float:
    total_weight = sum(row["weight"] for row in rows)
    if not rows or total_weight <= 0:
        return 0.0
    return sum(row["weight"] * (row.get(key) or 0) for row in rows) / total_weight


def create_text_report(data: dict, rows: list[dict], portfolio: list[dict], cash_weight: float) -> str:
    lines = []
    investor_names = ", ".join(INVESTOR_LABELS[i] for i in INVESTORS)
    avg_conf = round(sum(row["combined_confidence"] for row in portfolio) / len(portfolio)) if portfolio else 0
    avg_ret = weighted_avg(portfolio, "predicted_return_1y")
    unanimous = [row for row in portfolio if row["bullish_count"] == len(INVESTORS)]
    majority = [row for row in portfolio if row["bullish_count"] == 2]
    not_included = [row for row in rows if not row["included"]]
    cap_weights = defaultdict(float)
    for row in portfolio:
        cap_weights[row["market_cap"]["category"]] += row["weight"]
    if cash_weight:
        cap_weights["cash"] += cash_weight

    lines += [
        "══════════════════════════════════════════════════════════════════════════════════════════════",
        "📋 AI Hedge Fund 포트폴리오 리포트",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
        f"분석 일자    : {data['analysis_date']}",
        f"분석 대상    : KRX 상위 {TOP_N}개 종목",
        "분석 전략    : 하이브리드 (펀더멘털 70% + 모멘텀 30%)",
        f"투자자 관점  : {investor_names}",
        "데이터 소스  : DART + PyKRX + KRX Open API (predict) + predict 투자자 점수 기반 정성 분석",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
        "",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
        "📊 포트폴리오 구성",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
        "#   종목                     비중     시총       종합신호       신뢰도   예상수익률   P/E    ROE      합의",
        "─────────────────────────────────────────────────────────────────────────────────────────────",
    ]
    for i, row in enumerate(portfolio, 1):
        metrics = row.get("metrics", {})
        lines.append(
            f"{i:<3} {row['ticker']}({row['company_name']})".ljust(28)
            + f"{row['weight']:>6.1f}%   {row['market_cap']['display']:<9} "
            + f"{COMBINED_KR[row['combined_signal']]:<10} {row['combined_confidence']:>3}%"
            + f"     {fmt_pct(row.get('predicted_return_1y'), signed=True):>7}"
            + f"      {fmt_num(metrics.get('pe')):>5}  {fmt_pct(metrics.get('roe')):>7}"
            + f"   {row['bullish_count']}/{len(INVESTORS)}"
        )
    if cash_weight:
        lines.append(
            f"{len(portfolio)+1:<3} 현금/대기자금".ljust(28)
            + f"{cash_weight:>6.1f}%   {'N/A':<9} {'대기':<10} {'N/A':>3}"
            + f"     {'+0.0%':>7}      {'N/A':>5}  {'N/A':>7}   -"
        )
    lines += [
        "─────────────────────────────────────────────────────────────────────────────────────────────",
        f"    합계                    100.0%                              avg {avg_conf}%  avg {avg_ret:+.1f}%",
        "",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
        "👥 투자자별 종목 분석 매트릭스",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
        "종목                     W.Buffett        P.Lynch          P.Fisher",
        "─────────────────────────────────────────────────────────────────────────────────────────────",
    ]
    for row in rows:
        cells = []
        for inv in INVESTORS:
            analysis = row["investor_analysis"][inv]
            emoji = {"bullish": "🟢", "neutral": "🔵", "bearish": "🔴"}[analysis["signal"]]
            cells.append(f"{emoji} {analysis['signal']}({analysis['confidence']})")
        lines.append(f"{row['ticker']}({row['company_name']})".ljust(24) + f"{cells[0]:<17} {cells[1]:<17} {cells[2]}")

    lines += [
        "",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
        "💬 투자자별 핵심 분석 근거 (편입 상위 5개 종목)",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
    ]
    for row in portfolio[:5]:
        lines += [
            "",
            f"▸ {row['ticker']} ({row['company_name']}) — 종합: {COMBINED_KR[row['combined_signal']]} (신뢰도 {row['combined_confidence']}%)",
        ]
        for inv in INVESTORS:
            analysis = row["investor_analysis"][inv]
            emoji = {"bullish": "🟢", "neutral": "🔵", "bearish": "🔴"}[analysis["signal"]]
            lines.append(f"  {INVESTOR_LABELS[inv]:<15}: {emoji} ({analysis['confidence']}%) {analysis['reasoning']}")

    lines += [
        "",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
        "📈 포트폴리오 요약",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
        f"편입 종목 수         : {len(portfolio)}개 / 분석 {len(rows)}개 ({len(portfolio)/len(rows)*100:.1f}% 편입률)",
        f"현금/대기자금        : {cash_weight:.1f}% (15% 단일 종목, 35% 분류 집중도 제한 적용 후 잔여)",
        f"평균 신뢰도          : {avg_conf}%",
        f"평균 예상 수익률     : {avg_ret:+.1f}%",
        f"강력매수 비중        : {sum(row['weight'] for row in portfolio if row['combined_signal'] == 'strong_buy'):.1f}%",
        f"매수 비중            : {sum(row['weight'] for row in portfolio if row['combined_signal'] == 'buy'):.1f}%",
        "",
        "투자자 합의 분포:",
        f"  만장일치 (3/3)     : {len(unanimous)}개",
        f"  다수 합의 (2/3)    : {len(majority)}개 — " + ", ".join(row["ticker"] for row in majority[:8]) + (f" 외 {len(majority)-8}개" if len(majority) > 8 else ""),
        "",
        "시가총액 분포:",
    ]
    for category, weight in sorted(cap_weights.items(), key=lambda item: -item[1]):
        names = [f"{row['ticker']} {row['weight']:.1f}%" for row in portfolio if row["market_cap"]["category"] == category]
        if category == "cash":
            names = [f"현금 {cash_weight:.1f}%"]
        lines.append(f"  {CAP_LABELS[category]:<12}: {weight:>5.1f}% — " + ", ".join(names[:5]))

    lines += [
        "",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
        f"🚫 비편입 종목 ({len(not_included)}개)",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
        "종목                     순위   점수    사유",
        "─────────────────────────────────────────────────────────────────────────────────────────────",
    ]
    for row in not_included:
        reason = f"투자자 과반 미달 ({row['bullish_count']}/{len(INVESTORS)} bullish)"
        lines.append(f"{row['ticker']}({row['company_name']})".ljust(24) + f"{row['rank']:<6} {row['total_score']:<6.2f} {reason}")

    warnings = []
    for row in rows:
        for warning in row.get("investor_warnings") or []:
            warnings.append(f"{row['ticker']}({row['company_name']}) : {clean_warning(warning)}")
    top5_weight = sum(row["weight"] for row in portfolio[:5])
    max_stock = portfolio[0] if portfolio else None
    non_cash_weights = {k: v for k, v in cap_weights.items() if k != "cash"}
    max_cap_category, max_cap_weight = max(non_cash_weights.items(), key=lambda item: item[1]) if non_cash_weights else (None, 0)
    lines += [
        "",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
        "⚠️ 리스크 및 경고",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
        "포트폴리오 집중도:",
        f"  상위 5종목 비중     : {top5_weight:.1f}%",
        f"  최대 단일 종목      : {max_stock['ticker']} {max_stock['weight']:.1f}% (제한 15% 내)" if max_stock else "  최대 단일 종목      : N/A",
        f"  최대 비현금 분류 비중: {CAP_LABELS[max_cap_category]} {max_cap_weight:.1f}% (제한 35% 내)",
        "",
        "투자 철학 불일치 경고:",
    ]
    if warnings:
        lines.extend(f"  {warning}" for warning in warnings[:8])
        if len(warnings) > 8:
            lines.append(f"  ... 외 {len(warnings)-8}건")
    else:
        lines.append("  없음")

    lines += [
        "",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
        "💡 이 리포트는 교육/연구 목적이며 실제 투자 결정의 근거가 될 수 없습니다.",
        "   predict: DART + PyKRX + KRX Open API | investor-analysis: predict 투자자 점수 기반 정성 분석",
        f"   엑셀 리포트: {OUTPUT_XLSX}",
        "══════════════════════════════════════════════════════════════════════════════════════════════",
    ]
    return "\n".join(lines) + "\n"


def set_header(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22


def set_grid(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="맑은 고딕", size=10)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 18


def auto_width(ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 8), 30)


def signal_fill(signal: str) -> PatternFill:
    if signal in {"매수", "강력매수", "bullish"}:
        return BUY_FILL
    if signal in {"매도", "bearish"}:
        return SELL_FILL
    return NEUTRAL_FILL


def create_excel(data: dict, rows: list[dict], portfolio: list[dict], cash_weight: float) -> None:
    wb = Workbook()
    stock_last = len(portfolio) + 1
    portfolio_last = stock_last + (1 if cash_weight else 0)
    total_row = portfolio_last + 1

    ws1 = wb.active
    ws1.title = "요약"
    ws1.merge_cells("A1:B1")
    ws1["A1"] = "AI Hedge Fund 포트폴리오 리포트"
    ws1["A1"].fill = TITLE_FILL
    ws1["A1"].font = TITLE_FONT
    ws1["A1"].alignment = Alignment(horizontal="center")

    summary = [
        (3, "분석 일자", data["analysis_date"]),
        (4, "분석 대상", f"KRX 상위 {TOP_N}개"),
        (5, "분석 전략", "하이브리드"),
        (6, "투자자 관점", ", ".join(INVESTOR_SHORT[i] for i in INVESTORS)),
        (8, "포트폴리오 통계", ""),
        (9, "편입 종목 수", f'=COUNTA(\'포트폴리오\'!B2:B{stock_last})&" / {TOP_N} ("&TEXT(COUNTA(\'포트폴리오\'!B2:B{stock_last})/{TOP_N},"0.0%")&")"'),
        (10, "평균 신뢰도", f"=AVERAGE('포트폴리오'!F2:F{stock_last})/100"),
        (11, "평균 예상 수익률", f"=AVERAGE('포트폴리오'!G2:G{stock_last})"),
        (12, "강력매수 비중", f'=SUMIF(\'포트폴리오\'!E2:E{stock_last},"강력매수",\'포트폴리오\'!D2:D{stock_last})'),
        (13, "매수 비중", f'=SUMIF(\'포트폴리오\'!E2:E{stock_last},"매수",\'포트폴리오\'!D2:D{stock_last})'),
        (14, "현금/대기자금", cash_weight / 100),
        (16, "시가총액 분포", ""),
    ]
    for row, label, value in summary:
        ws1.cell(row=row, column=1, value=label)
        ws1.cell(row=row, column=2, value=value)
        ws1.cell(row=row, column=1).font = Font(name="맑은 고딕", bold=row in {8, 16})
        ws1.cell(row=row, column=1).border = THIN_BORDER
        ws1.cell(row=row, column=2).border = THIN_BORDER
    for row in [10, 11, 12, 13, 14]:
        ws1.cell(row=row, column=2).number_format = "0.0%"

    cap_weights = defaultdict(float)
    for row in portfolio:
        cap_weights[row["market_cap"]["category"]] += row["weight"]
    if cash_weight:
        cap_weights["cash"] += cash_weight
    dist_start = 17
    for offset, (category, weight) in enumerate(sorted(cap_weights.items(), key=lambda item: -item[1])):
        ws1.cell(row=dist_start + offset, column=1, value=CAP_LABELS[category])
        ws1.cell(row=dist_start + offset, column=2, value=weight / 100)
        ws1.cell(row=dist_start + offset, column=2).number_format = "0.0%"
        ws1.cell(row=dist_start + offset, column=1).border = THIN_BORDER
        ws1.cell(row=dist_start + offset, column=2).border = THIN_BORDER

    if cap_weights:
        chart = PieChart()
        chart.title = "시가총액 분포"
        data_ref = Reference(ws1, min_col=2, min_row=dist_start, max_row=dist_start + len(cap_weights) - 1)
        cats_ref = Reference(ws1, min_col=1, min_row=dist_start, max_row=dist_start + len(cap_weights) - 1)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        ws1.add_chart(chart, "D3")
    auto_width(ws1, 2)

    ws2 = wb.create_sheet("포트폴리오")
    headers = ["#", "종목코드", "회사명", "비중", "신호", "신뢰도", "예상수익률", "시가총액", "P/E", "ROE", "PEG", "합의", "섹터"]
    for col, header in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=header)
    set_header(ws2, 1, len(headers))
    for idx, row in enumerate(portfolio, 2):
        metrics = row.get("metrics", {})
        values = [
            idx - 1,
            row["ticker"],
            row["company_name"],
            row["weight"] / 100,
            COMBINED_KR[row["combined_signal"]],
            row["combined_confidence"],
            (row.get("predicted_return_1y") or 0) / 100,
            row["market_cap"]["display"],
            metrics.get("pe") if metrics.get("pe") is not None else "N/A",
            metrics.get("roe") / 100 if metrics.get("roe") is not None else "N/A",
            metrics.get("peg") if metrics.get("peg") is not None else "N/A",
            f"{row['bullish_count']}/{len(INVESTORS)}",
            CAP_LABELS[row["market_cap"]["category"]],
        ]
        for col, value in enumerate(values, 1):
            ws2.cell(row=idx, column=col, value=value)
        set_grid(ws2, idx, len(headers))
        for col in range(1, len(headers) + 1):
            ws2.cell(row=idx, column=col).fill = signal_fill(COMBINED_KR[row["combined_signal"]])
        ws2.cell(row=idx, column=4).number_format = "0.0%"
        ws2.cell(row=idx, column=7).number_format = "+0.0%;-0.0%"
        ws2.cell(row=idx, column=10).number_format = "0.0%"

    if cash_weight:
        cash_row = stock_last + 1
        cash_values = [len(portfolio) + 1, "CASH", "현금/대기자금", cash_weight / 100, "대기", "N/A", 0, "N/A", "N/A", "N/A", "N/A", "-", "현금"]
        for col, value in enumerate(cash_values, 1):
            ws2.cell(row=cash_row, column=col, value=value)
        set_grid(ws2, cash_row, len(headers))
        for col in range(1, len(headers) + 1):
            ws2.cell(row=cash_row, column=col).fill = NEUTRAL_FILL
        ws2.cell(row=cash_row, column=4).number_format = "0.0%"
        ws2.cell(row=cash_row, column=7).number_format = "+0.0%;-0.0%"

    total_values = ["", "", "합계/평균", f"=SUM(D2:D{portfolio_last})", "", f"=AVERAGE(F2:F{stock_last})", f"=AVERAGE(G2:G{stock_last})", "", "", "", "", "", ""]
    for col, value in enumerate(total_values, 1):
        ws2.cell(row=total_row, column=col, value=value)
        ws2.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws2.cell(row=total_row, column=col).font = Font(name="맑은 고딕", bold=True)
        ws2.cell(row=total_row, column=col).border = THIN_BORDER
    ws2.cell(row=total_row, column=4).number_format = "0.0%"
    ws2.cell(row=total_row, column=7).number_format = "+0.0%;-0.0%"
    ws2.auto_filter.ref = f"A1:M{portfolio_last}"
    ws2.freeze_panes = "A2"
    auto_width(ws2, len(headers))

    ws3 = wb.create_sheet("순위")
    rank_headers = ["순위", "종목코드", "회사명", "종합점수", "펀더멘털", "모멘텀", "앙상블", "신호", "예상수익률", "시가총액", "P/E", "P/B", "ROE", "매출성장률", "PEG", "편입여부"]
    for col, header in enumerate(rank_headers, 1):
        ws3.cell(row=1, column=col, value=header)
    set_header(ws3, 1, len(rank_headers))
    for idx, row in enumerate(rows, 2):
        metrics = row.get("metrics", {})
        scores = row.get("scores", {})
        values = [
            row["rank"],
            row["ticker"],
            row["company_name"],
            row["total_score"],
            scores.get("fundamental"),
            scores.get("enhanced_momentum"),
            row.get("ensemble_score"),
            row.get("signal"),
            (row.get("predicted_return_1y") or 0) / 100,
            row["market_cap"]["display"],
            metrics.get("pe") if metrics.get("pe") is not None else "N/A",
            metrics.get("pb") if metrics.get("pb") is not None else "N/A",
            metrics.get("roe") / 100 if metrics.get("roe") is not None else "N/A",
            metrics.get("revenue_growth") / 100 if metrics.get("revenue_growth") is not None else "N/A",
            metrics.get("peg") if metrics.get("peg") is not None else "N/A",
            "예" if row["included"] else "아니오",
        ]
        for col, value in enumerate(values, 1):
            ws3.cell(row=idx, column=col, value=value)
        set_grid(ws3, idx, len(rank_headers))
        if row["included"]:
            for col in range(1, len(rank_headers) + 1):
                ws3.cell(row=idx, column=col).fill = BUY_FILL
        for col in [9, 13, 14]:
            ws3.cell(row=idx, column=col).number_format = "+0.0%;-0.0%" if col in [9, 14] else "0.0%"
    ws3.auto_filter.ref = f"A1:P{len(rows)+1}"
    ws3.freeze_panes = "A2"
    auto_width(ws3, len(rank_headers))

    ws4 = wb.create_sheet("투자자 매트릭스")
    matrix_headers = ["종목코드", "회사명", *[INVESTOR_SHORT[i] for i in INVESTORS], "종합신호", "종합신뢰도"]
    for col, header in enumerate(matrix_headers, 1):
        ws4.cell(row=1, column=col, value=header)
    set_header(ws4, 1, len(matrix_headers))
    for idx, row in enumerate(rows, 2):
        ws4.cell(row=idx, column=1, value=row["ticker"])
        ws4.cell(row=idx, column=2, value=row["company_name"])
        for offset, inv in enumerate(INVESTORS, 3):
            analysis = row["investor_analysis"][inv]
            value = f"{SIGNAL_KR[analysis['signal']]}({analysis['confidence']})"
            cell = ws4.cell(row=idx, column=offset, value=value)
            cell.fill = signal_fill(analysis["signal"])
            cell.font = BUY_FONT if analysis["signal"] == "bullish" else SELL_FONT if analysis["signal"] == "bearish" else NEUTRAL_FONT
        ws4.cell(row=idx, column=6, value=COMBINED_KR[row["combined_signal"]])
        ws4.cell(row=idx, column=7, value=row["combined_confidence"])
        set_grid(ws4, idx, len(matrix_headers))
    ws4.auto_filter.ref = f"A1:G{len(rows)+1}"
    ws4.freeze_panes = "A2"
    auto_width(ws4, len(matrix_headers))

    ws5 = wb.create_sheet("투자자 상세")
    detail_headers = ["종목코드", "회사명", "투자자", "신호", "신뢰도", "분석근거"]
    for col, header in enumerate(detail_headers, 1):
        ws5.cell(row=1, column=col, value=header)
    set_header(ws5, 1, len(detail_headers))
    detail_row = 2
    for row in rows:
        for inv in INVESTORS:
            analysis = row["investor_analysis"][inv]
            values = [
                row["ticker"],
                row["company_name"],
                INVESTOR_LABELS[inv],
                SIGNAL_KR[analysis["signal"]],
                analysis["confidence"],
                analysis["reasoning"],
            ]
            for col, value in enumerate(values, 1):
                ws5.cell(row=detail_row, column=col, value=value)
            set_grid(ws5, detail_row, len(detail_headers))
            ws5.cell(row=detail_row, column=4).fill = signal_fill(analysis["signal"])
            ws5.cell(row=detail_row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
            detail_row += 1
    ws5.freeze_panes = "A2"
    auto_width(ws5, len(detail_headers))
    ws5.column_dimensions["F"].width = 90

    ws6 = wb.create_sheet("리스크 분석")
    ws6["A1"] = "리스크 분석"
    ws6["A1"].fill = TITLE_FILL
    ws6["A1"].font = TITLE_FONT
    non_cash_weights = {k: v for k, v in cap_weights.items() if k != "cash"}
    max_cap_category, max_cap_weight = max(non_cash_weights.items(), key=lambda item: item[1]) if non_cash_weights else (None, 0)
    risk_items = [
        (3, "집중도 지표", ""),
        (4, "상위 1종목 비중", "=SUM('포트폴리오'!D2:D2)"),
        (5, "상위 3종목 비중", f"=SUM('포트폴리오'!D2:D{min(4, stock_last)})"),
        (6, "상위 5종목 비중", f"=SUM('포트폴리오'!D2:D{min(6, stock_last)})"),
        (7, "상위 10종목 비중", f"=SUM('포트폴리오'!D2:D{min(11, portfolio_last)})"),
        (8, "HHI (허핀달 지수)", f"=SUMPRODUCT('포트폴리오'!D2:D{portfolio_last},'포트폴리오'!D2:D{portfolio_last})"),
        (10, "분류 집중도", ""),
        (11, "최대 비현금 분류", CAP_LABELS[max_cap_category] if non_cash_weights else "N/A"),
        (12, "최대 비현금 분류 비중", max_cap_weight / 100 if non_cash_weights else 0),
        (14, "투자자 합의 품질", ""),
        (15, "만장일치 비율", f'=COUNTIF(\'포트폴리오\'!L2:L{stock_last},"3/3")/COUNTA(\'포트폴리오\'!B2:B{stock_last})'),
        (16, "의견 분산 종목 수", sum(1 for row in rows if row.get("investor_consensus", {}).get("level") == "low")),
        (18, "비편입 사유 분포", ""),
        (19, "투자자 과반 미달", len([row for row in rows if not row["included"]])),
        (20, "최소 비중 미달", 0),
    ]
    for row, label, value in risk_items:
        ws6.cell(row=row, column=1, value=label)
        ws6.cell(row=row, column=2, value=value)
        ws6.cell(row=row, column=1).border = THIN_BORDER
        ws6.cell(row=row, column=2).border = THIN_BORDER
        if row in {3, 10, 14, 18}:
            ws6.cell(row=row, column=1).font = Font(name="맑은 고딕", bold=True)
    for row in [4, 5, 6, 7, 12, 15]:
        ws6.cell(row=row, column=2).number_format = "0.0%"
    ws6.cell(row=8, column=2).number_format = "0.000"
    auto_width(ws6, 2)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)


def main() -> None:
    if not PREDICT_JSON.exists():
        raise SystemExit(f"predict JSON not found: {PREDICT_JSON}")
    data = json.loads(PREDICT_JSON.read_text(encoding="utf-8"))
    rows = build_rows(data)
    portfolio, cash_weight = portfolio_rows(rows)
    report = create_text_report(data, rows, portfolio, cash_weight)
    OUTPUT_TXT.write_text(report, encoding="utf-8")
    create_excel(data, rows, portfolio, cash_weight)
    print(report)
    print(f"텍스트 리포트: {OUTPUT_TXT}")
    print(f"엑셀 리포트: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
