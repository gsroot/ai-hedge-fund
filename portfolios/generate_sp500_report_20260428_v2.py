#!/usr/bin/env python3
"""Generate S&P 500 portfolio report (2026-04-28 v2) using actual investor agent analyses.

v1 (generate_sp500_report_20260428.py) used algorithmic investor signals derived from
predict factor scores. v2 uses real LLM-based investor agent output (Buffett, Lynch,
Fisher) collected via the portfolio-report skill at 2026-04-28 ~11:10am KST.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_XLSX = ROOT / "portfolios" / "sp500_20260428_buffett_fisher_lynch.xlsx"
OUTPUT_TXT = ROOT / "portfolios" / "sp500_20260428_buffett_fisher_lynch.txt"

INVESTORS = ["buffett", "lynch", "fisher"]
INVESTOR_LABELS = {
    "buffett": "Warren Buffett",
    "lynch": "Peter Lynch",
    "fisher": "Phil Fisher",
}
INVESTOR_SHORT = {"buffett": "W.Buffett", "lynch": "P.Lynch", "fisher": "P.Fisher"}
SIGNAL_KR = {"bullish": "매수", "neutral": "중립", "bearish": "매도"}
COMBINED_KR = {"strong_buy": "강력매수", "buy": "매수", "hold": "보유", "sell": "매도"}


@dataclass(frozen=True)
class StockRow:
    rank: int
    ticker: str
    name: str
    market_cap: str
    market_cap_value: float
    sector: str
    total_score: float
    fundamental: float
    momentum: float
    ensemble: float
    signal: str
    expected_return: float
    pe: float | None
    roe: float | None
    revenue_growth: float | None
    peg: float | None
    bullish_investors: tuple[str, ...]


PREDICT_TOP30: tuple[StockRow, ...] = (
    StockRow(1, "MU", "Micron Technology", "$592B", 592, "Technology", 8.33, 8.0, 9.0, 8.9, "strong_buy", 18.7, 24.7, 40.0, 196.0, 0.29, ("buffett", "lynch", "druckenmiller")),
    StockRow(2, "ALL", "Allstate Corp.", "$55B", 55, "Financials", 8.27, 8.8, 5.5, 8.6, "strong_buy", 18.4, 5.6, 40.0, 5.1, 0.46, ("buffett", "lynch", "graham")),
    StockRow(3, "MTCH", "Match Group", "$8.7B", 8.7, "Communication Services", 8.04, 8.1, 8.0, 6.6, "strong_buy", 17.6, 15.5, None, 2.1, 0.33, ("lynch", "druckenmiller")),
    StockRow(4, "ADI", "Analog Devices", "$192B", 192, "Technology", 7.68, 6.6, 9.0, 7.0, "buy", 16.4, 71.8, 7.9, 30.0, 0.99, ("lynch", "druckenmiller", "fisher")),
    StockRow(5, "WDC", "Western Digital", "$137B", 137, "Technology", 7.66, 7.1, 9.0, 7.7, "buy", 16.3, 37.9, 41.1, 25.0, 0.81, ("lynch", "druckenmiller", "fisher")),
    StockRow(6, "HIG", "The Hartford", "$37B", 37, "Financials", 7.45, 8.2, 3.5, 8.2, "buy", 15.6, 9.6, 22.7, 6.1, 0.12, ("buffett", "lynch", "graham")),
    StockRow(7, "AVGO", "Broadcom", "$1,980B", 1980, "Technology", 7.35, 6.6, 9.0, 7.6, "buy", 15.2, 81.7, 33.4, 29.5, 0.92, ("lynch", "druckenmiller", "fisher")),
    StockRow(8, "CINF", "Cincinnati Financial", "$26B", 26, "Financials", 7.35, 7.8, 5.0, 8.1, "buy", 15.2, 10.9, 16.0, 21.8, 0.50, ("buffett", "lynch", "graham")),
    StockRow(9, "PYPL", "PayPal Holdings", "$45B", 45, "Financials", 7.28, 7.9, 4.0, 7.8, "buy", 15.0, 9.2, 25.7, 3.7, 0.91, ("buffett", "lynch", "graham")),
    StockRow(10, "NVDA", "NVIDIA", "$5,265B", 5265, "Technology", 7.18, 7.3, 7.0, 8.1, "buy", 14.6, 44.3, 101.5, 73.0, 0.74, ("buffett", "lynch", "druckenmiller")),
    StockRow(11, "GRMN", "Garmin", "$50B", 50, "Technology", 7.18, 6.4, 9.0, 7.4, "buy", 14.6, 30.0, 19.8, 16.6, 3.37, ("buffett", "druckenmiller", "fisher")),
    StockRow(12, "AMD", "Advanced Micro Devices", "$546B", 546, "Technology", 7.05, 5.5, 9.0, 6.3, "buy", 14.2, 128.2, 7.1, 34.0, 1.03, ("lynch", "druckenmiller", "fisher")),
    StockRow(13, "WST", "West Pharmaceutical", "$21B", 21, "Healthcare", 6.98, 6.1, 9.0, 6.8, "buy", 13.9, 40.5, 19.1, 21.0, 3.73, ("buffett", "druckenmiller", "fisher")),
    StockRow(14, "TER", "Teradyne", "$63B", 63, "Technology", 6.96, 6.1, 9.0, 6.9, "buy", 13.9, 115.9, 19.7, 44.0, 2.31, ("buffett", "druckenmiller", "fisher")),
    StockRow(15, "DELL", "Dell Technologies", "$140B", 140, "Technology", 6.95, 5.3, 9.0, 5.3, "buy", 13.8, 24.9, None, 39.5, 0.90, ("lynch", "druckenmiller")),
    StockRow(16, "TFX", "Teleflex", "$6.0B", 6.0, "Healthcare", 6.92, 5.2, 9.0, 4.3, "buy", 13.7, 103.1, 1.6, None, None, ("lynch", "druckenmiller")),
    StockRow(17, "GLW", "Corning", "$144B", 144, "Technology", 6.91, 4.4, 10.0, 4.9, "buy", 13.7, 91.8, 14.9, 20.4, 1.96, ("druckenmiller",)),
    StockRow(18, "STX", "Seagate Technology", "$133B", 133, "Technology", 6.91, 5.2, 9.0, 5.4, "buy", 13.7, 67.3, None, 21.5, 0.89, ("lynch", "druckenmiller")),
    StockRow(19, "FSLR", "First Solar", "$21B", 21, "Energy", 6.90, 7.6, 3.0, 7.8, "buy", 13.6, 13.9, 17.5, 11.1, 0.49, ("buffett", "lynch", "fisher")),
    StockRow(20, "INCY", "Incyte Corp.", "$19B", 19, "Healthcare", 6.81, 8.2, 3.5, 7.9, "buy", 13.3, 14.9, 30.0, 27.8, 0.34, ("buffett", "lynch", "fisher")),
    StockRow(21, "LRCX", "Lam Research", "$324B", 324, "Technology", 6.78, 6.3, 8.0, 7.4, "buy", 13.2, 49.0, 67.0, 23.8, 1.59, ("buffett", "druckenmiller", "fisher")),
    StockRow(22, "QRVO", "Qorvo", "$8.0B", 8.0, "Technology", 6.62, 6.9, 6.0, 6.3, "buy", 12.7, 23.6, 9.6, 8.4, 0.21, ("lynch",)),
    StockRow(23, "DVA", "DaVita", "$9.9B", 9.9, "Healthcare", 6.57, 7.0, 5.5, 5.8, "buy", 12.5, 15.8, 64.8, 9.9, 0.56, ("buffett", "lynch")),
    StockRow(24, "HWM", "Howmet Aerospace", "$97B", 97, "Industrials", 6.55, 6.2, 7.5, 6.6, "buy", 12.4, 65.1, 30.4, 14.6, 0.80, ("lynch", "druckenmiller")),
    StockRow(25, "TRV", "The Travelers", "$65B", 65, "Financials", 6.55, 6.8, 5.0, 7.1, "buy", 12.4, 9.1, 25.3, 1.0, 2.36, ("buffett", "graham")),
    StockRow(26, "NEM", "Newmont", "$124B", 124, "Materials", 6.48, 6.9, 5.5, 6.4, "buy", 12.2, 15.1, 25.8, 45.8, 2.78, ("graham", "fisher")),
    StockRow(27, "QCOM", "QUALCOMM", "$160B", 160, "Technology", 6.46, 6.7, 5.0, 6.6, "buy", 12.1, 30.3, 21.5, 5.0, 0.64, ("buffett", "lynch", "graham")),
    StockRow(28, "MPWR", "Monolithic Power", "$78B", 78, "Technology", 6.42, 5.3, 9.0, 6.4, "buy", 12.0, 123.2, 19.2, 20.8, 2.68, ("buffett", "druckenmiller", "fisher")),
    StockRow(29, "TXN", "Texas Instruments", "$245B", 245, "Technology", 6.40, 5.3, 9.0, 6.5, "buy", 11.9, 46.1, 32.3, 18.6, 1.57, ("buffett", "druckenmiller", "fisher")),
    StockRow(30, "HPE", "Hewlett Packard Enterprise", "$38B", 38, "Technology", 6.38, 5.2, 9.0, 5.3, "buy", 11.8, 10.5, -0.5, 18.4, 0.85, ("lynch", "graham", "druckenmiller")),
)


# Real investor agent analyses (signal/confidence/reasoning) from 2026-04-28 11:10am
INVESTOR_ANALYSES: dict[str, dict[str, dict]] = {
    "MU": {
        "buffett": {"signal": "neutral", "confidence": 45, "reasoning": "Micron은 메모리 반도체라는 극도로 사이클리컬한 원자재형 산업이며, ROE 40%와 낮은 부채(D/E 0.15)는 인상적이나 메모리 가격 하락 사이클 진입 시 수익이 급격히 무너지는 구조적 취약성과 P/E 24.8의 사이클 정점 가격이 우려된다."},
        "lynch": {"signal": "bullish", "confidence": 88, "reasoning": "사이클리컬이지만 AI 메모리 수요 급증으로 Fast Grower 국면; PEG 0.29의 극도 저평가, EPS 성장 770%, ROE 40%, D/E 0.15의 탄탄한 재무가 GARP 조건을 완벽 충족한다."},
        "fisher": {"signal": "bullish", "confidence": 82, "reasoning": "메모리 사이클 정점에서 196% 매출 성장과 58% 총마진은 탁월하며 R&D/Revenue 약 10%, HBM/DRAM 기술 리더십이 지속 중이다. 경영진은 AI 메모리 수요라는 구조적 동력을 일관되게 실행하고 있다."},
    },
    "ALL": {
        "buffett": {"signal": "bullish", "confidence": 72, "reasoning": "올스테이트는 내가 이해하는 손해보험으로 ROE 39.5%, FCF 89억 달러, D/E 0.25의 낮은 부채를 보유하며 P/E 5.6은 놀라울 정도로 저렴하다. float 활용 능력과 브랜드 인지도가 경쟁 우위를 제공한다."},
        "lynch": {"signal": "bullish", "confidence": 82, "reasoning": "보험 Stalwart로 PEG 0.46·PE 5.6은 명백한 저평가이며 EPS 99% 성장·ROE 40%·D/E 0.25는 보험업에서 보기 드문 수익성이다. $55B에서 텐배거는 어렵지만 안정적 GARP 플레이로 매력적이다."},
        "fisher": {"signal": "neutral", "confidence": 38, "reasoning": "보험업은 R&D 투자가 전무하고 차별화된 혁신 우위를 확인하기 어려우며, 양호한 ROE에도 불구하고 Scuttlebutt 관점에서 장기 10년 복리 성장 스토리가 부재하여 Fisher 기준 매력이 낮다."},
    },
    "MTCH": {
        "buffett": {"signal": "bearish", "confidence": 60, "reasoning": "Match Group은 ROE 데이터 부재와 매출 성장 2.1% 정체, 네트워크 효과 약화 환경이 우려된다. 내가 이해하기 어려운 소셜 플랫폼 비즈니스이며 지속적 moat가 불분명하다."},
        "lynch": {"signal": "bullish", "confidence": 79, "reasoning": "$8.7B 소형 Fast Grower 후보로 PEG 0.33은 가장 선호하는 구간이며 30% 영업이익률, FCF 수익률 10%가 뒷받침된다. 다만 매출 성장 2%로 성장 둔화는 리스크다."},
        "fisher": {"signal": "bearish", "confidence": 65, "reasoning": "매출 성장률 2.1%로 성장 동력 정체, 틱톡·인스타그램 경쟁 심화로 사용자 이탈 진행 중. R&D/Revenue 13%는 양호하나 ROE 음수, 성장 재가속 근거 희박."},
    },
    "ADI": {
        "buffett": {"signal": "bearish", "confidence": 55, "reasoning": "Analog Devices는 우수한 아날로그 반도체이나 ROE 7.9%는 15% 기준에 크게 못 미치고 P/E 71.8은 적정 가격 원칙에 위배된다. 30% 매출 성장은 고무적이나 현재 가격은 완벽한 미래를 이미 반영했다."},
        "lynch": {"signal": "bullish", "confidence": 72, "reasoning": "아날로그 반도체 Stalwart로 PEG 0.99는 적정가 상단이나 매출 30%·EPS 112% 성장 복귀는 사이클 회복 신호이며, 산업/자동차 아날로그라는 이해 가능한 비즈니스다."},
        "fisher": {"signal": "bullish", "confidence": 72, "reasoning": "아날로그 사이클 회복으로 매출 30% 성장, R&D/Revenue 16%의 두터운 연구 투자, 산업·자동차·통신 분야의 넓은 해자가 차별적 기술 우위를 증명한다."},
    },
    "WDC": {
        "buffett": {"signal": "neutral", "confidence": 40, "reasoning": "Western Digital은 HDD/플래시 스토리지 사이클리컬 하드웨어로 circle of competence 밖이다. ROE 41%는 사이클 고점 반영이며 P/E 37.9는 commodity 기업치고 고평가 우려가 있다."},
        "lynch": {"signal": "bullish", "confidence": 75, "reasoning": "HDD/NAND 사이클리컬에서 Fast Grower 전환 국면으로 PEG 0.81·ROE 41%·매출 25% 성장은 회복 사이클 초입 매수 원칙에 부합하며 D/E 0.65도 허용 가능하다."},
        "fisher": {"signal": "neutral", "confidence": 52, "reasoning": "HDD/NAND 스핀오프 후 구조 재편 중으로 매출 25% 성장과 ROE 41%는 사이클 반등 반영이나 R&D/Revenue 10%와 D/E 0.65는 중립이며 치킨게임적 경쟁 강도가 우려된다."},
    },
    "HIG": {
        "buffett": {"signal": "bullish", "confidence": 68, "reasoning": "The Hartford는 ROE 22.7%, D/E 0.23, P/E 9.6으로 Buffett이 선호하는 보험사 특성을 갖췄다. 상업용 보험과 근로자 보상 보험의 강력한 시장 지위와 낮은 밸류에이션이 안전마진을 제공한다."},
        "lynch": {"signal": "bullish", "confidence": 84, "reasoning": "보험 Stalwart로 PEG 0.12·PE 9.6은 선물 같은 가격의 저평가이며 ROE 23%·D/E 0.23의 건전한 재무로 안정적 복리 성장이 기대된다. 다만 매출 성장 6%는 Stalwart 수준이다."},
        "fisher": {"signal": "bearish", "confidence": 60, "reasoning": "보험사로서 R&D 투자가 전혀 없고 기술 혁신과 영업 조직 우월성을 평가할 근거가 부재하다. 매출 6%·ROE 23%는 사이클 호황 반영이며 장기 복리 성장 사업 모델 구별이 어렵다."},
    },
    "AVGO": {
        "buffett": {"signal": "neutral", "confidence": 45, "reasoning": "Broadcom은 반도체와 인프라 SW에서 강력한 moat와 ROE 33.4%, 영업마진 44.9%가 탁월하나 P/E 81.7과 D/E 0.83(VMware 부채)이 안전마진 원칙을 만족시키지 못하며 복잡한 사업이다."},
        "lynch": {"signal": "neutral", "confidence": 55, "reasoning": "AI 인프라 핵심 Fast Grower이지만 시가총액 $1.98T 메가캡으로 텐배거 가능성 0이며 PEG 0.92는 양호하나 PE 82·D/E 0.83은 부담이다. 큰 나무는 빨리 자라지 않는다."},
        "fisher": {"signal": "bullish", "confidence": 88, "reasoning": "AI 인프라(맞춤형 ASIC, 네트워킹) 수요 폭발로 매출 29.5% 성장, 총마진 76.7%, R&D/Revenue 17%의 집중 투자가 영업 조직+R&D 우월성 기준을 완벽 충족한다."},
    },
    "CINF": {
        "buffett": {"signal": "bullish", "confidence": 75, "reasoning": "Cincinnati Financial은 가장 좋아하는 유형이다. ROE 16%, D/E 0.056의 거의 무부채, P/E 10.9의 저평가, 60년 이상 배당 성장 기록의 배당 귀족주이며 독립 대리인 네트워크의 안정적 보험 moat를 갖췄다."},
        "lynch": {"signal": "bullish", "confidence": 71, "reasoning": "보험 Stalwart로 PE 10.9·ROE 16%·D/E 0.06의 극도 보수적 재무구조는 좋아하는 지루한 복리 기계이며 배당 전통과 결합하면 장기 보유 안전한 GARP 플레이다."},
        "fisher": {"signal": "bearish", "confidence": 58, "reasoning": "재산보험사로 R&D 투자 전무하며 기술 혁신, 연구 우월성, 차별적 제품 경쟁력을 전혀 확인할 수 없다. 매출 성장 21.8%는 보험료 인상과 사이클 효과로 구조적 성장 동력 미흡."},
    },
    "PYPL": {
        "buffett": {"signal": "neutral", "confidence": 50, "reasoning": "PayPal은 ROE 25.7%·P/E 9.2의 저렴한 가격이 매력적이나 Apple Pay·Google Pay 빅테크 공세로 네트워크 효과 moat가 잠식되고 있다. 매출 성장 3.7% 둔화는 구조적 경쟁 압력을 반영한다."},
        "lynch": {"signal": "bullish", "confidence": 76, "reasoning": "핀테크 Turnaround로 PEG 0.91·PE 9.2는 아무도 관심 없는 턴어라운드의 전형이며 ROE 26%·매출 4%는 미약하나 EPS 28%·FCF 수익률 10%가 내부 효율화 증명. $45B에서 3배주 가능."},
        "fisher": {"signal": "neutral", "confidence": 45, "reasoning": "매출 3.7%로 핀테크 혁신 기업임에도 성장 정체가 뚜렷하고 Apple Pay·Stripe 경쟁 격화로 해자 약화 중. R&D/Revenue 9%·ROE 26%는 무난하나 차별적 우월성과 성장 재가속 동력 불투명."},
    },
    "NVDA": {
        "buffett": {"signal": "neutral", "confidence": 50, "reasoning": "NVIDIA는 ROE 101.5%, 영업마진 65%, 최소 부채(D/E 0.07)로 본 어떤 기업보다 강력한 재무지표다. 그러나 P/E 44.3은 AI 수요 지속 전제이고 반도체 설계 변화 속도는 circle of competence를 벗어난다."},
        "lynch": {"signal": "neutral", "confidence": 48, "reasoning": "AI 시대 최대 수혜 Fast Grower이나 시가총액 $5.27T는 텐배거는커녕 2배도 구조적으로 어렵고 PEG 0.74는 매력적이나 이미 모든 기관이 보유한 종목에서 알파는 기대하기 어렵다."},
        "fisher": {"signal": "bullish", "confidence": 95, "reasoning": "매출 73% 성장, 영업마진 65%, ROE 102%, R&D/Revenue 8.6%로 15가지 체크리스트를 거의 완벽 충족하는 전설적 기업. Jensen Huang의 CUDA 생태계와 AI 컴퓨팅 패러다임 리더십은 평생 찾던 탁월한 기업의 전형이다."},
    },
    "GRMN": {
        "buffett": {"signal": "bullish", "confidence": 65, "reasoning": "Garmin은 항공·해양·스포츠용 GPS에서 강력한 브랜드 moat를 보유하며 ROE 19.8%·영업마진 28.9%·D/E 0.022의 거의 무부채 구조가 탁월하다. P/E 30은 다소 높지만 꾸준한 FCF와 16.6% 성장으로 적정 가격의 훌륭한 기업이다."},
        "lynch": {"signal": "neutral", "confidence": 52, "reasoning": "GPS·웨어러블 Stalwart로 이해하기 쉬운 제품 기업이지만 PEG 3.37은 과대평가 구간이며 매출 17% 성장은 양호하나 GARP 원칙 위반. PEG 1.5 이하 조정 후 재진입이 적절하다."},
        "fisher": {"signal": "bullish", "confidence": 70, "reasoning": "항공/마린/아웃도어/피트니스 전 분야 16.6% 매출 성장과 R&D/Revenue 15.5%의 지속 투자는 전문 기술 기업 특성. 무부채(D/E 0.02), 총마진 58.7%, 차별화된 GPS 소프트웨어 생태계가 장기 보유 근거를 제공한다."},
    },
    "AMD": {
        "buffett": {"signal": "bearish", "confidence": 65, "reasoning": "AMD는 ROE 7.1%로 15% 기준에 한참 못 미치며 P/E 128.2는 내재가치 대비 극도의 고평가다. NVIDIA와의 AI 반도체 경쟁 2위 포지션이 지속 가능한 moat를 의미하지 않으며 적정 가격 기준 미충족."},
        "lynch": {"signal": "neutral", "confidence": 50, "reasoning": "AI 반도체 Fast Grower로 매출 34%는 인상적이나 시가총액 $546B 준메가캡에서 텐배거 불가, PEG 1.03은 적정가나 PE 128·ROE 7%는 이익 없이 성장만 하는 주식 경고 프로파일이다."},
        "fisher": {"signal": "bullish", "confidence": 75, "reasoning": "데이터센터 AI 가속기와 EPYC CPU 수요로 매출 34% 성장, R&D/Revenue 23.4%의 강도 높은 연구 투자는 기술 혁신 우월성의 핵심 증거. Lisa Su 경영진 로드맵 실행 능력은 업계 최고 평가다."},
    },
    "WST": {
        "buffett": {"signal": "neutral", "confidence": 55, "reasoning": "West Pharmaceutical은 의약품 포장과 약물 전달 시스템의 틈새 moat를 가지며 ROE 19.1%·영업마진 21.7%·D/E 0.11이 양호하나 P/E 40.5는 안전마진이 없고 의료기기 원자재 공급업체 특성이 단순성을 저해한다."},
        "lynch": {"signal": "bearish", "confidence": 68, "reasoning": "제약 포장재 Stalwart로 비즈니스는 선호하는 지루하고 꼭 필요한 업종이지만 PEG 3.73은 명백한 과대평가, PE 40·ROE 19%의 조합에서 GARP 원칙이 완전히 깨진 상태다."},
        "fisher": {"signal": "neutral", "confidence": 55, "reasoning": "의약품 포장 특수 소재의 틈새 독점 지위와 21% 매출 성장은 인상적이나 R&D/Revenue 약 2.4%에 불과해 혁신 부족 경고 조건에 해당. 진입장벽이 규제·전환비용 기반이나 기술 혁신 부재가 선호도를 낮춘다."},
    },
    "TER": {
        "buffett": {"signal": "bearish", "confidence": 60, "reasoning": "Teradyne는 반도체 테스트 장비라는 매우 사이클리컬한 산업으로 P/E 115.9는 어떤 합리적 내재가치 계산으로도 정당화되기 어렵다. ROE 19.7%는 양호하나 지속 가능한 경쟁 우위 확인이 어렵다."},
        "lynch": {"signal": "bearish", "confidence": 65, "reasoning": "반도체 테스트 장비 Cyclical로 매출 44% 성장은 사이클 업턴 반영이나 PEG 2.31·PE 116은 과대평가 자동 bearish 조건. 사이클 정점 근처 고밸류에이션 매수는 가장 경계하는 패턴이다."},
        "fisher": {"signal": "bullish", "confidence": 68, "reasoning": "반도체 테스트 장비에서 44% 매출 성장과 R&D/Revenue 15.8%의 꾸준한 투자, 58% 총마진이 우수한 성장 기업 특성. 반도체 복잡도 증가와 AI 칩 테스트 수요는 장기 구조적 동력이며 경영진 일관성도 신뢰 충족."},
    },
    "DELL": {
        "buffett": {"signal": "neutral", "confidence": 40, "reasoning": "Dell은 PC와 서버에서 규모의 경제를 갖추나 ROE 데이터 결여와 영업마진 9.6% 박리다매 구조는 훌륭한 사업이 아님을 보여준다. AI 서버 매출 39.5% 성장은 commodity 조립 사업으로 pricing power 부재."},
        "lynch": {"signal": "bullish", "confidence": 73, "reasoning": "IT 인프라 Stalwart로 PEG 0.9·매출 40% 성장은 AI 서버 수요 폭발 GARP 조건 충족이며 PE 25는 합리적이다. $140B 시가총액으로 텐배거 불가하나 2-3배 Stalwart 플레이로 편안하게 보유할 종목이다."},
        "fisher": {"signal": "neutral", "confidence": 42, "reasoning": "AI 서버 수요로 39.5% 매출 성장은 인상적이나 총마진 20%·영업마진 9.6%의 낮은 수익성은 하드웨어 조립업체의 한계. R&D/Revenue 2.8%는 혁신 기준 미달이며 차별화된 기술 해자보다 공급망 효율성에 의존하는 구조."},
    },
    "TFX": {
        "buffett": {"signal": "bearish", "confidence": 70, "reasoning": "Teleflex는 의료기기 브랜드 지위가 있으나 ROE 1.6%는 자본 효율성이 극히 저조하고 P/E 103.1·D/E 0.90은 투자 매력을 완전히 소멸시킨다. 일관된 ROE 15% 이상 기준을 전혀 충족하지 못한다."},
        "lynch": {"signal": "bearish", "confidence": 72, "reasoning": "의료기기 기업으로 선호 업종이나 PE 103·ROE 2%·D/E 0.90의 조합은 고가에 저수익 고부채 최악 프로파일이며 EPS 성장 데이터 부재와 낮은 이익률은 즉시 통과(pass) 사유다."},
        "fisher": {"signal": "bearish", "confidence": 72, "reasoning": "의료기기 기업임에도 R&D/Revenue 7.3%는 업계 평균 이하, 순이익률 -45%의 대규모 손실, D/E 0.90의 부채 부담. 마진 악화·고부채·경영진 실행력 의문이 동시에 나타난 강한 기피 대상이다."},
    },
    "GLW": {
        "buffett": {"signal": "bearish", "confidence": 60, "reasoning": "Corning은 광섬유와 특수유리에서 기술 경쟁 우위를 가지나 ROE 14.9%는 15% 기준에 미치지 못하고 P/E 91.8은 심각한 고평가. 자본 집약적 사업과 D/E 0.76의 부채 수준은 투자 원칙에 부합하지 않는다."},
        "lynch": {"signal": "neutral", "confidence": 55, "reasoning": "광섬유·디스플레이 소재 Stalwart로 AI 데이터센터 광섬유 수요 성장 스토리는 명확하나 PEG 1.96·PE 92·D/E 0.76은 고평가 영역. 매출 20% 지속 시 재평가 여지는 있으나 현재 GARP 매력 부족."},
        "fisher": {"signal": "neutral", "confidence": 50, "reasoning": "광섬유/특수유리 분야의 기술 전문성과 20.4% 매출 성장, R&D/Revenue 7.1%는 긍정적이나 영업마진 16.6%·D/E 0.76은 우수 기업 기준에 미치지 못하며 장기 수십 년 성장 경로의 확신 불충분."},
    },
    "STX": {
        "buffett": {"signal": "bearish", "confidence": 75, "reasoning": "Seagate는 ROE 데이터 결여와 D/E 10.5의 극도 고부채 구조, P/E 67.3이 결합된 최악의 조합. HDD 사업은 SSD와 클라우드 스토리지에 의해 구조적으로 잠식되고 있어 절대 보유하지 않을 기업이다."},
        "lynch": {"signal": "bullish", "confidence": 68, "reasoning": "HDD 사이클리컬로 AI 데이터센터 대용량 스토리지 수요 폭발이 명확한 성장 스토리이며 PEG 0.89·매출 22% 성장은 GARP 조건 충족하나 D/E 10.5의 극단 부채는 최대 리스크다."},
        "fisher": {"signal": "neutral", "confidence": 45, "reasoning": "HDD 사이클 회복으로 21.5% 매출 성장과 영업마진 29.9%는 인상적이나 R&D/Revenue 8%와 극단 부채(D/E 10.5)는 기피 재무 구조. HDD 시장의 구조적 축소 위협과 과도한 레버리지가 장기 안정 성장 원칙과 충돌."},
    },
    "FSLR": {
        "buffett": {"signal": "bearish", "confidence": 65, "reasoning": "First Solar는 태양광 패널 제조로 P/E 13.9·낮은 부채(D/E 0.07)가 매력적이나 태양광 산업은 중국 경쟁과 보조금 의존도로 지속 가능한 경쟁 우위가 어렵다. 에너지·원자재 사업은 투자 원칙에 맞지 않는다."},
        "lynch": {"signal": "bullish", "confidence": 74, "reasoning": "태양광 Fast Grower로 PEG 0.49·PE 14·ROE 17%·D/E 0.07은 꿈꾸는 저부채 저PEG 성장주 프로파일이며 $21B 시가총액은 텐배거 잠재력 범위 내, 미국 에너지 전환이 강력한 성장 촉매다."},
        "fisher": {"signal": "neutral", "confidence": 55, "reasoning": "미국산 박막 태양광 패널의 IRA 혜택과 11.1% 매출 성장, 총마진 40.6%는 경쟁 우위이나 R&D/Revenue 4.5%는 충분치 않고 정책 리스크가 장기 불확실성 증가. 중국 패널 비용 경쟁이 해자 지속성 제한."},
    },
    "INCY": {
        "buffett": {"signal": "bullish", "confidence": 62, "reasoning": "Incyte는 ROE 29.9%·P/E 14.9·D/E 0.011의 무부채에 가까운 구조로 재무 건전성이 탁월하며 27.8% 매출 성장이 인상적. Jakafi 특허 보호의 혈액암 franchise가 moat이나 신약 파이프라인 의존성은 circle of competence 밖 리스크다."},
        "lynch": {"signal": "bullish", "confidence": 78, "reasoning": "바이오제약 Fast Grower로 PEG 0.34·PE 15·ROE 30%는 최애 조합이며 $19B 시가총액은 아직 텐배거 가능성, 매출 28%·D/E 0.01의 무부채 구조는 샴페인을 열 이유라 표현할 수준이다."},
        "fisher": {"signal": "bullish", "confidence": 73, "reasoning": "항암제/혈액암 임상 파이프라인 기반 27.8% 매출 성장과 R&D/Revenue 40%의 압도적 연구 투자는 제약·바이오에서 요구하는 혁신 역량의 최고 기준. 무부채(D/E 0.01)·ROE 30%·Jakafi 이후 파이프라인 다각화 전략이 장기 지속성을 뒷받침한다."},
    },
    "LRCX": {
        "buffett": {"signal": "neutral", "confidence": 50, "reasoning": "Lam Research는 반도체 식각 장비에서 과점적 지위와 ROE 66.8%·영업마진 35%가 매우 인상적이나 P/E 49.0과 강한 사이클리컬 특성, 중국 수출 규제 리스크가 안전마진 없는 가격을 정당화하기 어렵게 한다."},
        "lynch": {"signal": "neutral", "confidence": 58, "reasoning": "반도체 장비 Cyclical로 ROE 67%는 경이롭고 매출 24% 성장은 업사이클 반영이나 PEG 1.59·$324B 시가총액에서 합리적 가격 대비 규모가 너무 커 텐배거 불가, 사이클 전환 리스크 감안 시 중립."},
        "fisher": {"signal": "bullish", "confidence": 84, "reasoning": "반도체 식각 장비 세계 1위로 23.8% 매출 성장·R&D/Revenue 11.4%·영업마진 35%·ROE 67%는 15가지 체크리스트 대부분을 충족하는 탁월한 기업 지표. AI 미세화와 GAA 트랜지스터 전환이 구조적 동력이다."},
    },
    "QRVO": {
        "buffett": {"signal": "bearish", "confidence": 65, "reasoning": "Qorvo는 RF 반도체에서 스마트폰 OEM 의존도가 높고 ROE 9.6%의 낮은 자본 효율성, Qualcomm·Skyworks 치열한 경쟁으로 지속 가능한 moat가 없다. 회피하는 commodity 기술 사업이다."},
        "lynch": {"signal": "bullish", "confidence": 71, "reasoning": "RF 반도체 Turnaround·Cyclical로 PEG 0.21·시가총액 $8B는 소형 저평가 텐배거 후보 프로파일에 부합하며 EPS 298% 성장은 사이클 회복의 폭발적 레버리지. 다만 D/E 0.42와 매출 8% 회복세는 리스크다."},
        "fisher": {"signal": "bearish", "confidence": 62, "reasoning": "RF 반도체 분야 매출 8.4%로 부진하고 Apple 스마트폰 사이클 의존도가 높아 독립적 성장 동력 취약. R&D/Revenue 20%는 양호하나 영업마진 20%·ROE 10%의 낮은 수익성과 치열한 경쟁이 차별적 우위 약화."},
    },
    "DVA": {
        "buffett": {"signal": "neutral", "confidence": 48, "reasoning": "DaVita는 투석 센터 운영의 규모와 규제 장벽으로 moat가 있고 ROE 64.8%는 인상적이나 D/E 11.1의 극도 고부채가 핵심 리스크다. P/E 15.8은 적정이나 메디케어 의존도와 가정 투석 전환이 장기 전망을 불투명하게 한다."},
        "lynch": {"signal": "neutral", "confidence": 57, "reasoning": "투석 서비스 Stalwart로 이해하기 쉽고 반복 수입 모델이며 PEG 0.56·ROE 65%는 인상적이나 EPS 성장이 실제 -10%이고 D/E 11.1의 극단 레버리지는 부채로 만든 ROE 경고 상황이다."},
        "fisher": {"signal": "neutral", "confidence": 48, "reasoning": "투석 서비스 분야의 규제적 진입장벽과 9.9% 매출 성장, ROE 65%는 방어적 특성이나 R&D 투자 전무, D/E 11.1의 극단 레버리지가 재무 건전성 기준 심각 위배. 혁신 성장 부재 구조는 핵심 철학과 거리."},
    },
    "HWM": {
        "buffett": {"signal": "neutral", "confidence": 45, "reasoning": "Howmet Aerospace는 항공 엔진 부품의 강력한 기술 moat와 ROE 30.4%가 양호하나 P/E 65.1은 안전마진 없이 미래 성장을 과도하게 반영. 항공우주 부품 사업은 기술 복잡성으로 이해 범위를 벗어나며 사이클리컬이다."},
        "lynch": {"signal": "neutral", "confidence": 60, "reasoning": "항공우주 부품 Fast Grower로 PEG 0.8은 저평가 시사하나 PE 65·$97B 시가총액에서 텐배거 불가, 매출 15%·ROE 30%는 양호하나 D/E 0.60과 높은 PE의 조합에서 좋은 사업 비싼 가격으로 분류된다."},
        "fisher": {"signal": "neutral", "confidence": 55, "reasoning": "항공우주 엔진 부품 구조적 공급 부족으로 14.6% 성장과 영업마진 26.3%·ROE 30%는 양호하나 R&D/Revenue 0.45%에 불과해 가장 엄격한 혁신 기준 미충족. 기술 혁신이 아닌 특수 합금 제조 역량 의존."},
    },
    "TRV": {
        "buffett": {"signal": "bullish", "confidence": 78, "reasoning": "The Travelers는 ROE 25.3%·D/E 0.29·P/E 9.1로 Berkshire가 보유했던 보험사들과 유사한 탁월한 프로필. 미국 최대 상업 재산보험사 중 하나로서 강력한 underwriting 능력과 투자 포트폴리오의 float 수익이 지속적 경쟁 우위를 형성한다."},
        "lynch": {"signal": "neutral", "confidence": 55, "reasoning": "보험 Stalwart로 PE 9.1은 매력적이나 PEG 2.36은 EPS 성장의 일시성을 반영하며 매출 1%의 정체는 Slow Grower에 가깝고 배당 보유는 가능하나 성장 플레이로는 부적합하다."},
        "fisher": {"signal": "bearish", "confidence": 60, "reasoning": "대형 보험사로서 R&D 투자 전무하고 핵심으로 보는 기술 혁신·영업 우월성·차별적 제품 경쟁력 측정 불가. 매출 1.0%는 정체 수준이며 보험 사이클 의존 수익 구조는 일관되게 기피해온 전통 금융업의 전형이다."},
    },
    "NEM": {
        "buffett": {"signal": "bearish", "confidence": 70, "reasoning": "Newmont는 세계 최대 금광업체이지만 금 채굴은 평생 회피해온 commodity 원자재 사업이다. ROE 25.8%와 P/E 15.1이 매력적으로 보여도 금값 의존적 수익 구조와 광산 운영의 지질학적 리스크는 예측 가능한 수익을 원하는 나에게 맞지 않는다."},
        "lynch": {"signal": "bearish", "confidence": 62, "reasoning": "금광 Asset Play·Cyclical로 금값 상승 매출 46% 성장은 인상적이나 PEG 2.78·PE 15·$124B 시가총액은 이미 금 가격 상승 충분히 반영. 원자재 기업의 사이클 정점 매수를 극도로 경계, 광산주는 가장 싫어하는 카테고리다."},
        "fisher": {"signal": "bearish", "confidence": 65, "reasoning": "금 가격 급등으로 45.8% 매출 성장·총마진 67.4%이나 R&D/Revenue 0.7%에 불과한 채굴 기업은 명시적으로 경계하는 상품 가격 의존적 사업. 자원 고갈과 채굴 비용 상승 리스크로 영속적 경쟁 우위 부적합."},
    },
    "QCOM": {
        "buffett": {"signal": "neutral", "confidence": 55, "reasoning": "Qualcomm은 모바일 칩과 특허 라이선싱의 강력한 moat와 ROE 21.5%·영업마진 27.5%가 양호하나 Apple 자체 칩 개발, 중국 스마트폰 리스크, P/E 30.3 결합 시 충분한 안전마진 부재이며 기술 변화 속도가 우려된다."},
        "lynch": {"signal": "neutral", "confidence": 60, "reasoning": "통신 반도체 Stalwart로 PEG 0.64는 저평가 시사하나 EPS 성장이 실제 -5.5% 감소 중이며 매출 5% 정체와 D/E 0.64는 성장이 멈춘 거대주로 분류될 프로파일. AI 엣지 컴퓨팅 회복 촉매 가능하나 현재는 증거 부족."},
        "fisher": {"signal": "neutral", "confidence": 52, "reasoning": "스냅드래곤 AI 모바일 칩 확대와 R&D/Revenue 20.4%의 충분한 투자는 긍정적이나 매출 5% 정체와 중국 스마트폰 의존도, 특허 분쟁 리스크가 지속 가능한 성장 동력 신뢰를 낮춘다. ARM 자체칩 전환 트렌드도 우려."},
    },
    "MPWR": {
        "buffett": {"signal": "bearish", "confidence": 60, "reasoning": "Monolithic Power는 전력 관리 반도체에서 성장하는 틈새 기업이나 ROE 19.2%가 그다지 탁월하지 않고 P/E 123.5는 극도의 고평가다. 무부채(D/E 0.007)는 훌륭하나 현재 가격에서 어떤 합리적 내재가치 계산도 매수를 지지하지 않는다."},
        "lynch": {"signal": "bearish", "confidence": 70, "reasoning": "전력 반도체 Fast Grower로 비즈니스 품질은 탁월하나 PEG 2.68·PE 123·EPS 성장 실제 -86%의 역성장 조합은 성장 없는 고PEG 즉시 매도 원칙에 해당. $78B 규모에서 텐배거 불가에 밸류에이션까지 과도하다."},
        "fisher": {"signal": "bullish", "confidence": 78, "reasoning": "AI 데이터센터 전력 관리 IC 수요로 20.8% 매출 성장·R&D/Revenue 13.7%·영업마진 26.6%·총마진 55.2%는 선호하는 고마진 반도체 틈새 강자 특성. 무부채(D/E 0.007)와 AI 전력 효율화는 구조적 동력으로 장기 보유 부합."},
    },
    "TXN": {
        "buffett": {"signal": "bullish", "confidence": 70, "reasoning": "Texas Instruments는 아날로그·임베디드 반도체에서 수십 년의 고객 충성도와 광범위한 제품 포트폴리오로 훌륭한 moat이며 ROE 32.3%·영업마진 37.8%가 탁월하다. 산업·자동차 시장 집중과 자체 fab 보유로 경쟁 우위가 지속적이며 TI 품질에는 프리미엄이 정당하다."},
        "lynch": {"signal": "neutral", "confidence": 58, "reasoning": "아날로그 반도체 Stalwart로 ROE 32%·D/E 0.84·PEG 1.57은 좋은 사업이지만 약간 비싼 구간이며 매출 19%·EPS 31% 회복은 긍정적이나 $245B 메가캡 접근에서 텐배거 가능성 없고 Stalwart 복리 수익 보유 전략 적합."},
        "fisher": {"signal": "bullish", "confidence": 76, "reasoning": "아날로그·임베디드 반도체 분야 18.6% 매출 성장 회복과 R&D/Revenue 11.8%·영업마진 37.8%는 오랫동안 선호해온 기술 독점적 수익 구조. 자체 FAB 전략과 산업·자동차 다각화 고객 기반, 수십 년 일관된 사업 집중은 Hold Winners 원칙 부합."},
    },
    "HPE": {
        "buffett": {"signal": "bearish", "confidence": 72, "reasoning": "Hewlett Packard Enterprise는 ROE -0.5%의 손실 구조, 영업마진 7.6%, D/E 0.87의 결합된 회피 기업의 전형이다. 클라우드 전환 속에서 서버·스토리지 하드웨어 사업의 경쟁 우위가 지속 약화되고 있으며 적자 기업은 절대 매수하지 않는다."},
        "lynch": {"signal": "bearish", "confidence": 66, "reasoning": "IT 인프라 Slow Grower로 PEG 0.85·PE 10.5는 저렴해 보이나 ROE -0.5%·EPS 성장 -28%의 실적 역성장과 D/E 0.87 부채 부담은 싸다고 다 좋은 게 아니다, 떨어지는 칼은 잡지 말라 경고할 전형적 가치 함정이다."},
        "fisher": {"signal": "bearish", "confidence": 70, "reasoning": "AI 서버(GreenLake) 전환 노력으로 18.4% 매출 성장이나 순이익률 -0.33%로 사실상 손실 구조이고 ROE 음수다. R&D/Revenue 7.3%에도 하이브리드 클라우드 시장 Dell/AWS 경쟁 열위, 낮은 마진 구조는 영업 우월성 없는 성숙 IT 하드웨어 기업이다."},
    },
}


@dataclass
class Position:
    rank: int
    ticker: str
    name: str
    market_cap: str
    market_cap_value: float
    sector: str
    pe: float | None
    roe: float | None
    peg: float | None
    expected_return: float
    bullish_count: int
    bearish_count: int
    neutral_count: int
    combined_confidence: float
    weight: float = 0.0


def adjusted_score(signal: str, confidence: float) -> float:
    if signal == "bullish":
        return float(confidence)
    if signal == "bearish":
        return 100.0 - float(confidence)
    return 50.0


def build_portfolio() -> tuple[list[Position], dict[str, str], dict[str, dict]]:
    eligible: list[Position] = []
    rejection_reasons: dict[str, str] = {}
    investor_signal_summary: dict[str, dict] = {}

    for row in PREDICT_TOP30:
        analyses = INVESTOR_ANALYSES[row.ticker]
        bullish = sum(1 for a in analyses.values() if a["signal"] == "bullish")
        bearish = sum(1 for a in analyses.values() if a["signal"] == "bearish")
        neutral = 3 - bullish - bearish
        confidence = sum(adjusted_score(a["signal"], a["confidence"]) for a in analyses.values()) / 3.0
        investor_signal_summary[row.ticker] = {
            "bullish": bullish, "bearish": bearish, "neutral": neutral, "confidence": confidence,
        }
        if bullish >= 2:
            eligible.append(Position(
                rank=row.rank, ticker=row.ticker, name=row.name,
                market_cap=row.market_cap, market_cap_value=row.market_cap_value,
                sector=row.sector, pe=row.pe, roe=row.roe, peg=row.peg,
                expected_return=row.expected_return, bullish_count=bullish,
                bearish_count=bearish, neutral_count=neutral,
                combined_confidence=confidence,
            ))
        else:
            rejection_reasons[row.ticker] = f"투자자 과반 미달 ({bullish}/3 bullish)"

    total_conf = sum(p.combined_confidence for p in eligible)
    for p in eligible:
        p.weight = (p.combined_confidence / total_conf) * 100.0

    eligible = [p for p in eligible if p.weight >= 2.0]
    if eligible:
        total = sum(p.weight for p in eligible)
        for p in eligible:
            p.weight = p.weight * 100.0 / total

    apply_caps(eligible, single_cap=15.0, sector_cap=35.0)
    eligible.sort(key=lambda p: -p.weight)
    return eligible, rejection_reasons, investor_signal_summary


def apply_caps(positions: list[Position], single_cap: float, sector_cap: float) -> None:
    for _ in range(20):
        changed = False
        # Sector cap
        sectors: dict[str, list[Position]] = defaultdict(list)
        for p in positions:
            sectors[p.sector].append(p)
        for sec, members in sectors.items():
            sec_weight = sum(m.weight for m in members)
            if sec_weight > sector_cap + 0.01:
                scale = sector_cap / sec_weight
                for m in members:
                    m.weight *= scale
                changed = True
        # Single cap
        excess = 0.0
        capped: set[str] = set()
        for p in positions:
            if p.weight > single_cap + 0.01:
                excess += p.weight - single_cap
                p.weight = single_cap
                capped.add(p.ticker)
                changed = True
        # Redistribute excess proportionally to non-capped, non-sector-saturated positions
        if excess > 0.01:
            sectors_now: dict[str, list[Position]] = defaultdict(list)
            for p in positions:
                sectors_now[p.sector].append(p)
            free: list[Position] = []
            for p in positions:
                sec_total = sum(m.weight for m in sectors_now[p.sector])
                if p.ticker not in capped and sec_total < sector_cap - 0.01 and p.weight < single_cap - 0.01:
                    free.append(p)
            free_total = sum(p.weight for p in free)
            if free_total > 0.01:
                for p in free:
                    p.weight += excess * (p.weight / free_total)
                changed = True
        if not changed:
            return


# ---------------------------------------------------------------------------
# Excel rendering
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", size=12, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="2E75B6")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="FFFFFF")
BUY_FILL = PatternFill("solid", fgColor="E2EFDA")
SELL_FILL = PatternFill("solid", fgColor="FCE4EC")
NEUTRAL_FILL = PatternFill("solid", fgColor="F5F5F5")
TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
INCLUDED_FILL = PatternFill("solid", fgColor="EAF6E6")
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def write_header(ws, row: int, headers: list[str]) -> None:
    for idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN
    ws.row_dimensions[row].height = 22


def auto_width(ws, max_col: int, min_w: int = 8, max_w: int = 30) -> None:
    for col_idx in range(1, max_col + 1):
        col = get_column_letter(col_idx)
        widest = min_w
        for cell in ws[col]:
            try:
                length = len(str(cell.value)) if cell.value is not None else 0
                widest = max(widest, length + 2)
            except Exception:
                continue
        ws.column_dimensions[col].width = min(max_w, widest)


def signal_color(signal: str):
    if signal == "bullish":
        return BUY_FILL
    if signal == "bearish":
        return SELL_FILL
    return NEUTRAL_FILL


def build_excel(positions: list[Position], rejection_reasons: dict[str, str],
                investor_summary: dict[str, dict]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    cash_weight = max(0.0, 100.0 - sum(p.weight for p in positions))
    build_summary_sheet(wb, positions, cash_weight)
    build_portfolio_sheet(wb, positions, cash_weight)
    build_ranking_sheet(wb, positions)
    build_matrix_sheet(wb, investor_summary)
    build_detail_sheet(wb)
    build_risk_sheet(wb, positions, rejection_reasons, investor_summary)
    return wb


def build_summary_sheet(wb: Workbook, positions: list[Position], cash_weight: float) -> None:
    ws = wb.create_sheet("요약")
    ws["A1"] = "AI Hedge Fund 포트폴리오 리포트"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 26

    invested = sum(p.weight for p in positions)
    avg_conf = sum(p.combined_confidence * p.weight for p in positions) / invested if invested else 0
    avg_return = sum(p.expected_return * p.weight for p in positions) / invested if invested else 0
    strong_buy_pct = sum(p.weight for p in positions if p.bullish_count == 3)

    rows = [
        ("분석 일자", "2026-04-28"),
        ("분석 대상", "S&P 500 상위 30개"),
        ("분석 전략", "하이브리드 (펀더멘털 70% + 모멘텀 30%)"),
        ("투자자 관점", "W.Buffett, P.Lynch, P.Fisher"),
        ("", ""),
        ("포트폴리오 통계", ""),
        ("편입 종목 수", f"{len(positions)} / 30 ({len(positions)/30*100:.1f}%)"),
        ("평균 신뢰도", f"{avg_conf:.0f}%"),
        ("평균 예상 수익률", f"{avg_return:+.1f}%"),
        ("만장일치 (3/3) 비중", f"{strong_buy_pct:.1f}%"),
        ("현금 비중", f"{cash_weight:.1f}%"),
        ("", ""),
        ("시가총액 분포", ""),
    ]

    cap_buckets = defaultdict(float)
    for p in positions:
        if p.market_cap_value > 200:
            cap_buckets["메가캡 (>$200B)"] += p.weight
        elif p.market_cap_value > 10:
            cap_buckets["대형주 ($10-200B)"] += p.weight
        elif p.market_cap_value > 2:
            cap_buckets["중형주 ($2-10B)"] += p.weight
        else:
            cap_buckets["소형주 (<$2B)"] += p.weight
    for label, weight in sorted(cap_buckets.items(), key=lambda x: -x[1]):
        rows.append((label, f"{weight:.1f}%"))
    if cash_weight > 0.01:
        rows.append(("현금", f"{cash_weight:.1f}%"))

    rows.append(("", ""))
    rows.append(("섹터 분포", ""))
    sec_buckets = defaultdict(float)
    for p in positions:
        sec_buckets[p.sector] += p.weight
    for sec, weight in sorted(sec_buckets.items(), key=lambda x: -x[1]):
        rows.append((sec, f"{weight:.1f}%"))
    if cash_weight > 0.01:
        rows.append(("Cash", f"{cash_weight:.1f}%"))

    for offset, (label, value) in enumerate(rows, start=3):
        ws.cell(row=offset, column=1, value=label).border = THIN
        ws.cell(row=offset, column=2, value=value).border = THIN
        if label in {"포트폴리오 통계", "시가총액 분포", "섹터 분포"}:
            ws.cell(row=offset, column=1).font = Font(bold=True)
            ws.cell(row=offset, column=1).fill = TOTAL_FILL

    auto_width(ws, max_col=3, min_w=20, max_w=40)


def build_portfolio_sheet(wb: Workbook, positions: list[Position], cash_weight: float) -> None:
    ws = wb.create_sheet("포트폴리오")
    headers = ["#", "종목코드", "회사명", "비중", "신호", "신뢰도", "예상수익률",
               "시가총액", "P/E", "ROE", "PEG", "합의", "섹터"]
    write_header(ws, 1, headers)

    invested = sum(p.weight for p in positions)
    weighted_conf = sum(p.combined_confidence * p.weight for p in positions) / invested if invested else 0
    weighted_return = sum(p.expected_return * p.weight for p in positions) / invested if invested else 0

    for idx, p in enumerate(positions, 1):
        signal = "강력매수" if p.bullish_count == 3 else "매수"
        row_num = idx + 1
        ws.cell(row=row_num, column=1, value=idx).number_format = "0"
        ws.cell(row=row_num, column=2, value=p.ticker)
        ws.cell(row=row_num, column=3, value=p.name)
        ws.cell(row=row_num, column=4, value=p.weight / 100.0).number_format = "0.0%"
        ws.cell(row=row_num, column=5, value=signal)
        ws.cell(row=row_num, column=6, value=p.combined_confidence).number_format = "0"
        ws.cell(row=row_num, column=7, value=p.expected_return / 100.0).number_format = "+0.0%;-0.0%"
        ws.cell(row=row_num, column=8, value=p.market_cap)
        ws.cell(row=row_num, column=9, value=p.pe).number_format = "0.0"
        ws.cell(row=row_num, column=10, value=(p.roe or 0) / 100.0).number_format = "0.0%"
        ws.cell(row=row_num, column=11, value=p.peg).number_format = "0.0"
        ws.cell(row=row_num, column=12, value=f"{p.bullish_count}/3")
        ws.cell(row=row_num, column=13, value=p.sector)
        for col in range(1, 14):
            cell = ws.cell(row=row_num, column=col)
            cell.border = THIN
            cell.fill = INCLUDED_FILL
            cell.alignment = Alignment(horizontal="center" if col != 3 else "left", vertical="center")

    if cash_weight > 0.01:
        row_num = len(positions) + 2
        ws.cell(row=row_num, column=1, value=len(positions) + 1)
        ws.cell(row=row_num, column=2, value="CASH")
        ws.cell(row=row_num, column=3, value="현금")
        ws.cell(row=row_num, column=4, value=cash_weight / 100.0).number_format = "0.0%"
        for col in range(5, 14):
            ws.cell(row=row_num, column=col, value="-")
        for col in range(1, 14):
            ws.cell(row=row_num, column=col).border = THIN
            ws.cell(row=row_num, column=col).fill = NEUTRAL_FILL

    total_row = len(positions) + (3 if cash_weight > 0.01 else 2)
    ws.cell(row=total_row, column=2, value="합계")
    cell = ws.cell(row=total_row, column=4, value=(invested + cash_weight) / 100.0)
    cell.number_format = "0.0%"
    ws.cell(row=total_row, column=6, value=weighted_conf).number_format = "0"
    ws.cell(row=total_row, column=7, value=weighted_return / 100.0).number_format = "+0.0%;-0.0%"
    for col in range(1, 14):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).font = Font(bold=True)
        ws.cell(row=total_row, column=col).border = THIN

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{total_row}"
    auto_width(ws, max_col=13, min_w=8, max_w=30)


def build_ranking_sheet(wb: Workbook, positions: list[Position]) -> None:
    ws = wb.create_sheet("순위")
    headers = ["순위", "종목코드", "회사명", "종합점수", "펀더멘털", "모멘텀", "앙상블",
               "신호", "예상수익률", "시가총액", "P/E", "P/B", "ROE", "매출성장률",
               "PEG", "편입여부"]
    write_header(ws, 1, headers)

    included = {p.ticker for p in positions}
    for idx, row in enumerate(PREDICT_TOP30, 1):
        is_included = row.ticker in included
        signal_kr = COMBINED_KR.get(row.signal, row.signal)
        row_num = idx + 1
        ws.cell(row=row_num, column=1, value=row.rank)
        ws.cell(row=row_num, column=2, value=row.ticker)
        ws.cell(row=row_num, column=3, value=row.name)
        ws.cell(row=row_num, column=4, value=row.total_score).number_format = "0.00"
        ws.cell(row=row_num, column=5, value=row.fundamental).number_format = "0.0"
        ws.cell(row=row_num, column=6, value=row.momentum).number_format = "0.0"
        ws.cell(row=row_num, column=7, value=row.ensemble).number_format = "0.0"
        ws.cell(row=row_num, column=8, value=signal_kr)
        ws.cell(row=row_num, column=9, value=row.expected_return / 100.0).number_format = "+0.0%;-0.0%"
        ws.cell(row=row_num, column=10, value=row.market_cap)
        ws.cell(row=row_num, column=11, value=row.pe).number_format = "0.0"
        ws.cell(row=row_num, column=12, value="-")
        ws.cell(row=row_num, column=13, value=(row.roe or 0) / 100.0).number_format = "0.0%"
        rg = (row.revenue_growth or 0) / 100.0 if row.revenue_growth is not None else None
        ws.cell(row=row_num, column=14, value=rg).number_format = "0.0%"
        ws.cell(row=row_num, column=15, value=row.peg).number_format = "0.0"
        ws.cell(row=row_num, column=16, value="예" if is_included else "아니오")
        for col in range(1, 17):
            ws.cell(row=row_num, column=col).border = THIN
            if is_included:
                ws.cell(row=row_num, column=col).fill = INCLUDED_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{len(PREDICT_TOP30)+1}"
    auto_width(ws, max_col=16, min_w=8, max_w=30)


def build_matrix_sheet(wb: Workbook, investor_summary: dict[str, dict]) -> None:
    ws = wb.create_sheet("투자자 매트릭스")
    headers = ["종목코드", "회사명"] + [INVESTOR_SHORT[i] for i in INVESTORS] + ["종합신호", "종합신뢰도"]
    write_header(ws, 1, headers)

    for idx, row in enumerate(PREDICT_TOP30, 1):
        analyses = INVESTOR_ANALYSES[row.ticker]
        row_num = idx + 1
        ws.cell(row=row_num, column=1, value=row.ticker).border = THIN
        ws.cell(row=row_num, column=2, value=row.name).border = THIN
        for j, inv in enumerate(INVESTORS, 3):
            a = analyses[inv]
            cell = ws.cell(row=row_num, column=j, value=f"{SIGNAL_KR[a['signal']]}({a['confidence']})")
            cell.fill = signal_color(a["signal"])
            cell.border = THIN
            cell.alignment = Alignment(horizontal="center")
        summary = investor_summary[row.ticker]
        if summary["bullish"] >= 2:
            combined = "강력매수" if summary["bullish"] == 3 else "매수"
            combined_color = BUY_FILL
        elif summary["bearish"] >= 2:
            combined = "매도"
            combined_color = SELL_FILL
        else:
            combined = "보유"
            combined_color = NEUTRAL_FILL
        col = len(INVESTORS) + 3
        c = ws.cell(row=row_num, column=col, value=combined)
        c.fill = combined_color
        c.border = THIN
        c.alignment = Alignment(horizontal="center")
        c2 = ws.cell(row=row_num, column=col + 1, value=round(summary["confidence"]))
        c2.border = THIN
        c2.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(PREDICT_TOP30)+1}"
    auto_width(ws, max_col=len(headers), min_w=10, max_w=30)


def build_detail_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("투자자 상세")
    headers = ["종목코드", "회사명", "투자자", "신호", "신뢰도", "분석근거"]
    write_header(ws, 1, headers)

    row_num = 2
    for row in PREDICT_TOP30:
        for inv in INVESTORS:
            a = INVESTOR_ANALYSES[row.ticker][inv]
            ws.cell(row=row_num, column=1, value=row.ticker)
            ws.cell(row=row_num, column=2, value=row.name)
            ws.cell(row=row_num, column=3, value=INVESTOR_LABELS[inv])
            ws.cell(row=row_num, column=4, value=SIGNAL_KR[a["signal"]])
            ws.cell(row=row_num, column=5, value=a["confidence"])
            ws.cell(row=row_num, column=6, value=a["reasoning"]).alignment = Alignment(wrap_text=True, vertical="top")
            for col in range(1, 7):
                cell = ws.cell(row=row_num, column=col)
                cell.border = THIN
                if col == 4:
                    cell.fill = signal_color(a["signal"])
            ws.row_dimensions[row_num].height = 60
            row_num += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 90
    ws.freeze_panes = "A2"


def build_risk_sheet(wb: Workbook, positions: list[Position], rejection_reasons: dict[str, str],
                     investor_summary: dict[str, dict]) -> None:
    ws = wb.create_sheet("리스크 분석")
    ws["A1"] = "리스크 분석"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 26

    cash_weight = max(0.0, 100.0 - sum(p.weight for p in positions))
    weights = sorted([p.weight for p in positions], reverse=True)
    top1 = weights[0] if weights else 0
    top3 = sum(weights[:3])
    top5 = sum(weights[:5])
    top10 = sum(weights[:10])
    hhi = sum((w / 100.0) ** 2 for w in weights)

    sec_weights = defaultdict(float)
    for p in positions:
        sec_weights[p.sector] += p.weight
    sorted_sec = sorted(sec_weights.items(), key=lambda x: -x[1])
    max_sector = sorted_sec[0] if sorted_sec else (None, 0)
    top3_sec_sum = sum(w for _, w in sorted_sec[:3])

    unanimous = sum(1 for p in positions if p.bullish_count == 3)
    unanimous_pct = unanimous / len(positions) * 100 if positions else 0
    split_n = sum(1 for s in investor_summary.values() if s["bullish"] >= 1 and s["bearish"] >= 1)

    rejected_total = 30 - len(positions)

    rows = [
        ("집중도 지표", ""),
        ("상위 1종목 비중", f"{top1:.1f}%"),
        ("상위 3종목 비중", f"{top3:.1f}%"),
        ("상위 5종목 비중", f"{top5:.1f}%"),
        ("상위 10종목 비중", f"{top10:.1f}%"),
        ("HHI (허핀달 지수)", f"{hhi:.3f}"),
        ("현금 비중", f"{cash_weight:.1f}%"),
        ("", ""),
        ("섹터 집중도", ""),
        ("최대 섹터", f"{max_sector[0]} ({max_sector[1]:.1f}%)"),
        ("상위 3섹터 비중", f"{top3_sec_sum:.1f}%"),
        ("", ""),
        ("투자자 합의 품질", ""),
        ("만장일치 비율 (편입 종목)", f"{unanimous_pct:.1f}%"),
        ("의견 분산 종목 수 (전체 30)", f"{split_n}개"),
        ("", ""),
        ("비편입 사유 분포", ""),
        ("투자자 과반 미달", f"{rejected_total}개"),
    ]

    for offset, (label, value) in enumerate(rows, start=3):
        c1 = ws.cell(row=offset, column=1, value=label)
        c2 = ws.cell(row=offset, column=2, value=value)
        c1.border = THIN
        c2.border = THIN
        if label in {"집중도 지표", "섹터 집중도", "투자자 합의 품질", "비편입 사유 분포"}:
            c1.font = Font(bold=True)
            c1.fill = TOTAL_FILL

    auto_width(ws, max_col=2, min_w=20, max_w=50)


def render_text(positions: list[Position], rejection_reasons: dict[str, str],
                investor_summary: dict[str, dict]) -> str:
    bar = "═" * 100
    cash_weight = max(0.0, 100.0 - sum(p.weight for p in positions))
    invested = sum(p.weight for p in positions)
    avg_conf = sum(p.combined_confidence * p.weight for p in positions) / invested if invested else 0
    avg_return = sum(p.expected_return * p.weight for p in positions) / invested if invested else 0
    icon = {"bullish": "🟢", "neutral": "⚪", "bearish": "🔴"}

    lines = [
        bar, "📋 AI Hedge Fund 포트폴리오 리포트", bar,
        "분석 일자    : 2026-04-28",
        "분석 대상    : S&P 500 상위 30개 종목",
        "분석 전략    : 하이브리드 (펀더멘털 70% + 모멘텀 30%)",
        "투자자 관점  : Warren Buffett, Peter Lynch, Phil Fisher",
        "데이터 소스  : Yahoo Finance (predict) + LLM (investor-analysis)",
        bar, "",
        bar, "📊 포트폴리오 구성", bar,
        f"{'#':<3} {'종목':<26} {'비중':>6} {'시총':>9} {'종합신호':<10} {'신뢰도':>7} {'예상수익률':>10} {'P/E':>6} {'ROE':>7} {'합의':<5}",
        "-" * 100,
    ]
    for idx, p in enumerate(positions, 1):
        signal = "🟢 강력매수" if p.bullish_count == 3 else "🔵 매수    "
        roe_s = f"{p.roe:.1f}%" if p.roe is not None else "N/A"
        pe_s = f"{p.pe:.1f}" if p.pe is not None else "N/A"
        lines.append(
            f"{idx:<3} {p.ticker+'('+p.name[:18]+')':<26} {p.weight:>5.1f}% {p.market_cap:>9} "
            f"{signal:<10} {p.combined_confidence:>5.0f}%  {p.expected_return:>+9.1f}% {pe_s:>6} {roe_s:>7} {p.bullish_count}/3"
        )
    if cash_weight > 0.01:
        lines.append(f"{'':<3} {'CASH (현금)':<26} {cash_weight:>5.1f}% {'-':>9} {'-':<10} {'-':>6}  {'-':>9} {'-':>6} {'-':>7} {'-':<5}")
    lines.append("-" * 100)
    lines.append(f"{'   합계':<30} {invested+cash_weight:>5.1f}%   "
                 f"{'avg':>3} {avg_conf:>5.0f}%  avg {avg_return:+.1f}%")
    lines += ["", bar, "👥 투자자별 종목 분석 매트릭스", bar,
              f"{'종목':<26} {'W.Buffett':<22} {'P.Lynch':<22} {'P.Fisher':<22}", "-" * 100]
    for row in PREDICT_TOP30:
        a = INVESTOR_ANALYSES[row.ticker]
        cells = [f"{icon[a[inv]['signal']]} {a[inv]['signal']}({a[inv]['confidence']})" for inv in INVESTORS]
        lines.append(f"{row.ticker+'('+row.name[:18]+')':<26} {cells[0]:<22} {cells[1]:<22} {cells[2]:<22}")

    lines += ["", bar, "💬 투자자별 핵심 분석 근거 (편입 상위 5개 종목)", bar]
    for p in positions[:5]:
        signal = "🟢 강력매수" if p.bullish_count == 3 else "🔵 매수"
        lines.append("")
        lines.append(f"▸ {p.ticker} ({p.name}) — 종합: {signal} (신뢰도 {p.combined_confidence:.0f}%)")
        for inv in INVESTORS:
            a = INVESTOR_ANALYSES[p.ticker][inv]
            lines.append(f"  {INVESTOR_LABELS[inv]:<16}: {icon[a['signal']]} ({a['confidence']}%) {a['reasoning']}")

    unanimous = sum(1 for p in positions if p.bullish_count == 3)
    majority = sum(1 for p in positions if p.bullish_count == 2)
    sec_buckets = defaultdict(float)
    for p in positions:
        sec_buckets[p.sector] += p.weight
    lines += ["", bar, "📈 포트폴리오 요약", bar,
              f"편입 종목 수         : {len(positions)}개 / 분석 30개 ({len(positions)/30*100:.1f}% 편입률)",
              f"평균 신뢰도          : {avg_conf:.0f}%",
              f"평균 예상 수익률     : {avg_return:+.1f}%",
              f"강력매수(3/3) 비중   : {sum(p.weight for p in positions if p.bullish_count==3):.1f}%",
              f"매수(2/3) 비중       : {sum(p.weight for p in positions if p.bullish_count==2):.1f}%",
              f"현금 비중            : {cash_weight:.1f}%", "",
              "투자자 합의 분포:",
              f"  만장일치 (3/3)     : {unanimous}개 — " + ", ".join(p.ticker for p in positions if p.bullish_count == 3),
              f"  다수 합의 (2/3)    : {majority}개 — " + ", ".join(p.ticker for p in positions if p.bullish_count == 2),
              "", "섹터 분포:"]
    for sec, weight in sorted(sec_buckets.items(), key=lambda x: -x[1]):
        cap_note = " (제한 35% 내)" if 30 < weight <= 35.5 else ""
        lines.append(f"  {sec:<22}: {weight:.1f}%{cap_note}")
    if cash_weight > 0.01:
        lines.append(f"  {'Cash':<22}: {cash_weight:.1f}%")

    lines += ["", bar, f"🚫 비편입 종목 ({len(rejection_reasons)}개)", bar,
              f"{'종목':<26} {'순위':>4} {'점수':>6}  사유", "-" * 100]
    for row in PREDICT_TOP30:
        if row.ticker in rejection_reasons:
            lines.append(f"{row.ticker+'('+row.name[:18]+')':<26} {row.rank:>4} {row.total_score:>6.2f}  {rejection_reasons[row.ticker]}")

    lines += ["", bar, "⚠️ 리스크 및 경고", bar, "포트폴리오 집중도:",
              f"  상위 5종목 비중     : {sum(sorted([p.weight for p in positions], reverse=True)[:5]):.1f}%"]
    if positions:
        max_p = positions[0]
        lines.append(f"  최대 단일 종목      : {max_p.ticker} {max_p.weight:.1f}% (제한 15% 내)")
    sec_buckets_list = sorted(sec_buckets.items(), key=lambda x: -x[1])
    if sec_buckets_list:
        lines.append(f"  최대 섹터 비중      : {sec_buckets_list[0][0]} {sec_buckets_list[0][1]:.1f}% (제한 35% 내)")
    lines.append("")
    lines.append("투자자 의견 분산 종목 (1 bullish + 1 bearish 동시 존재):")
    for ticker, s in investor_summary.items():
        if s["bullish"] >= 1 and s["bearish"] >= 1:
            details = []
            for inv in INVESTORS:
                a = INVESTOR_ANALYSES[ticker][inv]
                if a["signal"] != "neutral":
                    details.append(f"{INVESTOR_SHORT[inv]} {icon[a['signal']]}{a['confidence']}")
            lines.append(f"  {ticker}: " + " vs ".join(details))
    lines += ["", bar, "💡 이 리포트는 교육/연구 목적이며 실제 투자 결정의 근거가 될 수 없습니다.",
              "   predict: Yahoo Finance | investor-analysis: LLM 기반 정성 분석",
              f"   엑셀 리포트: {OUTPUT_XLSX.relative_to(ROOT)}", bar]
    return "\n".join(lines) + "\n"


def main() -> None:
    positions, rejection_reasons, investor_summary = build_portfolio()
    wb = build_excel(positions, rejection_reasons, investor_summary)
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"Excel saved: {OUTPUT_XLSX}")
    text = render_text(positions, rejection_reasons, investor_summary)
    OUTPUT_TXT.write_text(text, encoding="utf-8")
    print(f"Text saved: {OUTPUT_TXT}")
    print()
    cash = max(0.0, 100.0 - sum(p.weight for p in positions))
    print(f"Portfolio size: {len(positions)} stocks (+ {cash:.1f}% cash)")
    for p in positions:
        print(f"  {p.ticker:5}  {p.weight:5.1f}%  {p.bullish_count}/3 bullish  conf {p.combined_confidence:.0f}")


if __name__ == "__main__":
    main()
