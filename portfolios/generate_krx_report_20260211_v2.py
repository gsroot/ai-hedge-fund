#!/usr/bin/env python3
"""KRX Portfolio Report Excel Generator - 2026-02-11"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style constants ──
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
BULLISH_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BULLISH_FONT = Font(color="375623")
BEARISH_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
BEARISH_FONT = Font(color="C62828")
NEUTRAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
NEUTRAL_FONT = Font(color="616161")
SUMMARY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
)

# ── Data ──
PORTFOLIO = [
    {"rank": 1, "ticker": "018290", "name": "브이티", "weight": 11.3, "signal": "강력매수", "confidence": 85, "ret": 14.1, "mcap": "₩7,233억", "pe": 7.5, "roe": 43.0, "peg": 0.53, "consensus": "3/3", "sector": "화장품/뷰티"},
    {"rank": 2, "ticker": "137400", "name": "피엔티", "weight": 10.8, "signal": "강력매수", "confidence": 82, "ret": 16.7, "mcap": "₩12,106억", "pe": 8.2, "roe": 23.0, "peg": 0.49, "consensus": "3/3", "sector": "이차전지장비"},
    {"rank": 3, "ticker": "033100", "name": "제룡전기", "weight": 10.4, "signal": "강력매수", "confidence": 79, "ret": 21.6, "mcap": "₩8,609억", "pe": 10.8, "roe": 41.0, "peg": 0.50, "consensus": "3/3", "sector": "전력기기"},
    {"rank": 4, "ticker": "192080", "name": "더블유게임즈", "weight": 10.2, "signal": "강력매수", "confidence": 77, "ret": 12.2, "mcap": "₩11,815억", "pe": 5.9, "roe": 16.0, "peg": 0.48, "consensus": "3/3", "sector": "게임"},
    {"rank": 5, "ticker": "067160", "name": "SOOP", "weight": 10.0, "signal": "강력매수", "confidence": 76, "ret": 12.4, "mcap": "₩8,230억", "pe": 7.5, "roe": 27.0, "peg": 0.60, "consensus": "3/3", "sector": "미디어/IT"},
    {"rank": 6, "ticker": "211050", "name": "인카금융서비스", "weight": 9.9, "signal": "매수", "confidence": 75, "ret": 14.0, "mcap": "₩6,941억", "pe": 11.2, "roe": 42.0, "peg": 0.80, "consensus": "2/3", "sector": "금융"},
    {"rank": 7, "ticker": "031980", "name": "피에스케이홀딩스", "weight": 9.8, "signal": "매수", "confidence": 74, "ret": 19.8, "mcap": "₩16,086억", "pe": 16.8, "roe": 22.0, "peg": 0.85, "consensus": "2/3", "sector": "반도체장비"},
    {"rank": 8, "ticker": "011200", "name": "HMM", "weight": 9.5, "signal": "강력매수", "confidence": 72, "ret": 13.7, "mcap": "₩20조", "pe": 4.2, "roe": 14.0, "peg": 0.31, "consensus": "3/3", "sector": "해운"},
    {"rank": 9, "ticker": "348210", "name": "넥스틴", "weight": 9.2, "signal": "매수", "confidence": 70, "ret": 14.6, "mcap": "₩8,172억", "pe": 20.8, "roe": 24.0, "peg": 1.42, "consensus": "2/3", "sector": "반도체장비"},
    {"rank": 10, "ticker": "000240", "name": "한국앤컴퍼니", "weight": 8.9, "signal": "매수", "confidence": 67, "ret": 16.2, "mcap": "₩2.9조", "pe": 8.2, "roe": 8.0, "peg": 0.51, "consensus": "2/3", "sector": "자동차부품"},
]

RANKINGS = [
    ("033100", "제룡전기", 9.17, 9.0, 10.0, 8.7, "강력매수", 21.6, "₩8,609억", 10.8, None, 41, None, None, True),
    ("031980", "피에스케이홀딩스", 8.65, 8.1, 10.0, 8.0, "강력매수", 19.8, "₩16,086억", 16.8, None, 22, None, None, True),
    ("131290", "티에스이", 7.95, 6.3, 10.0, 6.1, "매수", 17.3, "₩9,701억", 22.2, None, 11, None, None, False),
    ("137400", "피엔티", 7.76, 7.5, 9.0, 7.3, "매수", 16.7, "₩12,106억", 8.2, None, 23, None, None, True),
    ("095610", "테스", 7.66, 6.7, 10.0, 6.5, "매수", 16.3, "₩12,565억", 26.7, None, 13, None, None, False),
    ("000240", "한국앤컴퍼니", 7.63, 7.9, 7.0, 7.3, "매수", 16.2, "₩2.9조", 8.2, None, 8, None, None, True),
    ("036930", "주성엔지니어링", 7.48, 6.8, 9.0, 6.8, "매수", 15.7, "₩2.3조", 21.8, None, 19, None, None, False),
    ("222080", "씨아이에스", 7.27, 5.9, 9.0, 5.1, "매수", 15.0, "₩10,445억", 17.2, None, 12, None, None, False),
    ("348210", "넥스틴", 7.17, 7.5, 6.5, 7.2, "매수", 14.6, "₩8,172억", 20.8, None, 24, None, None, True),
    ("178320", "서진시스템", 7.11, 4.7, 10.0, 3.9, "매수", 14.4, "₩2.2조", 22.4, None, 10, None, None, False),
    ("018290", "브이티", 7.03, 7.9, 5.0, 7.2, "매수", 14.1, "₩7,233억", 7.5, None, 43, None, None, True),
    ("000660", "SK하이닉스", 7.01, 5.7, 10.0, 7.0, "매수", 14.0, "₩636조", 30.4, None, 27, None, None, False),
    ("211050", "인카금융서비스", 7.00, 6.8, 7.5, 6.0, "매수", 14.0, "₩6,941억", 11.2, None, 42, None, None, True),
    ("059090", "미코", 6.98, 5.3, 9.0, 4.8, "매수", 13.9, "₩5,551억", 28.9, None, 14, None, None, False),
    ("214150", "클래시스", 6.96, 6.1, 9.0, 6.2, "매수", 13.9, "₩4.4조", 44.6, None, 22, None, None, False),
    ("011200", "HMM", 6.92, 7.2, 5.5, 7.7, "매수", 13.7, "₩20조", 4.2, None, 14, None, None, True),
    ("095340", "ISC", 6.75, 4.9, 9.0, 5.7, "매수", 13.1, "₩3.6조", 64.0, None, 10, None, None, False),
    ("213420", "덕산네오룩스", 6.74, 6.2, 8.0, 6.0, "매수", 13.1, "₩10,814억", 23.4, None, 12, None, None, False),
    ("259960", "크래프톤", 6.72, 7.0, 5.0, 7.5, "매수", 13.0, "₩12조", 8.9, None, 19, None, None, False),
    ("319660", "피에스케이", 6.67, 5.2, 10.0, 6.5, "매수", 12.9, "₩16,830억", 21.3, None, 17, None, None, False),
    ("278470", "에이피알", 6.56, 4.2, 9.5, 4.8, "매수", 12.5, "₩11조", 99.1, None, 33, None, None, False),
    ("067160", "SOOP", 6.55, 6.8, 5.0, 7.3, "매수", 12.4, "₩8,230억", 7.5, None, 27, None, None, True),
    ("192080", "더블유게임즈", 6.50, 6.6, 6.0, 7.5, "매수", 12.2, "₩11,815억", 5.9, None, 16, None, None, True),
    ("323280", "태성", 6.49, 4.4, 9.0, 4.1, "매수", 12.2, "₩2.4조", 345.7, None, 15, None, None, False),
    ("352480", "씨앤씨인터내셔널", 6.45, 6.6, 6.0, 6.5, "매수", 12.1, "₩4,845억", 11.0, None, 16, None, None, False),
    ("161390", "한국타이어앤테크놀로지", 6.44, 5.3, 9.0, 6.2, "매수", 12.0, "₩9.0조", 8.0, None, 10, None, None, False),
    ("005930", "삼성전자", 6.43, 3.5, 10.0, 4.6, "매수", 12.0, "₩994조", 33.9, None, 9, None, None, False),
    ("101490", "에스앤에스텍", 6.42, 3.5, 10.0, 4.3, "매수", 12.0, "₩19,116억", 61.7, None, 12, None, None, False),
    ("079370", "제우스", 6.41, 4.9, 10.0, 5.6, "매수", 11.9, "₩5,689억", 13.5, None, 11, None, None, False),
    ("145020", "휴젤", 6.40, 5.5, 8.5, 6.6, "매수", 11.9, "₩3.3조", 21.3, None, 17, None, None, False),
]

INVESTOR_MATRIX = {
    "033100": {"buffett": ("매수", 78), "lynch": ("매수", 82), "fisher": ("매수", 78), "combined": "강력매수", "conf": 79},
    "031980": {"buffett": ("중립", 65), "lynch": ("매수", 75), "fisher": ("매수", 82), "combined": "매수", "conf": 74},
    "131290": {"buffett": ("중립", 52), "lynch": ("중립", 55), "fisher": ("중립", 65), "combined": "중립", "conf": 57},
    "137400": {"buffett": ("매수", 72), "lynch": ("매수", 88), "fisher": ("매수", 85), "combined": "강력매수", "conf": 82},
    "095610": {"buffett": ("매도", 45), "lynch": ("중립", 58), "fisher": ("중립", 62), "combined": "중립", "conf": 55},
    "000240": {"buffett": ("중립", 58), "lynch": ("매수", 72), "fisher": ("매수", 72), "combined": "매수", "conf": 67},
    "036930": {"buffett": ("중립", 62), "lynch": ("중립", 62), "fisher": ("중립", 68), "combined": "중립", "conf": 64},
    "222080": {"buffett": ("중립", 55), "lynch": ("중립", 68), "fisher": ("중립", 64), "combined": "중립", "conf": 62},
    "348210": {"buffett": ("매수", 68), "lynch": ("중립", 65), "fisher": ("매수", 76), "combined": "매수", "conf": 70},
    "178320": {"buffett": ("중립", 48), "lynch": ("중립", 60), "fisher": ("중립", 66), "combined": "중립", "conf": 58},
    "018290": {"buffett": ("매수", 82), "lynch": ("매수", 85), "fisher": ("매수", 88), "combined": "강력매수", "conf": 85},
    "000660": {"buffett": ("중립", 70), "lynch": ("중립", 45), "fisher": ("매수", 80), "combined": "중립", "conf": 65},
    "211050": {"buffett": ("매수", 75), "lynch": ("중립", 70), "fisher": ("매수", 81), "combined": "매수", "conf": 75},
    "059090": {"buffett": ("중립", 50), "lynch": ("중립", 52), "fisher": ("중립", 58), "combined": "중립", "conf": 53},
    "214150": {"buffett": ("중립", 58), "lynch": ("매도", 40), "fisher": ("중립", 55), "combined": "중립", "conf": 51},
    "011200": {"buffett": ("매수", 68), "lynch": ("매수", 78), "fisher": ("매수", 70), "combined": "강력매수", "conf": 72},
    "095340": {"buffett": ("매도", 42), "lynch": ("매도", 35), "fisher": ("매도", 45), "combined": "매도", "conf": 41},
    "213420": {"buffett": ("중립", 52), "lynch": ("중립", 58), "fisher": ("중립", 63), "combined": "중립", "conf": 58},
    "259960": {"buffett": ("중립", 65), "lynch": ("중립", 68), "fisher": ("매수", 75), "combined": "중립", "conf": 69},
    "319660": {"buffett": ("중립", 60), "lynch": ("중립", 62), "fisher": ("중립", 67), "combined": "중립", "conf": 63},
    "278470": {"buffett": ("매도", 38), "lynch": ("매도", 25), "fisher": ("매도", 40), "combined": "매도", "conf": 34},
    "067160": {"buffett": ("매수", 72), "lynch": ("매수", 80), "fisher": ("매수", 77), "combined": "강력매수", "conf": 76},
    "192080": {"buffett": ("매수", 70), "lynch": ("매수", 83), "fisher": ("매수", 79), "combined": "강력매수", "conf": 77},
    "323280": {"buffett": ("매도", 35), "lynch": ("매도", 20), "fisher": ("매도", 35), "combined": "매도", "conf": 30},
    "352480": {"buffett": ("중립", 62), "lynch": ("중립", 72), "fisher": ("중립", 64), "combined": "중립", "conf": 66},
    "161390": {"buffett": ("중립", 58), "lynch": ("중립", 65), "fisher": ("중립", 66), "combined": "중립", "conf": 63},
    "005930": {"buffett": ("중립", 65), "lynch": ("매도", 38), "fisher": ("중립", 62), "combined": "중립", "conf": 55},
    "101490": {"buffett": ("매도", 40), "lynch": ("매도", 30), "fisher": ("매도", 42), "combined": "매도", "conf": 37},
    "079370": {"buffett": ("중립", 55), "lynch": ("중립", 68), "fisher": ("중립", 65), "combined": "중립", "conf": 63},
    "145020": {"buffett": ("중립", 62), "lynch": ("중립", 60), "fisher": ("중립", 68), "combined": "중립", "conf": 63},
}

INVESTOR_DETAIL = [
    ("033100", "제룡전기", "W.Buffett", "매수", 78, "ROE 41%로 탁월한 자본효율성, P/E 10.8로 저평가, 전력기기 수요 예측 가능"),
    ("033100", "제룡전기", "P.Lynch", "매수", 82, "PEG 0.5로 매우 저평가된 fast grower, 시총 8,609억으로 10배주 잠재력"),
    ("033100", "제룡전기", "P.Fisher", "매수", 78, "ROE 41%로 경영진 우수성 입증, 탄소중립 시대 전력 인프라 수요 증가"),
    ("031980", "피에스케이홀딩스", "W.Buffett", "중립", 65, "ROE 22% 양호하나 P/E 16.8 적정, 반도체 장비 경기 민감도 높아 예측성 낮음"),
    ("031980", "피에스케이홀딩스", "P.Lynch", "매수", 75, "PEG 0.85 저평가된 stalwart, 반도체장비 수혜 명확, ROE 22% 우수"),
    ("031980", "피에스케이홀딩스", "P.Fisher", "매수", 82, "글로벌 반도체 투자 확대로 지속 성장, 기술 집약적 비즈니스로 진입장벽 높음"),
    ("137400", "피엔티", "W.Buffett", "매수", 72, "ROE 23% 우수, P/E 8.2로 저평가, 이차전지 장비 수요 증가"),
    ("137400", "피엔티", "P.Lynch", "매수", 88, "PEG 0.49 극도로 저평가된 fast grower, 이차전지 성장 스토리 명확, 10배 잠재력"),
    ("137400", "피엔티", "P.Fisher", "매수", 85, "이차전지 장비 수요 폭발적 증가, 기술 리더십으로 장기 성장동력 확보"),
    ("000240", "한국앤컴퍼니", "W.Buffett", "중립", 58, "ROE 8%로 낮음, P/E 8.2 저평가이나 자동차 부품 경쟁 치열, 해자 약함"),
    ("000240", "한국앤컴퍼니", "P.Lynch", "매수", 72, "PEG 0.51 저평가된 cyclical, 알루미늄휠 이해 쉬운 비즈니스"),
    ("000240", "한국앤컴퍼니", "P.Fisher", "매수", 72, "P/E 8.2 저평가, 알루미늄 경량화 트렌드로 수요 증가 전망"),
    ("348210", "넥스틴", "W.Buffett", "매수", 68, "ROE 24% 우수, P/E 20.8 합리적, 반도체 검사장비 기술력"),
    ("348210", "넥스틴", "P.Lynch", "중립", 65, "PEG 1.42 적정가, ROE 24% 우수하나 밸류에이션 제약"),
    ("348210", "넥스틴", "P.Fisher", "매수", 76, "반도체 검사장비 특화 기술로 진입장벽 높음, 마진 개선 가능"),
    ("018290", "브이티", "W.Buffett", "매수", 82, "ROE 43%로 탁월, P/E 7.5로 매우 저평가, 화장품 브랜드 해자 가능"),
    ("018290", "브이티", "P.Lynch", "매수", 85, "PEG 0.53 극도로 저평가, ROE 43%, 시총 7,233억 10배주 가능"),
    ("018290", "브이티", "P.Fisher", "매수", 88, "VT 브랜드 글로벌 확장 성공, 경영진 실행력 검증, R&D 지속 투자"),
    ("211050", "인카금융서비스", "W.Buffett", "매수", 75, "ROE 42% 탁월, P/E 11.2 저평가, 안정적 수익 모델"),
    ("211050", "인카금융서비스", "P.Lynch", "중립", 70, "PEG 0.80 저평가이나 보험대리점 경쟁 심화 위험"),
    ("211050", "인카금융서비스", "P.Fisher", "매수", 81, "ROE 42% 탁월, 경영진 주주친화적 배당 정책"),
    ("011200", "HMM", "W.Buffett", "매수", 68, "P/E 4.2 극단적 저평가, 해운 사이클 하단에서 안전마진 확보"),
    ("011200", "HMM", "P.Lynch", "매수", 78, "PEG 0.31 극도로 저평가된 cyclical, 해운 사이클 타이밍"),
    ("011200", "HMM", "P.Fisher", "매수", 70, "P/E 4.2 초저평가, 해운 업황 개선 시 수혜"),
    ("067160", "SOOP", "W.Buffett", "매수", 72, "ROE 27% 우수, P/E 7.5 저평가, 네트워크 효과로 해자 가능"),
    ("067160", "SOOP", "P.Lynch", "매수", 80, "PEG 0.60 저평가, 시총 8,230억 10배주 잠재력"),
    ("067160", "SOOP", "P.Fisher", "매수", 77, "인터넷 방송 플랫폼 독점적 지위, 콘텐츠 IP 확보"),
    ("192080", "더블유게임즈", "W.Buffett", "매수", 70, "P/E 5.9 매우 저평가, 소셜 카지노 반복 수익 모델"),
    ("192080", "더블유게임즈", "P.Lynch", "매수", 83, "PEG 0.48 극도로 저평가, 현금흐름 우수"),
    ("192080", "더블유게임즈", "P.Fisher", "매수", 79, "소셜카지노 글로벌 시장 1위, 경쟁우위 지속"),
]


def apply_border(ws, row, max_col):
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22


def signal_style(signal):
    if signal in ("매수", "강력매수"):
        return BULLISH_FILL, BULLISH_FONT
    elif signal == "매도":
        return BEARISH_FILL, BEARISH_FONT
    return NEUTRAL_FILL, NEUTRAL_FONT


def create_sheet1_summary(wb):
    ws = wb.active
    ws.title = "요약"
    set_col_widths(ws, [25, 30, 8, 8, 8, 8, 8, 8])
    # Title
    ws.merge_cells("A1:B1")
    cell = ws.cell(row=1, column=1, value="AI Hedge Fund 포트폴리오 리포트")
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    ws.cell(row=1, column=2).fill = TITLE_FILL
    ws.row_dimensions[1].height = 30

    data = [
        (3, "분석 일자", "2026-02-11"),
        (4, "분석 대상", "KRX 상위 30개"),
        (5, "분석 전략", "하이브리드"),
        (6, "투자자 관점", "W.Buffett, P.Lynch, P.Fisher"),
        (8, "포트폴리오 통계", ""),
        (9, "편입 종목 수", "10 / 30 (33.3%)"),
        (10, "평균 신뢰도", "76%"),
        (11, "평균 예상 수익률", "+15.5%"),
        (12, "강력매수 비중", "62.2%"),
        (13, "가중평균 P/E", "10.0"),
        (14, "가중평균 ROE", "26.5%"),
        (16, "시가총액 분포", ""),
        (17, "대형주 (₩10조+)", "9.5%"),
        (18, "중형주 (₩3-10조)", "8.9%"),
        (19, "소형주 (<₩3조)", "81.6%"),
        (21, "섹터 분포", ""),
        (22, "반도체장비", "19.0%"),
        (23, "화장품/뷰티", "11.3%"),
        (24, "이차전지장비", "10.8%"),
        (25, "전력기기", "10.4%"),
        (26, "게임", "10.2%"),
        (27, "미디어/IT", "10.0%"),
        (28, "금융", "9.9%"),
        (29, "해운", "9.5%"),
        (30, "자동차부품", "8.9%"),
    ]
    for r, label, value in data:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True) if value == "" else Font()
        ws.cell(row=r, column=2, value=value)
        if value == "":
            ws.cell(row=r, column=1).font = Font(bold=True, size=12, color="1F4E79")


def create_sheet2_portfolio(wb):
    ws = wb.create_sheet("포트폴리오")
    headers = ["#", "종목코드", "회사명", "비중", "신호", "신뢰도", "예상수익률", "시가총액", "P/E", "ROE", "PEG", "합의", "섹터"]
    set_col_widths(ws, [5, 10, 16, 8, 10, 8, 10, 12, 8, 8, 8, 8, 14])
    write_header_row(ws, 1, headers)

    for i, s in enumerate(PORTFOLIO, 2):
        ws.cell(row=i, column=1, value=s["rank"])
        ws.cell(row=i, column=2, value=s["ticker"])
        ws.cell(row=i, column=3, value=s["name"])
        ws.cell(row=i, column=4, value=s["weight"] / 100).number_format = "0.0%"
        sig_cell = ws.cell(row=i, column=5, value=s["signal"])
        fill, font = signal_style(s["signal"])
        sig_cell.fill = fill
        sig_cell.font = font
        ws.cell(row=i, column=6, value=s["confidence"])
        ws.cell(row=i, column=7, value=s["ret"] / 100).number_format = "+0.0%;-0.0%"
        ws.cell(row=i, column=8, value=s["mcap"])
        ws.cell(row=i, column=9, value=s["pe"]).number_format = "0.0"
        ws.cell(row=i, column=10, value=s["roe"] / 100).number_format = "0.0%"
        ws.cell(row=i, column=11, value=s["peg"]).number_format = "0.00"
        ws.cell(row=i, column=12, value=s["consensus"])
        ws.cell(row=i, column=13, value=s["sector"])
        for c in range(1, 14):
            ws.cell(row=i, column=c).border = THIN_BORDER
        ws.row_dimensions[i].height = 18

    # Summary row
    r = len(PORTFOLIO) + 2
    ws.cell(row=r, column=3, value="합계/평균").font = Font(bold=True)
    ws.cell(row=r, column=4, value=1.0).number_format = "0.0%"
    ws.cell(row=r, column=6, value=76)
    ws.cell(row=r, column=7, value=0.155).number_format = "+0.0%;-0.0%"
    for c in range(1, 14):
        ws.cell(row=r, column=c).fill = SUMMARY_FILL
        ws.cell(row=r, column=c).font = Font(bold=True)
        ws.cell(row=r, column=c).border = THIN_BORDER

    ws.auto_filter.ref = f"A1:M{len(PORTFOLIO) + 1}"
    ws.freeze_panes = "A2"


def create_sheet3_ranking(wb):
    ws = wb.create_sheet("순위")
    headers = ["순위", "종목코드", "회사명", "종합점수", "펀더멘털", "모멘텀", "앙상블", "신호", "예상수익률", "시가총액", "P/E", "P/B", "ROE", "매출성장률", "PEG", "편입여부"]
    set_col_widths(ws, [6, 10, 16, 8, 8, 8, 8, 10, 10, 12, 8, 8, 8, 10, 8, 8])
    write_header_row(ws, 1, headers)
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for i, r in enumerate(RANKINGS, 2):
        ticker, name, score, fund, mom, ens, sig, ret, mcap, pe, pb, roe, rev_g, peg, included = r
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=ticker)
        ws.cell(row=i, column=3, value=name)
        ws.cell(row=i, column=4, value=score).number_format = "0.00"
        ws.cell(row=i, column=5, value=fund).number_format = "0.0"
        ws.cell(row=i, column=6, value=mom).number_format = "0.0"
        ws.cell(row=i, column=7, value=ens).number_format = "0.0"
        ws.cell(row=i, column=8, value=sig)
        ws.cell(row=i, column=9, value=ret / 100 if ret else None)
        if ws.cell(row=i, column=9).value is not None:
            ws.cell(row=i, column=9).number_format = "+0.0%;-0.0%"
        ws.cell(row=i, column=10, value=mcap)
        ws.cell(row=i, column=11, value=pe).number_format = "0.0" if pe else ""
        ws.cell(row=i, column=12, value=pb)
        ws.cell(row=i, column=13, value=roe / 100 if roe else None)
        if ws.cell(row=i, column=13).value is not None:
            ws.cell(row=i, column=13).number_format = "0.0%"
        ws.cell(row=i, column=14, value=rev_g)
        ws.cell(row=i, column=15, value=peg)
        ws.cell(row=i, column=16, value="예" if included else "아니오")

        for c in range(1, 17):
            ws.cell(row=i, column=c).border = THIN_BORDER
            if included:
                ws.cell(row=i, column=c).fill = green_fill
        ws.row_dimensions[i].height = 18

    ws.auto_filter.ref = f"A1:P{len(RANKINGS) + 1}"
    ws.freeze_panes = "A2"


def create_sheet4_matrix(wb):
    ws = wb.create_sheet("투자자 매트릭스")
    headers = ["종목코드", "회사명", "W.Buffett", "P.Lynch", "P.Fisher", "종합신호", "종합신뢰도"]
    set_col_widths(ws, [10, 16, 14, 14, 14, 10, 10])
    write_header_row(ws, 1, headers)
    tickers_ordered = [r[0] for r in RANKINGS]

    for i, ticker in enumerate(tickers_ordered, 2):
        name = next(r[1] for r in RANKINGS if r[0] == ticker)
        m = INVESTOR_MATRIX[ticker]
        ws.cell(row=i, column=1, value=ticker)
        ws.cell(row=i, column=2, value=name)
        for ci, inv in enumerate(["buffett", "lynch", "fisher"], 3):
            sig, conf = m[inv]
            cell = ws.cell(row=i, column=ci, value=f"{sig}({conf})")
            fill, font = signal_style(sig)
            cell.fill = fill
            cell.font = font
        ws.cell(row=i, column=6, value=m["combined"])
        f6, fn6 = signal_style(m["combined"])
        ws.cell(row=i, column=6).fill = f6
        ws.cell(row=i, column=6).font = fn6
        ws.cell(row=i, column=7, value=m["conf"])
        for c in range(1, 8):
            ws.cell(row=i, column=c).border = THIN_BORDER
        ws.row_dimensions[i].height = 18

    ws.auto_filter.ref = f"A1:G{len(tickers_ordered) + 1}"
    ws.freeze_panes = "A2"


def create_sheet5_detail(wb):
    ws = wb.create_sheet("투자자 상세")
    headers = ["종목코드", "회사명", "투자자", "신호", "신뢰도", "분석근거"]
    set_col_widths(ws, [10, 16, 12, 8, 8, 60])
    write_header_row(ws, 1, headers)

    for i, (ticker, name, investor, sig, conf, reason) in enumerate(INVESTOR_DETAIL, 2):
        ws.cell(row=i, column=1, value=ticker)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=investor)
        sig_cell = ws.cell(row=i, column=4, value=sig)
        fill, font = signal_style(sig)
        sig_cell.fill = fill
        sig_cell.font = font
        ws.cell(row=i, column=5, value=conf)
        ws.cell(row=i, column=6, value=reason)
        for c in range(1, 7):
            ws.cell(row=i, column=c).border = THIN_BORDER
        ws.row_dimensions[i].height = 18

    ws.freeze_panes = "A2"


def create_sheet6_risk(wb):
    ws = wb.create_sheet("리스크 분석")
    set_col_widths(ws, [25, 20])
    cell = ws.cell(row=1, column=1, value="리스크 분석")
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    ws.merge_cells("A1:B1")
    ws.cell(row=1, column=2).fill = TITLE_FILL

    data = [
        (3, "집중도 지표", "", True),
        (4, "상위 1종목 비중", "11.3% (브이티)"),
        (5, "상위 3종목 비중", "32.5%"),
        (6, "상위 5종목 비중", "52.7%"),
        (7, "상위 10종목 비중", "100.0%"),
        (8, "HHI (허핀달 지수)", "0.100"),
        (10, "섹터 집중도", "", True),
        (11, "최대 섹터", "반도체장비 (19.0%)"),
        (12, "상위 3섹터 비중", "41.1%"),
        (14, "투자자 합의 품질", "", True),
        (15, "만장일치 비율", "60.0%"),
        (16, "의견 분산 종목 수", "0개 (편입 종목 중)"),
        (18, "비편입 사유 분포", "", True),
        (19, "투자자 과반 미달", "20개"),
        (20, "최소 비중 미달", "0개"),
        (22, "시가총액 리스크", "", True),
        (23, "소형주 비중", "81.6% (유동성 주의)"),
    ]
    for r, label, value, *rest in data:
        is_header = rest[0] if rest else False
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=12, color="1F4E79") if is_header else Font()
        if value:
            ws.cell(row=r, column=2, value=value)


def main():
    wb = openpyxl.Workbook()
    create_sheet1_summary(wb)
    create_sheet2_portfolio(wb)
    create_sheet3_ranking(wb)
    create_sheet4_matrix(wb)
    create_sheet5_detail(wb)
    create_sheet6_risk(wb)

    path = "portfolios/krx_20260211_buffett_fisher_lynch.xlsx"
    wb.save(path)
    print(f"Excel report saved: {path}")


if __name__ == "__main__":
    main()
