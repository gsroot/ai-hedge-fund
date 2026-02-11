#!/usr/bin/env python3
"""S&P 500 Portfolio Report Generator - 2026-02-11 (Analyst Team: Growth/Fundamentals/Sentiment/Technical)"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
BULLISH_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BULLISH_FONT = Font(color="375623")
BEARISH_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
BEARISH_FONT = Font(color="C62828")
NEUTRAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
NEUTRAL_FONT = Font(color="616161")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

portfolio = [
    {"ticker": "MU", "name": "Micron Technology", "weight": 17.5, "signal": "매수", "confidence": 77, "return": 15.8, "cap": "$420B", "pe": 36.5, "roe": 23, "peg": "N/A", "consensus": "3/4", "sector": "Technology"},
    {"ticker": "WDC", "name": "Western Digital", "weight": 17.5, "signal": "매수", "confidence": 77, "return": 11.1, "cap": "$90B", "pe": 27.0, "roe": 41, "peg": "N/A", "consensus": "3/4", "sector": "Technology"},
    {"ticker": "TRV", "name": "Travelers Companies", "weight": 17.0, "signal": "매수", "confidence": 72, "return": 13.4, "cap": "$67B", "pe": 10.7, "roe": 21, "peg": "N/A", "consensus": "3/4", "sector": "Financial"},
    {"ticker": "CF", "name": "CF Industries", "weight": 17.0, "signal": "매수", "confidence": 71, "return": 17.0, "cap": "$16B", "pe": 11.5, "roe": 22, "peg": "N/A", "consensus": "3/4", "sector": "Materials"},
    {"ticker": "DECK", "name": "Deckers Outdoor", "weight": 16.0, "signal": "매수", "confidence": 67, "return": 12.4, "cap": "$17B", "pe": 16.3, "roe": 40, "peg": "N/A", "consensus": "3/4", "sector": "Consumer Disc."},
    {"ticker": "FITB", "name": "Fifth Third Bancorp", "weight": 15.0, "signal": "매수", "confidence": 64, "return": 13.1, "cap": "$49B", "pe": 15.4, "roe": 12, "peg": "N/A", "consensus": "3/4", "sector": "Financial"},
]

rankings = [
    {"rank": 1, "ticker": "MOS", "name": "Mosaic Company", "total": 7.87, "fund": 7.8, "mom": 8.0, "ens": 7.1, "signal": "매수", "return": 17.1, "cap": "$9.6B", "pe": 7.7, "pb": "N/A", "roe": 10, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 2, "ticker": "CF", "name": "CF Industries", "total": 7.86, "fund": 8.0, "mom": 7.5, "ens": 7.8, "signal": "매수", "return": 17.0, "cap": "$16B", "pe": 11.5, "pb": "N/A", "roe": 22, "rev_growth": "N/A", "peg": "N/A", "included": True},
    {"rank": 3, "ticker": "MU", "name": "Micron Technology", "total": 7.51, "fund": 7.1, "mom": 8.5, "ens": 8.4, "signal": "매수", "return": 15.8, "cap": "$420B", "pe": 36.5, "pb": "N/A", "roe": 23, "rev_growth": "N/A", "peg": "N/A", "included": True},
    {"rank": 4, "ticker": "TER", "name": "Teradyne", "total": 7.42, "fund": 6.7, "mom": 9.0, "ens": 7.4, "signal": "매수", "return": 15.5, "cap": "$48B", "pe": 89.4, "pb": "N/A", "roe": 20, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 5, "ticker": "CTRA", "name": "Coterra Energy", "total": 7.38, "fund": 7.1, "mom": 8.0, "ens": 7.4, "signal": "매수", "return": 15.3, "cap": "$23B", "pe": 14.1, "pb": "N/A", "roe": 12, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 6, "ticker": "ADI", "name": "Analog Devices", "total": 7.15, "fund": 5.6, "mom": 9.0, "ens": 6.2, "signal": "매수", "return": 14.5, "cap": "$159B", "pe": 71.0, "pb": "N/A", "roe": 7, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 7, "ticker": "KEY", "name": "KeyCorp", "total": 7.12, "fund": 6.3, "mom": 9.0, "ens": 6.5, "signal": "매수", "return": 14.4, "cap": "$25B", "pe": 15.1, "pb": "N/A", "roe": 10, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 8, "ticker": "STX", "name": "Seagate Technology", "total": 6.95, "fund": 4.9, "mom": 9.5, "ens": 5.1, "signal": "매수", "return": 13.8, "cap": "$86B", "pe": 48.1, "pb": "N/A", "roe": "N/A", "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 9, "ticker": "NEM", "name": "Newmont Corporation", "total": 6.94, "fund": 6.7, "mom": 7.5, "ens": 6.9, "signal": "매수", "return": 13.8, "cap": "$133B", "pe": 18.8, "pb": "N/A", "roe": 23, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 10, "ticker": "CFG", "name": "Citizens Financial", "total": 6.90, "fund": 6.0, "mom": 9.0, "ens": 5.8, "signal": "매수", "return": 13.6, "cap": "$29B", "pe": 17.5, "pb": "N/A", "roe": 7, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 11, "ticker": "TRV", "name": "Travelers Companies", "total": 6.82, "fund": 7.1, "mom": 5.0, "ens": 7.1, "signal": "매수", "return": 13.4, "cap": "$67B", "pe": 10.7, "pb": "N/A", "roe": 21, "rev_growth": "N/A", "peg": "N/A", "included": True},
    {"rank": 12, "ticker": "PNC", "name": "PNC Financial", "total": 6.81, "fund": 5.9, "mom": 9.0, "ens": 6.1, "signal": "매수", "return": 13.3, "cap": "$97B", "pe": 14.5, "pb": "N/A", "roe": 12, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 13, "ticker": "GLW", "name": "Corning", "total": 6.76, "fund": 4.9, "mom": 9.0, "ens": 5.4, "signal": "매수", "return": 13.2, "cap": "$110B", "pe": 72.0, "pb": "N/A", "roe": 15, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 14, "ticker": "FITB", "name": "Fifth Third Bancorp", "total": 6.74, "fund": 5.3, "mom": 8.5, "ens": 5.6, "signal": "매수", "return": 13.1, "cap": "$49B", "pe": 15.4, "pb": "N/A", "roe": 12, "rev_growth": "N/A", "peg": "N/A", "included": True},
    {"rank": 15, "ticker": "PAYC", "name": "Paycom Software", "total": 6.71, "fund": 7.6, "mom": 1.5, "ens": 7.4, "signal": "매수", "return": 13.0, "cap": "$7.0B", "pe": 16.1, "pb": "N/A", "roe": 29, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 16, "ticker": "LMT", "name": "Lockheed Martin", "total": 6.70, "fund": 4.8, "mom": 9.0, "ens": 5.1, "signal": "매수", "return": 13.0, "cap": "$146B", "pe": 29.7, "pb": "N/A", "roe": 77, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 17, "ticker": "TDY", "name": "Teledyne Technologies", "total": 6.68, "fund": 4.0, "mom": 10.0, "ens": 4.3, "signal": "매수", "return": 12.9, "cap": "$31B", "pe": 34.7, "pb": "N/A", "roe": 9, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 18, "ticker": "ACGL", "name": "Arch Capital Group", "total": 6.62, "fund": 6.9, "mom": 5.0, "ens": 7.2, "signal": "매수", "return": 12.7, "cap": "$35B", "pe": 8.3, "pb": "N/A", "roe": 20, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 19, "ticker": "LRCX", "name": "Lam Research", "total": 6.58, "fund": 6.2, "mom": 7.5, "ens": 7.4, "signal": "매수", "return": 12.5, "cap": "$285B", "pe": 47.2, "pb": "N/A", "roe": 66, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 20, "ticker": "GILD", "name": "Gilead Sciences", "total": 6.55, "fund": 5.9, "mom": 8.0, "ens": 6.7, "signal": "매수", "return": 12.4, "cap": "$183B", "pe": 23.5, "pb": "N/A", "roe": 41, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 21, "ticker": "DECK", "name": "Deckers Outdoor", "total": 6.53, "fund": 6.1, "mom": 9.0, "ens": 6.4, "signal": "매수", "return": 12.4, "cap": "$17B", "pe": 16.3, "pb": "N/A", "roe": 40, "rev_growth": "N/A", "peg": "N/A", "included": True},
    {"rank": 22, "ticker": "BMY", "name": "Bristol-Myers Squibb", "total": 6.52, "fund": 5.9, "mom": 8.0, "ens": 6.0, "signal": "매수", "return": 12.3, "cap": "$124B", "pe": 17.6, "pb": "N/A", "roe": 40, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 23, "ticker": "HIG", "name": "Hartford Financial", "total": 6.51, "fund": 7.0, "mom": 4.0, "ens": 7.2, "signal": "매수", "return": 12.3, "cap": "$40B", "pe": 10.5, "pb": "N/A", "roe": 22, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 24, "ticker": "TPR", "name": "Tapestry", "total": 6.44, "fund": 5.2, "mom": 8.0, "ens": 5.5, "signal": "매수", "return": 12.0, "cap": "$31B", "pe": 59.8, "pb": "N/A", "roe": 55, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 25, "ticker": "OKE", "name": "ONEOK", "total": 6.41, "fund": 6.2, "mom": 7.0, "ens": 6.2, "signal": "매수", "return": 11.9, "cap": "$53B", "pe": 15.3, "pb": "N/A", "roe": 18, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 26, "ticker": "PYPL", "name": "PayPal Holdings", "total": 6.38, "fund": 7.2, "mom": 1.5, "ens": 7.1, "signal": "매수", "return": 11.8, "cap": "$39B", "pe": 7.6, "pb": "N/A", "roe": 26, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 27, "ticker": "JNJ", "name": "Johnson & Johnson", "total": 6.30, "fund": 6.0, "mom": 7.0, "ens": 6.8, "signal": "매수", "return": 11.6, "cap": "$574B", "pe": 21.6, "pb": "N/A", "roe": 36, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 28, "ticker": "WDC", "name": "Western Digital", "total": 6.17, "fund": 4.7, "mom": 9.5, "ens": 5.5, "signal": "매수", "return": 11.1, "cap": "$90B", "pe": 27.0, "pb": "N/A", "roe": 41, "rev_growth": "N/A", "peg": "N/A", "included": True},
    {"rank": 29, "ticker": "USB", "name": "U.S. Bancorp", "total": 6.13, "fund": 5.3, "mom": 8.0, "ens": 5.7, "signal": "매수", "return": 11.0, "cap": "$94B", "pe": 13.0, "pb": "N/A", "roe": 12, "rev_growth": "N/A", "peg": "N/A", "included": False},
    {"rank": 30, "ticker": "FAST", "name": "Fastenal Company", "total": 6.12, "fund": 5.3, "mom": 8.0, "ens": 6.1, "signal": "매수", "return": 10.9, "cap": "$54B", "pe": 42.8, "pb": "N/A", "roe": 33, "rev_growth": "N/A", "peg": "N/A", "included": False},
]

investor_matrix = [
    {"ticker": "MOS", "name": "Mosaic Company", "growth": ("bearish", 35), "fundamentals": ("neutral", 58), "sentiment": ("neutral", 50), "technical": ("bullish", 68), "combined": "-", "conf": "-"},
    {"ticker": "CF", "name": "CF Industries", "growth": ("neutral", 55), "fundamentals": ("bullish", 78), "sentiment": ("bullish", 62), "technical": ("bullish", 72), "combined": "매수", "conf": "71%"},
    {"ticker": "MU", "name": "Micron Technology", "growth": ("bullish", 85), "fundamentals": ("bullish", 72), "sentiment": ("bearish", 58), "technical": ("bullish", 75), "combined": "매수", "conf": "77%"},
    {"ticker": "TER", "name": "Teradyne", "growth": ("bullish", 72), "fundamentals": ("neutral", 62), "sentiment": ("neutral", 48), "technical": ("neutral", 55), "combined": "-", "conf": "-"},
    {"ticker": "CTRA", "name": "Coterra Energy", "growth": ("bearish", 40), "fundamentals": ("bullish", 70), "sentiment": ("neutral", 52), "technical": ("bullish", 65), "combined": "-", "conf": "-"},
    {"ticker": "ADI", "name": "Analog Devices", "growth": ("neutral", 58), "fundamentals": ("bearish", 48), "sentiment": ("bearish", 55), "technical": ("bullish", 70), "combined": "-", "conf": "-"},
    {"ticker": "KEY", "name": "KeyCorp", "growth": ("bearish", 30), "fundamentals": ("neutral", 55), "sentiment": ("bearish", 60), "technical": ("bullish", 68), "combined": "-", "conf": "-"},
    {"ticker": "STX", "name": "Seagate Technology", "growth": ("bullish", 78), "fundamentals": ("neutral", 60), "sentiment": ("bearish", 65), "technical": ("bullish", 78), "combined": "-", "conf": "-"},
    {"ticker": "NEM", "name": "Newmont Corporation", "growth": ("bullish", 80), "fundamentals": ("bullish", 75), "sentiment": ("bearish", 57), "technical": ("neutral", 50), "combined": "-", "conf": "-"},
    {"ticker": "CFG", "name": "Citizens Financial", "growth": ("neutral", 50), "fundamentals": ("neutral", 54), "sentiment": ("neutral", 45), "technical": ("bullish", 68), "combined": "-", "conf": "-"},
    {"ticker": "TRV", "name": "Travelers Companies", "growth": ("bullish", 68), "fundamentals": ("bullish", 80), "sentiment": ("bullish", 68), "technical": ("neutral", 45), "combined": "매수", "conf": "72%"},
    {"ticker": "PNC", "name": "PNC Financial", "growth": ("neutral", 52), "fundamentals": ("bullish", 68), "sentiment": ("neutral", 48), "technical": ("bullish", 70), "combined": "-", "conf": "-"},
    {"ticker": "GLW", "name": "Corning", "growth": ("neutral", 48), "fundamentals": ("bearish", 45), "sentiment": ("bearish", 62), "technical": ("bullish", 70), "combined": "-", "conf": "-"},
    {"ticker": "FITB", "name": "Fifth Third Bancorp", "growth": ("neutral", 54), "fundamentals": ("bullish", 70), "sentiment": ("bullish", 55), "technical": ("bullish", 67), "combined": "매수", "conf": "64%"},
    {"ticker": "PAYC", "name": "Paycom Software", "growth": ("bullish", 65), "fundamentals": ("bullish", 76), "sentiment": ("bearish", 58), "technical": ("bearish", 40), "combined": "-", "conf": "-"},
    {"ticker": "LMT", "name": "Lockheed Martin", "growth": ("neutral", 58), "fundamentals": ("neutral", 65), "sentiment": ("neutral", 50), "technical": ("bullish", 68), "combined": "-", "conf": "-"},
    {"ticker": "TDY", "name": "Teledyne Technologies", "growth": ("neutral", 55), "fundamentals": ("neutral", 62), "sentiment": ("neutral", 52), "technical": ("bullish", 75), "combined": "-", "conf": "-"},
    {"ticker": "ACGL", "name": "Arch Capital Group", "growth": ("bullish", 75), "fundamentals": ("bullish", 82), "sentiment": ("bearish", 60), "technical": ("neutral", 48), "combined": "-", "conf": "-"},
    {"ticker": "LRCX", "name": "Lam Research", "growth": ("bullish", 82), "fundamentals": ("bullish", 74), "sentiment": ("neutral", 48), "technical": ("neutral", 60), "combined": "-", "conf": "-"},
    {"ticker": "GILD", "name": "Gilead Sciences", "growth": ("neutral", 45), "fundamentals": ("neutral", 58), "sentiment": ("bearish", 55), "technical": ("neutral", 58), "combined": "-", "conf": "-"},
    {"ticker": "DECK", "name": "Deckers Outdoor", "growth": ("bullish", 62), "fundamentals": ("bullish", 72), "sentiment": ("neutral", 45), "technical": ("bullish", 68), "combined": "매수", "conf": "67%"},
    {"ticker": "BMY", "name": "Bristol-Myers Squibb", "growth": ("neutral", 42), "fundamentals": ("neutral", 60), "sentiment": ("bullish", 65), "technical": ("neutral", 55), "combined": "-", "conf": "-"},
    {"ticker": "HIG", "name": "Hartford Financial", "growth": ("bullish", 70), "fundamentals": ("bullish", 77), "sentiment": ("bearish", 70), "technical": ("bearish", 42), "combined": "-", "conf": "-"},
    {"ticker": "TPR", "name": "Tapestry", "growth": ("neutral", 48), "fundamentals": ("bearish", 42), "sentiment": ("bearish", 58), "technical": ("bullish", 65), "combined": "-", "conf": "-"},
    {"ticker": "OKE", "name": "ONEOK", "growth": ("bullish", 72), "fundamentals": ("neutral", 62), "sentiment": ("bullish", 60), "technical": ("neutral", 52), "combined": "-", "conf": "-"},
    {"ticker": "PYPL", "name": "PayPal Holdings", "growth": ("neutral", 52), "fundamentals": ("bullish", 78), "sentiment": ("bearish", 63), "technical": ("bearish", 35), "combined": "-", "conf": "-"},
    {"ticker": "JNJ", "name": "Johnson & Johnson", "growth": ("neutral", 55), "fundamentals": ("bullish", 73), "sentiment": ("neutral", 52), "technical": ("neutral", 52), "combined": "-", "conf": "-"},
    {"ticker": "WDC", "name": "Western Digital", "growth": ("bullish", 88), "fundamentals": ("bullish", 70), "sentiment": ("bearish", 68), "technical": ("bullish", 72), "combined": "매수", "conf": "77%"},
    {"ticker": "USB", "name": "U.S. Bancorp", "growth": ("neutral", 50), "fundamentals": ("neutral", 58), "sentiment": ("bearish", 62), "technical": ("neutral", 55), "combined": "-", "conf": "-"},
    {"ticker": "FAST", "name": "Fastenal Company", "growth": ("neutral", 58), "fundamentals": ("neutral", 55), "sentiment": ("neutral", 48), "technical": ("neutral", 58), "combined": "-", "conf": "-"},
]

investor_details = [
    ("MU", "Micron Technology", "Growth", "bullish", 85, "매출 +48.7% YoY 급증, AI 메모리 수요와 강력한 영업 레버리지로 수익 전환"),
    ("MU", "Micron Technology", "Fundamentals", "bullish", 72, "ROE 23%와 견고한 마진, 유동비율 2.46과 이자보상배율 21.2 우수, 56% 매출 성장 대비 P/E 36 합리적"),
    ("MU", "Micron Technology", "Sentiment", "bearish", 58, "경영진 $10M+ 대규모 매도, AI 반도체 낙관론에도 내부자 매도 압력"),
    ("MU", "Micron Technology", "Technical", "bullish", 75, "모멘텀 8.5, $102→$127 강한 상승 추세, 상승하는 50일 이평선 위 지속 거래"),
    ("WDC", "Western Digital", "Growth", "bullish", 88, "매출 +50.7% YoY 급증, 적자에서 흑자 전환, AI 스토리지 수요로 폭발적 회복"),
    ("WDC", "Western Digital", "Fundamentals", "bullish", 70, "ROE 41%와 마진 36% 우수, P/E 24.8 합리적, FCF 수익률 4.3%로 수익성 견고"),
    ("WDC", "Western Digital", "Sentiment", "bearish", 68, "경영진 $9.6M 대규모 매도, 스토리지 공급과잉 우려"),
    ("WDC", "Western Digital", "Technical", "bullish", 72, "모멘텀 9.5, $53→$77 스토리지 브레이크아웃, 반도체 랠리 수혜"),
    ("TRV", "Travelers Companies", "Growth", "bullish", 68, "매출 +12.4% YoY 꾸준한 성장, 순이익 +67.1% YoY 가속, 보험 가격 결정력과 마진 확대"),
    ("TRV", "Travelers Companies", "Fundamentals", "bullish", 80, "ROE 21%와 P/E 10.9 저평가, FCF 수익률 15.5%, 이자보상배율 16.8로 우수"),
    ("TRV", "Travelers Companies", "Sentiment", "bullish", 68, "경영진 낮은 행사가 옵션 행사로 자신감 표시, 보험 가격 강세 지속"),
    ("TRV", "Travelers Companies", "Technical", "neutral", 45, "모멘텀 5.0 약세, $260-$280 박스권, 보험 섹터 상대적 부진"),
    ("CF", "CF Industries", "Growth", "neutral", 55, "매출총이익률 36.9%이나 매출 -10.5% YoY 감소, 질소 가격 역풍과 사이클 둔화"),
    ("CF", "CF Industries", "Fundamentals", "bullish", 78, "ROE 22%와 P/E 11.5 매력적, 유동비율 2.27, FCF 수익률 8.7%와 이자보상배율 13.7 우수"),
    ("CF", "CF Industries", "Sentiment", "bullish", 62, "경영진 지분 보상 프로그램으로 리텐션 집중, 질소 가격 산업 내 개선 신호"),
    ("CF", "CF Industries", "Technical", "bullish", 72, "모멘텀 7.5, $90→$102 강한 상승 추세, 저항선 돌파와 거래량 확인"),
    ("DECK", "Deckers Outdoor", "Growth", "bullish", 62, "매출 +16.3% YoY, 순이익 +27.2% YoY 가속, 프리미엄 신발 브랜드 파워와 마진 확대"),
    ("DECK", "Deckers Outdoor", "Fundamentals", "bullish", 72, "ROE 40%와 마진 19% 우수, P/E 16.4 매력적, 유동비율 2.86으로 재무 건전"),
    ("DECK", "Deckers Outdoor", "Sentiment", "neutral", 45, "이사회 지분 보상만 존재, 의미있는 매수 없음, 신발 수요 신호 혼조"),
    ("DECK", "Deckers Outdoor", "Technical", "bullish", 68, "모멘텀 9.0, $105→$135 소비재 섹터 리더, 강한 상승 추세"),
    ("FITB", "Fifth Third Bancorp", "Growth", "neutral", 54, "3년 매출 성장 +10.4% 완만, 순이익 -4.2% YoY 감소, 은행 마진 압박"),
    ("FITB", "Fifth Third Bancorp", "Fundamentals", "bullish", 70, "ROE 12%와 P/E 15.4 매력적, 영업이익률 41%, 배당수익률 2.9% 건전"),
    ("FITB", "Fifth Third Bancorp", "Sentiment", "bullish", 55, "2026년 2월 대규모 지분 보상 프로그램 개시, 지역 은행 포지셔닝 개선"),
    ("FITB", "Fifth Third Bancorp", "Technical", "bullish", 67, "모멘텀 8.5, $37→$44 지역은행 회복, 거래량 지지 속 상승 추세 유지"),
]


def create_workbook():
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "요약"
    create_summary_sheet(ws_summary)
    ws_portfolio = wb.create_sheet("포트폴리오")
    create_portfolio_sheet(ws_portfolio)
    ws_ranking = wb.create_sheet("순위")
    create_ranking_sheet(ws_ranking)
    ws_matrix = wb.create_sheet("분석가 매트릭스")
    create_matrix_sheet(ws_matrix)
    ws_detail = wb.create_sheet("분석가 상세")
    create_detail_sheet(ws_detail)
    ws_risk = wb.create_sheet("리스크 분석")
    create_risk_sheet(ws_risk)
    return wb


def create_summary_sheet(ws):
    ws.merge_cells('A1:B1')
    ws['A1'] = "AI Hedge Fund 포트폴리오 리포트"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = Alignment(horizontal='center')
    info = [
        ("분석 일자", "2026-02-11"),
        ("분석 대상", "S&P 500 상위 30개"),
        ("분석 전략", "하이브리드 (펀더멘털 70% + 모멘텀 30%)"),
        ("분석가 관점", "Growth, Fundamentals, Sentiment, Technical"),
    ]
    for i, (key, val) in enumerate(info, start=3):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws['A8'] = "포트폴리오 통계"
    ws['A8'].font = Font(bold=True)
    stats = [
        ("편입 종목 수", "6 / 30 (20.0%)"),
        ("평균 신뢰도", "71%"),
        ("평균 예상 수익률", "+13.8%"),
        ("매수 비중", "100.0%"),
    ]
    for i, (key, val) in enumerate(stats, start=9):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws['A14'] = "시가총액 분포"
    ws['A14'].font = Font(bold=True)
    cap_dist = [
        ("메가캡 (>$200B)", "17.5% (MU)"),
        ("대형주 ($10-200B)", "82.5% (WDC, TRV, CF, DECK, FITB)"),
        ("중형주 ($2-10B)", "0%"),
    ]
    for i, (key, val) in enumerate(cap_dist, start=15):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws['A19'] = "섹터 분포"
    ws['A19'].font = Font(bold=True)
    sector_dist = [
        ("Technology", "35.0% (MU, WDC)"),
        ("Financial", "32.0% (TRV, FITB)"),
        ("Materials", "17.0% (CF)"),
        ("Consumer Discretionary", "16.0% (DECK)"),
    ]
    for i, (key, val) in enumerate(sector_dist, start=20):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50


def create_portfolio_sheet(ws):
    headers = ["#", "종목코드", "회사명", "비중", "신호", "신뢰도", "예상수익률", "시가총액", "P/E", "ROE", "PEG", "합의", "섹터"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    for row, item in enumerate(portfolio, 2):
        ws.cell(row=row, column=1, value=row - 1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=item['ticker']).border = THIN_BORDER
        ws.cell(row=row, column=3, value=item['name']).border = THIN_BORDER
        ws.cell(row=row, column=4, value=f"{item['weight']}%").border = THIN_BORDER
        signal_cell = ws.cell(row=row, column=5, value=f"🟢 {item['signal']}" if item['signal'] == '강력매수' else f"🔵 {item['signal']}")
        signal_cell.border = THIN_BORDER
        if item['signal'] == '강력매수':
            signal_cell.fill = BULLISH_FILL
        ws.cell(row=row, column=6, value=f"{item['confidence']}%").border = THIN_BORDER
        ws.cell(row=row, column=7, value=f"+{item['return']}%").border = THIN_BORDER
        ws.cell(row=row, column=8, value=item['cap']).border = THIN_BORDER
        ws.cell(row=row, column=9, value=item['pe']).border = THIN_BORDER
        ws.cell(row=row, column=10, value=f"{item['roe']}%").border = THIN_BORDER
        ws.cell(row=row, column=11, value=item['peg']).border = THIN_BORDER
        ws.cell(row=row, column=12, value=item['consensus']).border = THIN_BORDER
        ws.cell(row=row, column=13, value=item['sector']).border = THIN_BORDER
    total_row = len(portfolio) + 2
    for col in range(1, 14):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = THIN_BORDER
    ws.cell(row=total_row, column=2, value="합계").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value="100.0%").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value="avg 71%")
    ws.cell(row=total_row, column=7, value="avg +13.8%")
    widths = [5, 12, 22, 8, 14, 10, 12, 10, 8, 8, 8, 8, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:M{len(portfolio) + 1}"
    ws.freeze_panes = "A2"


def create_ranking_sheet(ws):
    headers = ["순위", "종목코드", "회사명", "종합점수", "펀더멘털", "모멘텀", "앙상블", "신호", "예상수익률", "시가총액", "P/E", "P/B", "ROE", "매출성장률", "PEG", "편입여부"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    for row, item in enumerate(rankings, 2):
        ws.cell(row=row, column=1, value=item['rank']).border = THIN_BORDER
        ws.cell(row=row, column=2, value=item['ticker']).border = THIN_BORDER
        ws.cell(row=row, column=3, value=item['name']).border = THIN_BORDER
        ws.cell(row=row, column=4, value=item['total']).border = THIN_BORDER
        ws.cell(row=row, column=5, value=item['fund']).border = THIN_BORDER
        ws.cell(row=row, column=6, value=item['mom']).border = THIN_BORDER
        ws.cell(row=row, column=7, value=item['ens']).border = THIN_BORDER
        ws.cell(row=row, column=8, value=item['signal']).border = THIN_BORDER
        ws.cell(row=row, column=9, value=f"+{item['return']}%").border = THIN_BORDER
        ws.cell(row=row, column=10, value=item['cap']).border = THIN_BORDER
        ws.cell(row=row, column=11, value=item['pe']).border = THIN_BORDER
        ws.cell(row=row, column=12, value=item['pb']).border = THIN_BORDER
        roe_val = f"{item['roe']}%" if item['roe'] != "N/A" else "N/A"
        ws.cell(row=row, column=13, value=roe_val).border = THIN_BORDER
        ws.cell(row=row, column=14, value=item['rev_growth']).border = THIN_BORDER
        ws.cell(row=row, column=15, value=item['peg']).border = THIN_BORDER
        included_cell = ws.cell(row=row, column=16, value="예" if item['included'] else "아니오")
        included_cell.border = THIN_BORDER
        if item['included']:
            for col_idx in range(1, 17):
                ws.cell(row=row, column=col_idx).fill = BULLISH_FILL
    widths = [6, 10, 22, 10, 10, 8, 8, 10, 12, 10, 8, 8, 8, 12, 8, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:P{len(rankings) + 1}"
    ws.freeze_panes = "A2"


def create_matrix_sheet(ws):
    headers = ["종목코드", "회사명", "Growth", "Fundamentals", "Sentiment", "Technical", "종합신호", "종합신뢰도"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    for row, item in enumerate(investor_matrix, 2):
        ws.cell(row=row, column=1, value=item['ticker']).border = THIN_BORDER
        ws.cell(row=row, column=2, value=item['name']).border = THIN_BORDER
        for col, inv in enumerate(['growth', 'fundamentals', 'sentiment', 'technical'], 3):
            signal, conf = item[inv]
            display = {"bullish": "매수", "bearish": "매도", "neutral": "중립"}[signal]
            cell = ws.cell(row=row, column=col, value=f"{display}({conf})")
            cell.border = THIN_BORDER
            if signal == 'bullish':
                cell.fill = BULLISH_FILL
                cell.font = BULLISH_FONT
            elif signal == 'bearish':
                cell.fill = BEARISH_FILL
                cell.font = BEARISH_FONT
            else:
                cell.fill = NEUTRAL_FILL
                cell.font = NEUTRAL_FONT
        ws.cell(row=row, column=7, value=item['combined']).border = THIN_BORDER
        ws.cell(row=row, column=8, value=item['conf']).border = THIN_BORDER
    widths = [10, 22, 15, 15, 15, 15, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:H{len(investor_matrix) + 1}"
    ws.freeze_panes = "A2"


def create_detail_sheet(ws):
    headers = ["종목코드", "회사명", "분석가", "신호", "신뢰도", "분석근거"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    for row, (ticker, name, analyst, signal, conf, reasoning) in enumerate(investor_details, 2):
        ws.cell(row=row, column=1, value=ticker).border = THIN_BORDER
        ws.cell(row=row, column=2, value=name).border = THIN_BORDER
        ws.cell(row=row, column=3, value=analyst).border = THIN_BORDER
        display = {"bullish": "매수", "bearish": "매도", "neutral": "중립"}[signal]
        signal_cell = ws.cell(row=row, column=4, value=display)
        signal_cell.border = THIN_BORDER
        if signal == 'bullish':
            signal_cell.fill = BULLISH_FILL
            signal_cell.font = BULLISH_FONT
        elif signal == 'bearish':
            signal_cell.fill = BEARISH_FILL
            signal_cell.font = BEARISH_FONT
        else:
            signal_cell.fill = NEUTRAL_FILL
            signal_cell.font = NEUTRAL_FONT
        ws.cell(row=row, column=5, value=conf).border = THIN_BORDER
        ws.cell(row=row, column=6, value=reasoning).border = THIN_BORDER
    widths = [10, 22, 15, 10, 10, 70]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def create_risk_sheet(ws):
    ws.merge_cells('A1:B1')
    ws['A1'] = "리스크 분석"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A3'] = "집중도 지표"
    ws['A3'].font = Font(bold=True)
    concentration = [
        ("상위 1종목 비중", "17.5% (MU / WDC)"),
        ("상위 3종목 비중", "51.5%"),
        ("상위 5종목 비중", "85.0%"),
        ("HHI (허핀달 지수)", "0.170"),
    ]
    for i, (key, val) in enumerate(concentration, 4):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws['A9'] = "섹터 집중도"
    ws['A9'].font = Font(bold=True)
    sector_conc = [
        ("최대 섹터", "Technology (35.0%)"),
        ("상위 2섹터 비중", "67.0%"),
    ]
    for i, (key, val) in enumerate(sector_conc, 10):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws['A13'] = "분석가 합의 품질"
    ws['A13'].font = Font(bold=True)
    consensus = [
        ("만장일치 비율", "0% (0개)"),
        ("다수 합의 (3/4)", "100% (6개: MU, WDC, TRV, CF, DECK, FITB)"),
        ("의견 분산 종목", "MU, WDC (Growth bullish vs Sentiment bearish)"),
    ]
    for i, (key, val) in enumerate(consensus, 14):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws['A18'] = "비편입 사유 분포"
    ws['A18'].font = Font(bold=True)
    exclusion = [
        ("분석가 과반 미달 (0/4)", "3개 (GILD, USB, FAST)"),
        ("분석가 과반 미달 (1/4)", "11개 (MOS, TER, ADI, KEY, CFG, GLW, LMT, TDY, BMY, TPR, PYPL)"),
        ("분석가 과반 미달 (2/4)", "10개 (CTRA, STX, NEM, PNC, PAYC, ACGL, LRCX, HIG, OKE, JNJ)"),
    ]
    for i, (key, val) in enumerate(exclusion, 19):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws['A23'] = "제약사항"
    ws['A23'].font = Font(bold=True)
    constraints = [
        ("개별종목 최대 비중", "17.5% (6종목 제약으로 15% 상한 초과 불가피)"),
        ("섹터 비중 제한", "Technology 35.0% (제한선 도달)"),
    ]
    for i, (key, val) in enumerate(constraints, 24):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = val
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 55


if __name__ == "__main__":
    wb = create_workbook()
    output_path = "/Users/sgkim/Projects/ai-hedge-fund/portfolios/sp500_20260211_analyst_team.xlsx"
    wb.save(output_path)
    print(f"Excel report saved: {output_path}")
