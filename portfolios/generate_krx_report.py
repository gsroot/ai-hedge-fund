"""KRX Portfolio Report Excel Generator - 2026-02-11"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
import os

wb = openpyxl.Workbook()

# === Styles ===
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
BUY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BUY_FONT = Font(name="맑은 고딕", color="375623")
SELL_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
SELL_FONT = Font(name="맑은 고딕", color="C62828")
NEUTRAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
NEUTRAL_FONT = Font(name="맑은 고딕", color="616161")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(name="맑은 고딕", bold=True)
DATA_FONT = Font(name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22

def style_data_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 18

def auto_width(ws, max_col):
    for col in range(1, max_col + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 30)


# === DATA ===
portfolio = [
    {"rank": 1, "ticker": "000660", "name": "SK하이닉스", "weight": 9.2, "signal": "매수", "confidence": 76, "ret": 8.0, "mcap": "N/A", "pe": "N/A", "roe": "26.8%", "peg": "N/A", "consensus": "2/3", "sector": "반도체 제조"},
    {"rank": 2, "ticker": "031980", "name": "피에스케이홀딩스", "weight": 8.8, "signal": "매수", "confidence": 73, "ret": 8.0, "mcap": "N/A", "pe": "16.8", "roe": "22.2%", "peg": "0.13", "consensus": "2/3", "sector": "반도체 장비"},
    {"rank": 3, "ticker": "403870", "name": "HPSP", "weight": 8.8, "signal": "매수", "confidence": 73, "ret": 8.0, "mcap": "N/A", "pe": "41.0", "roe": "31.0%", "peg": "N/A", "consensus": "2/3", "sector": "반도체 장비"},
    {"rank": 4, "ticker": "402340", "name": "SK스퀘어", "weight": 8.5, "signal": "매수", "confidence": 71, "ret": 8.2, "mcap": "₩71조", "pe": "19.3", "roe": "18.6%", "peg": "N/A", "consensus": "2/3", "sector": "지주회사"},
    {"rank": 5, "ticker": "319660", "name": "피에스케이", "weight": 8.4, "signal": "매수", "confidence": 70, "ret": 8.0, "mcap": "N/A", "pe": "21.2", "roe": "16.8%", "peg": "0.42", "consensus": "2/3", "sector": "반도체 장비"},
    {"rank": 6, "ticker": "064760", "name": "티씨케이", "weight": 8.3, "signal": "매수", "confidence": 69, "ret": 8.0, "mcap": "N/A", "pe": "34.0", "roe": "13.9%", "peg": "1.9", "consensus": "2/3", "sector": "반도체 소재"},
    {"rank": 7, "ticker": "095610", "name": "테스", "weight": 8.2, "signal": "매수", "confidence": 68, "ret": 8.0, "mcap": "₩1.2조", "pe": "26.4", "roe": "12.9%", "peg": "N/A", "consensus": "2/3", "sector": "반도체 장비"},
    {"rank": 8, "ticker": "005290", "name": "동진쎄미켐", "weight": 8.1, "signal": "매수", "confidence": 67, "ret": 8.0, "mcap": "N/A", "pe": "17.0", "roe": "14.1%", "peg": "1.06", "consensus": "2/3", "sector": "반도체 소재"},
    {"rank": 9, "ticker": "002380", "name": "KCC", "weight": 8.1, "signal": "매수", "confidence": 67, "ret": 8.0, "mcap": "N/A", "pe": "13.0", "roe": "5.8%", "peg": "N/A", "consensus": "2/3", "sector": "건자재/도료"},
    {"rank": 10, "ticker": "357780", "name": "솔브레인", "weight": 8.0, "signal": "매수", "confidence": 67, "ret": 8.0, "mcap": "N/A", "pe": "26.0", "roe": "11.5%", "peg": "N/A", "consensus": "2/3", "sector": "반도체 소재"},
    {"rank": 11, "ticker": "178320", "name": "서진시스템", "weight": 7.9, "signal": "매수", "confidence": 66, "ret": 8.0, "mcap": "N/A", "pe": "21.7", "roe": "N/A", "peg": "N/A", "consensus": "2/3", "sector": "전자부품"},
    {"rank": 12, "ticker": "009970", "name": "영원무역홀딩스", "weight": 7.6, "signal": "매수", "confidence": 63, "ret": 8.0, "mcap": "N/A", "pe": "7.8", "roe": "9.9%", "peg": "N/A", "consensus": "2/3", "sector": "섬유/의류"},
]

predict_all = [
    {"rank": 1, "ticker": "066570", "name": "LG전자", "score": 5.39, "fund": 1.6, "mom": 10.0, "ens": 1.7, "signal": "매수", "ret": 8.4, "mcap": "₩20조", "pe": "59.4", "roe": "2.3%", "rev_g": "6.6%", "peg": "N/A", "included": "아니오"},
    {"rank": 2, "ticker": "402340", "name": "SK스퀘어", "score": 5.34, "fund": 1.5, "mom": 10.0, "ens": 1.7, "signal": "매수", "ret": 8.2, "mcap": "₩71조", "pe": "19.3", "roe": "18.6%", "rev_g": "-16.2%", "peg": "N/A", "included": "예"},
    {"rank": 3, "ticker": "000660", "name": "SK하이닉스", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "N/A", "roe": "26.8%", "rev_g": "102%", "peg": "N/A", "included": "예"},
    {"rank": 4, "ticker": "240810", "name": "원익IPS", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "245", "roe": "2.3%", "rev_g": "8%", "peg": "N/A", "included": "아니오"},
    {"rank": 5, "ticker": "039030", "name": "이오테크닉스", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "107", "roe": "7.1%", "rev_g": "1.5%", "peg": "N/A", "included": "아니오"},
    {"rank": 6, "ticker": "140410", "name": "메지온", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "N/A", "roe": "-42%", "rev_g": "N/A", "peg": "N/A", "included": "아니오"},
    {"rank": 7, "ticker": "002380", "name": "KCC", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "13", "roe": "5.8%", "rev_g": "N/A", "peg": "N/A", "included": "예"},
    {"rank": 8, "ticker": "017800", "name": "현대엘리베이터", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "20", "roe": "14.4%", "rev_g": "11%", "peg": "N/A", "included": "아니오"},
    {"rank": 9, "ticker": "403870", "name": "HPSP", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "41", "roe": "31%", "rev_g": "7%", "peg": "N/A", "included": "예"},
    {"rank": 10, "ticker": "357780", "name": "솔브레인", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "26", "roe": "11.5%", "rev_g": "2%", "peg": "N/A", "included": "예"},
    {"rank": 11, "ticker": "237690", "name": "에스티팜", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "85", "roe": "N/A", "rev_g": "-3.9%", "peg": "N/A", "included": "아니오"},
    {"rank": 12, "ticker": "009970", "name": "영원무역홀딩스", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "7.8", "roe": "9.9%", "rev_g": "N/A", "peg": "N/A", "included": "예"},
    {"rank": 13, "ticker": "058610", "name": "에스피지", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "231", "roe": "N/A", "rev_g": "-1.3%", "peg": "N/A", "included": "아니오"},
    {"rank": 14, "ticker": "010060", "name": "OCI홀딩스", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "29", "roe": "N/A", "rev_g": "N/A", "peg": "N/A", "included": "아니오"},
    {"rank": 15, "ticker": "028670", "name": "팬오션", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "9.8", "roe": "N/A", "rev_g": "18%", "peg": "N/A", "included": "아니오"},
    {"rank": 16, "ticker": "005290", "name": "동진쎄미켐", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "17", "roe": "14.1%", "rev_g": "7.5%", "peg": "1.06", "included": "예"},
    {"rank": 17, "ticker": "064760", "name": "티씨케이", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "34", "roe": "13.9%", "rev_g": "21.7%", "peg": "1.9", "included": "예"},
    {"rank": 18, "ticker": "083650", "name": "비에이치아이", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "119", "roe": "N/A", "rev_g": "N/A", "peg": "N/A", "included": "아니오"},
    {"rank": 19, "ticker": "067310", "name": "하나마이크론", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "N/A", "roe": "N/A", "rev_g": "29%", "peg": "N/A", "included": "아니오"},
    {"rank": 20, "ticker": "065350", "name": "신성델타테크", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "N/A", "roe": "N/A", "rev_g": "N/A", "peg": "N/A", "included": "아니오"},
    {"rank": 21, "ticker": "178320", "name": "서진시스템", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "21.7", "roe": "N/A", "rev_g": "56%", "peg": "N/A", "included": "예"},
    {"rank": 22, "ticker": "101490", "name": "에스앤에스텍", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "61", "roe": "12.3%", "rev_g": "17%", "peg": "N/A", "included": "아니오"},
    {"rank": 23, "ticker": "003380", "name": "하림지주", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "58", "roe": "N/A", "rev_g": "N/A", "peg": "N/A", "included": "아니오"},
    {"rank": 24, "ticker": "085660", "name": "차바이오텍", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "N/A", "roe": "-2%", "rev_g": "N/A", "peg": "N/A", "included": "아니오"},
    {"rank": 25, "ticker": "319660", "name": "피에스케이", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "21.2", "roe": "16.8%", "rev_g": "13%", "peg": "0.42", "included": "예"},
    {"rank": 26, "ticker": "031980", "name": "피에스케이홀딩스", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "16.8", "roe": "22.2%", "rev_g": "127%", "peg": "0.13", "included": "예"},
    {"rank": 27, "ticker": "388720", "name": "유일로보틱스", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "N/A", "roe": "-9.5%", "rev_g": "N/A", "peg": "N/A", "included": "아니오"},
    {"rank": 28, "ticker": "095610", "name": "테스", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "₩1.2조", "pe": "26.4", "roe": "12.9%", "rev_g": "63%", "peg": "N/A", "included": "예"},
    {"rank": 29, "ticker": "131970", "name": "두산테스나", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "33.4", "roe": "8.4%", "rev_g": "10%", "peg": "N/A", "included": "아니오"},
    {"rank": 30, "ticker": "036540", "name": "SFA반도체", "score": 5.28, "fund": 1.4, "mom": 10.0, "ens": 1.5, "signal": "매수", "ret": 8.0, "mcap": "N/A", "pe": "55.5", "roe": "4.1%", "rev_g": "-3.9%", "peg": "N/A", "included": "아니오"},
]

investor_matrix = [
    {"ticker": "066570", "name": "LG전자", "buffett": ("neutral", 45), "lynch": ("bearish", 72), "fisher": ("neutral", 55), "combined": "보유", "comb_conf": 57},
    {"ticker": "402340", "name": "SK스퀘어", "buffett": ("bullish", 72), "lynch": ("neutral", 68), "fisher": ("bullish", 72), "combined": "매수", "comb_conf": 71},
    {"ticker": "000660", "name": "SK하이닉스", "buffett": ("neutral", 58), "lynch": ("bullish", 85), "fisher": ("bullish", 85), "combined": "매수", "comb_conf": 76},
    {"ticker": "240810", "name": "원익IPS", "buffett": ("bearish", 65), "lynch": ("bearish", 78), "fisher": ("neutral", 48), "combined": "매도", "comb_conf": 64},
    {"ticker": "039030", "name": "이오테크닉스", "buffett": ("neutral", 52), "lynch": ("neutral", 65), "fisher": ("neutral", 62), "combined": "보유", "comb_conf": 60},
    {"ticker": "140410", "name": "메지온", "buffett": ("bearish", 85), "lynch": ("bearish", 88), "fisher": ("bearish", 78), "combined": "매도", "comb_conf": 84},
    {"ticker": "002380", "name": "KCC", "buffett": ("bullish", 68), "lynch": ("bullish", 76), "fisher": ("neutral", 58), "combined": "매수", "comb_conf": 67},
    {"ticker": "017800", "name": "현대엘리베이터", "buffett": ("bullish", 75), "lynch": ("neutral", 58), "fisher": ("neutral", 60), "combined": "보유", "comb_conf": 64},
    {"ticker": "403870", "name": "HPSP", "buffett": ("bullish", 78), "lynch": ("neutral", 62), "fisher": ("bullish", 78), "combined": "매수", "comb_conf": 73},
    {"ticker": "357780", "name": "솔브레인", "buffett": ("bullish", 70), "lynch": ("neutral", 60), "fisher": ("bullish", 70), "combined": "매수", "comb_conf": 67},
    {"ticker": "237690", "name": "에스티팜", "buffett": ("neutral", 55), "lynch": ("neutral", 64), "fisher": ("neutral", 65), "combined": "보유", "comb_conf": 61},
    {"ticker": "009970", "name": "영원무역홀딩스", "buffett": ("bullish", 65), "lynch": ("bullish", 73), "fisher": ("neutral", 52), "combined": "매수", "comb_conf": 63},
    {"ticker": "058610", "name": "에스피지", "buffett": ("neutral", 50), "lynch": ("bearish", 70), "fisher": ("neutral", 50), "combined": "보유", "comb_conf": 57},
    {"ticker": "010060", "name": "OCI홀딩스", "buffett": ("neutral", 48), "lynch": ("bearish", 75), "fisher": ("bearish", 68), "combined": "매도", "comb_conf": 64},
    {"ticker": "028670", "name": "팬오션", "buffett": ("bearish", 62), "lynch": ("neutral", 66), "fisher": ("neutral", 58), "combined": "보유", "comb_conf": 62},
    {"ticker": "005290", "name": "동진쎄미켐", "buffett": ("neutral", 55), "lynch": ("bullish", 71), "fisher": ("bullish", 75), "combined": "매수", "comb_conf": 67},
    {"ticker": "064760", "name": "티씨케이", "buffett": ("neutral", 53), "lynch": ("bullish", 74), "fisher": ("bullish", 80), "combined": "매수", "comb_conf": 69},
    {"ticker": "083650", "name": "비에이치아이", "buffett": ("neutral", 50), "lynch": ("neutral", 55), "fisher": ("neutral", 50), "combined": "보유", "comb_conf": 52},
    {"ticker": "067310", "name": "하나마이크론", "buffett": ("neutral", 52), "lynch": ("bearish", 80), "fisher": ("bearish", 72), "combined": "매도", "comb_conf": 68},
    {"ticker": "065350", "name": "신성델타테크", "buffett": ("neutral", 48), "lynch": ("bearish", 72), "fisher": ("bearish", 65), "combined": "매도", "comb_conf": 62},
    {"ticker": "178320", "name": "서진시스템", "buffett": ("neutral", 50), "lynch": ("bullish", 79), "fisher": ("bullish", 68), "combined": "매수", "comb_conf": 66},
    {"ticker": "101490", "name": "에스앤에스텍", "buffett": ("neutral", 54), "lynch": ("neutral", 63), "fisher": ("bullish", 82), "combined": "보유", "comb_conf": 66},
    {"ticker": "003380", "name": "하림지주", "buffett": ("bullish", 66), "lynch": ("neutral", 52), "fisher": ("neutral", 45), "combined": "보유", "comb_conf": 54},
    {"ticker": "085660", "name": "차바이오텍", "buffett": ("bearish", 80), "lynch": ("bearish", 85), "fisher": ("bearish", 80), "combined": "매도", "comb_conf": 82},
    {"ticker": "319660", "name": "피에스케이", "buffett": ("neutral", 56), "lynch": ("bullish", 77), "fisher": ("bullish", 76), "combined": "매수", "comb_conf": 70},
    {"ticker": "031980", "name": "피에스케이홀딩스", "buffett": ("neutral", 54), "lynch": ("bullish", 82), "fisher": ("bullish", 83), "combined": "매수", "comb_conf": 73},
    {"ticker": "388720", "name": "유일로보틱스", "buffett": ("neutral", 50), "lynch": ("bearish", 83), "fisher": ("bearish", 75), "combined": "매도", "comb_conf": 69},
    {"ticker": "095610", "name": "테스", "buffett": ("neutral", 52), "lynch": ("bullish", 81), "fisher": ("bullish", 70), "combined": "매수", "comb_conf": 68},
    {"ticker": "131970", "name": "두산테스나", "buffett": ("neutral", 51), "lynch": ("neutral", 60), "fisher": ("neutral", 55), "combined": "보유", "comb_conf": 55},
    {"ticker": "036540", "name": "SFA반도체", "buffett": ("neutral", 53), "lynch": ("neutral", 58), "fisher": ("neutral", 52), "combined": "보유", "comb_conf": 54},
]


# === Sheet 1: 요약 ===
ws1 = wb.active
ws1.title = "요약"
ws1.merge_cells("A1:B1")
ws1.cell(row=1, column=1, value="AI Hedge Fund 포트폴리오 리포트").fill = TITLE_FILL
ws1.cell(row=1, column=1).font = TITLE_FONT
ws1.cell(row=1, column=1).alignment = Alignment(horizontal="center")

info = [
    (3, "분석 일자", "2026-02-11"),
    (4, "분석 대상", "KRX (KOSPI200+KOSDAQ150) 상위 30개"),
    (5, "분석 전략", "하이브리드 (펀더멘털 70% + 모멘텀 30%)"),
    (6, "투자자 관점", "W.Buffett, P.Lynch, P.Fisher"),
    (8, "포트폴리오 통계", ""),
    (9, "편입 종목 수", "12 / 30 (40.0%)"),
    (10, "평균 신뢰도", "69%"),
    (11, "평균 예상 수익률", "+8.0%"),
    (12, "매수 비중", "100.0%"),
    (14, "섹터 분포", ""),
    (15, "반도체 장비", "34.2%"),
    (16, "반도체 소재", "24.4%"),
    (17, "반도체 제조", "9.2%"),
    (18, "지주회사", "8.5%"),
    (19, "건자재/도료", "8.1%"),
    (20, "전자부품", "7.9%"),
    (21, "섬유/의류", "7.6%"),
]
for row, label, val in info:
    ws1.cell(row=row, column=1, value=label).font = Font(name="맑은 고딕", bold=True if row in (8, 14) else False, size=11)
    ws1.cell(row=row, column=2, value=val).font = DATA_FONT
    ws1.cell(row=row, column=1).border = THIN_BORDER
    ws1.cell(row=row, column=2).border = THIN_BORDER

# Pie chart
chart = PieChart()
chart.title = "섹터 분포"
chart.style = 10
data_ref = Reference(ws1, min_col=2, min_row=15, max_row=21)
cats_ref = Reference(ws1, min_col=1, min_row=15, max_row=21)
chart.add_data(data_ref, titles_from_data=False)
chart.set_categories(cats_ref)
ws1.add_chart(chart, "D3")
auto_width(ws1, 2)

# === Sheet 2: 포트폴리오 ===
ws2 = wb.create_sheet("포트폴리오")
headers2 = ["#", "종목코드", "회사명", "비중", "신호", "신뢰도", "예상수익률", "시가총액", "P/E", "ROE", "PEG", "합의", "섹터"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, len(headers2))

for idx, p in enumerate(portfolio, 2):
    vals = [p["rank"], p["ticker"], p["name"], f'{p["weight"]}%', p["signal"], p["confidence"], f'+{p["ret"]}%', p["mcap"], p["pe"], p["roe"], p["peg"], p["consensus"], p["sector"]]
    for col, v in enumerate(vals, 1):
        ws2.cell(row=idx, column=col, value=v)
    style_data_row(ws2, idx, len(headers2))
    for col in range(1, len(headers2) + 1):
        ws2.cell(row=idx, column=col).fill = BUY_FILL

total_row = len(portfolio) + 2
ws2.cell(row=total_row, column=3, value="합계/평균").font = TOTAL_FONT
ws2.cell(row=total_row, column=3).fill = TOTAL_FILL
ws2.cell(row=total_row, column=4, value="100.0%").font = TOTAL_FONT
ws2.cell(row=total_row, column=4).fill = TOTAL_FILL
ws2.cell(row=total_row, column=6, value=69).font = TOTAL_FONT
ws2.cell(row=total_row, column=6).fill = TOTAL_FILL
ws2.cell(row=total_row, column=7, value="+8.0%").font = TOTAL_FONT
ws2.cell(row=total_row, column=7).fill = TOTAL_FILL
for col in range(1, len(headers2) + 1):
    ws2.cell(row=total_row, column=col).border = THIN_BORDER

ws2.auto_filter.ref = f"A1:M{len(portfolio)+1}"
ws2.freeze_panes = "A2"
auto_width(ws2, len(headers2))

# === Sheet 3: 순위 ===
ws3 = wb.create_sheet("순위")
headers3 = ["순위", "종목코드", "회사명", "종합점수", "펀더멘털", "모멘텀", "앙상블", "신호", "예상수익률", "시가총액", "P/E", "ROE", "매출성장률", "PEG", "편입여부"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, len(headers3))

for idx, p in enumerate(predict_all, 2):
    vals = [p["rank"], p["ticker"], p["name"], p["score"], p["fund"], p["mom"], p["ens"], p["signal"], f'+{p["ret"]}%', p["mcap"], p["pe"], p["roe"], p["rev_g"], p["peg"], p["included"]]
    for col, v in enumerate(vals, 1):
        ws3.cell(row=idx, column=col, value=v)
    style_data_row(ws3, idx, len(headers3))
    if p["included"] == "예":
        for col in range(1, len(headers3) + 1):
            ws3.cell(row=idx, column=col).fill = BUY_FILL

ws3.auto_filter.ref = f"A1:O{len(predict_all)+1}"
ws3.freeze_panes = "A2"
auto_width(ws3, len(headers3))

# === Sheet 4: 투자자 매트릭스 ===
ws4 = wb.create_sheet("투자자 매트릭스")
headers4 = ["종목코드", "회사명", "W.Buffett", "P.Lynch", "P.Fisher", "종합신호", "종합신뢰도"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, len(headers4))

signal_kr = {"bullish": "매수", "neutral": "중립", "bearish": "매도"}
for idx, m in enumerate(investor_matrix, 2):
    ws4.cell(row=idx, column=1, value=m["ticker"])
    ws4.cell(row=idx, column=2, value=m["name"])
    for col, inv in [(3, "buffett"), (4, "lynch"), (5, "fisher")]:
        sig, conf = m[inv]
        cell = ws4.cell(row=idx, column=col, value=f'{signal_kr[sig]}({conf})')
        if sig == "bullish":
            cell.fill = BUY_FILL
            cell.font = BUY_FONT
        elif sig == "bearish":
            cell.fill = SELL_FILL
            cell.font = SELL_FONT
        else:
            cell.fill = NEUTRAL_FILL
            cell.font = NEUTRAL_FONT
    ws4.cell(row=idx, column=6, value=m["combined"])
    ws4.cell(row=idx, column=7, value=m["comb_conf"])
    style_data_row(ws4, idx, len(headers4))

ws4.auto_filter.ref = f"A1:G{len(investor_matrix)+1}"
ws4.freeze_panes = "A2"
auto_width(ws4, len(headers4))

# === Sheet 5: 투자자 상세 ===
ws5 = wb.create_sheet("투자자 상세")
headers5 = ["종목코드", "회사명", "투자자", "신호", "신뢰도", "분석근거"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, len(headers5))

reasons = {
    "066570": {"buffett": "ROE 2.3%로 낮고 P/E 60은 과대평가. 영업이익률 3.9%로 경쟁우위 불명확", "lynch": "PER 60으로 과대평가, 순이익 2년 연속 하락. Slow Grower인데 고성장주 가격", "fisher": "매출 성장률 6.6%로 완만, 순이익 48% 급감하여 수익성 악화"},
    "402340": {"buffett": "ROE 18.6%, 영업마진 우수. 부채비율 0.12로 낮고 지주회사 구조의 견고한 moat", "lynch": "지주회사로 본질가치 판단 어려움. 흑자전환하며 ROE 18.6%로 양호하나 매출 감소", "fisher": "순이익 흑자전환으로 경영진 개선 능력 입증. ROE 18.6%, 높은 수익성"},
    "000660": {"buffett": "ROE 26.8%로 우수하나 반도체 사이클 변동성이 영구 보유 원칙과 불일치", "lynch": "메모리 업황 회복으로 매출 2배 증가, PEG < 0.1 극도 저평가. 10배주 잠재력", "fisher": "매출 102% 폭증, 영업마진 35%, ROE 27%로 업계 최상위. 장기 성장 동력 확보"},
    "240810": {"buffett": "P/E 245로 심각한 고평가. ROE 2.3%, 영업마진 1.4%로 품질 낮음", "lynch": "PER 245로 극단적 고평가. 영업이익률 1.4%로 낮음", "fisher": "흑자전환했으나 영업마진 1.4%로 미약. 수익성 개선 추세 불확실"},
    "039030": {"buffett": "ROE 7.1%, P/E 107로 고평가. 수익성이 버핏 기준 15%+ ROE 미달", "lynch": "PER 107, PEG 6.3으로 과대평가. 매출/이익 감소 추세", "fisher": "매출 정체, 영업마진 9.7%로 평범. 레이저 장비 시장 내 차별화 제한적"},
    "140410": {"buffett": "ROE -42%로 심각한 적자. 바이오 업종은 예측 불가능하며 Circle of Competence 밖", "lynch": "4년 연속 적자, PBR 94로 극심한 고평가. 수익 없는 스토리주", "fisher": "4년 연속 영업손실, ROE -42%로 경영진 실행력 심각한 의문"},
    "002380": {"buffett": "P/E 13, P/B 0.8로 장부가 이하 거래. 안정적 건자재/도료 사업으로 안전마진 충분", "lynch": "PER 13, PEG < 0.1 극도 저평가. 2024년 순이익 3배 증가. 청산가치 이하 거래", "fisher": "순이익 3배 증가는 긍정적이나 영업마진 7%로 낮음. 장기 성장성 제한적"},
    "017800": {"buffett": "ROE 14.4%, 승강기 유지보수의 반복 수익 모델은 강한 moat. 배당수익률 5.4% 매력적", "lynch": "PER 20, 배당수익률 5.4%로 Stalwart. 그러나 순이익 37% 감소", "fisher": "매출 11% 성장, 영업마진 7.8%는 안정적이나 순이익 37% 감소로 수익성 악화"},
    "403870": {"buffett": "ROE 31%로 극히 우수. 부채비율 0.15, 유동비율 5.8로 탄탄한 재무. 반도체 장비 경쟁우위", "lynch": "PER 41, PEG 5.8로 고평가. ROE 31%는 우수하나 매출 성장 제한적", "fisher": "ROE 31%, 영업이익 940억원으로 고수익성. 기술 집약적 사업, 재무 건전성 우수"},
    "357780": {"buffett": "ROE 11.5%, 영업마진 19.5%로 양호. 반도체 소재의 높은 전환비용이 moat", "lynch": "PER 26, 순이익 감소. PEG > 3으로 고평가. 안정적이나 저평가 Fast Grower 아님", "fisher": "영업마진 19.4%, 순마진 13.9%로 고수익. 기술적 해자와 고객 다변화로 장기 경쟁력"},
    "237690": {"buffett": "CDMO 업종은 성장성 높으나 제약 계약 의존도로 예측성 낮음", "lynch": "PER 85, CDMO 기업으로 흑자 기저효과. PEG > 1로 적정가 상회", "fisher": "CDMO 사업으로 순이익 85% 증가는 긍정적이나 매출 4% 감소"},
    "009970": {"buffett": "섬유/의류 지주로 안정적 배당 기대. 글로벌 브랜드 납품은 장기 관계 기반 moat", "lynch": "PER 7.8, PBR 0.96으로 저평가. 배당수익률 2.4%. Value + Dividend play", "fisher": "순이익 31% 감소, 성숙 산업으로 R&D 투자 및 혁신 동력 부족"},
    "058610": {"buffett": "모터/정밀기기 업종은 기술 변화 빠름. 규모 작아 경쟁우위 판단 어려움", "lynch": "PER 231, PEG > 12 극단적 고평가. 영업이익률 3.2%로 낮음", "fisher": "매출 정체, 영업마진 3%로 매우 낮음. 차별화 약함"},
    "010060": {"buffett": "화학/태양광 지주. 폴리실리콘 가격 변동성 높아 예측 어려움", "lynch": "순이익 84% 급감. PER 29로 지주회사 치고 고평가. Cyclical 하락 사이클 진입", "fisher": "순이익 84% 급감. 수익성 변동성 극심. 장기 성장 비전 불명확"},
    "028670": {"buffett": "해운업은 운임 사이클 변동성 극심. 자본집약적이며 지속 가능 moat 부족", "lynch": "PER 9.8, PBR 0.46으로 저평가. 해운 Cyclical로 업황 정점 후 조정 가능성", "fisher": "매출 18% 증가로 안정적이나 해운업은 경기 순환적. 장기 경쟁우위 없음"},
    "005290": {"buffett": "반도체 소재 업체. 업종 특성상 중립", "lynch": "PER 17, PEG 1.06으로 적정가. ROE 14%, 안정적 성장 스토리", "fisher": "영업마진 14.8%, ROE 14.1%로 우수. 기술적 해자 보유"},
    "064760": {"buffett": "반도체 소재. 경쟁우위 불명확", "lynch": "PER 34, PEG 1.9로 약간 고평가. 그러나 매출 성장 21.7%, 영업이익률 29%로 우수", "fisher": "영업마진 29.3%, 순마진 26%로 초고수익. 반도체 소재 기술 리더십 확보"},
    "083650": {"buffett": "발전 설비는 정부 정책 및 입찰 의존. 예측 가능성 낮음", "lynch": "PER 119, 부채비율 3.5로 고위험. 턴어라운드 초기지만 재무 건전성 약함", "fisher": "적자에서 흑자전환했으나 순마진 4.8%로 낮음. 장기 성장성 제한적"},
    "067310": {"buffett": "반도체 패키징. 기술 변화 빠르고 자본지출 많음. 장기 경쟁력 판단 보류", "lynch": "2024년 적자전환. 부채비율 2.2로 높음. Cyclical 하락 국면", "fisher": "흑자에서 적자전환. 부채비율 221%로 높음. 경영진 실행력 의문"},
    "065350": {"buffett": "전자부품 업종. 대기업 납품 의존도 높으면 가격결정력 약함", "lynch": "순이익 63% 급감. PBR 9.4로 고평가. Slow Grower인데 성장주 가격", "fisher": "순이익 63% 급감, 영업마진 3%로 매우 낮음. 수익성 지속 악화 추세"},
    "178320": {"buffett": "전자부품/통신장비. 5G 수요 등 테마 의존도 높아 장기 예측성 낮음", "lynch": "PER 21.7, 흑자전환으로 성장률 470%. 매출 56% 급증. 10배주 잠재력", "fisher": "매출 56% 폭증, 적자에서 흑자전환. 영업마진 9%로 개선 여지 있음"},
    "101490": {"buffett": "블랭크마스크 특화 소재로 진입장벽 있으나 시장 규모 작음", "lynch": "PER 61, PEG 3.6으로 고평가. Fast Grower이나 밸류에이션 부담", "fisher": "매출 17% 성장, 영업마진 16.7%로 초고수익. 블랭크마스크 기술 독점적 지위"},
    "003380": {"buffett": "식품/사료 지주는 이해 가능한 비즈니스. 필수소비재로 안정적 수요", "lynch": "PER 58, PBR 0.5로 엇갈린 신호. 순이익 정체. Slow Grower", "fisher": "순이익 2% 감소, 영업마진 6%로 낮음. R&D 투자 및 혁신 동력 부재"},
    "085660": {"buffett": "바이오/줄기세포는 규제 및 임상 불확실성 높음. Circle of Competence 벗어남", "lynch": "4년 연속 적자, 부채비율 2.1, 유동비율 0.72로 재무 위험", "fisher": "4년 연속 적자, 영업손실 확대. ROE -2%, 부채비율 205%로 재무 취약"},
    "319660": {"buffett": "반도체 세정장비. 장비 업종은 사이클 타고 변동성 높음", "lynch": "PER 21.2, PEG 0.42로 극도 저평가. ROE 16.8%. 10배주 잠재력", "fisher": "영업마진 21%, 순마진 19.9%로 고수익. 반도체 세정장비 기술 경쟁력 우수"},
    "031980": {"buffett": "장비 업종 사이클 리스크. 지주 할인 존재 가능", "lynch": "PER 16.8, PEG 0.13으로 극도 저평가. ROE 22.2%, 영업이익률 41%. Fast Grower + 저평가 완벽 조합", "fisher": "매출 127% 폭증, 영업마진 41%, ROE 22%로 우수. 자회사 시너지 극대화"},
    "388720": {"buffett": "로봇 업종은 성장 초기 단계. 수익성 검증 부족하고 경쟁 치열", "lynch": "3년 연속 적자, PBR 14.4로 극단적 고평가. Speculative stock", "fisher": "3년 연속 적자, 순마진 -25.7%. ROE -9.5%로 경영진 실행력 의문"},
    "095610": {"buffett": "반도체 장비. 삼성전자 의존도 높으면 협상력 제한", "lynch": "PER 26.4, 흑자전환으로 성장률 2,623%. 매출 63% 급증. 턴어라운드 초기 포착", "fisher": "매출 63% 증가, 영업마진 16%, 순마진 17.8%로 고수익. 업황 회복 수혜"},
    "131970": {"buffett": "반도체 테스트. 장비 업종 공통 리스크. 성장성 불투명", "lynch": "PER 33.4, 순이익 25% 감소. Cyclical로 업황 정점 통과 가능성", "fisher": "매출 10% 성장은 안정적이나 순이익 25% 감소. 경쟁 심화"},
    "036540": {"buffett": "반도체 패키징. 후공정은 마진 낮고 경쟁 치열", "lynch": "PER 55.5, 흑자전환했으나 영업이익 거의 0원. 밸류에이션 부담", "fisher": "흑자전환은 긍정적이나 영업이익 거의 0원. 순마진 5%로 매우 낮음"},
}

row = 2
for m in investor_matrix:
    t = m["ticker"]
    for inv_name, inv_key in [("W.Buffett", "buffett"), ("P.Lynch", "lynch"), ("P.Fisher", "fisher")]:
        sig, conf = m[inv_key]
        ws5.cell(row=row, column=1, value=t)
        ws5.cell(row=row, column=2, value=m["name"])
        ws5.cell(row=row, column=3, value=inv_name)
        ws5.cell(row=row, column=4, value=signal_kr[sig])
        ws5.cell(row=row, column=5, value=conf)
        ws5.cell(row=row, column=6, value=reasons.get(t, {}).get(inv_key, ""))
        style_data_row(ws5, row, len(headers5))
        row += 1

ws5.freeze_panes = "A2"
auto_width(ws5, len(headers5))
ws5.column_dimensions["F"].width = 60

# === Sheet 6: 리스크 분석 ===
ws6 = wb.create_sheet("리스크 분석")
ws6.cell(row=1, column=1, value="리스크 분석").fill = TITLE_FILL
ws6.cell(row=1, column=1).font = TITLE_FONT
ws6.merge_cells("A1:B1")

risk_data = [
    (3, "집중도 지표", ""),
    (4, "상위 1종목 비중", "9.2% (SK하이닉스)"),
    (5, "상위 3종목 비중", "26.8%"),
    (6, "상위 5종목 비중", "43.7%"),
    (7, "HHI (허핀달 지수)", "0.070"),
    (9, "섹터 집중도", ""),
    (10, "반도체 관련 전체 비중", "76.3% (12종목 중 9종목)"),
    (11, "반도체 장비", "34.2%"),
    (12, "반도체 소재", "24.4%"),
    (13, "반도체 제조", "9.2%"),
    (15, "투자자 합의 품질", ""),
    (16, "만장일치 비율 (3/3)", "0% (없음)"),
    (17, "다수 합의 (2/3)", "100% (12개 전체)"),
    (19, "비편입 사유 분포", ""),
    (20, "투자자 과반 미달 (0/3)", "15개"),
    (21, "투자자 과반 미달 (1/3)", "3개"),
]
for row, label, val in risk_data:
    ws6.cell(row=row, column=1, value=label).font = Font(name="맑은 고딕", bold=True if row in (3, 9, 15, 19) else False, size=11)
    ws6.cell(row=row, column=2, value=val).font = DATA_FONT
    ws6.cell(row=row, column=1).border = THIN_BORDER
    ws6.cell(row=row, column=2).border = THIN_BORDER

auto_width(ws6, 2)

# Save
output_path = os.path.join(os.path.dirname(__file__), "krx_20260211_buffett_fisher_lynch.xlsx")
wb.save(output_path)
print(f"Excel saved: {output_path}")
