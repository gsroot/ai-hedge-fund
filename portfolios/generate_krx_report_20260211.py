"""KRX Portfolio Report Excel Generator - 2026-02-11 (Updated)"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
import os

wb = openpyxl.Workbook()

# === Styles ===
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
BUY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BUY_FONT = Font(name="맑은 고딕", color="375623")
SELL_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
SELL_FONT = Font(name="맑은 고딕", color="C62828")
NEUTRAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
NEUTRAL_FONT = Font(name="맑은 고딕", color="616161")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(name="맑은 고딕", bold=True)
DATA_FONT = Font(name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22

def style_data_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 18

def auto_width(ws, max_col):
    for col in range(1, max_col + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 30)


# === DATA ===
portfolio = [
    {"rank": 1, "ticker": "033100", "name": "제룡전기", "weight": 12.0, "signal": "강력매수", "confidence": 86, "ret": 21.6, "mcap": "₩8,609억", "pe": "10.8", "roe": "41.0%", "peg": "N/A", "consensus": "3/3", "sector": "산업재"},
    {"rank": 2, "ticker": "018290", "name": "브이티", "weight": 11.7, "signal": "강력매수", "confidence": 84, "ret": 14.1, "mcap": "₩7,233억", "pe": "7.5", "roe": "43.0%", "peg": "N/A", "consensus": "3/3", "sector": "소비재"},
    {"rank": 3, "ticker": "067160", "name": "SOOP", "weight": 10.6, "signal": "매수", "confidence": 76, "ret": 12.4, "mcap": "₩8,230억", "pe": "7.5", "roe": "27.0%", "peg": "N/A", "consensus": "2/3", "sector": "커뮤니케이션"},
    {"rank": 4, "ticker": "011200", "name": "HMM", "weight": 10.6, "signal": "강력매수", "confidence": 76, "ret": 13.7, "mcap": "₩20조", "pe": "4.2", "roe": "14.0%", "peg": "N/A", "consensus": "3/3", "sector": "산업재"},
    {"rank": 5, "ticker": "259960", "name": "크래프톤", "weight": 10.4, "signal": "매수", "confidence": 75, "ret": 12.8, "mcap": "₩12조", "pe": "8.9", "roe": "19.0%", "peg": "N/A", "consensus": "2/3", "sector": "커뮤니케이션"},
    {"rank": 6, "ticker": "211050", "name": "인카금융서비스", "weight": 9.8, "signal": "매수", "confidence": 70, "ret": 14.0, "mcap": "₩6,941억", "pe": "11.2", "roe": "42.0%", "peg": "N/A", "consensus": "2/3", "sector": "금융"},
    {"rank": 7, "ticker": "031980", "name": "피에스케이홀딩스", "weight": 7.4, "signal": "강력매수", "confidence": 82, "ret": 19.8, "mcap": "₩16,086억", "pe": "16.8", "roe": "22.0%", "peg": "0.13", "consensus": "3/3", "sector": "IT"},
    {"rank": 8, "ticker": "000660", "name": "SK하이닉스", "weight": 7.2, "signal": "매수", "confidence": 79, "ret": 14.0, "mcap": "₩636조", "pe": "30.4", "roe": "27.0%", "peg": "N/A", "consensus": "2/3", "sector": "IT"},
    {"rank": 9, "ticker": "036930", "name": "주성엔지니어링", "weight": 7.0, "signal": "강력매수", "confidence": 77, "ret": 15.7, "mcap": "₩2.3조", "pe": "21.8", "roe": "19.0%", "peg": "N/A", "consensus": "3/3", "sector": "IT"},
    {"rank": 10, "ticker": "183300", "name": "코미코", "weight": 6.9, "signal": "강력매수", "confidence": 76, "ret": 11.8, "mcap": "₩11,193억", "pe": "19.8", "roe": "23.0%", "peg": "N/A", "consensus": "3/3", "sector": "IT"},
    {"rank": 11, "ticker": "348210", "name": "넥스틴", "weight": 6.5, "signal": "매수", "confidence": 72, "ret": 14.6, "mcap": "₩8,172억", "pe": "20.8", "roe": "24.0%", "peg": "N/A", "consensus": "2/3", "sector": "IT"},
]

predict_all = [
    {"rank": 1, "ticker": "033100", "name": "제룡전기", "score": 9.17, "fund": 9.0, "mom": 10.0, "ens": 8.7, "signal": "강력매수", "ret": 21.6, "mcap": "₩8,609억", "pe": "10.8", "pb": "N/A", "roe": "41%", "rev_g": "N/A", "peg": "N/A", "included": "예"},
    {"rank": 2, "ticker": "031980", "name": "피에스케이홀딩스", "score": 8.65, "fund": 8.1, "mom": 10.0, "ens": 8.0, "signal": "강력매수", "ret": 19.8, "mcap": "₩16,086억", "pe": "16.8", "pb": "N/A", "roe": "22%", "rev_g": "128%", "peg": "0.13", "included": "예"},
    {"rank": 3, "ticker": "131290", "name": "티에스이", "score": 7.95, "fund": 6.3, "mom": 10.0, "ens": 6.1, "signal": "매수", "ret": 17.3, "mcap": "₩9,701억", "pe": "22.2", "pb": "N/A", "roe": "11%", "rev_g": "40%", "peg": "N/A", "included": "아니오"},
    {"rank": 4, "ticker": "137400", "name": "피엔티", "score": 7.76, "fund": 7.5, "mom": 9.0, "ens": 7.3, "signal": "매수", "ret": 16.7, "mcap": "₩12,106억", "pe": "8.2", "pb": "N/A", "roe": "23%", "rev_g": "90%", "peg": "N/A", "included": "아니오"},
    {"rank": 5, "ticker": "095610", "name": "테스", "score": 7.66, "fund": 6.7, "mom": 10.0, "ens": 6.5, "signal": "매수", "ret": 16.3, "mcap": "₩12,565억", "pe": "26.7", "pb": "N/A", "roe": "13%", "rev_g": "63%", "peg": "N/A", "included": "아니오"},
    {"rank": 6, "ticker": "000240", "name": "한국앤컴퍼니", "score": 7.63, "fund": 7.9, "mom": 7.0, "ens": 7.3, "signal": "매수", "ret": 16.2, "mcap": "₩2.9조", "pe": "8.2", "pb": "N/A", "roe": "8%", "rev_g": "28%", "peg": "N/A", "included": "아니오"},
    {"rank": 7, "ticker": "036930", "name": "주성엔지니어링", "score": 7.48, "fund": 6.8, "mom": 9.0, "ens": 6.8, "signal": "매수", "ret": 15.7, "mcap": "₩2.3조", "pe": "21.8", "pb": "N/A", "roe": "19%", "rev_g": "44%", "peg": "N/A", "included": "예"},
    {"rank": 8, "ticker": "222080", "name": "씨아이에스", "score": 7.27, "fund": 5.9, "mom": 9.0, "ens": 5.1, "signal": "매수", "ret": 15.0, "mcap": "₩10,445억", "pe": "17.2", "pb": "N/A", "roe": "12%", "rev_g": "64%", "peg": "N/A", "included": "아니오"},
    {"rank": 9, "ticker": "348210", "name": "넥스틴", "score": 7.17, "fund": 7.5, "mom": 6.5, "ens": 7.2, "signal": "매수", "ret": 14.6, "mcap": "₩8,172억", "pe": "20.8", "pb": "N/A", "roe": "24%", "rev_g": "N/A", "peg": "N/A", "included": "예"},
    {"rank": 10, "ticker": "178320", "name": "서진시스템", "score": 7.11, "fund": 4.7, "mom": 10.0, "ens": 3.9, "signal": "매수", "ret": 14.4, "mcap": "₩2.2조", "pe": "22.4", "pb": "N/A", "roe": "10%", "rev_g": "56%", "peg": "N/A", "included": "아니오"},
    {"rank": 11, "ticker": "018290", "name": "브이티", "score": 7.03, "fund": 7.9, "mom": 5.0, "ens": 7.2, "signal": "매수", "ret": 14.1, "mcap": "₩7,233억", "pe": "7.5", "pb": "N/A", "roe": "43%", "rev_g": "46%", "peg": "N/A", "included": "예"},
    {"rank": 12, "ticker": "000660", "name": "SK하이닉스", "score": 7.01, "fund": 5.7, "mom": 10.0, "ens": 7.0, "signal": "매수", "ret": 14.0, "mcap": "₩636조", "pe": "30.4", "pb": "N/A", "roe": "27%", "rev_g": "102%", "peg": "N/A", "included": "예"},
    {"rank": 13, "ticker": "211050", "name": "인카금융서비스", "score": 7.00, "fund": 6.8, "mom": 7.5, "ens": 6.0, "signal": "매수", "ret": 14.0, "mcap": "₩6,941억", "pe": "11.2", "pb": "N/A", "roe": "42%", "rev_g": "49%", "peg": "N/A", "included": "예"},
    {"rank": 14, "ticker": "059090", "name": "미코", "score": 6.98, "fund": 5.3, "mom": 9.0, "ens": 4.8, "signal": "매수", "ret": 13.9, "mcap": "₩5,551억", "pe": "28.9", "pb": "N/A", "roe": "14%", "rev_g": "42%", "peg": "N/A", "included": "아니오"},
    {"rank": 15, "ticker": "214150", "name": "클래시스", "score": 6.96, "fund": 6.1, "mom": 9.0, "ens": 6.2, "signal": "매수", "ret": 13.9, "mcap": "₩4.4조", "pe": "44.6", "pb": "N/A", "roe": "22%", "rev_g": "35%", "peg": "N/A", "included": "아니오"},
    {"rank": 16, "ticker": "011200", "name": "HMM", "score": 6.92, "fund": 7.2, "mom": 5.5, "ens": 7.7, "signal": "매수", "ret": 13.7, "mcap": "₩20조", "pe": "4.2", "pb": "N/A", "roe": "14%", "rev_g": "39%", "peg": "N/A", "included": "예"},
    {"rank": 17, "ticker": "095340", "name": "ISC", "score": 6.75, "fund": 4.9, "mom": 9.0, "ens": 5.7, "signal": "매수", "ret": 13.1, "mcap": "₩3.6조", "pe": "64.0", "pb": "N/A", "roe": "10%", "rev_g": "24%", "peg": "N/A", "included": "아니오"},
    {"rank": 18, "ticker": "213420", "name": "덕산네오룩스", "score": 6.74, "fund": 6.2, "mom": 8.0, "ens": 6.0, "signal": "매수", "ret": 13.1, "mcap": "₩10,814억", "pe": "23.4", "pb": "N/A", "roe": "12%", "rev_g": "30%", "peg": "N/A", "included": "아니오"},
    {"rank": 19, "ticker": "319660", "name": "피에스케이", "score": 6.67, "fund": 5.2, "mom": 10.0, "ens": 6.5, "signal": "매수", "ret": 12.9, "mcap": "₩16,830억", "pe": "21.3", "pb": "N/A", "roe": "17%", "rev_g": "13%", "peg": "0.42", "included": "아니오"},
    {"rank": 20, "ticker": "259960", "name": "크래프톤", "score": 6.65, "fund": 6.9, "mom": 5.0, "ens": 7.5, "signal": "매수", "ret": 12.8, "mcap": "₩12조", "pe": "8.9", "pb": "N/A", "roe": "19%", "rev_g": "42%", "peg": "N/A", "included": "예"},
    {"rank": 21, "ticker": "278470", "name": "에이피알", "score": 6.56, "fund": 4.2, "mom": 9.5, "ens": 4.8, "signal": "매수", "ret": 12.5, "mcap": "₩11조", "pe": "99.1", "pb": "N/A", "roe": "33%", "rev_g": "N/A", "peg": "N/A", "included": "아니오"},
    {"rank": 22, "ticker": "067160", "name": "SOOP", "score": 6.55, "fund": 6.8, "mom": 5.0, "ens": 7.3, "signal": "매수", "ret": 12.4, "mcap": "₩8,230억", "pe": "7.5", "pb": "N/A", "roe": "27%", "rev_g": "20%", "peg": "N/A", "included": "예"},
    {"rank": 23, "ticker": "323280", "name": "태성", "score": 6.49, "fund": 4.4, "mom": 9.0, "ens": 4.1, "signal": "매수", "ret": 12.2, "mcap": "₩2.4조", "pe": "345.7", "pb": "N/A", "roe": "15%", "rev_g": "78%", "peg": "N/A", "included": "아니오"},
    {"rank": 24, "ticker": "352480", "name": "씨앤씨인터내셔널", "score": 6.45, "fund": 6.6, "mom": 6.0, "ens": 6.5, "signal": "매수", "ret": 12.1, "mcap": "₩4,845억", "pe": "11.0", "pb": "N/A", "roe": "16%", "rev_g": "28%", "peg": "N/A", "included": "아니오"},
    {"rank": 25, "ticker": "161390", "name": "한국타이어앤테크놀로지", "score": 6.44, "fund": 5.3, "mom": 9.0, "ens": 6.2, "signal": "매수", "ret": 12.0, "mcap": "₩9.0조", "pe": "8.0", "pb": "N/A", "roe": "10%", "rev_g": "5%", "peg": "N/A", "included": "아니오"},
    {"rank": 26, "ticker": "005930", "name": "삼성전자", "score": 6.43, "fund": 3.5, "mom": 10.0, "ens": 4.6, "signal": "매수", "ret": 12.0, "mcap": "₩994조", "pe": "33.9", "pb": "N/A", "roe": "9%", "rev_g": "16%", "peg": "N/A", "included": "아니오"},
    {"rank": 27, "ticker": "101490", "name": "에스앤에스텍", "score": 6.42, "fund": 3.5, "mom": 10.0, "ens": 4.3, "signal": "매수", "ret": 12.0, "mcap": "₩19,116억", "pe": "61.7", "pb": "N/A", "roe": "12%", "rev_g": "17%", "peg": "N/A", "included": "아니오"},
    {"rank": 28, "ticker": "079370", "name": "제우스", "score": 6.41, "fund": 4.9, "mom": 10.0, "ens": 5.6, "signal": "매수", "ret": 11.9, "mcap": "₩5,689억", "pe": "13.5", "pb": "N/A", "roe": "11%", "rev_g": "22%", "peg": "N/A", "included": "아니오"},
    {"rank": 29, "ticker": "183300", "name": "코미코", "score": 6.39, "fund": 6.3, "mom": 6.5, "ens": 6.1, "signal": "매수", "ret": 11.8, "mcap": "₩11,193억", "pe": "19.8", "pb": "N/A", "roe": "23%", "rev_g": "65%", "peg": "N/A", "included": "예"},
    {"rank": 30, "ticker": "036620", "name": "감성코퍼레이션", "score": 6.39, "fund": 5.3, "mom": 9.0, "ens": 6.3, "signal": "매수", "ret": 11.9, "mcap": "₩6,381억", "pe": "22.0", "pb": "N/A", "roe": "26%", "rev_g": "24%", "peg": "N/A", "included": "아니오"},
]

investor_matrix = [
    {"ticker": "033100", "name": "제룡전기", "buffett": ("bullish", 85), "lynch": ("bullish", 85), "fisher": ("bullish", 88), "combined": "강력매수", "comb_conf": 86},
    {"ticker": "031980", "name": "피에스케이홀딩스", "buffett": ("bullish", 82), "lynch": ("bullish", 85), "fisher": ("bullish", 80), "combined": "강력매수", "comb_conf": 82},
    {"ticker": "131290", "name": "티에스이", "buffett": ("neutral", 55), "lynch": ("bullish", 75), "fisher": ("neutral", 72), "combined": "보유", "comb_conf": 67},
    {"ticker": "137400", "name": "피엔티", "buffett": ("neutral", 65), "lynch": ("bullish", 85), "fisher": ("neutral", 70), "combined": "보유", "comb_conf": 73},
    {"ticker": "095610", "name": "테스", "buffett": ("neutral", 58), "lynch": ("bullish", 75), "fisher": ("neutral", 75), "combined": "보유", "comb_conf": 69},
    {"ticker": "000240", "name": "한국앤컴퍼니", "buffett": ("neutral", 68), "lynch": ("bullish", 70), "fisher": ("neutral", 72), "combined": "보유", "comb_conf": 70},
    {"ticker": "036930", "name": "주성엔지니어링", "buffett": ("bullish", 72), "lynch": ("bullish", 80), "fisher": ("bullish", 80), "combined": "강력매수", "comb_conf": 77},
    {"ticker": "222080", "name": "씨아이에스", "buffett": ("neutral", 62), "lynch": ("neutral", 65), "fisher": ("neutral", 67), "combined": "보유", "comb_conf": 65},
    {"ticker": "348210", "name": "넥스틴", "buffett": ("bullish", 78), "lynch": ("neutral", 70), "fisher": ("bullish", 68), "combined": "매수", "comb_conf": 72},
    {"ticker": "178320", "name": "서진시스템", "buffett": ("neutral", 52), "lynch": ("neutral", 60), "fisher": ("bearish", 49), "combined": "보유", "comb_conf": 54},
    {"ticker": "018290", "name": "브이티", "buffett": ("bullish", 88), "lynch": ("bullish", 75), "fisher": ("bullish", 88), "combined": "강력매수", "comb_conf": 84},
    {"ticker": "000660", "name": "SK하이닉스", "buffett": ("bullish", 75), "lynch": ("neutral", 80), "fisher": ("bullish", 83), "combined": "매수", "comb_conf": 79},
    {"ticker": "211050", "name": "인카금융서비스", "buffett": ("bearish", 45), "lynch": ("bullish", 85), "fisher": ("bullish", 80), "combined": "매수", "comb_conf": 70},
    {"ticker": "059090", "name": "미코", "buffett": ("neutral", 48), "lynch": ("neutral", 60), "fisher": ("neutral", 65), "combined": "보유", "comb_conf": 58},
    {"ticker": "214150", "name": "클래시스", "buffett": ("neutral", 68), "lynch": ("neutral", 55), "fisher": ("bullish", 73), "combined": "보유", "comb_conf": 65},
    {"ticker": "011200", "name": "HMM", "buffett": ("bullish", 80), "lynch": ("bullish", 75), "fisher": ("bullish", 73), "combined": "강력매수", "comb_conf": 76},
    {"ticker": "095340", "name": "ISC", "buffett": ("neutral", 58), "lynch": ("neutral", 65), "fisher": ("neutral", 75), "combined": "보유", "comb_conf": 66},
    {"ticker": "213420", "name": "덕산네오룩스", "buffett": ("neutral", 65), "lynch": ("neutral", 60), "fisher": ("neutral", 72), "combined": "보유", "comb_conf": 66},
    {"ticker": "319660", "name": "피에스케이", "buffett": ("bullish", 72), "lynch": ("neutral", 60), "fisher": ("neutral", 72), "combined": "보유", "comb_conf": 68},
    {"ticker": "259960", "name": "크래프톤", "buffett": ("bullish", 78), "lynch": ("neutral", 70), "fisher": ("bullish", 76), "combined": "매수", "comb_conf": 75},
    {"ticker": "278470", "name": "에이피알", "buffett": ("bearish", 40), "lynch": ("bearish", 40), "fisher": ("neutral", 75), "combined": "매도", "comb_conf": 52},
    {"ticker": "067160", "name": "SOOP", "buffett": ("bullish", 82), "lynch": ("bullish", 75), "fisher": ("neutral", 72), "combined": "매수", "comb_conf": 76},
    {"ticker": "323280", "name": "태성", "buffett": ("bearish", 35), "lynch": ("neutral", 55), "fisher": ("neutral", 65), "combined": "보유", "comb_conf": 52},
    {"ticker": "352480", "name": "씨앤씨인터내셔널", "buffett": ("bullish", 75), "lynch": ("neutral", 65), "fisher": ("bearish", 49), "combined": "보유", "comb_conf": 63},
    {"ticker": "161390", "name": "한국타이어앤테크놀로지", "buffett": ("neutral", 68), "lynch": ("neutral", 60), "fisher": ("bearish", 45), "combined": "보유", "comb_conf": 58},
    {"ticker": "005930", "name": "삼성전자", "buffett": ("neutral", 62), "lynch": ("neutral", 55), "fisher": ("neutral", 70), "combined": "보유", "comb_conf": 62},
    {"ticker": "101490", "name": "에스앤에스텍", "buffett": ("neutral", 60), "lynch": ("neutral", 40), "fisher": ("neutral", 70), "combined": "보유", "comb_conf": 57},
    {"ticker": "079370", "name": "제우스", "buffett": ("neutral", 63), "lynch": ("bullish", 70), "fisher": ("neutral", 65), "combined": "보유", "comb_conf": 66},
    {"ticker": "183300", "name": "코미코", "buffett": ("bullish", 70), "lynch": ("bullish", 80), "fisher": ("bullish", 78), "combined": "강력매수", "comb_conf": 76},
    {"ticker": "036620", "name": "감성코퍼레이션", "buffett": ("bullish", 76), "lynch": ("neutral", 60), "fisher": ("neutral", 75), "combined": "보유", "comb_conf": 70},
]

reasons = {
    "033100": {"buffett": "ROE 41%와 영업마진 37%로 최고 수준 수익성, 부채비율 0.23으로 안전, P/E 10.7배 매력적", "lynch": "PEG 매력적(P/E 10.8, 성장률 43%), 소형주(₩8,609억)로 10배 잠재력", "fisher": "탁월한 자본효율성(ROE 41%)과 합리적 밸류에이션(PER 10.8), 장기 보유 최적"},
    "031980": {"buffett": "ROE 22%와 영업마진 41%로 탁월한 수익성, P/E 17배 합리적 수준, 반도체 장비 경쟁우위", "lynch": "고성장(매출+128%), P/E 16.8로 합리적, 소형주(₩16,086억)로 GARP 전략에 부합", "fisher": "폭발적 성장(매출 128%)과 고마진(41%) 동시 달성, 스케일업 입증"},
    "131290": {"buffett": "ROE 11%로 기준 미달이나 반도체 장비 성장성 존재, 2023년 적자 이력으로 수익 안정성 우려", "lynch": "성장률 40% vs P/E 22로 PEG 양호, 강한 모멘텀(3M +59%), 소형주 10배 가능성", "fisher": "매출·이익 고성장(40%, 2249%), 지속가능 성장 모델 확립 필요"},
    "137400": {"buffett": "ROE 23%로 우수하나 부채비율 1.59로 과도, 2차전지 성장 잠재력은 있으나 재무 레버리지 부담", "lynch": "PEG 우수(P/E 8.2, 성장률 90%), ROE 23%, 린치 이상형의 저평가 성장주", "fisher": "매출·이익 고성장(90%, 108%), 지속가능 성장 모델 확립 필요"},
    "095610": {"buffett": "2023년 적자 후 회복세지만 수익 변동성 크고 P/E 26배로 고평가, ROE 13%는 평범", "lynch": "고성장(매출+63%), 모멘텀 강함, 부채 낮음, 소형주 10배 잠재력", "fisher": "매출·이익 고성장(63%, 2624%), 지속가능 성장 모델 확립 필요"},
    "000240": {"buffett": "ROE 8%로 낮으나 안정적 배당주, P/E 8배로 저평가이나 성장성 제한적", "lynch": "P/E 8.2로 저평가, 성장률 28%, 영업마진 30%로 우량", "fisher": "초고마진(30%)으로 가격결정력 탁월, 안정적 현금창출"},
    "036930": {"buffett": "ROE 19%와 영업마진 24%로 양호, 반도체 장비 수혜주이나 수익 변동성 우려", "lynch": "고성장(매출+44%), ROE 19%, 강한 모멘텀(3M +64%), 성장 여력 있음", "fisher": "매출·이익 고성장(44%, 214%), 지속가능 성장 모델 확립"},
    "222080": {"buffett": "ROE 12%로 평범, 부채비율 0.53으로 다소 높음, 디스플레이 장비 경쟁우위 불명확", "lynch": "성장률 64%로 우수하나 P/E 17로 약간 비쌈, 모멘텀 과열(RSI 72)", "fisher": "매출·이익 고성장(64%, 96%), 지속가능 성장 모델 확립 필요"},
    "348210": {"buffett": "ROE 24%와 영업마진 41%로 탁월, 부채비율 0.20으로 매우 건전, P/E 20배 합리적", "lynch": "ROE 24%, 영업마진 41%로 우량하나 모멘텀 약함, GARP 기준 애매", "fisher": "초고마진(41%)으로 가격결정력 탁월, 안정적 현금창출"},
    "178320": {"buffett": "ROE 10%로 낮고 부채비율 1.40으로 과도, 2023년 적자 이력, 재무 안정성 미흡", "lynch": "고성장(매출+56%)이나 ROE 10%로 낮음, 재무 건전성 부족", "fisher": "매출·이익 고성장이나 ROE 10%, 부채비율 과도로 Fisher 기준 미달"},
    "018290": {"buffett": "ROE 43%로 최상위, 영업마진 26% 우수, 부채비율 0.49 양호, P/E 7.5배 큰 안전마진", "lynch": "P/E 7.5로 저평가, ROE 43%, 성장률 46%, 펀더멘털 탄탄", "fisher": "탁월한 자본효율성(ROE 43%)과 합리적 밸류에이션(PER 7.5), 장기 보유 최적"},
    "000660": {"buffett": "ROE 27%와 영업마진 35%로 탁월하나 P/E 30배로 고평가, 메모리 경쟁우위 명확", "lynch": "메가캡(₩636조)으로 10배 불가능, 성장률 102%로 우수하나 P/E 30으로 비쌈", "fisher": "폭발적 성장(매출 102%)과 고마진(35%) 동시 달성, 스케일업 입증"},
    "211050": {"buffett": "ROE 42%로 높으나 부채비율 3.80으로 과도, 금융서비스업 레버리지 리스크", "lynch": "P/E 11, ROE 42%, 성장률 49%로 GARP 완벽, 소형주 10배 가능", "fisher": "탁월한 자본효율성(ROE 42%)과 합리적 밸류에이션(PER 11.2), 장기 보유 적합"},
    "059090": {"buffett": "ROE 14%로 평범, 부채비율 1.61 과다, 2023년 적자 이력, P/E 29배로 고평가", "lynch": "성장률 42%로 우수하나 P/E 29로 비쌈, 유동성 위험", "fisher": "매출·이익 고성장(42%, 3857%), 지속가능 성장 모델 검증 필요"},
    "214150": {"buffett": "ROE 22%와 영업마진 50%로 탁월하나 P/E 44배로 과도 고평가", "lynch": "P/E 45로 고평가, 성장률 35%로 PEG 불량, 모멘텀 의존형", "fisher": "매출·이익 고성장(35%, 31%), 지속가능 성장 모델 확립"},
    "011200": {"buffett": "ROE 14% 양호, 영업마진 30% 우수, P/E 4.2배로 극단적 저평가", "lynch": "P/E 4.2로 초저평가, 성장률 39%, 영업마진 30%, 해운 경기순환 리스크", "fisher": "매출·이익 고성장(39%, 290%), 지속가능 성장 모델 확립"},
    "095340": {"buffett": "ROE 10%로 낮고 P/E 64배로 극단적 고평가, 가격 대비 품질 불균형", "lynch": "P/E 64로 고평가, 성장률 24%로 PEG 나쁨, 모멘텀 과열", "fisher": "강력한 성장세(매출 24%) 지속, 시장점유율 확대 중"},
    "213420": {"buffett": "ROE 11%로 평범, 영업마진 25% 양호, P/E 23배 적정, ROE 개선 필요", "lynch": "성장률 30%, P/E 23로 보통, 모멘텀 양호하나 GARP 기준 매력 부족", "fisher": "강력한 성장세(매출 30%) 지속, 시장점유율 확대 중"},
    "319660": {"buffett": "ROE 17%와 영업마진 21%로 양호, 부채비율 0.21로 건전, P/E 21배 합리적", "lynch": "모멘텀 강하나 성장률 13%로 낮음, P/E 21로 비쌈, GARP 부적합", "fisher": "성장·수익성 지표 혼재, Fisher 기준 추가 검증 필요"},
    "259960": {"buffett": "ROE 19%와 영업마진 44%로 탁월, P/E 9배로 저평가, 게임 IP 경쟁우위 명확", "lynch": "P/E 8.9로 저평가, 성장률 42%, 대형주(₩12조)로 10배 불가능", "fisher": "폭발적 성장(매출 42%)과 고마진(44%) 동시 달성, 스케일업 입증"},
    "278470": {"buffett": "ROE 33%로 높으나 P/E 98배로 극단적 고평가, 안전마진 전무", "lynch": "P/E 99로 심각한 고평가, ROE 33% 우수하나 밸류에이션 과도", "fisher": "뛰어난 수익성(ROE 33%)과 지속 가능한 경쟁 우위, 경영진 품질 우수"},
    "067160": {"buffett": "ROE 27%와 영업마진 27% 우수, P/E 7.5배로 큰 저평가, 플랫폼 경쟁우위", "lynch": "P/E 7.5로 저평가, ROE 27%, 성장률 20%, 소형주 벨류+성장 조화", "fisher": "강력한 성장세(매출 20%) 지속, 시장점유율 확대 중"},
    "323280": {"buffett": "ROE 15%이나 적자 이력, P/E 344배로 극단적 고평가, 수익 변동성 심각", "lynch": "성장률 78%로 우수하나 P/E 346으로 극단적 고평가, 투기적", "fisher": "매출·이익 고성장(78%, 521%), 지속가능 성장 모델 확립 필요"},
    "352480": {"buffett": "ROE 16%와 영업마진 10% 양호, P/E 11배로 저평가, 화장품 브랜드 경쟁우위", "lynch": "P/E 11로 저평가, 성장률 28%로 양호, Graham 선호주", "fisher": "강력한 성장세(매출 28%) 지속, 시장점유율 확대 중이나 마진 개선 필요"},
    "161390": {"buffett": "ROE 10%로 낮으나 P/E 8배로 저평가, 타이어 산업은 자본집약적", "lynch": "P/E 8.0으로 저평가이나 성장률 5%로 매우 낮음, Slow Grower", "fisher": "Fisher 15가지 체크리스트 미달, 투자 매력 제한적"},
    "005930": {"buffett": "ROE 9%로 낮고 P/E 34배로 고평가, 메모리 경쟁우위는 명확하나 가격 매력 부족", "lynch": "메가캡(₩994조)으로 10배 불가능, P/E 34로 비쌈, 모멘텀 주도", "fisher": "성장·수익성 지표 혼재, Fisher 기준 추가 검증 필요"},
    "101490": {"buffett": "ROE 12%로 평범하고 P/E 62배로 과도 고평가, 가격 대비 품질 불균형", "lynch": "P/E 62로 고평가, 성장률 17%로 PEG 나쁨", "fisher": "성장·수익성 지표 혼재, Fisher 기준 추가 검증 필요"},
    "079370": {"buffett": "ROE 11%로 평범, 부채비율 0.68로 다소 높음, P/E 13배 적정", "lynch": "P/E 13.5, 성장률 22%로 GARP 양호, 모멘텀 강함, 소형주 10배 가능", "fisher": "강력한 성장세(매출 22%) 지속, 시장점유율 확대 중"},
    "183300": {"buffett": "ROE 23%와 영업마진 22% 우수, P/E 20배 합리적, 반도체 부품 경쟁우위", "lynch": "고성장(매출+65%), ROE 23%, P/E 20으로 GARP 우수, 소형주 10배 가능", "fisher": "매출·이익 고성장(65%, 93%), 지속가능 성장 모델 확립"},
    "036620": {"buffett": "ROE 26%와 영업마진 16% 우수, P/E 22배 적정, 부채비율 0.37로 건전", "lynch": "ROE 26% 우수하나 성장률 24%에 P/E 22로 약간 비쌈", "fisher": "강력한 성장세(매출 24%) 지속, 시장점유율 확대 중"},
}


# === Sheet 1: 요약 ===
ws1 = wb.active
ws1.title = "요약"
ws1.merge_cells("A1:B1")
ws1.cell(row=1, column=1, value="AI Hedge Fund 포트폴리오 리포트").fill = TITLE_FILL
ws1.cell(row=1, column=1).font = TITLE_FONT
ws1.cell(row=1, column=1).alignment = Alignment(horizontal="center")

info = [
    (3, "분석 일자", "2026-02-11"),
    (4, "분석 대상", "KRX (KOSPI200+KOSDAQ150) 상위 30개"),
    (5, "분석 전략", "하이브리드 (펀더멘털 70% + 모멘텀 30%)"),
    (6, "투자자 관점", "W.Buffett, P.Lynch, P.Fisher"),
    (8, "포트폴리오 통계", ""),
    (9, "편입 종목 수", "11 / 30 (36.7%)"),
    (10, "평균 신뢰도", "78%"),
    (11, "평균 예상 수익률", "+15.0%"),
    (12, "강력매수 비중", "55.6%"),
    (13, "매수 비중", "44.4%"),
    (15, "시가총액 분포", ""),
    (16, "메가캡 (>₩200조)", "7.2%"),
    (17, "대형주 (₩1-200조)", "28.0%"),
    (18, "중형주 (₩1-2조)", "14.3%"),
    (19, "소형주 (<₩1조)", "50.4%"),
    (21, "섹터 분포", ""),
    (22, "IT", "35.0%"),
    (23, "산업재", "22.6%"),
    (24, "커뮤니케이션", "21.0%"),
    (25, "소비재", "11.7%"),
    (26, "금융", "9.8%"),
]
for row, label, val in info:
    ws1.cell(row=row, column=1, value=label).font = Font(name="맑은 고딕", bold=True if row in (8, 15, 21) else False, size=11)
    ws1.cell(row=row, column=2, value=val).font = DATA_FONT
    ws1.cell(row=row, column=1).border = THIN_BORDER
    ws1.cell(row=row, column=2).border = THIN_BORDER

# Pie chart for sector
chart = PieChart()
chart.title = "섹터 분포"
chart.style = 10
data_ref = Reference(ws1, min_col=2, min_row=22, max_row=26)
cats_ref = Reference(ws1, min_col=1, min_row=22, max_row=26)
chart.add_data(data_ref, titles_from_data=False)
chart.set_categories(cats_ref)
ws1.add_chart(chart, "D3")
auto_width(ws1, 2)

# === Sheet 2: 포트폴리오 ===
ws2 = wb.create_sheet("포트폴리오")
headers2 = ["#", "종목코드", "회사명", "비중", "신호", "신뢰도", "예상수익률", "시가총액", "P/E", "ROE", "PEG", "합의", "섹터"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, len(headers2))

for idx, p in enumerate(portfolio, 2):
    ret_str = f'+{p["ret"]}%' if p["ret"] > 0 else f'{p["ret"]}%'
    vals = [p["rank"], p["ticker"], p["name"], f'{p["weight"]}%', p["signal"], p["confidence"], ret_str, p["mcap"], p["pe"], p["roe"], p["peg"], p["consensus"], p["sector"]]
    for col, v in enumerate(vals, 1):
        ws2.cell(row=idx, column=col, value=v)
    style_data_row(ws2, idx, len(headers2))
    if "강력매수" in p["signal"]:
        for col in range(1, len(headers2) + 1):
            ws2.cell(row=idx, column=col).fill = BUY_FILL
            ws2.cell(row=idx, column=col).font = BUY_FONT
    elif "매수" in p["signal"]:
        for col in range(1, len(headers2) + 1):
            ws2.cell(row=idx, column=col).fill = PatternFill(start_color="F0F7E6", end_color="F0F7E6", fill_type="solid")

total_row = len(portfolio) + 2
ws2.cell(row=total_row, column=3, value="합계/평균").font = TOTAL_FONT
ws2.cell(row=total_row, column=3).fill = TOTAL_FILL
ws2.cell(row=total_row, column=4, value="100.0%").font = TOTAL_FONT
ws2.cell(row=total_row, column=4).fill = TOTAL_FILL
ws2.cell(row=total_row, column=6, value=78).font = TOTAL_FONT
ws2.cell(row=total_row, column=6).fill = TOTAL_FILL
ws2.cell(row=total_row, column=7, value="+15.0%").font = TOTAL_FONT
ws2.cell(row=total_row, column=7).fill = TOTAL_FILL
for col in range(1, len(headers2) + 1):
    ws2.cell(row=total_row, column=col).border = THIN_BORDER

ws2.auto_filter.ref = f"A1:M{len(portfolio)+1}"
ws2.freeze_panes = "A2"
auto_width(ws2, len(headers2))

# === Sheet 3: 순위 ===
ws3 = wb.create_sheet("순위")
headers3 = ["순위", "종목코드", "회사명", "종합점수", "펀더멘털", "모멘텀", "앙상블", "신호", "예상수익률", "시가총액", "P/E", "ROE", "매출성장률", "PEG", "편입여부"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, len(headers3))

for idx, p in enumerate(predict_all, 2):
    ret_str = f'+{p["ret"]}%' if p["ret"] > 0 else f'{p["ret"]}%'
    vals = [p["rank"], p["ticker"], p["name"], p["score"], p["fund"], p["mom"], p["ens"], p["signal"], ret_str, p["mcap"], p["pe"], p["roe"], p["rev_g"], p["peg"], p["included"]]
    for col, v in enumerate(vals, 1):
        ws3.cell(row=idx, column=col, value=v)
    style_data_row(ws3, idx, len(headers3))
    if p["included"] == "예":
        for col in range(1, len(headers3) + 1):
            ws3.cell(row=idx, column=col).fill = BUY_FILL

ws3.auto_filter.ref = f"A1:O{len(predict_all)+1}"
ws3.freeze_panes = "A2"
auto_width(ws3, len(headers3))

# === Sheet 4: 투자자 매트릭스 ===
ws4 = wb.create_sheet("투자자 매트릭스")
headers4 = ["종목코드", "회사명", "W.Buffett", "P.Lynch", "P.Fisher", "종합신호", "종합신뢰도"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, len(headers4))

signal_kr = {"bullish": "매수", "neutral": "중립", "bearish": "매도"}
for idx, m in enumerate(investor_matrix, 2):
    ws4.cell(row=idx, column=1, value=m["ticker"])
    ws4.cell(row=idx, column=2, value=m["name"])
    for col, inv in [(3, "buffett"), (4, "lynch"), (5, "fisher")]:
        sig, conf = m[inv]
        cell = ws4.cell(row=idx, column=col, value=f'{signal_kr[sig]}({conf})')
        if sig == "bullish":
            cell.fill = BUY_FILL
            cell.font = BUY_FONT
        elif sig == "bearish":
            cell.fill = SELL_FILL
            cell.font = SELL_FONT
        else:
            cell.fill = NEUTRAL_FILL
            cell.font = NEUTRAL_FONT
    ws4.cell(row=idx, column=6, value=m["combined"])
    ws4.cell(row=idx, column=7, value=m["comb_conf"])
    style_data_row(ws4, idx, len(headers4))

ws4.auto_filter.ref = f"A1:G{len(investor_matrix)+1}"
ws4.freeze_panes = "A2"
auto_width(ws4, len(headers4))

# === Sheet 5: 투자자 상세 ===
ws5 = wb.create_sheet("투자자 상세")
headers5 = ["종목코드", "회사명", "투자자", "신호", "신뢰도", "분석근거"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, len(headers5))

row = 2
for m in investor_matrix:
    t = m["ticker"]
    for inv_name, inv_key in [("W.Buffett", "buffett"), ("P.Lynch", "lynch"), ("P.Fisher", "fisher")]:
        sig, conf = m[inv_key]
        ws5.cell(row=row, column=1, value=t)
        ws5.cell(row=row, column=2, value=m["name"])
        ws5.cell(row=row, column=3, value=inv_name)
        ws5.cell(row=row, column=4, value=signal_kr[sig])
        ws5.cell(row=row, column=5, value=conf)
        ws5.cell(row=row, column=6, value=reasons.get(t, {}).get(inv_key, ""))
        style_data_row(ws5, row, len(headers5))
        row += 1

ws5.freeze_panes = "A2"
auto_width(ws5, len(headers5))
ws5.column_dimensions["F"].width = 60

# === Sheet 6: 리스크 분석 ===
ws6 = wb.create_sheet("리스크 분석")
ws6.cell(row=1, column=1, value="리스크 분석").fill = TITLE_FILL
ws6.cell(row=1, column=1).font = TITLE_FONT
ws6.merge_cells("A1:B1")

risk_data = [
    (3, "집중도 지표", ""),
    (4, "상위 1종목 비중", "12.0% (제룡전기)"),
    (5, "상위 3종목 비중", "34.3%"),
    (6, "상위 5종목 비중", "55.3%"),
    (7, "상위 10종목 비중", "93.5%"),
    (8, "HHI (허핀달 지수)", "0.096"),
    (10, "섹터 집중도", ""),
    (11, "최대 섹터", "IT (35.0%)"),
    (12, "상위 3섹터 비중", "78.6%"),
    (14, "투자자 합의 품질", ""),
    (15, "만장일치 비율 (3/3)", "54.5% (6개)"),
    (16, "다수 합의 (2/3)", "45.5% (5개)"),
    (17, "의견 분산 종목 수", "1개 (인카금융서비스)"),
    (19, "비편입 사유 분포", ""),
    (20, "투자자 과반 미달", "19개"),
    (22, "추가 리스크", ""),
    (23, "소형주 편중", "50.4% (유동성 리스크)"),
    (24, "해운업 경기민감도", "HMM 10.6%"),
]
for row, label, val in risk_data:
    ws6.cell(row=row, column=1, value=label).font = Font(name="맑은 고딕", bold=True if row in (3, 10, 14, 19, 22) else False, size=11)
    ws6.cell(row=row, column=2, value=val).font = DATA_FONT
    ws6.cell(row=row, column=1).border = THIN_BORDER
    ws6.cell(row=row, column=2).border = THIN_BORDER

auto_width(ws6, 2)

# Save
output_path = os.path.join(os.path.dirname(__file__), "krx_20260211_buffett_fisher_lynch.xlsx")
wb.save(output_path)
print(f"Excel saved: {output_path}")
