#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path
from typing import Any

import yfinance as yf
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
BULLISH_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BEARISH_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
NEUTRAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

INVESTOR_CONFIG = {
    "buffett": {
        "display": "Warren Buffett",
        "short": "W.Buffett",
        "weight": 1.0,
    },
    "lynch": {
        "display": "Peter Lynch",
        "short": "P.Lynch",
        "weight": 0.85,
    },
    "fisher": {
        "display": "Phil Fisher",
        "short": "P.Fisher",
        "weight": 0.82,
    },
}


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def fmt_pct(value: float | None, digits: int = 1, signed: bool = False) -> str:
    if value is None:
        return "N/A"
    if signed:
        return f"{value:+.{digits}f}%"
    return f"{value:.{digits}f}%"


def fmt_num(value: float | None, digits: int = 1) -> str:
    if value is None:
        return "N/A"
    return f"{value:.{digits}f}"


def ticker_label(stock: dict[str, Any], max_name: int = 18) -> str:
    name = (stock.get("company_name") or stock["ticker"]).strip()
    if name == stock["ticker"]:
        return stock["ticker"]
    if len(name) > max_name:
        name = name[: max_name - 1] + "…"
    return f"{stock['ticker']}({name})"


def signal_cell_text(signal: str, confidence: int) -> str:
    icon = {"bullish": "🟢", "neutral": "🔵", "bearish": "🔴"}[signal]
    return f"{icon} {signal}({confidence})"


def portfolio_signal_text(signal: str) -> str:
    return {"strong_buy": "🟢 강력매수", "buy": "🔵 매수"}[signal]


def load_predict_results(path: Path, top_n: int) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return payload, payload["rankings"][:top_n]


def normalize_sector(sector: str | None) -> str:
    if not sector:
        return "Unknown"
    mapping = {
        "Financial Services": "Financial",
        "Consumer Cyclical": "Consumer Disc.",
        "Consumer Defensive": "Consumer Staples",
        "Basic Materials": "Materials",
        "Communication Services": "Communication",
        "Healthcare": "Healthcare",
        "Technology": "Technology",
        "Energy": "Energy",
        "Industrials": "Industrials",
        "Real Estate": "Real Estate",
        "Utilities": "Utilities",
    }
    return mapping.get(sector, sector)


def fetch_sector(ticker: str) -> str:
    try:
        info = yf.Ticker(ticker).info
        return normalize_sector(info.get("sector"))
    except Exception:
        return "Unknown"


def fetch_sectors(stocks: list[dict[str, Any]]) -> dict[str, str]:
    sectors: dict[str, str] = {}
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(fetch_sector, stock["ticker"]): stock["ticker"] for stock in stocks}
        for future, ticker in futures.items():
            try:
                sectors[ticker] = future.result()
            except Exception:
                sectors[ticker] = "Unknown"
    return sectors


def score_to_signal(score: float) -> str:
    if score >= 7.0:
        return "bullish"
    if score <= 4.0:
        return "bearish"
    return "neutral"


def score_to_confidence(score: float, signal: str) -> int:
    if signal == "bullish":
        return int(clamp(round(48 + (score - 5) * 13), 55, 95))
    if signal == "bearish":
        return int(clamp(round(48 + (5 - score) * 13), 55, 95))
    return int(clamp(round(48 + abs(score - 5.5) * 8), 45, 70))


def choose_warning(stock: dict[str, Any], keywords: list[str]) -> str | None:
    for warning in stock.get("investor_warnings", []):
        lowered = warning.lower()
        if any(keyword in lowered for keyword in keywords):
            return warning.replace("⚠️ ", "")
    return None


def synthesize_reasoning(investor: str, stock: dict[str, Any], signal: str) -> str:
    pe = stock.get("metrics", {}).get("pe")
    roe = stock.get("metrics", {}).get("roe")
    rev_growth = stock.get("metrics", {}).get("revenue_growth")
    peg = stock.get("metrics", {}).get("peg")
    cap = stock.get("market_cap", {}).get("display", "N/A")
    cap_category = stock.get("market_cap", {}).get("category")
    momentum = stock.get("scores", {}).get("enhanced_momentum")
    pieces: list[str] = []

    if investor == "buffett":
        if roe and roe >= 20:
            pieces.append(f"ROE {roe:.1f}%")
        if pe and pe <= 20:
            pieces.append(f"P/E {pe:.1f}")
        if signal == "bullish":
            warning = choose_warning(stock, ["원자재", "과대평가"])
            if warning:
                pieces.append(warning)
            pieces.append("경제적 해자와 수익성 방어력에 점수")
        elif signal == "bearish":
            if pe and pe > 35:
                pieces.append(f"P/E {pe:.1f}로 안전마진 부족")
            if roe and roe < 10:
                pieces.append(f"ROE {roe:.1f}%로 자본효율 낮음")
            warning = choose_warning(stock, ["원자재", "메가캡", "과대평가"])
            if warning:
                pieces.append(warning)
        else:
            pieces.append("품질은 있으나 가격 메리트가 애매함")

    elif investor == "lynch":
        if rev_growth is not None:
            pieces.append(f"매출 성장률 {rev_growth:.1f}%")
        if peg is not None:
            pieces.append(f"PEG {peg:.1f}")
        if cap_category == "mega":
            pieces.append(f"메가캡({cap})이라 10배주 여지는 제한적")
        if signal == "bullish":
            pieces.append("GARP 기준에서 성장 대비 가격이 수용 가능")
        elif signal == "bearish":
            if rev_growth is not None and rev_growth < 5:
                pieces.append("성장 스토리가 약함")
            pieces.append("린치식 10배주 조건과 거리가 있음")
        else:
            pieces.append("성장성은 보이지만 확신할 만큼 싸지 않음")

    elif investor == "fisher":
        if rev_growth is not None:
            pieces.append(f"매출 성장률 {rev_growth:.1f}%")
        if roe is not None:
            pieces.append(f"ROE {roe:.1f}%")
        if momentum is not None:
            pieces.append(f"모멘텀 {momentum:.1f}")
        if signal == "bullish":
            pieces.append("성장 품질과 실행력이 장기 보유 기준에 부합")
        elif signal == "bearish":
            pieces.append("혁신·성장 품질이 Fisher 기준에 못 미침")
        else:
            pieces.append("장기 성장성은 있으나 질적 우위가 뚜렷하지 않음")

    unique_pieces = []
    for piece in pieces:
        if piece and piece not in unique_pieces:
            unique_pieces.append(piece)
    return ", ".join(unique_pieces[:3])


def synthesize_investor_analysis(stock: dict[str, Any], investors: list[str]) -> dict[str, dict[str, Any]]:
    analyses: dict[str, dict[str, Any]] = {}
    for investor in investors:
        score = stock.get("investor_scores", {}).get(investor, 5.0)
        signal = score_to_signal(score)
        confidence = score_to_confidence(score, signal)
        analyses[investor] = {
            "signal": signal,
            "confidence": confidence,
            "reasoning": synthesize_reasoning(investor, stock, signal),
            "score": score,
        }
    return analyses


def combined_confidence(analyses: dict[str, dict[str, Any]], investors: list[str]) -> int:
    total_weight = sum(INVESTOR_CONFIG[investor]["weight"] for investor in investors)
    weighted = sum(analyses[investor]["confidence"] * INVESTOR_CONFIG[investor]["weight"] for investor in investors)
    return int(round(weighted / total_weight)) if total_weight else 0


def majority_threshold(count: int) -> int:
    return math.floor(count / 2) + 1


def build_candidates(stocks: list[dict[str, Any]], investors: list[str], sectors: dict[str, str]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    candidates = []
    excluded = []
    need = majority_threshold(len(investors))
    for stock in stocks:
        analyses = synthesize_investor_analysis(stock, investors)
        bullish = sum(1 for investor in investors if analyses[investor]["signal"] == "bullish")
        bearish = sum(1 for investor in investors if analyses[investor]["signal"] == "bearish")
        confidence = combined_confidence(analyses, investors)
        consensus_ratio = bullish / len(investors)
        portfolio_signal = "strong_buy" if bullish == len(investors) else "buy"
        enriched = {
            **stock,
            "sector": sectors.get(stock["ticker"], "Unknown"),
            "investor_analysis": analyses,
            "bullish_count": bullish,
            "bearish_count": bearish,
            "consensus_ratio": consensus_ratio,
            "combined_confidence": confidence,
            "combined_signal": portfolio_signal,
            "raw_weight_score": confidence * max(consensus_ratio, 0.01),
        }
        if bullish >= need:
            candidates.append(enriched)
        else:
            excluded.append({**enriched, "exclusion_reason": f"투자자 과반 미달 ({bullish}/{len(investors)} bullish)"})
    return candidates, excluded


def cap_and_redistribute(weights: dict[str, float], cap: float) -> dict[str, float]:
    locked: dict[str, float] = {}
    active = dict(weights)
    while active:
        total_active = sum(active.values())
        if total_active <= 0:
            break
        normalized = {ticker: value / total_active * (100 - sum(locked.values())) for ticker, value in active.items()}
        over = {ticker: value for ticker, value in normalized.items() if value > cap + 1e-9}
        if not over:
            locked.update(normalized)
            break
        for ticker, value in over.items():
            locked[ticker] = cap
            active.pop(ticker, None)
    return locked


def apply_sector_cap(positions: list[dict[str, Any]], sector_cap: float = 35.0) -> list[dict[str, Any]]:
    remaining = {position["ticker"]: position["weight"] for position in positions}
    meta = {position["ticker"]: position for position in positions}
    changed = True
    while changed:
        changed = False
        sector_weights: dict[str, float] = defaultdict(float)
        for ticker, weight in remaining.items():
            sector_weights[meta[ticker]["sector"]] += weight
        excess_sectors = {sector: weight for sector, weight in sector_weights.items() if weight > sector_cap + 1e-9}
        if not excess_sectors:
            break
        for sector, weight in excess_sectors.items():
            members = [ticker for ticker in remaining if meta[ticker]["sector"] == sector]
            scale = sector_cap / weight
            freed = 0.0
            for ticker in members:
                old = remaining[ticker]
                new = old * scale
                remaining[ticker] = new
                freed += old - new
            other_members = [ticker for ticker in remaining if meta[ticker]["sector"] != sector and remaining[ticker] < 15 - 1e-9]
            if other_members and freed > 0:
                total_other = sum(remaining[ticker] for ticker in other_members)
                for ticker in other_members:
                    headroom = 15 - remaining[ticker]
                    bump = freed * (remaining[ticker] / total_other) if total_other else 0
                    remaining[ticker] += min(headroom, bump)
            changed = True
    total = sum(remaining.values())
    if total > 0:
        remaining = {ticker: value / total * 100 for ticker, value in remaining.items()}
    for position in positions:
        position["weight"] = remaining[position["ticker"]]
    return positions


def allocate_weights(candidates: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if not candidates:
        return [], []

    raw = {candidate["ticker"]: candidate["raw_weight_score"] for candidate in candidates}
    capped = cap_and_redistribute(raw, 15.0)
    for candidate in candidates:
        candidate["weight"] = capped.get(candidate["ticker"], 0.0)

    candidates = apply_sector_cap(candidates, 35.0)

    included: list[dict[str, Any]] = []
    excluded: list[dict[str, Any]] = []
    for candidate in candidates:
        if candidate["weight"] < 2.0:
            candidate["exclusion_reason"] = "최소 비중 2% 미달"
            excluded.append(candidate)
        else:
            included.append(candidate)

    if not included:
        return [], excluded

    total = sum(item["weight"] for item in included)
    for item in included:
        item["weight"] = item["weight"] / total * 100

    re_capped = True
    while re_capped:
        re_capped = False
        over = [item for item in included if item["weight"] > 15.0 + 1e-9]
        if over:
            excess = 0.0
            for item in over:
                excess += item["weight"] - 15.0
                item["weight"] = 15.0
            under = [item for item in included if item["weight"] < 15.0 - 1e-9]
            under_total = sum(item["weight"] for item in under)
            for item in under:
                headroom = 15.0 - item["weight"]
                bump = excess * (item["weight"] / under_total) if under_total else 0.0
                item["weight"] += min(headroom, bump)
            re_capped = True

    total = sum(item["weight"] for item in included)
    for item in included:
        item["weight"] = round(item["weight"] / total * 100, 1)

    delta = round(100.0 - sum(item["weight"] for item in included), 1)
    if included and abs(delta) >= 0.1:
        included[0]["weight"] = round(included[0]["weight"] + delta, 1)

    included.sort(key=lambda item: (-item["weight"], item["rank"]))
    return included, excluded


def summarize_portfolio(included: list[dict[str, Any]], top_stocks: list[dict[str, Any]], investors: list[str]) -> dict[str, Any]:
    total_names = len(top_stocks)
    included_tickers = {item["ticker"] for item in included}
    weighted_conf = sum(item["weight"] * item["combined_confidence"] for item in included) / 100 if included else 0.0
    weighted_return = sum(item["weight"] * item["predicted_return_1y"] for item in included) / 100 if included else 0.0
    strong_buy_weight = sum(item["weight"] for item in included if item["combined_signal"] == "strong_buy")
    buy_weight = sum(item["weight"] for item in included if item["combined_signal"] == "buy")

    agreement = Counter(item["bullish_count"] for item in included)
    cap_weights: dict[str, float] = defaultdict(float)
    sector_weights: dict[str, float] = defaultdict(float)
    for item in included:
        cap_weights[item["market_cap"]["category"]] += item["weight"]
        sector_weights[item["sector"]] += item["weight"]

    return {
        "included_count": len(included),
        "analyzed_count": total_names,
        "included_rate": (len(included) / total_names * 100) if total_names else 0.0,
        "avg_confidence": weighted_conf,
        "avg_return": weighted_return,
        "strong_buy_weight": strong_buy_weight,
        "buy_weight": buy_weight,
        "agreement": agreement,
        "cap_weights": cap_weights,
        "sector_weights": dict(sorted(sector_weights.items(), key=lambda item: item[1], reverse=True)),
        "top5_weight": sum(item["weight"] for item in included[:5]),
        "max_name": included[0]["ticker"] if included else "N/A",
        "max_weight": included[0]["weight"] if included else 0.0,
        "max_sector": max(sector_weights.items(), key=lambda item: item[1]) if sector_weights else ("N/A", 0.0),
        "included_tickers": included_tickers,
        "need_consensus": majority_threshold(len(investors)),
    }


def auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 30)


def apply_header(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def apply_grid(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER


def create_summary_sheet(ws, analysis_date: str, index_label: str, investor_labels: str, summary: dict[str, Any]) -> None:
    ws.merge_cells("A1:B1")
    ws["A1"] = "AI Hedge Fund 포트폴리오 리포트"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    info = [
        ("분석 일자", analysis_date),
        ("분석 대상", index_label),
        ("분석 전략", "하이브리드 (펀더멘털 70% + 모멘텀 30%)"),
        ("투자자 관점", investor_labels),
    ]
    for row, (label, value) in enumerate(info, start=3):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value

    ws["A8"] = "포트폴리오 통계"
    ws["A8"].font = Font(bold=True)
    stats = [
        ("편입 종목 수", f"{summary['included_count']} / {summary['analyzed_count']} ({summary['included_rate']:.1f}%)"),
        ("평균 신뢰도", f"{summary['avg_confidence']:.0f}%"),
        ("평균 예상 수익률", fmt_pct(summary["avg_return"], signed=True)),
        ("강력매수 비중", fmt_pct(summary["strong_buy_weight"])),
    ]
    for row, (label, value) in enumerate(stats, start=9):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value

    ws["A14"] = "시가총액 분포"
    ws["A14"].font = Font(bold=True)
    cap_labels = {
        "mega": "메가캡 (>$200B)",
        "large": "대형주 ($10B-$200B)",
        "mid": "중형주 ($2B-$10B)",
        "small": "소형주 (<$2B)",
        None: "Unknown",
    }
    chart_start = 15
    for idx, (category, label) in enumerate(cap_labels.items(), start=chart_start):
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = round(summary["cap_weights"].get(category, 0.0) / 100, 4)
        ws[f"B{idx}"].number_format = "0.0%"

    ws["A20"] = "섹터 분포"
    ws["A20"].font = Font(bold=True)
    sector_start = 21
    for idx, (sector, weight) in enumerate(summary["sector_weights"].items(), start=sector_start):
        ws[f"A{idx}"] = sector
        ws[f"B{idx}"] = round(weight / 100, 4)
        ws[f"B{idx}"].number_format = "0.0%"

    pie = PieChart()
    pie.title = "섹터 분포"
    data = Reference(ws, min_col=2, min_row=sector_start, max_row=sector_start + max(len(summary["sector_weights"]) - 1, 0))
    labels = Reference(ws, min_col=1, min_row=sector_start, max_row=sector_start + max(len(summary["sector_weights"]) - 1, 0))
    if summary["sector_weights"]:
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(labels)
        pie.height = 7
        pie.width = 10
        ws.add_chart(pie, "D14")

    apply_grid(ws)
    auto_width(ws)


def create_portfolio_sheet(ws, included: list[dict[str, Any]], investors: list[str]) -> None:
    headers = ["#", "종목코드", "회사명", "비중", "신호", "신뢰도", "예상수익률", "시가총액", "P/E", "ROE", "PEG", "합의", "섹터"]
    apply_header(ws, headers)
    for row, item in enumerate(included, start=2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=item["ticker"])
        ws.cell(row=row, column=3, value=item.get("company_name"))
        ws.cell(row=row, column=4, value=item["weight"] / 100).number_format = "0.0%"
        ws.cell(row=row, column=5, value=portfolio_signal_text(item["combined_signal"]))
        ws.cell(row=row, column=6, value=item["combined_confidence"])
        ws.cell(row=row, column=7, value=item["predicted_return_1y"] / 100).number_format = "+0.0%;-0.0%"
        ws.cell(row=row, column=8, value=item["market_cap"]["display"])
        ws.cell(row=row, column=9, value=item["metrics"].get("pe"))
        roe = item["metrics"].get("roe")
        if roe is not None:
            ws.cell(row=row, column=10, value=roe / 100).number_format = "0.0%"
        else:
            ws.cell(row=row, column=10, value="N/A")
        ws.cell(row=row, column=11, value=item["metrics"].get("peg"))
        ws.cell(row=row, column=12, value=f"{item['bullish_count']}/{len(investors)}")
        ws.cell(row=row, column=13, value=item["sector"])

    total_row = len(included) + 2
    for col in range(1, 14):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
    ws.cell(row=total_row, column=2, value="합계/평균").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=1.0).number_format = "0.0%"
    avg_conf = sum(item["weight"] * item["combined_confidence"] for item in included) / 100 if included else 0
    avg_ret = sum(item["weight"] * item["predicted_return_1y"] for item in included) / 100 if included else 0
    ws.cell(row=total_row, column=6, value=round(avg_conf))
    ws.cell(row=total_row, column=7, value=avg_ret / 100).number_format = "+0.0%;-0.0%"

    apply_grid(ws)
    auto_width(ws)
    ws.auto_filter.ref = f"A1:M{len(included) + 1}"
    ws.freeze_panes = "A2"


def create_ranking_sheet(ws, top_stocks: list[dict[str, Any]], included_tickers: set[str]) -> None:
    headers = ["순위", "종목코드", "회사명", "종합점수", "펀더멘털", "모멘텀", "앙상블", "신호", "예상수익률", "시가총액", "P/E", "P/B", "ROE", "매출성장률", "PEG", "편입여부"]
    apply_header(ws, headers)
    for row, item in enumerate(top_stocks, start=2):
        metrics = item.get("metrics", {})
        ws.cell(row=row, column=1, value=item["rank"])
        ws.cell(row=row, column=2, value=item["ticker"])
        ws.cell(row=row, column=3, value=item.get("company_name"))
        ws.cell(row=row, column=4, value=item["total_score"])
        ws.cell(row=row, column=5, value=item["scores"].get("fundamental"))
        ws.cell(row=row, column=6, value=item["scores"].get("enhanced_momentum"))
        ws.cell(row=row, column=7, value=item.get("ensemble_score"))
        ws.cell(row=row, column=8, value=item["signal"])
        ws.cell(row=row, column=9, value=item["predicted_return_1y"] / 100).number_format = "+0.0%;-0.0%"
        ws.cell(row=row, column=10, value=item["market_cap"]["display"])
        ws.cell(row=row, column=11, value=metrics.get("pe"))
        ws.cell(row=row, column=12, value=metrics.get("pb"))
        roe = metrics.get("roe")
        if roe is not None:
            ws.cell(row=row, column=13, value=roe / 100).number_format = "0.0%"
        else:
            ws.cell(row=row, column=13, value="N/A")
        rev = metrics.get("revenue_growth")
        if rev is not None:
            ws.cell(row=row, column=14, value=rev / 100).number_format = "0.0%"
        else:
            ws.cell(row=row, column=14, value="N/A")
        ws.cell(row=row, column=15, value=metrics.get("peg"))
        included = item["ticker"] in included_tickers
        included_cell = ws.cell(row=row, column=16, value="예" if included else "아니오")
        if included:
            for col in range(1, 17):
                ws.cell(row=row, column=col).fill = BULLISH_FILL
        included_cell.border = THIN_BORDER

    apply_grid(ws)
    auto_width(ws)
    ws.auto_filter.ref = f"A1:P{len(top_stocks) + 1}"
    ws.freeze_panes = "A2"


def create_matrix_sheet(ws, top_stocks: list[dict[str, Any]], investors: list[str]) -> None:
    headers = ["종목코드", "회사명"] + [INVESTOR_CONFIG[investor]["display"] for investor in investors] + ["종합신호", "종합신뢰도"]
    apply_header(ws, headers)
    for row, item in enumerate(top_stocks, start=2):
        ws.cell(row=row, column=1, value=item["ticker"])
        ws.cell(row=row, column=2, value=item.get("company_name"))
        for idx, investor in enumerate(investors, start=3):
            analysis = item["investor_analysis"][investor]
            cell = ws.cell(row=row, column=idx, value={"bullish": "매수", "neutral": "중립", "bearish": "매도"}[analysis["signal"]] + f"({analysis['confidence']})")
            cell.fill = {"bullish": BULLISH_FILL, "neutral": NEUTRAL_FILL, "bearish": BEARISH_FILL}[analysis["signal"]]
        ws.cell(row=row, column=3 + len(investors), value=item["combined_signal"] if item.get("combined_signal") else "-")
        ws.cell(row=row, column=4 + len(investors), value=item.get("combined_confidence", "-"))

    apply_grid(ws)
    auto_width(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(top_stocks) + 1}"
    ws.freeze_panes = "A2"


def create_detail_sheet(ws, top_stocks: list[dict[str, Any]], investors: list[str]) -> None:
    headers = ["종목코드", "회사명", "투자자", "신호", "신뢰도", "분석근거"]
    apply_header(ws, headers)
    row = 2
    for item in top_stocks:
        start_row = row
        for investor in investors:
            analysis = item["investor_analysis"][investor]
            ws.cell(row=row, column=1, value=item["ticker"])
            ws.cell(row=row, column=2, value=item.get("company_name"))
            ws.cell(row=row, column=3, value=INVESTOR_CONFIG[investor]["display"])
            ws.cell(row=row, column=4, value=analysis["signal"])
            ws.cell(row=row, column=5, value=analysis["confidence"])
            ws.cell(row=row, column=6, value=analysis["reasoning"])
            row += 1
        for col in range(1, 7):
            ws.cell(row=start_row, column=col).border = THIN_BORDER
    apply_grid(ws)
    auto_width(ws)
    ws.freeze_panes = "A2"


def create_risk_sheet(ws, summary: dict[str, Any], excluded_all: list[dict[str, Any]], investors: list[str]) -> None:
    ws["A1"] = "리스크 분석"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:B1")

    top_weights = sorted([item["weight"] for item in excluded_all if item.get("weight")], reverse=True)
    rows = [
        ("상위 1종목 비중", fmt_pct(summary["max_weight"])),
        ("상위 3종목 비중", fmt_pct(sum(item for item in top_weights[:3]))),
        ("상위 5종목 비중", fmt_pct(summary["top5_weight"])),
        ("최대 섹터", f"{summary['max_sector'][0]} ({summary['max_sector'][1]:.1f}%)"),
        ("만장일치 비율", fmt_pct(summary["agreement"].get(len(investors), 0) / max(summary["included_count"], 1) * 100 if summary["included_count"] else 0)),
        ("비편입 종목 수", f"{sum(1 for item in excluded_all if item.get('exclusion_reason'))}개"),
    ]
    for row, (label, value) in enumerate(rows, start=4):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
    apply_grid(ws)
    auto_width(ws)


def write_workbook(
    output_path: Path,
    analysis_date: str,
    index_label: str,
    investors: list[str],
    top_stocks: list[dict[str, Any]],
    included: list[dict[str, Any]],
    summary: dict[str, Any],
    excluded_all: list[dict[str, Any]],
) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "요약"
    create_summary_sheet(ws_summary, analysis_date, index_label, ", ".join(INVESTOR_CONFIG[investor]["short"] for investor in investors), summary)

    ws_portfolio = wb.create_sheet("포트폴리오")
    create_portfolio_sheet(ws_portfolio, included, investors)

    ws_ranking = wb.create_sheet("순위")
    create_ranking_sheet(ws_ranking, top_stocks, summary["included_tickers"])

    ws_matrix = wb.create_sheet("투자자 매트릭스")
    create_matrix_sheet(ws_matrix, top_stocks, investors)

    ws_detail = wb.create_sheet("투자자 상세")
    create_detail_sheet(ws_detail, top_stocks, investors)

    ws_risk = wb.create_sheet("리스크 분석")
    create_risk_sheet(ws_risk, summary, excluded_all, investors)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def print_report(
    analysis_date: str,
    index_name: str,
    investors: list[str],
    top_stocks: list[dict[str, Any]],
    included: list[dict[str, Any]],
    excluded_all: list[dict[str, Any]],
    summary: dict[str, Any],
    output_path: Path | None,
) -> None:
    investor_names = ", ".join(INVESTOR_CONFIG[investor]["display"] for investor in investors)
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("📋 AI Hedge Fund 포트폴리오 리포트")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print(f"분석 일자    : {analysis_date}")
    print(f"분석 대상    : {index_name} 상위 {len(top_stocks)}개 종목")
    print("분석 전략    : 하이브리드 (펀더멘털 70% + 모멘텀 30%)")
    print(f"투자자 관점  : {investor_names}")
    print("데이터 소스  : Yahoo Finance (predict) + persona synthesis")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print()
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("📊 포트폴리오 구성")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print(" #  종목                  비중    시총      종합신호     신뢰도  예상수익률  P/E    ROE      합의")
    print("──────────────────────────────────────────────────────────────────────────────────────────────")
    for idx, item in enumerate(included, start=1):
        print(
            f"{idx:<2}  {ticker_label(item, 16):<20} {item['weight']:>5.1f}%  {item['market_cap']['display']:<8} "
            f"{portfolio_signal_text(item['combined_signal']):<11} {item['combined_confidence']:>4}%  "
            f"{fmt_pct(item['predicted_return_1y'], signed=True):>9}  {fmt_num(item['metrics'].get('pe')):>5}  "
            f"{fmt_pct(item['metrics'].get('roe')):>7}  {item['bullish_count']}/{len(investors)}"
        )
    print("──────────────────────────────────────────────────────────────────────────────────────────────")
    print(f"    합계                 100.0%                           avg {summary['avg_confidence']:.0f}%  avg {fmt_pct(summary['avg_return'], signed=True)}")
    print()
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("👥 투자자별 종목 분석 매트릭스")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    headers = "  ".join(f"{INVESTOR_CONFIG[investor]['short']:<18}" for investor in investors)
    print(f"종목                  {headers}")
    print("──────────────────────────────────────────────────────────────────────────────────────────────")
    for item in top_stocks:
        cells = "  ".join(f"{signal_cell_text(item['investor_analysis'][investor]['signal'], item['investor_analysis'][investor]['confidence']):<18}" for investor in investors)
        print(f"{ticker_label(item, 16):<20} {cells}")
    print()
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("💬 투자자별 핵심 분석 근거 (상위 5개 종목)")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    for item in included[:5]:
        print(f"▸ {item['ticker']} — 종합: {portfolio_signal_text(item['combined_signal'])} (신뢰도 {item['combined_confidence']}%)")
        for investor in investors:
            analysis = item["investor_analysis"][investor]
            icon = {"bullish": "🟢", "neutral": "🔵", "bearish": "🔴"}[analysis["signal"]]
            print(f"  {INVESTOR_CONFIG[investor]['display']:<16}: {icon} ({analysis['confidence']}%) {analysis['reasoning']}")
        print()
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("📈 포트폴리오 요약")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print(f"편입 종목 수         : {summary['included_count']}개 / 분석 {summary['analyzed_count']}개 ({summary['included_rate']:.1f}% 편입률)")
    print(f"평균 신뢰도          : {summary['avg_confidence']:.0f}%")
    print(f"평균 예상 수익률     : {fmt_pct(summary['avg_return'], signed=True)}")
    print(f"강력매수 비중        : {fmt_pct(summary['strong_buy_weight'])}")
    print(f"매수 비중            : {fmt_pct(summary['buy_weight'])}")
    unanimous = summary["agreement"].get(len(investors), 0)
    majority = summary["agreement"].get(summary["need_consensus"], 0) - unanimous if len(investors) != summary["need_consensus"] else 0
    print()
    print("투자자 합의 분포:")
    print(f"  만장일치 ({len(investors)}/{len(investors)})     : {unanimous}개")
    print(f"  다수 합의 ({summary['need_consensus']}/{len(investors)})    : {summary['included_count'] - unanimous}개")
    print()
    print("시가총액 분포:")
    cap_labels = {"mega": "메가캡", "large": "대형주", "mid": "중형주", "small": "소형주", None: "Unknown"}
    for key in ["mega", "large", "mid", "small", None]:
        weight = summary["cap_weights"].get(key, 0.0)
        if weight > 0:
            print(f"  {cap_labels[key]:<18}: {weight:>4.1f}%")
    print()
    print("섹터 분포:")
    for sector, weight in list(summary["sector_weights"].items())[:6]:
        print(f"  {sector:<18}: {weight:>4.1f}%")
    print()
    non_included = [item for item in excluded_all if item.get("exclusion_reason")]
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print(f"🚫 비편입 종목 ({len(non_included)}개)")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("종목                  순위   점수    사유")
    print("──────────────────────────────────────────────────────────────────────────────────────────────")
    for item in non_included[:12]:
        print(f"{ticker_label(item, 16):<20} {item['rank']:>3}   {item['total_score']:>5.2f}   {item['exclusion_reason']}")
    print()
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("⚠️ 리스크 및 경고")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("포트폴리오 집중도:")
    print(f"  상위 5종목 비중     : {summary['top5_weight']:.1f}%")
    print(f"  최대 단일 종목      : {summary['max_name']} {summary['max_weight']:.1f}% (제한 15% 내)")
    print(f"  최대 섹터 비중      : {summary['max_sector'][0]} {summary['max_sector'][1]:.1f}% (제한 35% 내)")
    print()
    warnings = []
    for item in top_stocks:
        for warning in item.get("investor_warnings", []):
            warnings.append((item["ticker"], warning.replace("⚠️ ", "")))
    if warnings:
        print("투자 철학 불일치 경고:")
        for ticker, warning in warnings[:6]:
            print(f"  {ticker:<6}: {warning}")
    print()
    low_consensus = [item for item in top_stocks if item.get("investor_consensus", {}).get("level") == "low"]
    if low_consensus:
        print("투자자 의견 분산 종목:")
        for item in low_consensus[:5]:
            print(f"  {item['ticker']:<6}: predict 합의도 {item['investor_consensus']['level']} (std={item['investor_consensus']['std']})")
    print()
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("💡 이 리포트는 교육/연구 목적이며 실제 투자 결정의 근거가 될 수 없습니다.")
    print("   predict: Yahoo Finance | investor-analysis: predict 기반 페르소나 합성")
    if output_path:
        print(f"   엑셀 리포트: {output_path}")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--predict-json", required=True)
    parser.add_argument("--index", default="sp500")
    parser.add_argument("--top", type=int, default=30)
    parser.add_argument("--investors", default="buffett,lynch,fisher")
    parser.add_argument("--xlsx", default="yes")
    parser.add_argument("--output-dir", default="portfolios")
    args = parser.parse_args()

    investors = [token.strip() for token in args.investors.split(",") if token.strip()]
    payload, top_stocks = load_predict_results(Path(args.predict_json), args.top)
    analysis_date = payload.get("analysis_date") or datetime.now().strftime("%Y-%m-%d")
    sectors = fetch_sectors(top_stocks)

    candidates, excluded_majority = build_candidates(top_stocks, investors, sectors)
    included, excluded_weight = allocate_weights(candidates)
    included_map = {item["ticker"]: item for item in included}

    enriched_top = []
    for stock in top_stocks:
        base = next(candidate for candidate in candidates + excluded_majority if candidate["ticker"] == stock["ticker"])
        if stock["ticker"] in included_map:
            base = included_map[stock["ticker"]]
        enriched_top.append(base)

    summary = summarize_portfolio(included, top_stocks, investors)
    output_path = None
    if args.xlsx.lower() in {"yes", "true", "1", "excel", "xlsx"}:
        investor_suffix = "_".join(sorted(investors))
        output_path = Path(args.output_dir) / f"{args.index}_{analysis_date.replace('-', '')}_{investor_suffix}.xlsx"
        write_workbook(
            output_path=output_path,
            analysis_date=analysis_date,
            index_label=f"{args.index.upper()} 상위 {args.top}개" if args.index != "sp500" else f"S&P 500 상위 {args.top}개",
            investors=investors,
            top_stocks=enriched_top,
            included=included,
            summary=summary,
            excluded_all=excluded_majority + excluded_weight,
        )

    print_report(
        analysis_date=analysis_date,
        index_name="S&P 500" if args.index == "sp500" else args.index.upper(),
        investors=investors,
        top_stocks=enriched_top,
        included=included,
        excluded_all=excluded_majority + excluded_weight,
        summary=summary,
        output_path=output_path,
    )


if __name__ == "__main__":
    main()
