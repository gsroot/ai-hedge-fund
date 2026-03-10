#!/usr/bin/env python3
"""S&P 500 투자자 분석 포트폴리오 리포트 생성"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()

# 색상 정의
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
BULLISH_FILL = PatternFill('solid', fgColor='C6EFCE')
BEARISH_FILL = PatternFill('solid', fgColor='FFC7CE')
NEUTRAL_FILL = PatternFill('solid', fgColor='FFEB9C')
LIGHT_BLUE = PatternFill('solid', fgColor='D6EAF8')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# 분석 데이터 (이전 세션에서 수집한 결과)
analysis_data = [
    # 종목, 순위점수, 버핏, 린치, 피셔, 버핏신뢰도, 린치신뢰도, 피셔신뢰도, 섹터
    ("LRCX", 7.83, "neutral", "bullish", "neutral", 65, 78, 62, "Technology"),
    ("CF", 7.82, "bullish", "bullish", "neutral", 75, 82, 55, "Materials"),
    ("MU", 7.64, "neutral", "bullish", "bullish", 58, 85, 72, "Technology"),
    ("NEM", 7.41, "bullish", "bullish", "neutral", 72, 76, 48, "Materials"),
    ("MOS", 7.41, "bullish", "bullish", "neutral", 70, 78, 52, "Materials"),
    ("STX", 7.31, "neutral", "bullish", "bullish", 55, 80, 68, "Technology"),
    ("ADI", 7.20, "neutral", "bullish", "bullish", 62, 75, 78, "Technology"),
    ("AKAM", 6.95, "bullish", "bullish", "neutral", 68, 72, 58, "Technology"),
    ("INCY", 6.86, "neutral", "bullish", "bullish", 52, 78, 70, "Healthcare"),
    ("CTRA", 6.82, "neutral", "bullish", "neutral", 48, 75, 55, "Energy"),
    ("MRK", 6.78, "bullish", "neutral", "bullish", 78, 65, 72, "Healthcare"),
    ("APH", 6.72, "bullish", "bullish", "bullish", 82, 78, 85, "Technology"),
    ("ALL", 6.68, "bullish", "bullish", "neutral", 85, 75, 62, "Financials"),
    ("GOOGL", 6.65, "neutral", "bullish", "bullish", 68, 82, 78, "Technology"),
    ("ENPH", 6.62, "bearish", "neutral", "neutral", 85, 58, 35, "Technology"),
    ("DXCM", 6.58, "bearish", "neutral", "neutral", 72, 72, 62, "Healthcare"),
    ("HIG", 6.55, "bullish", "bullish", "neutral", 80, 72, 58, "Financials"),
    ("PAYC", 6.52, "neutral", "neutral", "bullish", 68, 72, 88, "Technology"),
    ("HII", 6.48, "bearish", "neutral", "bearish", 82, 62, 72, "Industrials"),
    ("KLAC", 6.45, "neutral", "bullish", "bullish", 62, 78, 75, "Technology"),
    ("BMY", 6.42, "bullish", "neutral", "neutral", 75, 58, 55, "Healthcare"),
    ("TRV", 6.38, "bullish", "bullish", "neutral", 82, 70, 58, "Financials"),
    ("ACGL", 6.35, "bullish", "bullish", "bullish", 92, 85, 78, "Financials"),
    ("JNJ", 6.32, "bullish", "neutral", "neutral", 78, 62, 58, "Healthcare"),
    ("AFL", 6.28, "neutral", "bullish", "bearish", 55, 72, 72, "Financials"),
    ("CFG", 6.25, "bearish", "neutral", "bearish", 78, 55, 75, "Financials"),
    ("OKE", 6.22, "neutral", "neutral", "bearish", 45, 62, 65, "Energy"),
    ("AMAT", 6.18, "neutral", "bullish", "bullish", 65, 80, 75, "Technology"),
    ("IVZ", 6.15, "bearish", "bearish", "bearish", 78, 47, 82, "Financials"),
    ("DVN", 6.12, "bearish", "bearish", "bearish", 75, 75, 75, "Energy"),
]

# ===== Sheet 1: Summary =====
ws1 = wb.active
ws1.title = "Summary"

ws1['A1'] = "AI Hedge Fund - Portfolio Report"
ws1['A1'].font = Font(bold=True, size=18, color='1F4E79')
ws1.merge_cells('A1:E1')

ws1['A3'] = "분석 개요"
ws1['A3'].font = Font(bold=True, size=14)
ws1['A4'] = "분석 날짜:"
ws1['B4'] = datetime.now().strftime("%Y-%m-%d")
ws1['A5'] = "분석 대상:"
ws1['B5'] = "S&P 500 상위 30개 종목"
ws1['A6'] = "분석 전략:"
ws1['B6'] = "Hybrid (Fundamental 70% + Momentum 30%)"
ws1['A7'] = "투자자 페르소나:"
ws1['B7'] = "Warren Buffett, Peter Lynch, Phil Fisher"

ws1['A9'] = "포트폴리오 요약"
ws1['A9'].font = Font(bold=True, size=14)
ws1['A10'] = "총 분석 종목:"
ws1['B10'] = 30
ws1['A11'] = "포트폴리오 편입 종목:"
ws1['B11'] = "=COUNTIF(Portfolio!E:E,\">0\")"
ws1['A12'] = "평균 Confidence:"
ws1['B12'] = "=AVERAGE(Portfolio!D:D)"
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 40

# ===== Sheet 2: Portfolio =====
ws2 = wb.create_sheet("Portfolio")
headers2 = ["Rank", "Ticker", "Signal", "Avg Confidence", "Weight (%)", "Sector", "Buffett", "Lynch", "Fisher"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# 포트폴리오 편입 조건: 2/3 이상 bullish
portfolio_stocks = []
for i, d in enumerate(analysis_data):
    ticker, score, buf, lyn, fish, buf_c, lyn_c, fish_c, sector = d
    bullish_count = sum(1 for s in [buf, lyn, fish] if s == "bullish")
    if bullish_count >= 2:
        avg_conf = (buf_c + lyn_c + fish_c) / 3
        portfolio_stocks.append((i+1, ticker, "BUY", avg_conf, sector, buf, lyn, fish))

# 비중 계산 (confidence 가중)
total_conf = sum(s[3] for s in portfolio_stocks)
for row, stock in enumerate(portfolio_stocks, 2):
    rank, ticker, signal, avg_conf, sector, buf, lyn, fish = stock
    weight = min(15, (avg_conf / total_conf) * 100 * len(portfolio_stocks) / 10)  # 최대 15%

    ws2.cell(row=row, column=1, value=rank).border = thin_border
    ws2.cell(row=row, column=2, value=ticker).border = thin_border
    cell = ws2.cell(row=row, column=3, value=signal)
    cell.fill = BULLISH_FILL
    cell.border = thin_border
    ws2.cell(row=row, column=4, value=round(avg_conf, 1)).border = thin_border
    ws2.cell(row=row, column=5, value=round(weight, 2)).border = thin_border
    ws2.cell(row=row, column=6, value=sector).border = thin_border

    for col, sig in enumerate([buf, lyn, fish], 7):
        cell = ws2.cell(row=row, column=col, value=sig)
        cell.fill = BULLISH_FILL if sig == "bullish" else (BEARISH_FILL if sig == "bearish" else NEUTRAL_FILL)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

for col in range(1, 10):
    ws2.column_dimensions[get_column_letter(col)].width = 14

# ===== Sheet 3: Ranking =====
ws3 = wb.create_sheet("Ranking")
headers3 = ["Rank", "Ticker", "Score", "Signal", "Sector", "Buffett", "Lynch", "Fisher", "Consensus"]
for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for row, d in enumerate(analysis_data, 2):
    ticker, score, buf, lyn, fish, buf_c, lyn_c, fish_c, sector = d
    bullish_count = sum(1 for s in [buf, lyn, fish] if s == "bullish")
    bearish_count = sum(1 for s in [buf, lyn, fish] if s == "bearish")

    if bullish_count >= 2:
        consensus = "BUY"
        signal_fill = BULLISH_FILL
    elif bearish_count >= 2:
        consensus = "SELL"
        signal_fill = BEARISH_FILL
    else:
        consensus = "HOLD"
        signal_fill = NEUTRAL_FILL

    ws3.cell(row=row, column=1, value=row-1).border = thin_border
    ws3.cell(row=row, column=2, value=ticker).border = thin_border
    ws3.cell(row=row, column=3, value=score).border = thin_border
    cell = ws3.cell(row=row, column=4, value=consensus)
    cell.fill = signal_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')
    ws3.cell(row=row, column=5, value=sector).border = thin_border

    for col, sig in enumerate([buf, lyn, fish], 6):
        cell = ws3.cell(row=row, column=col, value=sig)
        cell.fill = BULLISH_FILL if sig == "bullish" else (BEARISH_FILL if sig == "bearish" else NEUTRAL_FILL)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    cell = ws3.cell(row=row, column=9, value=consensus)
    cell.fill = signal_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

for col in range(1, 10):
    ws3.column_dimensions[get_column_letter(col)].width = 12

# ===== Sheet 4: Investor Matrix =====
ws4 = wb.create_sheet("Investor Matrix")
headers4 = ["Ticker", "Buffett Signal", "Buffett Conf", "Lynch Signal", "Lynch Conf", "Fisher Signal", "Fisher Conf", "Avg Conf"]
for col, h in enumerate(headers4, 1):
    cell = ws4.cell(row=1, column=col, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for row, d in enumerate(analysis_data, 2):
    ticker, score, buf, lyn, fish, buf_c, lyn_c, fish_c, sector = d
    ws4.cell(row=row, column=1, value=ticker).border = thin_border

    cell = ws4.cell(row=row, column=2, value=buf)
    cell.fill = BULLISH_FILL if buf == "bullish" else (BEARISH_FILL if buf == "bearish" else NEUTRAL_FILL)
    cell.border = thin_border
    ws4.cell(row=row, column=3, value=buf_c).border = thin_border

    cell = ws4.cell(row=row, column=4, value=lyn)
    cell.fill = BULLISH_FILL if lyn == "bullish" else (BEARISH_FILL if lyn == "bearish" else NEUTRAL_FILL)
    cell.border = thin_border
    ws4.cell(row=row, column=5, value=lyn_c).border = thin_border

    cell = ws4.cell(row=row, column=6, value=fish)
    cell.fill = BULLISH_FILL if fish == "bullish" else (BEARISH_FILL if fish == "bearish" else NEUTRAL_FILL)
    cell.border = thin_border
    ws4.cell(row=row, column=7, value=fish_c).border = thin_border

    ws4.cell(row=row, column=8, value=f"=AVERAGE(C{row},E{row},G{row})").border = thin_border

for col in range(1, 9):
    ws4.column_dimensions[get_column_letter(col)].width = 14

# ===== Sheet 5: Investor Detail =====
ws5 = wb.create_sheet("Investor Detail")
ws5['A1'] = "투자자별 상세 분석 요약"
ws5['A1'].font = Font(bold=True, size=14)
ws5.merge_cells('A1:F1')

detail_data = [
    ("Warren Buffett", "가치 투자, 경제적 해자, 안전 마진 중심"),
    ("", "강점: 보험(ACGL, ALL, HIG, TRV), 헬스케어(MRK, JNJ)"),
    ("", "약점: 고성장 기술주(ENPH, DXCM) - 밸류에이션 우려"),
    ("", ""),
    ("Peter Lynch", "GARP(Growth at Reasonable Price), PEG 비율 중심"),
    ("", "강점: 반도체(MU, LRCX, KLAC), 기술(GOOGL, APH)"),
    ("", "약점: 가치함정 주의(IVZ, DVN) - 성장 둔화"),
    ("", ""),
    ("Phil Fisher", "장기 성장, R&D 투자, 경영진 품질 중심"),
    ("", "강점: 고품질 기술주(APH, ADI, PAYC)"),
    ("", "약점: 원자재/에너지(CF, NEM, DVN) - 혁신 부재"),
]

for row, (investor, detail) in enumerate(detail_data, 3):
    ws5.cell(row=row, column=1, value=investor).font = Font(bold=True) if investor else Font()
    ws5.cell(row=row, column=2, value=detail)

ws5.column_dimensions['A'].width = 18
ws5.column_dimensions['B'].width = 60

# ===== Sheet 6: Risk Analysis =====
ws6 = wb.create_sheet("Risk Analysis")
ws6['A1'] = "섹터 분포 분석"
ws6['A1'].font = Font(bold=True, size=14)

headers6 = ["Sector", "Count", "Percentage", "Max Allowed", "Status"]
for col, h in enumerate(headers6, 1):
    cell = ws6.cell(row=3, column=col, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = thin_border

sectors = {}
for d in analysis_data:
    sector = d[8]
    sectors[sector] = sectors.get(sector, 0) + 1

for row, (sector, count) in enumerate(sorted(sectors.items(), key=lambda x: -x[1]), 4):
    pct = count / 30 * 100
    ws6.cell(row=row, column=1, value=sector).border = thin_border
    ws6.cell(row=row, column=2, value=count).border = thin_border
    ws6.cell(row=row, column=3, value=f"{pct:.1f}%").border = thin_border
    ws6.cell(row=row, column=4, value="35%").border = thin_border
    status = "OK" if pct <= 35 else "OVER"
    cell = ws6.cell(row=row, column=5, value=status)
    cell.fill = BULLISH_FILL if status == "OK" else BEARISH_FILL
    cell.border = thin_border

ws6['A12'] = "리스크 지표"
ws6['A12'].font = Font(bold=True, size=14)
ws6['A13'] = "단일 종목 최대 비중:"
ws6['B13'] = "15%"
ws6['A14'] = "섹터 최대 비중:"
ws6['B14'] = "35%"
ws6['A15'] = "포트폴리오 집중도:"
ws6['B15'] = "=COUNTIF(Portfolio!E:E,\">10\")"

for col in range(1, 6):
    ws6.column_dimensions[get_column_letter(col)].width = 15

# 저장
output_path = "/home/ubuntu/projects/ai-hedge-fund/portfolios/sp500_portfolio_report_20260128.xlsx"
wb.save(output_path)
print(f"리포트 생성 완료: {output_path}")
