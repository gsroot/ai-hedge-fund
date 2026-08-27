#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path
from typing import Any

import yfinance as yf
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
BULLISH_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BEARISH_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
NEUTRAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

INVESTOR_CONFIG = {
    "buffett": {"display": "Warren Buffett", "short": "W.Buffett", "weight": 1.0},
    "munger": {"display": "Charlie Munger", "short": "C.Munger", "weight": 0.95},
    "damodaran": {"display": "Aswath Damodaran", "short": "A.Damodaran", "weight": 0.90},
    "lynch": {"display": "Peter Lynch", "short": "P.Lynch", "weight": 0.85},
    "graham": {"display": "Ben Graham", "short": "B.Graham", "weight": 0.85},
    "fisher": {"display": "Phil Fisher", "short": "P.Fisher", "weight": 0.82},
    "druckenmiller": {"display": "Stanley Druckenmiller", "short": "S.Druckenmiller", "weight": 0.80},
    "pabrai": {"display": "Mohnish Pabrai", "short": "M.Pabrai", "weight": 0.78},
    "burry": {"display": "Michael Burry", "short": "M.Burry", "weight": 0.75},
    "ackman": {"display": "Bill Ackman", "short": "B.Ackman", "weight": 0.75},
    "jhunjhunwala": {"display": "Rakesh Jhunjhunwala", "short": "R.Jhunjhunwala", "weight": 0.72},
    "wood": {"display": "Cathie Wood", "short": "C.Wood", "weight": 0.70},
}

INVESTOR_ALIASES = {
    "버핏": "buffett", "멍거": "munger", "다모다란": "damodaran",
    "린치": "lynch", "그레이엄": "graham", "피셔": "fisher",
    "드러켄밀러": "druckenmiller", "파브라이": "pabrai", "버리": "burry",
    "애크먼": "ackman", "준준왈라": "jhunjhunwala", "우드": "wood",
}

AGENT_TO_INVESTOR = {
    "warren-buffett-analyst": "buffett", "charlie-munger-analyst": "munger",
    "aswath-damodaran-analyst": "damodaran", "peter-lynch-analyst": "lynch",
    "ben-graham-analyst": "graham", "phil-fisher-analyst": "fisher",
    "stanley-druckenmiller-analyst": "druckenmiller", "mohnish-pabrai-analyst": "pabrai",
    "michael-burry-analyst": "burry", "bill-ackman-analyst": "ackman",
    "rakesh-jhunjhunwala-analyst": "jhunjhunwala", "cathie-wood-analyst": "wood",
}


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def fmt_pct(value: float | None, digits: int = 1, signed: bool = False) -> str:
    if value is None:
        return "N/A"
    if signed:
        return f"{value:+.{digits}f}%"
    return f"{value:.{digits}f}%"


def fmt_num(value: float | None, digits: int = 1) -> str:
    if value is None:
        return "N/A"
    return f"{value:.{digits}f}"


def ticker_label(stock: dict[str, Any], max_name: int = 18) -> str:
    name = (stock.get("company_name") or stock["ticker"]).strip()
    if name == stock["ticker"]:
        return stock["ticker"]
    if len(name) > max_name:
        name = name[: max_name - 1] + "…"
    return f"{stock['ticker']}({name})"


def signal_cell_text(signal: str, confidence: int) -> str:
    icon = {"bullish": "🟢", "neutral": "🔵", "bearish": "🔴"}[signal]
    return f"{icon} {signal}({confidence})"


def portfolio_signal_text(signal: str) -> str:
    return {"strong_buy": "🟢 강력매수", "buy": "🔵 매수"}[signal]


def load_predict_results(path: Path, top_n: int) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return payload, payload["rankings"][:top_n]


def load_investor_analyses(
    path: Path,
    expected_analysis_date: str,
) -> dict[str, dict[str, dict[str, Any]]]:
    """독립 investor-analysis 결과를 표준 ticker → investor 구조로 읽는다."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    raw = payload.get("analyses", payload)
    payload_date = payload.get("analysis_date") if isinstance(payload, dict) else None
    if payload_date and payload_date != expected_analysis_date:
        raise ValueError(
            f"investor-analysis 기준일({payload_date})이 predict 기준일({expected_analysis_date})과 다릅니다."
        )
    normalized: dict[str, dict[str, dict[str, Any]]] = {}
    for ticker, analyses in raw.items():
        normalized[ticker.upper()] = {}
        for investor, result in analyses.items():
            canonical = AGENT_TO_INVESTOR.get(investor, INVESTOR_ALIASES.get(investor, investor))
            result_date = result.get("analysis_date") or payload_date
            if result_date != expected_analysis_date:
                raise ValueError(
                    f"{ticker}/{canonical}: analysis_date가 없거나 predict 기준일과 다릅니다."
                )
            normalized[ticker.upper()][canonical] = result
    return normalized


def load_risk_snapshot(path: Path, expected_analysis_date: str) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    snapshot_date = payload.get("analysis_date")
    if snapshot_date != expected_analysis_date:
        raise ValueError(
            f"risk snapshot 기준일({snapshot_date})이 predict 기준일({expected_analysis_date})과 다릅니다."
        )
    if not isinstance(payload.get("annualized_volatility"), dict):
        raise ValueError("risk snapshot에 annualized_volatility가 없습니다.")
    if not isinstance(payload.get("correlation"), dict):
        raise ValueError("risk snapshot에 correlation이 없습니다.")
    market_regime = payload.get("market_regime")
    if not isinstance(market_regime, dict):
        raise ValueError("risk snapshot에 market_regime이 없습니다.")
    target_cash = market_regime.get("target_cash_weight")
    if not isinstance(target_cash, (int, float)) or not 0 <= target_cash <= 1:
        raise ValueError("market_regime.target_cash_weight는 0~1이어야 합니다.")
    return payload


def apply_risk_adjustment(
    candidates: list[dict[str, Any]],
    risk_snapshot: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """변동성과 후보 간 양의 상관을 원시 비중 점수에 반영한다."""
    volatilities = risk_snapshot["annualized_volatility"]
    correlations = risk_snapshot["correlation"]
    adjusted: list[dict[str, Any]] = []
    excluded: list[dict[str, Any]] = []
    volatility_eligible = []
    for item in candidates:
        ticker = item["ticker"]
        volatility = volatilities.get(ticker)
        if not isinstance(volatility, (int, float)) or volatility <= 0:
            item["exclusion_reason"] = "연환산 변동성 데이터 부족"
            excluded.append(item)
            continue
        volatility_eligible.append(item)

    candidate_tickers = [item["ticker"] for item in volatility_eligible]
    for item in volatility_eligible:
        ticker = item["ticker"]
        volatility = volatilities[ticker]
        peer_correlations = []
        missing_peer = None
        for peer in candidate_tickers:
            if peer == ticker:
                continue
            value = correlations.get(ticker, {}).get(peer)
            if not isinstance(value, (int, float)):
                missing_peer = peer
                break
            peer_correlations.append(clamp(float(value), -1.0, 1.0))
        if missing_peer:
            item["exclusion_reason"] = f"상관 데이터 부족 ({missing_peer})"
            excluded.append(item)
            continue
        avg_positive_correlation = (
            sum(max(value, 0.0) for value in peer_correlations) / len(peer_correlations)
            if peer_correlations else 0.0
        )
        diversification_multiplier = 1.0 - 0.5 * avg_positive_correlation
        item["annualized_volatility"] = float(volatility)
        item["avg_positive_correlation"] = avg_positive_correlation
        item["raw_weight_score"] = (
            item["raw_weight_score"]
            / max(float(volatility), 0.10)
            * diversification_multiplier
        )
        adjusted.append(item)
    return adjusted, excluded


def score_implied_return(stock: dict[str, Any]) -> float:
    """검증된 기대수익률이 아닌 predict 점수 환산값을 읽는다."""
    value = stock.get("score_implied_return_pct")
    if value is None:
        raise ValueError(f"{stock.get('ticker', 'unknown')}: score_implied_return_pct가 없습니다.")
    return float(value or 0)


def normalize_sector(sector: str | None) -> str:
    if not sector:
        return "Unknown"
    mapping = {
        "Financial Services": "Financial",
        "Consumer Cyclical": "Consumer Disc.",
        "Consumer Defensive": "Consumer Staples",
        "Basic Materials": "Materials",
        "Communication Services": "Communication",
        "Healthcare": "Healthcare",
        "Technology": "Technology",
        "Energy": "Energy",
        "Industrials": "Industrials",
        "Real Estate": "Real Estate",
        "Utilities": "Utilities",
    }
    return mapping.get(sector, sector)


def fetch_sector(ticker: str) -> str:
    try:
        info = yf.Ticker(ticker).info
        return normalize_sector(info.get("sector"))
    except Exception:
        return "Unknown"


def fetch_sectors(stocks: list[dict[str, Any]], analysis_date: str) -> dict[str, str]:
    sectors = {
        stock["ticker"]: normalize_sector(stock.get("metrics", {}).get("sector"))
        for stock in stocks
        if stock.get("metrics", {}).get("sector")
    }
    # 과거 리포트에 현재 Yahoo 섹터를 섞지 않는다.
    if analysis_date < datetime.now().strftime("%Y-%m-%d"):
        return {stock["ticker"]: sectors.get(stock["ticker"], "Unknown") for stock in stocks}
    missing = [stock for stock in stocks if stock["ticker"] not in sectors]
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(fetch_sector, stock["ticker"]): stock["ticker"] for stock in missing}
        for future, ticker in futures.items():
            try:
                sectors[ticker] = future.result()
            except Exception:
                sectors[ticker] = "Unknown"
    return sectors


def combined_confidence(analyses: dict[str, dict[str, Any]], investors: list[str]) -> int:
    total_weight = sum(INVESTOR_CONFIG[investor]["weight"] for investor in investors)
    weighted = sum(analyses[investor]["confidence"] * INVESTOR_CONFIG[investor]["weight"] for investor in investors)
    return int(round(weighted / total_weight)) if total_weight else 0


def majority_threshold(count: int) -> int:
    return math.floor(count / 2) + 1


def build_candidates(
    stocks: list[dict[str, Any]],
    investors: list[str],
    sectors: dict[str, str],
    independent_analyses: dict[str, dict[str, dict[str, Any]]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    candidates = []
    excluded = []
    need = majority_threshold(len(investors))
    for stock in stocks:
        ticker_analyses = independent_analyses.get(stock["ticker"].upper(), {})
        missing = [investor for investor in investors if investor not in ticker_analyses]
        if missing:
            raise ValueError(f"{stock['ticker']}: 독립 투자자 분석 누락 - {', '.join(missing)}")
        analyses = {investor: ticker_analyses[investor] for investor in investors}
        insufficient_investors = []
        for investor, analysis in analyses.items():
            signal = analysis.get("signal")
            confidence_value = analysis.get("confidence")
            if signal not in {"bullish", "neutral", "bearish"}:
                raise ValueError(f"{stock['ticker']}/{investor}: 잘못된 signal {signal!r}")
            if not isinstance(confidence_value, (int, float)) or not 0 <= confidence_value <= 100:
                raise ValueError(f"{stock['ticker']}/{investor}: confidence는 0~100이어야 합니다.")
            if not isinstance(analysis.get("reasoning"), str) or not analysis["reasoning"].strip():
                raise ValueError(f"{stock['ticker']}/{investor}: reasoning이 비어 있습니다.")
            data_quality = analysis.get("data_quality")
            if data_quality not in {"complete", "partial", "insufficient"}:
                raise ValueError(
                    f"{stock['ticker']}/{investor}: data_quality는 complete/partial/insufficient 중 하나여야 합니다."
                )
            if data_quality == "insufficient":
                insufficient_investors.append(investor)
        bullish = sum(1 for investor in investors if analyses[investor]["signal"] == "bullish")
        bearish = sum(1 for investor in investors if analyses[investor]["signal"] == "bearish")
        confidence = combined_confidence(analyses, investors)
        consensus_ratio = bullish / len(investors)
        portfolio_signal = "strong_buy" if bullish == len(investors) else "buy"
        enriched = {
            **stock,
            "sector": sectors.get(stock["ticker"], "Unknown"),
            "investor_analysis": analyses,
            "bullish_count": bullish,
            "bearish_count": bearish,
            "consensus_ratio": consensus_ratio,
            "combined_confidence": confidence,
            "combined_signal": portfolio_signal,
            "analysis_source": "independent_investor_analysis",
            "raw_weight_score": max(float(stock.get("total_score", 0)), 0.0) * (confidence / 100) * max(consensus_ratio, 0.01),
        }
        if insufficient_investors:
            excluded.append({
                **enriched,
                "exclusion_reason": f"투자자 분석 데이터 부족 ({', '.join(insufficient_investors)})",
            })
        elif bullish >= need:
            candidates.append(enriched)
        else:
            excluded.append({**enriched, "exclusion_reason": f"투자자 과반 미달 ({bullish}/{len(investors)} bullish)"})
    return candidates, excluded


def allocate_weights(
    candidates: list[dict[str, Any]],
    name_cap: float = 15.0,
    sector_cap: float = 35.0,
    min_weight: float = 2.0,
    target_cash_weight: float = 0.0,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], float]:
    if not 0.0 <= target_cash_weight <= 100.0:
        raise ValueError("target_cash_weight는 0~100이어야 합니다.")
    if not candidates:
        return [], [], 100.0

    def constrained_allocation(active_candidates: list[dict[str, Any]]) -> dict[str, float]:
        allocations = {item["ticker"]: 0.0 for item in active_candidates}
        active = {item["ticker"] for item in active_candidates if item["raw_weight_score"] > 0}
        meta = {item["ticker"]: item for item in active_candidates}
        remaining = 100.0 - target_cash_weight
        while active and remaining > 1e-9:
            sector_used: dict[str, float] = defaultdict(float)
            for ticker, weight in allocations.items():
                sector_used[meta[ticker]["sector"]] += weight
            active = {
                ticker for ticker in active
                if allocations[ticker] < name_cap - 1e-9
                and sector_used[meta[ticker]["sector"]] < sector_cap - 1e-9
            }
            if not active:
                break
            raw_total = sum(meta[ticker]["raw_weight_score"] for ticker in active)
            if raw_total <= 0:
                break
            proposed = {
                ticker: remaining * meta[ticker]["raw_weight_score"] / raw_total
                for ticker in active
            }
            fraction = 1.0
            for ticker, amount in proposed.items():
                if amount > 0:
                    fraction = min(fraction, (name_cap - allocations[ticker]) / amount)
            by_sector: dict[str, float] = defaultdict(float)
            for ticker, amount in proposed.items():
                by_sector[meta[ticker]["sector"]] += amount
            for sector, amount in by_sector.items():
                if amount > 0:
                    fraction = min(fraction, (sector_cap - sector_used[sector]) / amount)
            fraction = clamp(fraction, 0.0, 1.0)
            added = 0.0
            for ticker, amount in proposed.items():
                increment = amount * fraction
                allocations[ticker] += increment
                added += increment
            if added <= 1e-9:
                break
            remaining -= added
            if fraction >= 1.0 - 1e-9:
                break
        return allocations

    eligible = list(candidates)
    excluded: list[dict[str, Any]] = []
    while eligible:
        allocations = constrained_allocation(eligible)
        below = [item for item in eligible if allocations.get(item["ticker"], 0.0) < min_weight - 1e-9]
        if not below:
            break
        for item in below:
            item["exclusion_reason"] = f"최소 비중 {min_weight:g}% 미달"
            excluded.append(item)
        eligible = [item for item in eligible if item not in below]
    else:
        return [], excluded, 100.0

    for item in eligible:
        # 내림 반올림으로 이름/섹터 상한을 절대 넘지 않게 한다.
        item["weight"] = math.floor(allocations[item["ticker"]] * 10 + 1e-9) / 10
    included = [item for item in eligible if item["weight"] >= min_weight]
    included.sort(key=lambda item: (-item["weight"], item["rank"]))
    cash_weight = round(
        max(target_cash_weight, 100.0 - sum(item["weight"] for item in included)),
        1,
    )
    return included, excluded, cash_weight


def summarize_portfolio(
    included: list[dict[str, Any]],
    top_stocks: list[dict[str, Any]],
    investors: list[str],
    cash_weight: float,
    market_regime: dict[str, Any] | None = None,
) -> dict[str, Any]:
    total_names = len(top_stocks)
    included_tickers = {item["ticker"] for item in included}
    invested_weight = sum(item["weight"] for item in included)
    weighted_conf = sum(item["weight"] * item["combined_confidence"] for item in included) / invested_weight if invested_weight else 0.0
    weighted_return = sum(item["weight"] * score_implied_return(item) for item in included) / 100 if included else 0.0
    strong_buy_weight = sum(item["weight"] for item in included if item["combined_signal"] == "strong_buy")
    buy_weight = sum(item["weight"] for item in included if item["combined_signal"] == "buy")

    agreement = Counter(item["bullish_count"] for item in included)
    cap_weights: dict[str, float] = defaultdict(float)
    sector_weights: dict[str, float] = defaultdict(float)
    for item in included:
        cap_weights[item["market_cap"]["category"]] += item["weight"]
        sector_weights[item["sector"]] += item["weight"]

    return {
        "included_count": len(included),
        "analyzed_count": total_names,
        "included_rate": (len(included) / total_names * 100) if total_names else 0.0,
        "avg_confidence": weighted_conf,
        "score_implied_return_contribution": weighted_return,
        "invested_weight": invested_weight,
        "cash_weight": cash_weight,
        "market_cash_target": (
            float(market_regime["target_cash_weight"]) * 100.0
            if market_regime else 0.0
        ),
        "market_regime": market_regime,
        "strong_buy_weight": strong_buy_weight,
        "buy_weight": buy_weight,
        "agreement": agreement,
        "cap_weights": cap_weights,
        "sector_weights": dict(sorted(sector_weights.items(), key=lambda item: item[1], reverse=True)),
        "top3_weight": sum(item["weight"] for item in included[:3]),
        "top5_weight": sum(item["weight"] for item in included[:5]),
        "max_name": included[0]["ticker"] if included else "N/A",
        "max_weight": included[0]["weight"] if included else 0.0,
        "max_sector": max(sector_weights.items(), key=lambda item: item[1]) if sector_weights else ("N/A", 0.0),
        "included_tickers": included_tickers,
        "need_consensus": majority_threshold(len(investors)),
    }


def write_portfolio_json(
    output_path: Path,
    analysis_date: str,
    included: list[dict[str, Any]],
    cash_weight: float,
    market_regime: dict[str, Any] | None = None,
) -> None:
    payload = {
        "analysis_date": analysis_date,
        "weights": {item["ticker"]: round(item["weight"] / 100, 6) for item in included},
        "cash_weight": round(cash_weight / 100, 6),
        "methodology": "predict_consensus_risk_adjusted_caps_dynamic_market_cash",
        "market_regime": market_regime,
        "constraints": {
            "max_name": 0.15,
            "max_sector": 0.35,
            "min_name": 0.02,
            "market_cash_target": (
                float(market_regime["target_cash_weight"])
                if market_regime else 0.0
            ),
        },
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 30)


def apply_header(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def apply_grid(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER


def create_summary_sheet(ws, analysis_date: str, index_label: str, investor_labels: str, summary: dict[str, Any]) -> None:
    ws.merge_cells("A1:B1")
    ws["A1"] = "AI Hedge Fund 포트폴리오 리포트"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    info = [
        ("분석 일자", analysis_date),
        ("분석 대상", index_label),
        ("분석 전략", "하이브리드 (펀더멘털 70% + 모멘텀 30%)"),
        ("투자자 관점", investor_labels),
    ]
    for row, (label, value) in enumerate(info, start=3):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value

    ws["A8"] = "포트폴리오 통계"
    ws["A8"].font = Font(bold=True)
    stats = [
        ("편입 종목 수", f"{summary['included_count']} / {summary['analyzed_count']} ({summary['included_rate']:.1f}%)"),
        ("평균 신뢰도", f"{summary['avg_confidence']:.0f}%"),
        ("점수 환산 기여값", fmt_pct(summary["score_implied_return_contribution"], signed=True)),
        ("강력매수 비중", fmt_pct(summary["strong_buy_weight"])),
        ("현금 비중", fmt_pct(summary["cash_weight"])),
        (
            "시장 국면",
            summary["market_regime"]["regime"]
            if summary.get("market_regime") else "N/A",
        ),
        ("시장 목표 현금", fmt_pct(summary.get("market_cash_target", 0.0))),
    ]
    stats_start = 9
    for row, (label, value) in enumerate(stats, start=stats_start):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value

    cap_header_row = stats_start + len(stats) + 1
    ws[f"A{cap_header_row}"] = "시가총액 분포"
    ws[f"A{cap_header_row}"].font = Font(bold=True)
    cap_labels = {
        "mega": "메가캡 (>$200B)",
        "large": "대형주 ($10B-$200B)",
        "mid": "중형주 ($2B-$10B)",
        "small": "소형주 (<$2B)",
        None: "Unknown",
    }
    chart_start = cap_header_row + 1
    for idx, (category, label) in enumerate(cap_labels.items(), start=chart_start):
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = round(summary["cap_weights"].get(category, 0.0) / 100, 4)
        ws[f"B{idx}"].number_format = "0.0%"

    sector_header_row = chart_start + len(cap_labels) + 1
    ws[f"A{sector_header_row}"] = "섹터 분포"
    ws[f"A{sector_header_row}"].font = Font(bold=True)
    sector_start = sector_header_row + 1
    for idx, (sector, weight) in enumerate(summary["sector_weights"].items(), start=sector_start):
        ws[f"A{idx}"] = sector
        ws[f"B{idx}"] = round(weight / 100, 4)
        ws[f"B{idx}"].number_format = "0.0%"

    pie = PieChart()
    pie.title = "섹터 분포"
    data = Reference(ws, min_col=2, min_row=sector_start, max_row=sector_start + max(len(summary["sector_weights"]) - 1, 0))
    labels = Reference(ws, min_col=1, min_row=sector_start, max_row=sector_start + max(len(summary["sector_weights"]) - 1, 0))
    if summary["sector_weights"]:
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(labels)
        pie.height = 7
        pie.width = 10
        ws.add_chart(pie, f"D{cap_header_row}")

    apply_grid(ws)
    auto_width(ws)


def create_portfolio_sheet(ws, included: list[dict[str, Any]], investors: list[str]) -> None:
    headers = ["#", "종목코드", "회사명", "비중", "신호", "신뢰도", "점수 환산값", "시가총액", "P/E", "ROE", "PEG", "합의", "섹터"]
    apply_header(ws, headers)
    for row, item in enumerate(included, start=2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=item["ticker"])
        ws.cell(row=row, column=3, value=item.get("company_name"))
        ws.cell(row=row, column=4, value=item["weight"] / 100).number_format = "0.0%"
        ws.cell(row=row, column=5, value=portfolio_signal_text(item["combined_signal"]))
        ws.cell(row=row, column=6, value=item["combined_confidence"])
        ws.cell(row=row, column=7, value=score_implied_return(item) / 100).number_format = "+0.0%;-0.0%"
        ws.cell(row=row, column=8, value=item["market_cap"]["display"])
        ws.cell(row=row, column=9, value=item["metrics"].get("pe"))
        roe = item["metrics"].get("roe")
        if roe is not None:
            ws.cell(row=row, column=10, value=roe / 100).number_format = "0.0%"
        else:
            ws.cell(row=row, column=10, value="N/A")
        ws.cell(row=row, column=11, value=item["metrics"].get("peg"))
        ws.cell(row=row, column=12, value=f"{item['bullish_count']}/{len(investors)}")
        ws.cell(row=row, column=13, value=item["sector"])

    cash_row = len(included) + 2
    cash_weight = max(0.0, 100.0 - sum(item["weight"] for item in included))
    ws.cell(row=cash_row, column=2, value="CASH")
    ws.cell(row=cash_row, column=3, value="현금")
    ws.cell(row=cash_row, column=4, value=cash_weight / 100).number_format = "0.0%"
    ws.cell(row=cash_row, column=5, value="cash")

    total_row = cash_row + 1
    for col in range(1, 14):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
    ws.cell(row=total_row, column=2, value="합계/평균").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=1.0).number_format = "0.0%"
    invested_weight = sum(item["weight"] for item in included)
    avg_conf = sum(item["weight"] * item["combined_confidence"] for item in included) / invested_weight if invested_weight else 0
    avg_ret = sum(item["weight"] * score_implied_return(item) for item in included) / 100 if included else 0
    ws.cell(row=total_row, column=6, value=round(avg_conf))
    ws.cell(row=total_row, column=7, value=avg_ret / 100).number_format = "+0.0%;-0.0%"

    apply_grid(ws)
    auto_width(ws)
    ws.auto_filter.ref = f"A1:M{len(included) + 1}"
    ws.freeze_panes = "A2"


def create_ranking_sheet(ws, top_stocks: list[dict[str, Any]], included_tickers: set[str]) -> None:
    headers = ["순위", "종목코드", "회사명", "종합점수", "펀더멘털", "모멘텀", "앙상블", "신호", "점수 환산값", "시가총액", "P/E", "P/B", "ROE", "매출성장률", "PEG", "편입여부"]
    apply_header(ws, headers)
    for row, item in enumerate(top_stocks, start=2):
        metrics = item.get("metrics", {})
        ws.cell(row=row, column=1, value=item["rank"])
        ws.cell(row=row, column=2, value=item["ticker"])
        ws.cell(row=row, column=3, value=item.get("company_name"))
        ws.cell(row=row, column=4, value=item["total_score"])
        ws.cell(row=row, column=5, value=item["scores"].get("fundamental"))
        ws.cell(row=row, column=6, value=item["scores"].get("enhanced_momentum"))
        ws.cell(row=row, column=7, value=item.get("ensemble_score"))
        ws.cell(row=row, column=8, value=item["signal"])
        ws.cell(row=row, column=9, value=score_implied_return(item) / 100).number_format = "+0.0%;-0.0%"
        ws.cell(row=row, column=10, value=item["market_cap"]["display"])
        ws.cell(row=row, column=11, value=metrics.get("pe"))
        ws.cell(row=row, column=12, value=metrics.get("pb"))
        roe = metrics.get("roe")
        if roe is not None:
            ws.cell(row=row, column=13, value=roe / 100).number_format = "0.0%"
        else:
            ws.cell(row=row, column=13, value="N/A")
        rev = metrics.get("revenue_growth")
        if rev is not None:
            ws.cell(row=row, column=14, value=rev / 100).number_format = "0.0%"
        else:
            ws.cell(row=row, column=14, value="N/A")
        ws.cell(row=row, column=15, value=metrics.get("peg"))
        included = item["ticker"] in included_tickers
        included_cell = ws.cell(row=row, column=16, value="예" if included else "아니오")
        if included:
            for col in range(1, 17):
                ws.cell(row=row, column=col).fill = BULLISH_FILL
        included_cell.border = THIN_BORDER

    apply_grid(ws)
    auto_width(ws)
    ws.auto_filter.ref = f"A1:P{len(top_stocks) + 1}"
    ws.freeze_panes = "A2"


def create_matrix_sheet(ws, top_stocks: list[dict[str, Any]], investors: list[str]) -> None:
    headers = ["종목코드", "회사명"] + [INVESTOR_CONFIG[investor]["display"] for investor in investors] + ["종합신호", "종합신뢰도"]
    apply_header(ws, headers)
    for row, item in enumerate(top_stocks, start=2):
        ws.cell(row=row, column=1, value=item["ticker"])
        ws.cell(row=row, column=2, value=item.get("company_name"))
        for idx, investor in enumerate(investors, start=3):
            analysis = item["investor_analysis"][investor]
            cell = ws.cell(row=row, column=idx, value={"bullish": "매수", "neutral": "중립", "bearish": "매도"}[analysis["signal"]] + f"({analysis['confidence']})")
            cell.fill = {"bullish": BULLISH_FILL, "neutral": NEUTRAL_FILL, "bearish": BEARISH_FILL}[analysis["signal"]]
        ws.cell(row=row, column=3 + len(investors), value=item["combined_signal"] if item.get("combined_signal") else "-")
        ws.cell(row=row, column=4 + len(investors), value=item.get("combined_confidence", "-"))

    apply_grid(ws)
    auto_width(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(top_stocks) + 1}"
    ws.freeze_panes = "A2"


def create_detail_sheet(ws, top_stocks: list[dict[str, Any]], investors: list[str]) -> None:
    headers = ["종목코드", "회사명", "투자자", "신호", "신뢰도", "분석근거"]
    apply_header(ws, headers)
    row = 2
    for item in top_stocks:
        start_row = row
        for investor in investors:
            analysis = item["investor_analysis"][investor]
            ws.cell(row=row, column=1, value=item["ticker"])
            ws.cell(row=row, column=2, value=item.get("company_name"))
            ws.cell(row=row, column=3, value=INVESTOR_CONFIG[investor]["display"])
            ws.cell(row=row, column=4, value=analysis["signal"])
            ws.cell(row=row, column=5, value=analysis["confidence"])
            ws.cell(row=row, column=6, value=analysis["reasoning"])
            row += 1
        for col in range(1, 7):
            ws.cell(row=start_row, column=col).border = THIN_BORDER
    apply_grid(ws)
    auto_width(ws)
    ws.freeze_panes = "A2"


def create_risk_sheet(ws, summary: dict[str, Any], excluded_all: list[dict[str, Any]], investors: list[str]) -> None:
    ws["A1"] = "리스크 분석"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:B1")

    rows = [
        ("상위 1종목 비중", fmt_pct(summary["max_weight"])),
        ("상위 3종목 비중", fmt_pct(summary["top3_weight"])),
        ("상위 5종목 비중", fmt_pct(summary["top5_weight"])),
        ("최대 섹터", f"{summary['max_sector'][0]} ({summary['max_sector'][1]:.1f}%)"),
        ("만장일치 비율", fmt_pct(summary["agreement"].get(len(investors), 0) / max(summary["included_count"], 1) * 100 if summary["included_count"] else 0)),
        ("비편입 종목 수", f"{sum(1 for item in excluded_all if item.get('exclusion_reason'))}개"),
    ]
    regime = summary.get("market_regime")
    if regime:
        scores = regime.get("scores", {})
        rows.extend(
            [
                ("시장 국면", regime["regime"]),
                ("시장 목표 현금", fmt_pct(summary["market_cash_target"])),
                ("실제 현금", fmt_pct(summary["cash_weight"])),
                ("과열 점수", fmt_num(scores.get("overheat"), 2)),
                ("공포 점수", fmt_num(scores.get("fear"), 2)),
                ("전망 점수", fmt_num(scores.get("outlook"), 2)),
            ]
        )
    for row, (label, value) in enumerate(rows, start=4):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
    apply_grid(ws)
    auto_width(ws)


def write_workbook(
    output_path: Path,
    analysis_date: str,
    index_label: str,
    investors: list[str],
    top_stocks: list[dict[str, Any]],
    included: list[dict[str, Any]],
    summary: dict[str, Any],
    excluded_all: list[dict[str, Any]],
) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "요약"
    create_summary_sheet(ws_summary, analysis_date, index_label, ", ".join(INVESTOR_CONFIG[investor]["short"] for investor in investors), summary)

    ws_portfolio = wb.create_sheet("포트폴리오")
    create_portfolio_sheet(ws_portfolio, included, investors)

    ws_ranking = wb.create_sheet("순위")
    create_ranking_sheet(ws_ranking, top_stocks, summary["included_tickers"])

    ws_matrix = wb.create_sheet("투자자 매트릭스")
    create_matrix_sheet(ws_matrix, top_stocks, investors)

    ws_detail = wb.create_sheet("투자자 상세")
    create_detail_sheet(ws_detail, top_stocks, investors)

    ws_risk = wb.create_sheet("리스크 분석")
    create_risk_sheet(ws_risk, summary, excluded_all, investors)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def print_report(
    analysis_date: str,
    index_name: str,
    investors: list[str],
    top_stocks: list[dict[str, Any]],
    included: list[dict[str, Any]],
    excluded_all: list[dict[str, Any]],
    summary: dict[str, Any],
    output_path: Path | None,
) -> None:
    investor_names = ", ".join(INVESTOR_CONFIG[investor]["display"] for investor in investors)
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("📋 AI Hedge Fund 포트폴리오 리포트")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print(f"분석 일자    : {analysis_date}")
    print(f"분석 대상    : {index_name} 상위 {len(top_stocks)}개 종목")
    print("분석 전략    : 하이브리드 (펀더멘털 70% + 모멘텀 30%)")
    print(f"투자자 관점  : {investor_names}")
    print("데이터 소스  : predict + 독립 investor-analysis + 가격·시장국면 risk snapshot")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print()
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("📊 포트폴리오 구성")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print(" #  종목                  비중    시총      종합신호     신뢰도  점수환산값  P/E    ROE      합의")
    print("──────────────────────────────────────────────────────────────────────────────────────────────")
    for idx, item in enumerate(included, start=1):
        print(
            f"{idx:<2}  {ticker_label(item, 16):<20} {item['weight']:>5.1f}%  {item['market_cap']['display']:<8} "
            f"{portfolio_signal_text(item['combined_signal']):<11} {item['combined_confidence']:>4}%  "
            f"{fmt_pct(score_implied_return(item), signed=True):>9}  {fmt_num(item['metrics'].get('pe')):>5}  "
            f"{fmt_pct(item['metrics'].get('roe')):>7}  {item['bullish_count']}/{len(investors)}"
        )
    print("──────────────────────────────────────────────────────────────────────────────────────────────")
    print(f"    주식 합계            {summary['invested_weight']:>5.1f}%                           avg {summary['avg_confidence']:.0f}%")
    print(f"    현금                  {summary['cash_weight']:>5.1f}%                           점수환산 기여 {fmt_pct(summary['score_implied_return_contribution'], signed=True)}")
    print()
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("👥 투자자별 종목 분석 매트릭스")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    headers = "  ".join(f"{INVESTOR_CONFIG[investor]['short']:<18}" for investor in investors)
    print(f"종목                  {headers}")
    print("──────────────────────────────────────────────────────────────────────────────────────────────")
    for item in top_stocks:
        cells = "  ".join(f"{signal_cell_text(item['investor_analysis'][investor]['signal'], item['investor_analysis'][investor]['confidence']):<18}" for investor in investors)
        print(f"{ticker_label(item, 16):<20} {cells}")
    print()
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("💬 투자자별 핵심 분석 근거 (상위 5개 종목)")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    for item in included[:5]:
        print(f"▸ {item['ticker']} — 종합: {portfolio_signal_text(item['combined_signal'])} (신뢰도 {item['combined_confidence']}%)")
        for investor in investors:
            analysis = item["investor_analysis"][investor]
            icon = {"bullish": "🟢", "neutral": "🔵", "bearish": "🔴"}[analysis["signal"]]
            print(f"  {INVESTOR_CONFIG[investor]['display']:<16}: {icon} ({analysis['confidence']}%) {analysis['reasoning']}")
        print()
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("📈 포트폴리오 요약")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print(f"편입 종목 수         : {summary['included_count']}개 / 분석 {summary['analyzed_count']}개 ({summary['included_rate']:.1f}% 편입률)")
    print(f"평균 신뢰도          : {summary['avg_confidence']:.0f}%")
    print(f"점수 환산 기여값     : {fmt_pct(summary['score_implied_return_contribution'], signed=True)} (예상수익률 아님)")
    if summary.get("market_regime"):
        regime = summary["market_regime"]
        print(f"시장 국면            : {regime['regime']} ({regime['benchmark']}, {regime['as_of_date']})")
        print(f"시장 목표 현금       : {fmt_pct(summary['market_cash_target'])}")
    print(f"현금 비중            : {fmt_pct(summary['cash_weight'])}")
    print(f"강력매수 비중        : {fmt_pct(summary['strong_buy_weight'])}")
    print(f"매수 비중            : {fmt_pct(summary['buy_weight'])}")
    unanimous = summary["agreement"].get(len(investors), 0)
    majority = summary["agreement"].get(summary["need_consensus"], 0) - unanimous if len(investors) != summary["need_consensus"] else 0
    print()
    print("투자자 합의 분포:")
    print(f"  만장일치 ({len(investors)}/{len(investors)})     : {unanimous}개")
    print(f"  다수 합의 ({summary['need_consensus']}/{len(investors)})    : {summary['included_count'] - unanimous}개")
    print()
    print("시가총액 분포:")
    cap_labels = {"mega": "메가캡", "large": "대형주", "mid": "중형주", "small": "소형주", None: "Unknown"}
    for key in ["mega", "large", "mid", "small", None]:
        weight = summary["cap_weights"].get(key, 0.0)
        if weight > 0:
            print(f"  {cap_labels[key]:<18}: {weight:>4.1f}%")
    print()
    print("섹터 분포:")
    for sector, weight in list(summary["sector_weights"].items())[:6]:
        print(f"  {sector:<18}: {weight:>4.1f}%")
    print()
    non_included = [item for item in excluded_all if item.get("exclusion_reason")]
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print(f"🚫 비편입 종목 ({len(non_included)}개)")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("종목                  순위   점수    사유")
    print("──────────────────────────────────────────────────────────────────────────────────────────────")
    for item in non_included[:12]:
        print(f"{ticker_label(item, 16):<20} {item['rank']:>3}   {item['total_score']:>5.2f}   {item['exclusion_reason']}")
    print()
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("⚠️ 리스크 및 경고")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("포트폴리오 집중도:")
    print(f"  상위 5종목 비중     : {summary['top5_weight']:.1f}%")
    print(f"  최대 단일 종목      : {summary['max_name']} {summary['max_weight']:.1f}% (제한 15% 내)")
    print(f"  최대 섹터 비중      : {summary['max_sector'][0]} {summary['max_sector'][1]:.1f}% (제한 35% 내)")
    print()
    warnings = []
    for item in top_stocks:
        for warning in item.get("investor_warnings", []):
            warnings.append((item["ticker"], warning.replace("⚠️ ", "")))
    if warnings:
        print("투자 철학 불일치 경고:")
        for ticker, warning in warnings[:6]:
            print(f"  {ticker:<6}: {warning}")
    print()
    low_consensus = [item for item in top_stocks if item.get("investor_consensus", {}).get("level") == "low"]
    if low_consensus:
        print("투자자 의견 분산 종목:")
        for item in low_consensus[:5]:
            print(f"  {item['ticker']:<6}: predict 합의도 {item['investor_consensus']['level']} (std={item['investor_consensus']['std']})")
    print()
    print("══════════════════════════════════════════════════════════════════════════════════════════════")
    print("💡 이 리포트는 교육/연구 목적이며 실제 투자 결정의 근거가 될 수 없습니다.")
    print("   predict: 다중 팩터 순위 | investor-analysis: 독립 입력 결과")
    if output_path:
        print(f"   엑셀 리포트: {output_path}")
    print("══════════════════════════════════════════════════════════════════════════════════════════════")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--predict-json", required=True)
    parser.add_argument("--investor-json", required=True, help="종목별 독립 investor-analysis 결과 JSON")
    parser.add_argument("--risk-json", required=True, help="동일 기준일의 변동성·상관 risk snapshot JSON")
    parser.add_argument("--index", default="sp500")
    parser.add_argument("--top", type=int, default=30)
    parser.add_argument("--investors", default="buffett,lynch,fisher")
    parser.add_argument("--xlsx", default="yes")
    parser.add_argument("--output-dir", default="portfolios")
    parser.add_argument("--portfolio-json", help="백테스트용 목표비중 JSON 저장 경로")
    args = parser.parse_args()

    requested = [token.strip() for token in args.investors.split(",") if token.strip()]
    if not requested:
        parser.error("--investors에는 최소 한 명을 지정해야 합니다.")
    if args.top <= 0:
        parser.error("--top은 1 이상이어야 합니다.")
    if requested == ["all"]:
        investors = list(INVESTOR_CONFIG)
    else:
        investors = [INVESTOR_ALIASES.get(token, token) for token in requested]
    unknown = [investor for investor in investors if investor not in INVESTOR_CONFIG]
    if unknown:
        parser.error(f"알 수 없는 투자자: {', '.join(unknown)}")
    payload, top_stocks = load_predict_results(Path(args.predict_json), args.top)
    analysis_date = payload.get("analysis_date")
    if not analysis_date:
        parser.error("predict JSON에 analysis_date가 없습니다.")
    sectors = fetch_sectors(top_stocks, analysis_date)
    independent_analyses = load_investor_analyses(Path(args.investor_json), analysis_date)

    candidates, excluded_majority = build_candidates(top_stocks, investors, sectors, independent_analyses)
    risk_snapshot = load_risk_snapshot(Path(args.risk_json), analysis_date)
    candidates, excluded_risk = apply_risk_adjustment(candidates, risk_snapshot)
    market_regime = risk_snapshot["market_regime"]
    included, excluded_weight, cash_weight = allocate_weights(
        candidates,
        target_cash_weight=float(market_regime["target_cash_weight"]) * 100.0,
    )
    included_map = {item["ticker"]: item for item in included}

    enriched_top = []
    for stock in top_stocks:
        base = next(
            candidate for candidate in candidates + excluded_risk + excluded_majority
            if candidate["ticker"] == stock["ticker"]
        )
        if stock["ticker"] in included_map:
            base = included_map[stock["ticker"]]
        enriched_top.append(base)

    summary = summarize_portfolio(
        included, top_stocks, investors, cash_weight, market_regime
    )
    if args.portfolio_json:
        write_portfolio_json(
            Path(args.portfolio_json),
            analysis_date,
            included,
            cash_weight,
            market_regime,
        )
    output_path = None
    if args.xlsx.lower() in {"yes", "true", "1", "excel", "xlsx"}:
        investor_suffix = "_".join(sorted(investors))
        output_path = Path(args.output_dir) / f"{args.index}_{analysis_date.replace('-', '')}_{investor_suffix}.xlsx"
        write_workbook(
            output_path=output_path,
            analysis_date=analysis_date,
            index_label=f"{args.index.upper()} 상위 {args.top}개" if args.index != "sp500" else f"S&P 500 상위 {args.top}개",
            investors=investors,
            top_stocks=enriched_top,
            included=included,
            summary=summary,
            excluded_all=excluded_majority + excluded_risk + excluded_weight,
        )

    print_report(
        analysis_date=analysis_date,
        index_name="S&P 500" if args.index == "sp500" else args.index.upper(),
        investors=investors,
        top_stocks=enriched_top,
        included=included,
        excluded_all=excluded_majority + excluded_risk + excluded_weight,
        summary=summary,
        output_path=output_path,
    )
    if args.portfolio_json:
        print(f"백테스트용 목표비중 JSON: {args.portfolio_json}")


if __name__ == "__main__":
    main()
